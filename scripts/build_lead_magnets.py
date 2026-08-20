"""Build, upload, and verify every CapVeri lead-magnet resource."""

# ruff: noqa: E501

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from PyPDF2 import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import (
    Flowable,
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

ROOT = Path(__file__).resolve().parents[1]
GENERATED_DIR = ROOT / "generated" / "lead-magnets"
DOCS_ASSETS_DIR = ROOT / "docs" / "assets"
R2_BUCKET = "capveri-lead-magnets"
R2_OBJECT_PREFIX = "lead-magnets/2026-06-25/"
MANIFEST_PATH = GENERATED_DIR / "manifest.json"
PUBLIC_KNOWLEDGE_PATH = ROOT / "knowledge" / "generated" / "public-knowledge.json"
PUBLIC_KNOWLEDGE = json.loads(PUBLIC_KNOWLEDGE_PATH.read_text(encoding="utf-8"))
APP_BASE_URL = PUBLIC_KNOWLEDGE["company"]["appUrl"].rstrip("/")
REGISTER_URL = f"{APP_BASE_URL}/auth/register"
FIXED_BUILD_TIME = datetime(2000, 1, 1, tzinfo=timezone.utc)
FIXED_ZIP_TIME = (2000, 1, 1, 0, 0, 0)

NAVY = "1B2F4E"
TEAL = "0D9488"
LIGHT = "F0FDFA"
WHITE = "FFFFFF"
AMBER = "D97706"
GRAY = "E5E7EB"
INPUT = "FFF9C4"


def hex_color(value: str) -> colors.Color:
    return colors.HexColor(value if value.startswith("#") else f"#{value}")


@dataclass(frozen=True)
class AssetSpec:
    slug: str
    name: str
    filename: str
    kind: str
    theme: str
    use_case: str
    required_phrases: tuple[str, ...]

    @property
    def storage_path(self) -> str:
        return f"{R2_OBJECT_PREFIX}{self.filename}"


ASSETS: tuple[AssetSpec, ...] = (
    AssetSpec(
        "cam-gross-up-calculator",
        "CAM Gross-Up Scenario Calculator",
        "cam-gross-up-calculator.xlsx",
        "xlsx",
        "gross-up modeling",
        "Model variable CAM expense gross-up scenarios before tenant statements go out.",
        ("CAM Gross-Up Scenario Calculator", "Gross-Up", "CapVeri"),
    ),
    AssetSpec(
        "lease-abstract-matrix",
        "Lease Abstract Discrepancy Matrix",
        "lease-abstract-matrix.xlsx",
        "xlsx",
        "lease abstract drift",
        "Compare signed lease clauses, abstracts, and reconciliation inputs in one review grid.",
        ("Lease Abstract Discrepancy Matrix", "Abstract", "CapVeri"),
    ),
    AssetSpec(
        "cam-reconciliation-checklist",
        "CAM Reconciliation Review Checklist",
        "cam-reconciliation-checklist.pdf",
        "pdf",
        "pre-send reconciliation QA",
        "Catch high-risk CAM errors before statements reach tenants.",
        ("CAM Reconciliation Review Checklist", "High-Risk CAM Checks", "CapVeri"),
    ),
    AssetSpec(
        "boma-2024-calculator",
        "BOMA 2024 Calculator Companion",
        "boma-2024-calculator.pdf",
        "pdf",
        "BOMA 2024 remeasurement",
        "Use alongside the in-browser calculator to document rentable-area assumptions.",
        ("BOMA 2024", "Remeasurement", "CapVeri"),
    ),
    AssetSpec(
        "fixed-cam-vs-traditional",
        "Fixed CAM vs Traditional Comparison Companion",
        "fixed-cam-vs-traditional.pdf",
        "pdf",
        "fixed CAM comparison",
        "Document the tradeoffs between fixed CAM and traditional reconciled CAM structures.",
        ("Fixed CAM", "Traditional", "CapVeri"),
    ),
    AssetSpec(
        "admin-fee-calculator",
        "Admin Fee Calculator",
        "admin-fee-calculator.xlsx",
        "xlsx",
        "admin fee methods",
        "Compare gross, net, capped, and excluded-base admin fee methods.",
        ("Admin Fee Calculator", "Admin Fee", "CapVeri"),
    ),
    AssetSpec(
        "cam-estimate-forecaster",
        "CAM Estimate Forecaster",
        "cam-estimate-forecaster.xlsx",
        "xlsx",
        "CAM budget forecasting",
        "Forecast next-year estimates by category using CPI and historical trends.",
        ("CAM Estimate Forecaster", "Forecast", "CapVeri"),
    ),
    AssetSpec(
        "boma-remeasurement-impact",
        "BOMA Remeasurement Impact Analyzer",
        "boma-remeasurement-impact.xlsx",
        "xlsx",
        "BOMA remeasurement ROI",
        "Quantify how a measurement standard change affects recoverable CAM and NOI.",
        ("BOMA Remeasurement Impact Analyzer", "BOMA", "CapVeri"),
    ),
    AssetSpec(
        "cam-cap-calculator",
        "CAM Cap Calculator",
        "cam-cap-calculator.xlsx",
        "xlsx",
        "CAM cap modeling",
        "Model cumulative and non-cumulative cap structures over a lease term.",
        ("CAM Cap Calculator", "Cap", "CapVeri"),
    ),
    AssetSpec(
        "base-year-escalation",
        "Base Year Escalation Calculator",
        "base-year-escalation.xlsx",
        "xlsx",
        "base year escalation",
        "Calculate excess expenses above a base year or expense stop.",
        ("Base Year Escalation Calculator", "Base Year", "CapVeri"),
    ),
    AssetSpec(
        "reconciliation-statement-generator",
        "Reconciliation Statement Generator",
        "reconciliation-statement-generator.xlsx",
        "xlsx",
        "statement package creation",
        "Prepare an audit-ready tenant reconciliation summary with support notes.",
        ("Reconciliation Statement Generator", "Statement", "CapVeri"),
    ),
    AssetSpec(
        "recovery-gap-analyzer",
        "Recovery Gap Analyzer",
        "recovery-gap-analyzer.xlsx",
        "xlsx",
        "recovery leakage",
        "Estimate unrecovered CAM and translate the gap into NOI and value impact.",
        ("Recovery Gap Analyzer", "Recovery Gap", "CapVeri"),
    ),
    AssetSpec(
        "pro-rata-calculator",
        "Pro-Rata Share Calculator",
        "pro-rata-calculator.xlsx",
        "xlsx",
        "pro-rata allocation",
        "Compare tenant share allocations under different denominator methods.",
        ("Pro-Rata Share Calculator", "Pro-Rata", "CapVeri"),
    ),
    AssetSpec(
        "hcad-tax-normalizer",
        "HCAD Tax Normalizer",
        "hcad-tax-normalizer.xlsx",
        "xlsx",
        "property tax normalization",
        "Normalize parcel-level tax changes before passing through CAM tax pools.",
        ("HCAD Tax Normalizer", "Tax", "CapVeri"),
    ),
    AssetSpec(
        "noi-impact-calculator",
        "NOI Impact Calculator",
        "noi-impact-calculator.xlsx",
        "xlsx",
        "NOI sensitivity",
        "Translate CAM under-recovery, caps, and leakage into NOI valuation impact.",
        ("NOI Impact Calculator", "NOI", "CapVeri"),
    ),
    AssetSpec(
        "cam-leakage-estimator",
        "CAM Leakage Estimator",
        "cam-leakage-estimator.xlsx",
        "xlsx",
        "CAM leakage triage",
        "Estimate likely recovery leakage across common reconciliation failure modes.",
        ("CAM Leakage Estimator", "Leakage", "CapVeri"),
    ),
    AssetSpec(
        "cam-overcharge-calculator",
        "Tenant Challenge Exposure Calculator",
        "cam-overcharge-calculator.pdf",
        "pdf",
        "pre-send tenant challenge triage",
        "Estimate which tenant statements are most likely to draw objections before packets go out.",
        ("Tenant Challenge Exposure Calculator", "Challenge", "CapVeri"),
    ),
    AssetSpec(
        "audit-risk-scorecard",
        "Pre-Send Audit Exposure Scorecard",
        "audit-risk-scorecard.pdf",
        "pdf",
        "tenant audit risk",
        "Score tenant, lease, variance, and documentation risk before statements go out.",
        ("Pre-Send Audit Exposure Scorecard", "Risk", "CapVeri"),
    ),
    AssetSpec(
        "sb-1103-checker",
        "SB 1103 Compliance Checker",
        "sb-1103-checker.pdf",
        "pdf",
        "California SB 1103 compliance",
        "Review California CAM disclosure language and supporting documentation gaps.",
        ("SB 1103 Compliance Checker", "Compliance", "CapVeri"),
    ),
    AssetSpec(
        "audit-risk-quiz",
        "Pre-Send Audit Exposure Quiz",
        "audit-risk-quiz.pdf",
        "pdf",
        "audit risk self-assessment",
        "Turn quiz responses into a practical CAM audit exposure action list.",
        ("Pre-Send Audit Exposure Quiz", "Exposure", "CapVeri"),
    ),
    AssetSpec(
        "cam-reconciliation-statement",
        "Tenant CAM Statement Outline",
        "cam-reconciliation-statement.pdf",
        "pdf",
        "statement template",
        "Use a structured tenant statement layout with transparent support sections.",
        ("Tenant CAM Statement Outline", "Statement", "CapVeri"),
    ),
    AssetSpec(
        "cam-reconciliation-excel",
        "CAM Reconciliation Excel Template",
        "cam-reconciliation-excel.xlsx",
        "xlsx",
        "reconciliation workbook",
        "Build a complete reconciliation workbook from GL totals, pools, and tenant shares.",
        ("CAM Reconciliation Excel Template", "Reconciliation", "CapVeri"),
    ),
    AssetSpec(
        "tenant-cam-reconciliation-letter",
        "Landlord CAM Reconciliation Cover Letter",
        "tenant-cam-reconciliation-letter.pdf",
        "pdf",
        "tenant communication",
        "Draft a concise tenant-facing reconciliation cover letter with support references.",
        ("Landlord CAM Reconciliation Cover Letter", "Tenant", "CapVeri"),
    ),
    AssetSpec(
        "cam-reconciliation-california",
        "California CAM Packet Starter",
        "cam-reconciliation-california.pdf",
        "pdf",
        "California CAM disclosure",
        "Prepare a California-specific reconciliation package checklist and disclosure outline.",
        ("California CAM Packet Starter", "California", "CapVeri"),
    ),
    AssetSpec(
        "cam-reconciliation-texas",
        "Texas CAM Packet Starter",
        "cam-reconciliation-texas.pdf",
        "pdf",
        "Texas CAM package",
        "Prepare a Texas CAM reconciliation package with tax and operating expense support.",
        ("Texas CAM Packet Starter", "Texas", "CapVeri"),
    ),
    AssetSpec(
        "cam-reconciliation-florida",
        "Florida CAM Packet Starter",
        "cam-reconciliation-florida.pdf",
        "pdf",
        "Florida CAM package",
        "Prepare a Florida CAM reconciliation package with insurance and tax support prompts.",
        ("Florida CAM Packet Starter", "Florida", "CapVeri"),
    ),
    AssetSpec(
        "nnn-lease-cam-reconciliation",
        "NNN Lease CAM Reconciliation Template",
        "nnn-lease-cam-reconciliation.pdf",
        "pdf",
        "NNN reconciliation",
        "Structure a triple-net reconciliation with pools, exclusions, and tenant credits.",
        ("NNN Lease CAM Reconciliation Template", "NNN", "CapVeri"),
    ),
    AssetSpec(
        "cam-dispute-response-template",
        "CAM Dispute Response Template",
        "cam-dispute-response-template.pdf",
        "pdf",
        "tenant dispute response",
        "Respond to tenant CAM objections with a documented evidence checklist.",
        ("CAM Dispute Response Template", "Dispute", "CapVeri"),
    ),
    AssetSpec(
        "cam-estimate-letter",
        "CAM Estimate / Budget Letter",
        "cam-estimate-letter.pdf",
        "pdf",
        "estimate communication",
        "Send next-year CAM estimates with budget assumptions and variance explanations.",
        ("CAM Estimate / Budget Letter", "Estimate", "CapVeri"),
    ),
    AssetSpec(
        "cumulative-cap-bank-calculator",
        "Cumulative CAM Cap Bank Calculator",
        "cumulative-cap-bank-calculator.xlsx",
        "xlsx",
        "cumulative cap bank tracking",
        "Track unused annual CAM cap capacity and future recoverability by lease year.",
        ("Cumulative CAM Cap Bank Calculator", "Cap Bank", "CapVeri"),
    ),
    AssetSpec(
        "cam-pre-send-packet-checklist",
        "CAM Pre-Send Packet Checklist",
        "cam-pre-send-packet-checklist.pdf",
        "pdf",
        "pre-send close package",
        "Confirm every tenant packet has calculation support before distribution.",
        ("CAM Pre-Send Packet Checklist", "Packet", "CapVeri"),
    ),
    AssetSpec(
        "yardi-export-qa-checklist",
        "Yardi Export Error Checklist",
        "yardi-export-qa-checklist.pdf",
        "pdf",
        "Yardi export QA",
        "Validate Yardi GL and rent roll exports before CAM reconciliation work begins.",
        ("Yardi Export Error Checklist", "Yardi", "CapVeri"),
    ),
    AssetSpec(
        "mri-recovery-billing-qa-checklist",
        "MRI Recovery Billing Error Checklist",
        "mri-recovery-billing-qa-checklist.pdf",
        "pdf",
        "MRI recovery billing QA",
        "Validate MRI recovery billing exports and tenant setup before reconciliation.",
        ("MRI Recovery Billing Error Checklist", "MRI", "CapVeri"),
    ),
    AssetSpec(
        "multi-state-cam-disclosure-matrix",
        "Multi-State CAM Packet Review Checklist",
        "multi-state-cam-disclosure-matrix.pdf",
        "pdf",
        "multi-state disclosure review",
        "Compare CAM notice, support, and counsel-review needs across state-specific packages.",
        ("Multi-State CAM Packet Review Checklist", "Disclosure", "CapVeri"),
    ),
    AssetSpec(
        "cam-recovery-ratio-worksheet",
        "CAM Recovery Ratio Benchmark Worksheet",
        "cam-recovery-ratio-worksheet.xlsx",
        "xlsx",
        "recovery ratio benchmarking",
        "Benchmark billed CAM recovery against recoverable expense pools and portfolio targets.",
        ("CAM Recovery Ratio Benchmark Worksheet", "Recovery Ratio", "CapVeri"),
    ),
    AssetSpec(
        "property-tax-appeal-recovery-calculator",
        "Property Tax Appeal Recovery Calculator",
        "property-tax-appeal-recovery-calculator.xlsx",
        "xlsx",
        "tax appeal recovery modeling",
        "Model tenant credits, rebills, and NOI impact after a property tax appeal.",
        ("Property Tax Appeal Recovery Calculator", "Tax Appeal", "CapVeri"),
    ),
    AssetSpec(
        "tenant-dispute-response-letter-template",
        "Tenant CAM Dispute Response Letter",
        "tenant-dispute-response-letter-template.pdf",
        "pdf",
        "tenant dispute response",
        "Draft a structured response letter for tenant CAM objections and backup requests.",
        ("Tenant CAM Dispute Response Letter", "Dispute Letter", "CapVeri"),
    ),
    AssetSpec(
        "audit-defense-packet-builder",
        "Audit Defense Packet Builder",
        "audit-defense-packet-builder.pdf",
        "pdf",
        "audit defense packet",
        "Assemble a tenant-audit defense packet with calculations, support, and sign-offs.",
        ("Audit Defense Packet Builder", "Audit Defense", "CapVeri"),
    ),
    AssetSpec(
        "lease-clause-extraction-matrix",
        "Lease Clause Extraction Matrix",
        "lease-clause-extraction-matrix.xlsx",
        "xlsx",
        "lease clause extraction",
        "Extract CAM recovery clauses into a structured review matrix for reconciliation.",
        ("Lease Clause Extraction Matrix", "Clause Extraction", "CapVeri"),
    ),
)


class Rule(Flowable):
    def wrap(self, avail_width: float, avail_height: float) -> tuple[float, float]:
        return avail_width, 8

    def draw(self) -> None:
        self.canv.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.canv.line(0, 4, self.width, 4)


def styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "eyebrow": ParagraphStyle(
            "Eyebrow",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=hex_color(TEAL),
            spaceAfter=5,
        ),
        "title": ParagraphStyle(
            "Title",
            parent=base["Title"],
            fontName="Helvetica-Bold",
            fontSize=22,
            leading=27,
            textColor=hex_color(NAVY),
            spaceAfter=8,
        ),
        "body": ParagraphStyle(
            "Body",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=13.5,
            textColor=colors.HexColor("#1F2937"),
        ),
        "h2": ParagraphStyle(
            "Section",
            parent=base["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=12.5,
            leading=15,
            textColor=hex_color(NAVY),
            spaceBefore=10,
            spaceAfter=6,
        ),
        "small": ParagraphStyle(
            "Small",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=10.5,
            textColor=colors.HexColor("#64748B"),
        ),
        "header": ParagraphStyle(
            "Header",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=10.5,
            textColor=colors.white,
        ),
    }


def deterministic_canvas(*args, **kwargs) -> Canvas:  # type: ignore[no-untyped-def]
    kwargs["invariant"] = 1
    return Canvas(*args, **kwargs)


def pdf_workflow_for(spec: AssetSpec) -> list[tuple[str, str, str]]:
    workflows: dict[str, list[tuple[str, str, str]]] = {
        "audit-risk-quiz": [
            (
                "1",
                "Answer the five exposure questions",
                "Score lease clarity, variance size, exclusions, documentation, and tenant history.",
            ),
            (
                "2",
                "Total the exposure score",
                "Use the score band to decide whether to send, review, or escalate before statements go out.",
            ),
            (
                "3",
                "Write one mitigation action",
                "Document the specific backup or lease citation that would reduce each high-risk answer.",
            ),
            (
                "4",
                "Assign an owner",
                "Route accounting, asset management, or counsel follow-up before the tenant receives the package.",
            ),
            (
                "5",
                "Retest after remediation",
                "Update the score after backup is attached and exceptions are resolved.",
            ),
        ],
        "sb-1103-checker": [
            (
                "1",
                "Identify covered California space",
                "Confirm whether the tenant and premises fall under California SB 1103 disclosure rules.",
            ),
            (
                "2",
                "Check fee transparency",
                "List CAM, tax, insurance, admin, and other operating charge categories separately.",
            ),
            (
                "3",
                "Attach support",
                "Keep budget, prior reconciliation, lease language, and written explanations with the disclosure file.",
            ),
            (
                "4",
                "Review timing",
                "Confirm required notices and estimate updates are sent before billing deadlines.",
            ),
            (
                "5",
                "Escalate ambiguity",
                "Send unclear lease or statutory questions to counsel before issuing the package.",
            ),
        ],
        "cam-dispute-response-template": [
            (
                "1",
                "Acknowledge the tenant issue",
                "Confirm receipt, deadline, disputed categories, and any audit-right language.",
            ),
            (
                "2",
                "Map each objection",
                "Tie every tenant claim to a lease clause, source report, invoice, or calculation schedule.",
            ),
            (
                "3",
                "Separate conceded items",
                "Remove or credit confirmed errors without weakening supported recoveries.",
            ),
            (
                "4",
                "Prepare evidence bundle",
                "Attach GL detail, pool mapping, rent roll, invoices, and reviewer notes.",
            ),
            (
                "5",
                "Close with next step",
                "State the proposed resolution, response date, and escalation path.",
            ),
        ],
        "cam-reconciliation-statement": [
            (
                "1",
                "Summarize the tenant position",
                "Show prior estimates, actual share, credits, balance due, and payment instructions.",
            ),
            (
                "2",
                "Disclose pool totals",
                "Break out recoverable taxes, insurance, utilities, maintenance, admin, and excluded costs.",
            ),
            (
                "3",
                "Show allocation math",
                "Present denominator, tenant square feet, pro-rata share, caps, and gross-up adjustments.",
            ),
            (
                "4",
                "Explain variances",
                "Write tenant-ready notes for material year-over-year changes.",
            ),
            (
                "5",
                "Attach support index",
                "Reference source reports without overwhelming the tenant-facing statement.",
            ),
        ],
        "tenant-cam-reconciliation-letter": [
            (
                "1",
                "Open with the result",
                "State whether the tenant has a credit or balance due and the covered reconciliation period.",
            ),
            (
                "2",
                "Explain the driver",
                "Name the top two changes behind the result in plain language.",
            ),
            (
                "3",
                "Reference attachments",
                "Point to statement, support schedule, lease clause, and payment instructions.",
            ),
            (
                "4",
                "Invite questions",
                "Give a contact, response window, and document request process.",
            ),
            (
                "5",
                "Preserve rights",
                "Use neutral language that does not waive lease remedies or audit rights.",
            ),
        ],
        "cam-estimate-letter": [
            (
                "1",
                "State the new estimate",
                "Show the monthly and annual estimated CAM amount and effective date.",
            ),
            (
                "2",
                "Explain budget assumptions",
                "Summarize tax, insurance, utilities, contract, and occupancy drivers.",
            ),
            (
                "3",
                "Tie to lease authority",
                "Reference the lease section that permits estimates and later true-up.",
            ),
            (
                "4",
                "Show variance context",
                "Compare prior estimate, current estimate, and expected reconciliation effect.",
            ),
            (
                "5",
                "Give tenant next steps",
                "Include payment timing, contact, and expected year-end reconciliation schedule.",
            ),
        ],
        "boma-2024-calculator": [
            (
                "1",
                "Capture current rentable area",
                "Record the current standard, measurement date, and source plan set.",
            ),
            (
                "2",
                "Enter BOMA 2024 assumptions",
                "Document common-area allocations, exclusions, and measurement treatment.",
            ),
            (
                "3",
                "Compare tenant share",
                "Calculate denominator and share changes before changing CAM statements.",
            ),
            (
                "4",
                "Model NOI effect",
                "Translate area change into recoverable CAM and valuation impact.",
            ),
            (
                "5",
                "Approve implementation",
                "Keep broker, asset manager, and legal sign-off with the measurement file.",
            ),
        ],
        "fixed-cam-vs-traditional": [
            (
                "1",
                "Define the fixed CAM offer",
                "Record starting rate, annual escalator, exclusions, and controllable/uncontrollable treatment.",
            ),
            (
                "2",
                "Build traditional baseline",
                "Use prior actuals and budget assumptions to model normal reconciled CAM.",
            ),
            (
                "3",
                "Compare volatility",
                "Show tenant predictability versus landlord leakage or upside.",
            ),
            (
                "4",
                "Stress-test inflation",
                "Model high tax, insurance, utilities, and service contract scenarios.",
            ),
            (
                "5",
                "Set renewal position",
                "Decide whether fixed CAM, hybrid caps, or traditional true-up is the better lease term.",
            ),
        ],
        "cam-pre-send-packet-checklist": [
            (
                "1",
                "Confirm packet inventory",
                "Verify each tenant packet includes statement, variance notes, and support index.",
            ),
            (
                "2",
                "Check calculation trace",
                "Confirm every balance due ties from GL pool to denominator to tenant share.",
            ),
            (
                "3",
                "Review exception log",
                "Attach approvals for exclusions, caps, credits, gross-up, and admin fee treatment.",
            ),
            (
                "4",
                "Validate delivery list",
                "Match tenant contacts, notice addresses, and email recipients to lease records.",
            ),
            (
                "5",
                "Archive final copy",
                "Store the sent packet and reviewer sign-off before distribution.",
            ),
        ],
        "yardi-export-qa-checklist": [
            (
                "1",
                "Export complete GL detail",
                "Pull Yardi accounts, properties, dates, books, and descriptions for the full recovery period.",
            ),
            (
                "2",
                "Tie to trial balance",
                "Reconcile export totals to Yardi financial statements before mapping pools.",
            ),
            (
                "3",
                "Validate account mapping",
                "Separate recoverable, excluded, capital, tax, insurance, and utility accounts.",
            ),
            (
                "4",
                "Check tenant setup",
                "Compare Yardi lease recovery setup to signed lease terms and amendments.",
            ),
            (
                "5",
                "Lock source files",
                "Save raw exports so reviewers can reproduce the reconciliation.",
            ),
        ],
        "mri-recovery-billing-qa-checklist": [
            (
                "1",
                "Validate MRI export scope",
                "Confirm entity, lease, recovery group, period, and book filters before export.",
            ),
            (
                "2",
                "Review recovery setup",
                "Compare MRI recovery billing settings to lease caps, exclusions, and admin fees.",
            ),
            (
                "3",
                "Tie rent roll denominator",
                "Check occupied, rentable, and excluded square footage against the rent roll.",
            ),
            (
                "4",
                "Inspect exception tenants",
                "Review anchor tenants, fixed CAM deals, base-year leases, and manual overrides.",
            ),
            (
                "5",
                "Archive export package",
                "Store MRI source reports with the final tenant billing package.",
            ),
        ],
        "multi-state-cam-disclosure-matrix": [
            (
                "1",
                "List covered states",
                "Identify each property state and any local disclosure or timing review needs.",
            ),
            (
                "2",
                "Map required support",
                "Track estimate, actual, tax, insurance, admin, and reconciliation support by state.",
            ),
            (
                "3",
                "Add counsel notes",
                "Record legal review status for tenant notices and disclosure language.",
            ),
            (
                "4",
                "Compare packet gaps",
                "Flag properties missing state-specific backup or notice evidence.",
            ),
            (
                "5",
                "Finalize send approval",
                "Confirm state-sensitive packages are approved before tenant delivery.",
            ),
        ],
        "tenant-dispute-response-letter-template": [
            (
                "1",
                "Acknowledge the dispute",
                "Identify the tenant, period, disputed amount, and response deadline.",
            ),
            (
                "2",
                "Summarize landlord position",
                "Tie the response to lease sections, calculations, and source support.",
            ),
            (
                "3",
                "Address each objection",
                "Use a table for tenant claim, support reference, decision, and next step.",
            ),
            (
                "4",
                "Attach backup",
                "Include only relevant support and offer a structured follow-up process.",
            ),
            (
                "5",
                "Reserve rights",
                "Avoid informal language that waives lease remedies or collection rights.",
            ),
        ],
        "audit-defense-packet-builder": [
            (
                "1",
                "Create evidence index",
                "List each source file needed to defend CAM billing under the lease.",
            ),
            (
                "2",
                "Trace the calculation",
                "Show GL total, pool mapping, denominator, tenant share, caps, and credits.",
            ),
            (
                "3",
                "Document judgments",
                "Explain exclusions, capital treatment, gross-up, and admin fee decisions.",
            ),
            (
                "4",
                "Assign reviewers",
                "Capture accounting, asset management, and legal approvals where needed.",
            ),
            (
                "5",
                "Package the response",
                "Prepare a clean tenant-audit folder with version-controlled support.",
            ),
        ],
    }
    if spec.slug in workflows:
        return workflows[spec.slug]
    if spec.slug in {
        "cam-reconciliation-california",
        "cam-reconciliation-texas",
        "cam-reconciliation-florida",
    }:
        return [
            (
                "1",
                "Confirm local package requirements",
                f"Adapt the reconciliation cover and support checklist for {spec.theme}.",
            ),
            (
                "2",
                "Tie charges to lease language",
                "Separate recoverable, excluded, capped, and notice-sensitive expense categories.",
            ),
            (
                "3",
                "Build tenant statement",
                "Show estimates, actuals, credits, balance due, and support references.",
            ),
            (
                "4",
                "Review tax and insurance drivers",
                "Document changes that often create tenant questions in this market.",
            ),
            (
                "5",
                "Archive support",
                "Keep the final package, source reports, and approvals together.",
            ),
        ]
    if spec.slug == "multi-state-cam-disclosure-matrix":
        return [
            (
                "1",
                "Build state inventory",
                f"List each state in scope for {spec.theme}.",
            ),
            (
                "2",
                "Compare disclosure timing",
                "Track notice, estimate, and actual reconciliation timing requirements.",
            ),
            (
                "3",
                "Review support language",
                "Confirm tenant-facing descriptions are clear and counsel-reviewed.",
            ),
            (
                "4",
                "Flag missing evidence",
                "Identify packets missing tax, insurance, or operating cost support.",
            ),
            (
                "5",
                "Approve send list",
                "Confirm each state-specific package has final review sign-off.",
            ),
        ]
    if spec.slug == "nnn-lease-cam-reconciliation":
        return [
            (
                "1",
                "Classify NNN recoveries",
                "Separate taxes, insurance, CAM, utilities, and direct bill items.",
            ),
            (
                "2",
                "Validate exclusions",
                "Remove landlord overhead, capital, and non-property costs unless the lease allows them.",
            ),
            (
                "3",
                "Calculate tenant share",
                "Use the NNN denominator and any anchor or vacant-area rules.",
            ),
            (
                "4",
                "Apply credits",
                "Net estimates, direct payments, abatements, and prior corrections.",
            ),
            (
                "5",
                "Package support",
                "Attach the schedule tenants need to trace the NNN true-up.",
            ),
        ]
    return [
        (
            "1",
            "Confirm lease authority",
            "Tie the promised recovery method to signed lease language.",
        ),
        (
            "2",
            "Validate source data",
            "Reconcile GL exports, rent rolls, and prior-year true-up data before billing.",
        ),
        (
            "3",
            "Review allocation logic",
            "Check pools, denominators, exclusions, caps, gross-up, and admin fee bases.",
        ),
        (
            "4",
            "Package support",
            "Attach variance notes and source support before a tenant has to ask.",
        ),
        (
            "5",
            "Escalate exceptions",
            "Route legal, compliance, or material-dollar issues to the right reviewer.",
        ),
    ]


def pdf_focus_sections_for(spec: AssetSpec) -> list[tuple[str, list[str]]]:
    sections: dict[str, list[tuple[str, list[str]]]] = {
        "audit-risk-quiz": [
            (
                "Quiz Questions",
                [
                    "Is any recoverable expense category missing a clear lease citation?",
                    "Did total recoverable CAM increase more than 10% from the prior year?",
                    "Are exclusions, caps, or gross-up provisions manually adjusted outside the workbook?",
                    "Would a tenant auditor be able to trace the balance due from source reports?",
                    "Has this tenant disputed CAM or requested backup in the past two years?",
                ],
            ),
            (
                "Score Bands",
                [
                    "0-3 points: low exposure; send after normal reviewer sign-off.",
                    "4-7 points: medium exposure; add variance notes and source support before sending.",
                    "8+ points: high exposure; escalate exceptions before the tenant package goes out.",
                ],
            ),
            (
                "Action Log",
                [
                    "For every yes answer, write the owner, mitigation, evidence location, and due date.",
                    "Retain the final score with the reconciliation approval packet.",
                ],
            ),
        ],
        "sb-1103-checker": [
            (
                "California Disclosure Checks",
                [
                    "Identify mandatory fees separately from optional services and pass-through estimates.",
                    "State whether CAM, taxes, insurance, and admin charges are estimates or fixed amounts.",
                    "Keep copies of written explanations provided before lease execution or renewal.",
                ],
            ),
            (
                "Evidence File",
                [
                    "Lease or amendment language supporting each operating cost category.",
                    "Budget, prior-year actuals, property tax notices, insurance renewal, and vendor contract changes.",
                ],
            ),
            (
                "Caveat",
                [
                    "This is not legal advice; confirm SB 1103 applicability and disclosure language with California counsel.",
                ],
            ),
        ],
        "cam-dispute-response-template": [
            (
                "Response Outline",
                [
                    "Thank the tenant, identify the disputed period, and reserve all lease rights.",
                    "Answer each objection in a table with landlord position, support, and proposed resolution.",
                    "Separate accepted corrections from items supported by the governing lease.",
                ],
            ),
            (
                "Evidence Checklist",
                [
                    "Signed lease excerpt, amendment history, GL detail, invoices, rent roll, denominator support, and prior credits.",
                    "Internal reviewer sign-off for settlement offers or material concessions.",
                ],
            ),
            (
                "Suggested Closing",
                [
                    "Offer a meeting, state the payment or credit next step, and give a response deadline.",
                ],
            ),
        ],
        "cam-reconciliation-statement": [
            (
                "Statement Sections",
                [
                    "Tenant summary: estimates paid, actual CAM share, credits, and balance due.",
                    "Pool detail: recoverable expenses by category with exclusions shown separately.",
                    "Calculation bridge: denominator, share, caps, gross-up, admin fee, and final adjustment.",
                ],
            ),
            (
                "Tenant-Ready Variance Notes",
                [
                    "Explain material tax, insurance, utility, snow removal, security, or contract changes.",
                    "Avoid internal account codes without a plain-English label.",
                ],
            ),
            (
                "Audit Trail",
                [
                    "Store source reports and approval notes even when the tenant-facing package is concise.",
                ],
            ),
        ],
        "tenant-cam-reconciliation-letter": [
            (
                "Letter Template Blocks",
                [
                    "Opening result paragraph with covered period and balance or credit.",
                    "Plain-English driver paragraph covering the largest changes.",
                    "Attachment and payment paragraph with contact information.",
                ],
            ),
            (
                "Language Guardrails",
                [
                    "Do not characterize disputed or estimated amounts as final unless the lease supports it.",
                    "Avoid waiving audit, collection, or correction rights in informal email language.",
                ],
            ),
            (
                "Support References",
                [
                    "Include the statement, variance summary, and document request contact.",
                ],
            ),
        ],
        "cam-estimate-letter": [
            (
                "Estimate Letter Blocks",
                [
                    "New monthly estimate, annualized amount, effective date, and payment start.",
                    "Budget assumptions by category: taxes, insurance, utilities, contracts, occupancy.",
                    "True-up reminder explaining how estimates reconcile against actuals.",
                ],
            ),
            (
                "Variance Explanation Prompts",
                [
                    "What changed in the tax assessment or appeal status?",
                    "Which insurance, utility, or service contracts reset?",
                    "Which costs are non-controllable under the lease?",
                ],
            ),
            (
                "Approval Notes",
                [
                    "Retain budget owner approval and final tenant notice copy.",
                ],
            ),
        ],
        "cam-pre-send-packet-checklist": [
            (
                "Packet Completeness",
                [
                    "Tenant statement, variance notes, calculation summary, support index, and delivery contact are present.",
                    "No packet leaves accounting without an owner and final reviewer sign-off.",
                ],
            ),
            (
                "Calculation Trace",
                [
                    "Every tenant balance ties from GL export to pool mapping to denominator to final amount due.",
                    "Caps, credits, fixed CAM, gross-up, and admin fee decisions are separately documented.",
                ],
            ),
            (
                "Send Controls",
                [
                    "Use the checklist before email or portal distribution and archive a final sent copy.",
                ],
            ),
        ],
        "yardi-export-qa-checklist": [
            (
                "Yardi Export Checks",
                [
                    "Confirm property, book, period, account range, and recovery account filters before exporting.",
                    "Tie Yardi export totals to the financial statement or trial balance.",
                ],
            ),
            (
                "Mapping Review",
                [
                    "Flag excluded accounts, capital items, tax, insurance, utilities, and admin fee bases separately.",
                    "Compare Yardi recovery setup against lease abstract fields before billing.",
                ],
            ),
            (
                "Audit Trail",
                [
                    "Store raw Yardi exports, mapping decisions, and reviewer sign-off with the reconciliation.",
                ],
            ),
        ],
        "mri-recovery-billing-qa-checklist": [
            (
                "MRI Export Checks",
                [
                    "Validate entity, lease, recovery group, period, and book before running reports.",
                    "Tie MRI recovery billing totals to the source GL and rent roll denominator.",
                ],
            ),
            (
                "Tenant Setup Review",
                [
                    "Review manual overrides, anchors, base-year tenants, fixed CAM tenants, and capped leases.",
                    "Confirm admin fee, tax, insurance, and utility settings match the lease.",
                ],
            ),
            (
                "Close Package",
                [
                    "Archive raw MRI exports and final tenant statements together.",
                ],
            ),
        ],
        "multi-state-cam-disclosure-matrix": [
            (
                "Matrix Columns",
                [
                    "State, property, lease type, estimate notice, actual reconciliation, tax support, insurance support, and counsel review.",
                    "Use status values for missing, draft, reviewed, and approved.",
                ],
            ),
            (
                "Disclosure Caveat",
                [
                    "This matrix is an operational tracker, not legal advice. Confirm state-specific requirements with counsel.",
                ],
            ),
            (
                "Portfolio Workflow",
                [
                    "Review state-sensitive packets before tenant delivery and keep evidence with the final package.",
                ],
            ),
        ],
        "tenant-dispute-response-letter-template": [
            (
                "Letter Blocks",
                [
                    "Use the Dispute Letter structure to keep the response factual, lease-grounded, and easy to audit.",
                    "Acknowledgement, landlord position, response table, support references, and proposed next step.",
                    "Use neutral language and preserve all rights under the lease.",
                ],
            ),
            (
                "Response Table Fields",
                [
                    "Tenant objection, lease reference, source evidence, landlord response, accepted correction, and open issue.",
                ],
            ),
            (
                "Approval Notes",
                [
                    "Route concessions, settlement offers, and legal interpretations for approval before sending.",
                ],
            ),
        ],
        "audit-defense-packet-builder": [
            (
                "Defense Packet Tabs",
                [
                    "Lease authority, source reports, pool mapping, calculation trace, correspondence, and reviewer approvals.",
                    "Keep the packet tenant-specific so support is easy to inspect.",
                ],
            ),
            (
                "High-Risk Evidence",
                [
                    "Excluded expense decisions, gross-up assumptions, cap treatment, denominator changes, and admin fee base.",
                ],
            ),
            (
                "Audit Caveat",
                [
                    "This resource organizes support; legal strategy and privilege decisions should be reviewed with counsel.",
                ],
            ),
        ],
    }
    if spec.slug in sections:
        return sections[spec.slug]
    if spec.slug in {
        "cam-reconciliation-california",
        "cam-reconciliation-texas",
        "cam-reconciliation-florida",
    }:
        return [
            (
                "Package Sections",
                [
                    f"Use this {spec.theme} outline to organize tenant summary, expense pool detail, variance notes, and support references.",
                    "Show statutory or market-sensitive items as review prompts, not legal conclusions.",
                ],
            ),
            (
                "State-Specific Review Prompts",
                [
                    "Confirm tax assessment, insurance renewal, and notice timing assumptions.",
                    "Have counsel review any local-law disclosure or tenant-right language before sending.",
                ],
            ),
            (
                "Final File Checklist",
                [
                    "Signed lease excerpts, source reports, reviewer sign-off, and tenant-facing copy.",
                ],
            ),
        ]
    return [
        (
            "High-Risk CAM Checks",
            [
                "Confirm excluded expenses are removed before tenant share calculations.",
                "Separate fixed costs from variable costs before applying gross-up language.",
                "Document year-over-year variances that would draw a tenant auditor's attention.",
            ],
        ),
        (
            "Evidence To Keep With The File",
            [
                "Signed lease or amendment excerpt supporting the recovery method.",
                "GL export, allocation pool summary, rent roll, and supporting invoice references.",
                "Reviewer sign-off showing who approved exceptions and why.",
            ],
        ),
        (
            "Reviewer Sign-Off Prompts",
            [
                "What changed materially from the prior year, and is the explanation tenant-ready?",
                "Which line items are most likely to be challenged, and where is the support stored?",
                "Does the final package show the calculation path from source totals to tenant amount due?",
            ],
        ),
        (
            "Common Failure Modes",
            [
                "A lease exclusion is mapped to the wrong GL account and recovered anyway.",
                "A denominator changes after a suite remeasurement without an approval trail.",
                "A cap or expense stop is applied after gross-up instead of at the lease-defined step.",
            ],
        ),
        (
            "Worked Example",
            [
                "If recoverable CAM rises from $390,000 to $450,000 and the tenant share is 8.5%, the tenant-facing variance note should explain the $5,100 share movement before the statement is issued.",
                "Attach the GL export, rent roll denominator, lease excerpt, and reviewer approval so the number can be traced without rebuilding the file.",
            ],
        ),
    ]


def mobile_summary_for(spec: AssetSpec) -> list[str]:
    return [
        f"Use this when you need to {spec.use_case[0].lower() + spec.use_case[1:]}",
        "Start with the first two checks on your phone. Finish the workbook or packet on desktop.",
        "Keep the lease excerpt, GL export, rent roll, and reviewer approval with the final file.",
        "Next step: run the same check in CapVeri during a 30-day trial.",
    ]


def footer(canvas, doc) -> None:  # type: ignore[no-untyped-def]
    canvas.saveState()
    width, height = letter
    canvas.setFillColor(hex_color(TEAL))
    canvas.rect(0, height - 0.16 * inch, width, 0.16 * inch, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor("#64748B"))
    canvas.setFont("Helvetica", 8)
    canvas.drawString(doc.leftMargin, 0.38 * inch, "CapVeri CAM review resource")
    canvas.drawRightString(width - doc.rightMargin, 0.38 * inch, f"Page {doc.page}")
    canvas.restoreState()


def build_pdf(spec: AssetSpec, out_path: Path) -> None:
    s = styles()
    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=letter,
        leftMargin=0.62 * inch,
        rightMargin=0.62 * inch,
        topMargin=0.62 * inch,
        bottomMargin=0.68 * inch,
        title=spec.name,
        author="CapVeri",
    )
    rows = pdf_workflow_for(spec)
    story: list[Flowable] = [
        Paragraph("CAPVERI RESOURCE", s["eyebrow"]),
        Paragraph(spec.name, s["title"]),
        Paragraph(spec.use_case, s["body"]),
        Spacer(1, 8),
        KeepTogether(
            [
                Paragraph("Phone Summary", s["h2"]),
                *[
                    Paragraph(f"- {item}", s["body"])
                    for item in mobile_summary_for(spec)
                ],
            ]
        ),
        Spacer(1, 6),
        Table(
            [
                [
                    Paragraph("<b>Best use</b><br/>" + spec.theme, s["small"]),
                    Paragraph(
                        "<b>Resource type</b><br/>" + spec.kind.upper(), s["small"]
                    ),
                    Paragraph("<b>Review time</b><br/>15-30 minutes", s["small"]),
                ]
            ],
            colWidths=[2.1 * inch, 2.2 * inch, 2.2 * inch],
            style=TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), hex_color(LIGHT)),
                    ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("PADDING", (0, 0), (-1, -1), 9),
                ]
            ),
        ),
        Spacer(1, 8),
        Paragraph(
            "This resource is an operational aid, not legal advice. Confirm final CAM billing positions with the governing lease and counsel where required.",
            s["small"],
        ),
        Rule(),
        Paragraph("Practical Review Workflow", s["h2"]),
        Table(
            [
                [
                    Paragraph("Step", s["header"]),
                    Paragraph("Checkpoint", s["header"]),
                    Paragraph("What good looks like", s["header"]),
                ]
            ]
            + [
                [
                    Paragraph(num, s["body"]),
                    Paragraph(label, s["body"]),
                    Paragraph(detail, s["body"]),
                ]
                for num, label, detail in rows
            ],
            colWidths=[0.55 * inch, 1.85 * inch, 4.1 * inch],
            style=TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), hex_color(NAVY)),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("PADDING", (0, 0), (-1, -1), 7),
                ]
            ),
        ),
    ]
    focus_sections = [
        *pdf_focus_sections_for(spec),
        (
            "Quality Gate",
            [
                "Before sending, confirm the workbook or packet ties to source reports, lease authority, reviewer initials, and a dated final archive copy.",
                "Escalate unresolved exclusions, manual overrides, legal interpretation, or tenant-specific concessions before relying on the output.",
            ],
        ),
        (
            "CapVeri Follow-Up",
            [
                "Use CapVeri when you need deterministic CAM math, lease-aware validation, and audit-ready export packages across a portfolio.",
                f"Start a 30-day trial at {REGISTER_URL}.",
            ],
        ),
    ]
    for title, bullets in focus_sections:
        story.append(
            KeepTogether(
                [Paragraph(title, s["h2"])]
                + [Paragraph(f"- {item}", s["body"]) for item in bullets]
            )
        )
    doc.build(
        story, onFirstPage=footer, onLaterPages=footer, canvasmaker=deterministic_canvas
    )


def cell_header(cell, value: str, fill: str = NAVY) -> None:
    cell.value = value
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def cell_label(cell, value: str, bold: bool = False) -> None:
    cell.value = value
    cell.font = Font(bold=bold, color="111827")
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def cell_input(cell, value, number_format: str = "General") -> None:
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=INPUT)
    cell.protection = Protection(locked=False)
    cell.number_format = number_format


def cell_formula(cell, formula: str, number_format: str = "General") -> None:
    cell.value = formula
    cell.fill = PatternFill("solid", fgColor=LIGHT)
    cell.font = Font(bold=True, color=NAVY)
    cell.number_format = number_format


def xlsx_model_for(spec: AssetSpec) -> tuple[
    list[tuple[str, float, str]],
    list[tuple[str, str, str, str, str]],
    list[tuple[str, str]],
]:
    models: dict[
        str,
        tuple[
            list[tuple[str, float, str]],
            list[tuple[str, str, str, str, str]],
            list[tuple[str, str]],
        ],
    ] = {
        "cam-gross-up-calculator": (
            [
                ("Variable recoverable CAM", 310000, '"$"#,##0'),
                ("Fixed recoverable CAM", 140000, '"$"#,##0'),
                ("Actual occupancy", 0.82, "0.0%"),
                ("Lease gross-up occupancy", 0.95, "0.0%"),
                ("Tenant pro-rata share", 0.085, "0.00%"),
                ("Excluded variable costs", 25000, '"$"#,##0'),
                ("Prior-year tenant charge", 36000, '"$"#,##0'),
            ],
            [
                (
                    "Gross-up factor",
                    "=B8",
                    "=B8/B7",
                    "=C16-B16",
                    "Use only for variable costs permitted by lease.",
                ),
                (
                    "Adjusted CAM pool",
                    "=B5+B6",
                    "=(B5-B10)*C16+B10+B6",
                    "=C17-B17",
                    "Fixed costs should not be grossed up.",
                ),
                (
                    "Tenant recoverable charge",
                    "=B17*B9",
                    "=C17*B9",
                    "=C18-B18",
                    "Tenant impact of gross-up treatment.",
                ),
                (
                    "Variance vs prior year",
                    "=B18-B11",
                    "=C18-B11",
                    "=C19-B19",
                    "Explain material changes before billing.",
                ),
            ],
            [
                ("Gross-up above actual occupancy", '=IF(B8>B7,"VERIFY LEASE","OK")'),
                ("Excluded variable costs", '=IF(B10>0,"CHECK EXCLUSIONS","OK")'),
            ],
        ),
        "lease-abstract-matrix": (
            [
                ("Signed lease recovery share", 0.085, "0.00%"),
                ("Abstract recovery share", 0.0875, "0.00%"),
                ("Lease cap rate", 0.05, "0.0%"),
                ("Abstract cap rate", 0.0, "0.0%"),
                ("Recoverable CAM pool", 450000, '"$"#,##0'),
                ("Excluded expenses in lease", 65000, '"$"#,##0'),
                ("Excluded expenses in abstract", 25000, '"$"#,##0'),
            ],
            [
                (
                    "Share variance",
                    "=B6-B5",
                    "=ABS(B6-B5)",
                    "=C16-B16",
                    "Any share mismatch needs source review.",
                ),
                (
                    "Cap variance",
                    "=B8-B7",
                    "=ABS(B8-B7)",
                    "=C17-B17",
                    "Missing cap language creates billing exposure.",
                ),
                (
                    "Exclusion variance",
                    "=B11-B10",
                    "=ABS(B11-B10)",
                    "=C18-B18",
                    "Compare excluded category mapping.",
                ),
                (
                    "Potential exposure",
                    "=B9*(B16+B17)+B18",
                    "=B9*C16+B18",
                    "=C19-B19",
                    "Estimate dollar impact of abstract drift.",
                ),
            ],
            [
                ("Share mismatch", '=IF(C16>0.0001,"REVIEW","OK")'),
                ("Missing cap", '=IF(AND(B7>0,B8=0),"REVIEW","OK")'),
            ],
        ),
        "admin-fee-calculator": (
            [
                ("Recoverable CAM before admin", 450000, '"$"#,##0'),
                ("Excluded expense base", 65000, '"$"#,##0'),
                ("Admin fee rate", 0.15, "0.0%"),
                ("Fee cap amount", 50000, '"$"#,##0'),
                ("Prior-year admin fee", 42000, '"$"#,##0'),
                ("Tenant pro-rata share", 0.085, "0.00%"),
                ("Negotiated reduction", 0.0, "0.0%"),
            ],
            [
                (
                    "Gross fee",
                    "=B5*B7",
                    "=MIN(B5*B7,B8)",
                    "=C16-B16",
                    "Compare uncapped and capped admin fee.",
                ),
                (
                    "Net-of-exclusions fee",
                    "=(B5-B6)*B7",
                    "=MIN((B5-B6)*B7,B8)",
                    "=C17-B17",
                    "Use when lease excludes tax or insurance from fee base.",
                ),
                (
                    "Tenant share",
                    "=B16*B10",
                    "=C17*B10*(1-B11)",
                    "=C18-B18",
                    "Apply share and negotiated reduction.",
                ),
                (
                    "Year-over-year variance",
                    "=B16-B9",
                    "=C17-B9",
                    "=C19-B19",
                    "Explain material fee movement.",
                ),
            ],
            [
                ("Fee base restriction", '=IF(B6>0,"REVIEW EXCLUSIONS","OK")'),
                ("Cap binding", '=IF(B16>B8,"CAP APPLIES","OK")'),
            ],
        ),
        "cam-estimate-forecaster": (
            [
                ("Current-year CAM budget", 450000, '"$"#,##0'),
                ("Tax growth assumption", 0.06, "0.0%"),
                ("Insurance growth assumption", 0.12, "0.0%"),
                ("Contract growth assumption", 0.04, "0.0%"),
                ("Variable cost share", 0.42, "0.0%"),
                ("Occupancy assumption", 0.93, "0.0%"),
                ("Tenant pro-rata share", 0.085, "0.00%"),
            ],
            [
                (
                    "Base forecast",
                    "=B5*(1+B8)",
                    "=B5*(1+B6+B7+B8)/3",
                    "=C16-B16",
                    "Blend budget category assumptions.",
                ),
                (
                    "Variable occupancy adjustment",
                    "=B16*B9",
                    "=B16*B9/B10",
                    "=C17-B17",
                    "Adjust variable costs for occupancy.",
                ),
                (
                    "Tenant annual estimate",
                    "=B16*B11",
                    "=C16*B11",
                    "=C18-B18",
                    "Translate forecast to tenant estimate.",
                ),
                (
                    "Monthly estimate",
                    "=B18/12",
                    "=C18/12",
                    "=C19-B19",
                    "Use in tenant estimate letter.",
                ),
            ],
            [
                ("High estimate change", '=IF(ABS(C18-B18)/B18>0.1,"REVIEW","OK")'),
                ("Low occupancy", '=IF(B10<0.9,"REVIEW","OK")'),
            ],
        ),
        "boma-remeasurement-impact": (
            [
                ("Current building RSF", 100000, "#,##0"),
                ("BOMA 2024 building RSF", 106500, "#,##0"),
                ("Tenant current RSF", 8500, "#,##0"),
                ("Tenant BOMA 2024 RSF", 9100, "#,##0"),
                ("Recoverable CAM pool", 450000, '"$"#,##0'),
                ("NOI cap rate", 0.065, "0.0%"),
                ("Implementation cost", 35000, '"$"#,##0'),
            ],
            [
                (
                    "Current share",
                    "=B7/B5",
                    "=B8/B6",
                    "=C16-B16",
                    "Compare tenant share before and after remeasurement.",
                ),
                (
                    "Recoverable CAM",
                    "=B9*B16",
                    "=B9*C16",
                    "=C17-B17",
                    "Estimate annual recovery movement.",
                ),
                (
                    "Value impact",
                    "=B17/B10",
                    "=C17/B10",
                    "=C18-B18",
                    "Cap-rate translation of recovery change.",
                ),
                (
                    "Payback",
                    "=B11/MAX(C17-B17,1)",
                    "=B11/MAX(C17-B17,1)",
                    "=C19-B19",
                    "Years to recover implementation cost.",
                ),
            ],
            [
                ("Area increase", '=IF(C16-B16>0.005,"REVIEW NOTICE","OK")'),
                ("Payback", '=IF(B19>2,"REVIEW","OK")'),
            ],
        ),
        "cam-cap-calculator": (
            [
                ("Base-year controllable CAM", 390000, '"$"#,##0'),
                ("Current controllable CAM", 455000, '"$"#,##0'),
                ("Annual cap rate", 0.05, "0.0%"),
                ("Lease year", 3, "0"),
                ("Tenant pro-rata share", 0.085, "0.00%"),
                ("Non-controllable CAM", 120000, '"$"#,##0'),
                ("Prior-year billed amount", 42000, '"$"#,##0'),
            ],
            [
                (
                    "Cumulative cap limit",
                    "=B5*(1+B7)",
                    "=B5*(1+B7)^B8",
                    "=C16-B16",
                    "Compare annual and compounded cap treatment.",
                ),
                (
                    "Allowed controllable CAM",
                    "=MIN(B6,B16)",
                    "=MIN(B6,C16)",
                    "=C17-B17",
                    "Apply cap before adding non-controllables.",
                ),
                (
                    "Tenant charge",
                    "=(B17+B10)*B9",
                    "=(C17+B10)*B9",
                    "=C18-B18",
                    "Tenant-level cap effect.",
                ),
                (
                    "Variance vs prior bill",
                    "=B18-B11",
                    "=C18-B11",
                    "=C19-B19",
                    "Explain changes from prior-year billed amount.",
                ),
            ],
            [
                ("Cap exceeded", '=IF(B6>C16,"CAP CREDIT","OK")'),
                ("Large variance", '=IF(ABS(C19)/B11>0.1,"REVIEW","OK")'),
            ],
        ),
        "base-year-escalation": (
            [
                ("Base-year CAM", 390000, '"$"#,##0'),
                ("Current-year CAM", 455000, '"$"#,##0'),
                ("Tenant pro-rata share", 0.085, "0.00%"),
                ("Base occupancy", 0.88, "0.0%"),
                ("Current occupancy", 0.94, "0.0%"),
                ("Allowed gross-up occupancy", 0.95, "0.0%"),
                ("Prior-year excess bill", 4800, '"$"#,##0'),
            ],
            [
                (
                    "Base year adjusted",
                    "=B5",
                    "=B5/B8*B10",
                    "=C16-B16",
                    "Normalize base year if lease permits.",
                ),
                (
                    "Current year adjusted",
                    "=B6",
                    "=B6/B9*B10",
                    "=C17-B17",
                    "Normalize current variable expenses.",
                ),
                (
                    "Excess expense",
                    "=MAX(B17-B16,0)",
                    "=MAX(C17-C16,0)",
                    "=C18-B18",
                    "Bill only excess over base.",
                ),
                (
                    "Tenant excess share",
                    "=B18*B7",
                    "=C18*B7",
                    "=C19-B19",
                    "Tenant-level amount due.",
                ),
            ],
            [
                ("Negative excess", '=IF(C18=0,"NO BILL","OK")'),
                ("Large change", '=IF(ABS(C19-B11)/MAX(B11,1)>0.1,"REVIEW","OK")'),
            ],
        ),
        "reconciliation-statement-generator": (
            [
                ("Estimates billed", 385000, '"$"#,##0'),
                ("Actual recoverable CAM", 450000, '"$"#,##0'),
                ("Tenant pro-rata share", 0.085, "0.00%"),
                ("Tenant estimates paid", 32725, '"$"#,##0'),
                ("Prior-year credit applied", 1500, '"$"#,##0'),
                ("Admin fee", 0.05, "0.0%"),
                ("Material variance threshold", 0.1, "0.0%"),
            ],
            [
                (
                    "Tenant actual share",
                    "=B6*B7",
                    "=B6*B7*(1+B10)",
                    "=C16-B16",
                    "Show CAM share and admin treatment.",
                ),
                (
                    "Net balance due",
                    "=B16-B8-B9",
                    "=C16-B8-B9",
                    "=C17-B17",
                    "Tenant-facing amount due or credit.",
                ),
                (
                    "Pool variance",
                    "=B6-B5",
                    "=(B6-B5)/B5",
                    "=C18-B18",
                    "Use percentage for variance note trigger.",
                ),
                (
                    "Statement total",
                    "=B17",
                    "=C17",
                    "=C19-B19",
                    "Tie final statement to calculation support.",
                ),
            ],
            [
                ("Variance note required", '=IF(ABS(C18)>B11,"WRITE NOTE","OK")'),
                ("Credit balance", '=IF(C17<0,"CREDIT","DUE")'),
            ],
        ),
        "recovery-gap-analyzer": (
            [
                ("Recoverable per lease", 450000, '"$"#,##0'),
                ("Actually billed", 410000, '"$"#,##0'),
                ("Tenant share", 0.085, "0.00%"),
                ("Number of similar tenants", 12, "0"),
                ("NOI cap rate", 0.065, "0.0%"),
                ("Collection probability", 0.75, "0.0%"),
                ("Review cost", 8500, '"$"#,##0'),
            ],
            [
                (
                    "Portfolio leakage",
                    "=B5-B6",
                    "=(B5-B6)*B8",
                    "=C16-B16",
                    "Scale one-tenant gap across similar tenants.",
                ),
                (
                    "Recoverable tenant gap",
                    "=B16*B7",
                    "=C16*B7*B10",
                    "=C17-B17",
                    "Probability-weight expected recovery.",
                ),
                (
                    "NOI value impact",
                    "=B17/B9",
                    "=C17/B9",
                    "=C18-B18",
                    "Estimate valuation effect.",
                ),
                (
                    "Net opportunity",
                    "=B17-B11",
                    "=C17-B11",
                    "=C19-B19",
                    "Compare recovery to review cost.",
                ),
            ],
            [
                ("Material gap", '=IF(C19>25000,"PRIORITIZE","MONITOR")'),
                ("Low probability", '=IF(B10<0.6,"REVIEW","OK")'),
            ],
        ),
        "pro-rata-calculator": (
            [
                ("Tenant RSF", 8500, "#,##0"),
                ("Building RSF", 100000, "#,##0"),
                ("Occupied RSF", 92000, "#,##0"),
                ("Excluded anchor RSF", 12000, "#,##0"),
                ("Recoverable CAM pool", 450000, '"$"#,##0'),
                ("Lease-stated share", 0.085, "0.00%"),
                ("Prior-year tenant charge", 36000, '"$"#,##0'),
            ],
            [
                (
                    "Building denominator share",
                    "=B5/B6",
                    "=B5/(B6-B8)",
                    "=C16-B16",
                    "Compare full building and anchor-excluded denominator.",
                ),
                (
                    "Occupied denominator share",
                    "=B5/B7",
                    "=B5/(B7-B8)",
                    "=C17-B17",
                    "Use only if lease permits occupancy denominator.",
                ),
                (
                    "Tenant charge",
                    "=B9*B10",
                    "=B9*C16",
                    "=C18-B18",
                    "Compare lease-stated share and calculated share.",
                ),
                (
                    "Variance vs prior year",
                    "=B18-B11",
                    "=C18-B11",
                    "=C19-B19",
                    "Explain denominator-driven changes.",
                ),
            ],
            [
                ("Share mismatch", '=IF(ABS(B10-C16)>0.001,"REVIEW","OK")'),
                ("Anchor exclusion", '=IF(B8>0,"VERIFY LEASE","OK")'),
            ],
        ),
        "hcad-tax-normalizer": (
            [
                ("Prior tax assessment", 8500000, '"$"#,##0'),
                ("Current HCAD assessment", 9300000, '"$"#,##0'),
                ("Tax rate", 0.0215, "0.0000%"),
                ("Protest reduction estimate", 0.08, "0.0%"),
                ("Tenant share", 0.085, "0.00%"),
                ("Prior-year tax recoveries", 152000, '"$"#,##0'),
                ("Escrow/estimate billed", 160000, '"$"#,##0'),
            ],
            [
                (
                    "Prior tax amount",
                    "=B5*B7",
                    "=B6*B7",
                    "=C16-B16",
                    "Calculate raw assessment-driven change.",
                ),
                (
                    "Normalized current tax",
                    "=C16",
                    "=B6*(1-B8)*B7",
                    "=C17-B17",
                    "Model expected protest effect.",
                ),
                (
                    "Tenant tax share",
                    "=B17*B9",
                    "=C17*B9",
                    "=C18-B18",
                    "Tenant impact after normalization.",
                ),
                (
                    "True-up vs estimate",
                    "=B17-B11",
                    "=C17-B11",
                    "=C19-B19",
                    "Compare actual/normalized tax to billed estimate.",
                ),
            ],
            [
                ("Assessment jump", '=IF((B6-B5)/B5>0.1,"REVIEW","OK")'),
                ("Protest assumption", '=IF(B8>0,"DOCUMENT","OK")'),
            ],
        ),
        "noi-impact-calculator": (
            [
                ("Annual CAM leakage", 40000, '"$"#,##0'),
                ("Recoverable cap credit", 12000, '"$"#,##0'),
                ("Collection probability", 0.75, "0.0%"),
                ("NOI cap rate", 0.065, "0.0%"),
                ("Properties affected", 5, "0"),
                ("Implementation cost", 25000, '"$"#,##0'),
                ("Annual software/process cost", 12000, '"$"#,##0'),
            ],
            [
                (
                    "Gross NOI opportunity",
                    "=B5+B6",
                    "=(B5+B6)*B9",
                    "=C16-B16",
                    "Scale by affected properties.",
                ),
                (
                    "Probability-weighted NOI",
                    "=B16*B7",
                    "=C16*B7",
                    "=C17-B17",
                    "Risk-adjust recoverable opportunity.",
                ),
                (
                    "Value impact",
                    "=B17/B8",
                    "=C17/B8",
                    "=C18-B18",
                    "Cap-rate valuation estimate.",
                ),
                (
                    "Net first-year benefit",
                    "=B17-B10-B11",
                    "=C17-B10-B11",
                    "=C19-B19",
                    "Subtract implementation and process cost.",
                ),
            ],
            [
                ("Positive ROI", '=IF(C19>0,"GO","REVIEW")'),
                ("Large value impact", '=IF(C18>500000,"PRIORITIZE","OK")'),
            ],
        ),
        "cam-leakage-estimator": (
            [
                ("Unmapped GL expense", 22000, '"$"#,##0'),
                ("Missed admin fee base", 180000, '"$"#,##0'),
                ("Admin fee rate", 0.15, "0.0%"),
                ("Incorrect denominator loss", 15000, '"$"#,##0'),
                ("Unbilled tax/insurance", 28000, '"$"#,##0'),
                ("Collection probability", 0.7, "0.0%"),
                ("Tenant challenge reserve", 0.15, "0.0%"),
            ],
            [
                (
                    "Direct leakage",
                    "=B5+B8+B9",
                    "=B5+B8+B9",
                    "=C16-B16",
                    "Unmapped and unbilled costs.",
                ),
                (
                    "Admin leakage",
                    "=B6*B7",
                    "=B6*B7",
                    "=C17-B17",
                    "Fee base recovery gap.",
                ),
                (
                    "Expected recovery",
                    "=(B16+B17)*B10",
                    "=(C16+C17)*B10*(1-B11)",
                    "=C18-B18",
                    "Risk-adjust for challenge reserve.",
                ),
                (
                    "Total leakage estimate",
                    "=B16+B17",
                    "=C16+C17",
                    "=C19-B19",
                    "Gross exposure before probability weighting.",
                ),
            ],
            [
                ("High leakage", '=IF(C19>50000,"PRIORITIZE","OK")'),
                ("Challenge reserve", '=IF(B11>0.1,"DOCUMENT","OK")'),
            ],
        ),
        "cam-reconciliation-excel": (
            [
                ("GL recoverable total", 450000, '"$"#,##0'),
                ("Excluded GL total", 65000, '"$"#,##0'),
                ("Tenant pro-rata share", 0.085, "0.00%"),
                ("Estimates paid", 32725, '"$"#,##0'),
                ("Admin fee rate", 0.05, "0.0%"),
                ("Cap credit", 2500, '"$"#,##0'),
                ("Prior correction", 1200, '"$"#,##0'),
            ],
            [
                (
                    "Net recoverable pool",
                    "=B5-B6",
                    "=B5-B6",
                    "=C16-B16",
                    "Remove excluded GL before allocation.",
                ),
                (
                    "Tenant CAM share",
                    "=B16*B7",
                    "=C16*B7",
                    "=C17-B17",
                    "Allocate net recoverable pool.",
                ),
                (
                    "Admin and cap adjustment",
                    "=B17*B9-B10",
                    "=C17*B9-B10",
                    "=C18-B18",
                    "Apply admin and cap credit.",
                ),
                (
                    "Balance due / credit",
                    "=B17+B18-B8-B11",
                    "=C17+C18-B8-B11",
                    "=C19-B19",
                    "Final tenant statement amount.",
                ),
            ],
            [
                ("Excluded costs present", '=IF(B6>0,"VERIFY MAPPING","OK")'),
                ("Credit result", '=IF(C19<0,"CREDIT","DUE")'),
            ],
        ),
        "cumulative-cap-bank-calculator": (
            [
                ("Prior-year controllable CAM", 390000, '"$"#,##0'),
                ("Current-year controllable CAM", 425000, '"$"#,##0'),
                ("Annual cap rate", 0.05, "0.0%"),
                ("Lease year", 4, "0"),
                ("Previously unused cap bank", 18000, '"$"#,##0'),
                ("Tenant pro-rata share", 0.085, "0.00%"),
                ("Non-controllable CAM", 95000, '"$"#,##0'),
            ],
            [
                (
                    "Current cap limit",
                    "=B5*(1+B7)",
                    "=B5*(1+B7)^B8",
                    "=C16-B16",
                    "Compare annual and cumulative cap math.",
                ),
                (
                    "Unused cap capacity",
                    "=MAX(B16-B6,0)",
                    "=MAX(C16-B6,0)",
                    "=C17-B17",
                    "Track available cap bank capacity.",
                ),
                (
                    "Recoverable with bank",
                    "=MIN(B6,B16+B9)",
                    "=MIN(B6,C16+B9)",
                    "=C18-B18",
                    "Apply prior unused cap bank if lease permits.",
                ),
                (
                    "Tenant impact",
                    "=(B18+B11)*B10",
                    "=(C18+B11)*B10",
                    "=C19-B19",
                    "Tenant charge after cap bank treatment.",
                ),
            ],
            [
                ("Bank used", '=IF(C18>C16,"DOCUMENT","OK")'),
                ("Cap exceeded", '=IF(B6>C16,"CAP REVIEW","OK")'),
            ],
        ),
        "cam-recovery-ratio-worksheet": (
            [
                ("Recoverable CAM pool", 450000, '"$"#,##0'),
                ("CAM billed to tenants", 408000, '"$"#,##0'),
                ("Target recovery ratio", 0.97, "0.0%"),
                ("Known non-recoverable leakage", 12000, '"$"#,##0'),
                ("Number of properties", 8, "0"),
                ("NOI cap rate", 0.065, "0.0%"),
                ("Collection probability", 0.75, "0.0%"),
            ],
            [
                (
                    "Actual recovery ratio",
                    "=B6/B5",
                    "=(B6+B8)/B5",
                    "=C16-B16",
                    "Compare raw and adjusted recovery.",
                ),
                (
                    "Recovery shortfall",
                    "=MAX(B5*B7-B6,0)",
                    "=MAX(B5*B7-(B6+B8),0)",
                    "=C17-B17",
                    "Target-dollar gap.",
                ),
                (
                    "Portfolio opportunity",
                    "=B17*B9",
                    "=C17*B9*B11",
                    "=C18-B18",
                    "Scale across similar properties.",
                ),
                (
                    "Value impact",
                    "=B18/B10",
                    "=C18/B10",
                    "=C19-B19",
                    "Translate recovery gap into value.",
                ),
            ],
            [
                ("Below target", '=IF(B16<B7,"REVIEW","OK")'),
                ("Material opportunity", '=IF(C18>50000,"PRIORITIZE","OK")'),
            ],
        ),
        "property-tax-appeal-recovery-calculator": (
            [
                ("Original tax bill", 210000, '"$"#,##0'),
                ("Appealed tax bill", 185000, '"$"#,##0'),
                ("Tenant tax share", 0.085, "0.00%"),
                ("Taxes previously billed", 205000, '"$"#,##0'),
                ("Refund timing reserve", 0.1, "0.0%"),
                ("Admin handling fee", 0.02, "0.0%"),
                ("NOI cap rate", 0.065, "0.0%"),
            ],
            [
                (
                    "Tax savings",
                    "=B5-B6",
                    "=B5-B6",
                    "=C16-B16",
                    "Gross appeal savings.",
                ),
                (
                    "Tenant credit",
                    "=B16*B7",
                    "=B16*B7*(1-B9)",
                    "=C17-B17",
                    "Reserve timing or collection uncertainty.",
                ),
                (
                    "Admin recovery",
                    "=B16*B10",
                    "=C16*B10",
                    "=C18-B18",
                    "Only if lease permits appeal cost recovery.",
                ),
                (
                    "NOI / value impact",
                    "=(B16-B17+B18)/B11",
                    "=(C16-C17+C18)/B11",
                    "=C19-B19",
                    "Estimate retained savings value.",
                ),
            ],
            [
                ("Refund due", '=IF(C17>0,"CREDIT TRACKING","OK")'),
                ("Admin fee", '=IF(B10>0,"VERIFY LEASE","OK")'),
            ],
        ),
        "lease-clause-extraction-matrix": (
            [
                ("Lease clauses reviewed", 12, "0"),
                ("Missing source citations", 2, "0"),
                ("Conflicting abstract fields", 3, "0"),
                ("High-risk recovery clauses", 4, "0"),
                ("Tenants in population", 18, "0"),
                ("Average CAM exposure", 42000, '"$"#,##0'),
                ("Review completion", 0.72, "0.0%"),
            ],
            [
                (
                    "Citation completeness",
                    "=(B5-B6)/B5",
                    "=B11",
                    "=C16-B16",
                    "Every extracted clause needs source text.",
                ),
                (
                    "Conflict rate",
                    "=B7/B5",
                    "=B7/B5",
                    "=C17-B17",
                    "Abstract drift requiring review.",
                ),
                (
                    "Exposure estimate",
                    "=B8*B10",
                    "=B8*B10*B9",
                    "=C18-B18",
                    "High-risk clauses scaled to tenant exposure.",
                ),
                (
                    "Completion gap",
                    "=1-B11",
                    "=1-B11",
                    "=C19-B19",
                    "Remaining review work.",
                ),
            ],
            [
                ("Missing citations", '=IF(B6>0,"REVIEW","OK")'),
                ("Completion", '=IF(B11<0.9,"INCOMPLETE","OK")'),
            ],
        ),
    }
    if spec.slug in models:
        return models[spec.slug]
    return (
        [
            ("Annual recoverable CAM pool", 450000, '"$"#,##0'),
            ("Tenant rentable square feet", 8500, "#,##0"),
            ("Building rentable square feet", 100000, "#,##0"),
            ("Excluded / anchor square feet", 12000, "#,##0"),
            ("Occupancy or recovery factor", 0.92, "0.0%"),
            ("Admin fee or cap rate", 0.15, "0.0%"),
            ("Prior-year baseline", 390000, '"$"#,##0'),
        ],
        [
            (
                "Tenant share",
                "=B6/(B7-B8)",
                "=B16",
                "=C16-B16",
                "Verify denominator against lease.",
            ),
            (
                "Recoverable charge",
                "=B5*B16",
                "=B5*C16*B9",
                "=C17-B17",
                "Variable costs only when grossing up.",
            ),
            (
                "Admin or cap impact",
                "=B17*B10",
                "=C17*B10",
                "=C18-B18",
                "Confirm fee base excludes restricted items.",
            ),
            (
                "Variance from baseline",
                "=C17-B11",
                "=C17+B18-B11",
                "=C19-B19",
                "Explain material changes before sending.",
            ),
        ],
        [
            ("Excluded denominator share", '=IF(B8/B7>0.15,"REVIEW","OK")'),
            ("Large CAM variance", '=IF(ABS(C19)/B11>0.1,"REVIEW","OK")'),
        ],
    )


def build_xlsx(spec: AssetSpec, out_path: Path) -> None:
    wb = Workbook()
    wb.properties.creator = "CapVeri"
    wb.properties.lastModifiedBy = "CapVeri"
    wb.properties.created = FIXED_BUILD_TIME
    wb.properties.modified = FIXED_BUILD_TIME
    wb.properties.title = spec.name
    wb.properties.subject = spec.theme
    ws = wb.active
    ws.title = "Model"
    for col, width in {"A": 34, "B": 18, "C": 18, "D": 18, "E": 20}.items():
        ws.column_dimensions[col].width = width
    ws.merge_cells("A1:E1")
    cell_header(ws["A1"], f"{spec.name} | CapVeri")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:E2")
    cell_label(ws["A2"], spec.use_case)
    ws["A2"].font = Font(italic=True, color="475569")

    ws.merge_cells("A4:E4")
    cell_header(ws["A4"], "Inputs", TEAL)
    inputs, rows, flags = xlsx_model_for(spec)
    for index, (label, value, fmt) in enumerate(inputs, start=5):
        cell_label(ws[f"A{index}"], label)
        cell_input(ws[f"B{index}"], value, fmt)

    ws.merge_cells("A14:E14")
    cell_header(ws["A14"], "Worked Example", TEAL)
    headers = ["Metric", "Base", "Adjusted", "Variance", "Review note"]
    for col_num, header in enumerate(headers, start=1):
        cell_header(ws.cell(row=15, column=col_num), header, GRAY)
        ws.cell(row=15, column=col_num).font = Font(bold=True, color="111827")
    for row_num, row in enumerate(rows, start=16):
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            if isinstance(value, str) and value.startswith("="):
                fmt = "0.00%" if row_num == 16 else '"$"#,##0;[Red]-"$"#,##0'
                cell_formula(cell, value, fmt)
            else:
                cell_label(cell, value, bold=col_num == 1)

    ws.merge_cells("A22:E22")
    cell_header(ws["A22"], "Risk Flags", AMBER)
    flags = [
        *flags,
        (
            "Documentation packet",
            '="Attach lease excerpt, GL export, rent roll, and reviewer sign-off"',
        ),
    ]
    for index, (label, formula) in enumerate(flags, start=23):
        cell_label(ws[f"A{index}"], label, bold=True)
        cell_formula(ws[f"B{index}"], formula)
        ws.merge_cells(start_row=index, start_column=2, end_row=index, end_column=5)

    notes = wb.create_sheet("Instructions")
    notes.column_dimensions["A"].width = 96
    instruction_lines = [
        f"{spec.name} - How to use",
        "Yellow cells are editable inputs. Blue cells calculate automatically.",
        spec.use_case,
        "Replace the sample numbers with your property-level CAM pool, tenant area, building area, exclusions, and lease-specific factor.",
        "Use the worked example section to document the position you plan to send to tenants.",
        "Keep the source GL export, rent roll, signed lease excerpt, and reviewer approval with the final reconciliation package.",
        "This workbook is an operating aid, not legal advice. Confirm final positions against the governing lease.",
        f"Need portfolio-scale CAM validation? Start a 30-day trial at {REGISTER_URL}.",
    ]
    for row_num, text in enumerate(instruction_lines, start=1):
        cell_label(notes[f"A{row_num}"], text, bold=row_num == 1)
        notes[f"A{row_num}"].alignment = Alignment(wrap_text=True, vertical="top")
    mobile = wb.create_sheet("Phone Summary", 0)
    mobile.column_dimensions["A"].width = 28
    mobile.column_dimensions["B"].width = 78
    cell_header(mobile["A1"], "Phone Summary | CapVeri", TEAL)
    mobile.merge_cells("A1:B1")
    mobile.row_dimensions[1].height = 28
    mobile_rows = [
        ("Best use", spec.use_case),
        ("First check", "Confirm the lease allows the recovery method before modeling dollars."),
        ("Second check", "Tie the GL export, rent roll, and prior-year support to the same period."),
        ("Desktop step", "Use the Model tab to enter property numbers and review the formulas."),
        ("Evidence", "Keep the lease excerpt, source reports, reviewer initials, and final archive copy."),
        ("Next step", f"Run the same check in CapVeri during a 30-day trial: {REGISTER_URL}"),
    ]
    for row_num, (label, value) in enumerate(mobile_rows, start=3):
        cell_label(mobile[f"A{row_num}"], label, bold=True)
        cell_label(mobile[f"B{row_num}"], value)
        mobile.row_dimensions[row_num].height = 34
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical="center",
                    wrap_text=True,
                )
        sheet.protection.sheet = True
        sheet.protection.password = "capveri"
    for row in range(1, 28):
        ws.row_dimensions[row].height = 22
    wb.save(out_path)
    normalize_xlsx_zip(out_path)


def normalize_xlsx_zip(path: Path) -> None:
    temp_path = path.with_suffix(f"{path.suffix}.tmp")
    with zipfile.ZipFile(path, "r") as source:
        with zipfile.ZipFile(
            temp_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as target:
            for name in source.namelist():
                info = zipfile.ZipInfo(name, FIXED_ZIP_TIME)
                info.compress_type = zipfile.ZIP_DEFLATED
                info.external_attr = 0o600 << 16
                payload = source.read(name)
                if name == "docProps/core.xml":
                    payload = re.sub(
                        rb"<dcterms:modified\b[^>]*>.*?</dcterms:modified>",
                        b'<dcterms:modified xmlns:dcterms="http://purl.org/dc/terms/" '
                        b'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
                        b'xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:modified>',
                        payload,
                    )
                target.writestr(info, payload)
    temp_path.replace(path)


def verify_xlsx(path: Path, spec: AssetSpec) -> None:
    if not path.exists() or path.stat().st_size < 5_000:
        raise RuntimeError(f"Missing or too-small workbook: {path}")
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        text = "\n".join(
            str(cell.value)
            for sheet in wb.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        if len(wb.sheetnames) < 2:
            raise RuntimeError(f"Workbook needs model and instructions tabs: {path}")
        formulas = [
            cell.value
            for row in wb["Model"].iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        if len(formulas) < 8:
            raise RuntimeError(f"Workbook needs meaningful formulas: {path}")
        missing = [phrase for phrase in spec.required_phrases if phrase not in text]
        if missing:
            raise RuntimeError(
                f"Workbook text verification failed for {path}: {missing}"
            )
    finally:
        wb.close()


def verify_pdf(path: Path, spec: AssetSpec) -> None:
    if not path.exists() or path.stat().st_size < 3_500:
        raise RuntimeError(f"Missing or too-small PDF: {path}")
    reader = PdfReader(str(path))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    normalized_text = text.lower()
    missing = [
        phrase
        for phrase in spec.required_phrases
        if phrase.lower() not in normalized_text
    ]
    if missing:
        raise RuntimeError(f"PDF text verification failed for {path}: {missing}")


def verify_asset(path: Path, spec: AssetSpec) -> None:
    if spec.kind == "xlsx":
        verify_xlsx(path, spec)
    else:
        verify_pdf(path, spec)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def xlsx_semantic_fingerprint(path: Path) -> tuple[tuple[str, tuple[tuple[str, ...], ...]], ...]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        sheets: list[tuple[str, tuple[tuple[str, ...], ...]]] = []
        for sheet in wb.worksheets:
            rows: list[tuple[str, ...]] = []
            for row in sheet.iter_rows():
                values = tuple("" if cell.value is None else str(cell.value) for cell in row)
                if any(values):
                    rows.append(values)
            sheets.append((sheet.title, tuple(rows)))
        return tuple(sheets)
    finally:
        wb.close()


def remote_matches_source(destination: Path, source: Path, spec: AssetSpec) -> bool:
    if destination.stat().st_size == source.stat().st_size and sha256_file(destination) == sha256_file(source):
        return True
    if spec.kind != "xlsx":
        return False
    return xlsx_semantic_fingerprint(destination) == xlsx_semantic_fingerprint(source)


def write_manifest(remote_verified: bool) -> None:
    manifest = {
        "bucket": R2_BUCKET,
        "remote_verified": remote_verified,
        "assets": [
            {
                "slug": spec.slug,
                "filename": spec.filename,
                "storage_path": spec.storage_path,
                "kind": spec.kind,
                "sha256": sha256_file(GENERATED_DIR / spec.filename),
                "size_bytes": (GENERATED_DIR / spec.filename).stat().st_size,
                "required_phrases": list(spec.required_phrases),
            }
            for spec in ASSETS
        ],
    }
    MANIFEST_PATH.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n")


def run_wrangler(args: list[str], *, allow_failure: bool = False) -> None:
    npx = shutil.which("npx") or shutil.which("npx.cmd")
    if not npx:
        raise RuntimeError("npx is required to upload or verify R2 objects.")
    command = [npx, "wrangler", *args]
    result = subprocess.run(
        command,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        check=False,
    )
    if result.stdout:
        print_console_safe(result.stdout)
    if result.stderr:
        print_console_safe(result.stderr)
    if result.returncode != 0 and not allow_failure:
        raise RuntimeError(f"Wrangler command failed: {' '.join(command)}")


def print_console_safe(text: str) -> None:
    output = text.strip()
    if not output:
        return
    encoding = sys.stdout.encoding or "utf-8"
    safe = output.encode(encoding, errors="replace").decode(encoding)
    print(safe)


def upload_asset(spec: AssetSpec) -> None:
    run_wrangler(
        [
            "r2",
            "object",
            "delete",
            f"{R2_BUCKET}/{spec.storage_path}",
            "--remote",
        ],
        allow_failure=True,
    )
    run_wrangler(
        [
            "r2",
            "object",
            "put",
            f"{R2_BUCKET}/{spec.storage_path}",
            "--file",
            str(GENERATED_DIR / spec.filename),
            "--remote",
        ]
    )


def verify_remote_assets() -> None:
    with tempfile.TemporaryDirectory(prefix="capveri-lead-magnets-") as temp_dir:
        temp = Path(temp_dir)
        for spec in ASSETS:
            destination = temp / spec.filename
            source = GENERATED_DIR / spec.filename
            run_wrangler(
                [
                    "r2",
                    "object",
                    "get",
                    f"{R2_BUCKET}/{spec.storage_path}",
                    "--file",
                    str(destination),
                    "--remote",
                ]
            )
            verify_asset(destination, spec)
            if not remote_matches_source(destination, source, spec):
                raise RuntimeError(f"Remote checksum mismatch for {spec.filename}")


def build_all() -> None:
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    DOCS_ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    for spec in ASSETS:
        path = GENERATED_DIR / spec.filename
        if spec.kind == "xlsx":
            build_xlsx(spec, path)
        else:
            build_pdf(spec, path)
        verify_asset(path, spec)
        shutil.copy2(path, DOCS_ASSETS_DIR / spec.filename)
        print(f"Built and verified {spec.filename}: {path.stat().st_size} bytes")
    write_manifest(remote_verified=False)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--upload", action="store_true", help="Upload all generated assets to R2."
    )
    parser.add_argument(
        "--verify-remote",
        action="store_true",
        help="Fetch and verify all known R2 assets.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    build_all()
    if args.upload:
        for spec in ASSETS:
            upload_asset(spec)
    if args.upload or args.verify_remote:
        verify_remote_assets()
        write_manifest(remote_verified=True)
        print(f"Verified {len(ASSETS)} remote R2 lead magnets in {R2_BUCKET}.")


if __name__ == "__main__":
    main()
