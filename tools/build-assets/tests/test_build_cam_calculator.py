"""Tests for CAM Gross-Up Calculator Excel generator."""

import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).parent.parent.parent.parent
SCRIPT_PATH = Path(__file__).parent.parent / "build_cam_calculator.py"
OUTPUT_PATH = REPO_ROOT / "docs" / "assets" / "cam-gross-up-calculator.xlsx"

EXPECTED_SHEET_NAMES = [
    "Instructions",
    "Calculator",
    "Tenant Allocation",
    "Scenario Comparison",
    "Sample Data",
]


@pytest.fixture(scope="module")
def generated_workbook():
    """Run the generator script and return the loaded workbook."""
    result = subprocess.run(
        [sys.executable, str(SCRIPT_PATH)],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, f"Script failed:\n{result.stdout}\n{result.stderr}"
    assert OUTPUT_PATH.exists(), f"Output file not created at {OUTPUT_PATH}"
    return load_workbook(str(OUTPUT_PATH))


def test_generates_xlsx_file(generated_workbook):
    """Script runs without errors and creates a valid xlsx file."""
    assert OUTPUT_PATH.exists()
    assert OUTPUT_PATH.suffix == ".xlsx"
    assert OUTPUT_PATH.stat().st_size > 1000  # Non-trivial file


def test_workbook_has_five_tabs(generated_workbook):
    """Workbook contains exactly the 5 required tabs in the correct order."""
    wb = generated_workbook
    assert wb.sheetnames == EXPECTED_SHEET_NAMES


def test_calculator_has_formulas_not_hardcoded_values(generated_workbook):
    """Key calculated cells on the Calculator tab contain formulas, not literals."""
    ws = generated_workbook["Calculator"]

    # Occupancy Rate cell (D6) should be a formula =C6/B6
    occupancy_cell = ws["D6"]
    assert occupancy_cell.value is not None
    assert str(occupancy_cell.value).startswith(
        "="
    ), f"Occupancy rate cell D6 should be a formula, got: {occupancy_cell.value!r}"

    # Gross-Up Multiplier cell (F6) should be a formula using MIN
    multiplier_cell = ws["F6"]
    assert multiplier_cell.value is not None
    assert str(multiplier_cell.value).startswith(
        "="
    ), f"Gross-Up Multiplier cell F6 should be a formula, got: {multiplier_cell.value!r}"
    assert (
        "MIN" in str(multiplier_cell.value).upper()
    ), f"Multiplier formula should use MIN(), got: {multiplier_cell.value!r}"


def test_sample_data_is_populated(generated_workbook):
    """Sample Data tab contains realistic pre-populated data."""
    ws = generated_workbook["Sample Data"]

    # Building name should be "Meridian Office Center"
    found_building = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "Meridian" in str(cell.value):
                found_building = True
                break

    assert found_building, "Sample Data tab should contain 'Meridian Office Center'"

    # GLA of 85,000 SF should be present
    found_gla = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 85000:
                found_gla = True
                break

    assert found_gla, "Sample Data tab should contain GLA value 85000"


def test_scenario_comparison_links_to_calculator(generated_workbook):
    """Scenario Comparison tab cells reference Calculator tab data."""
    ws = generated_workbook["Scenario Comparison"]

    # At least one cell should reference the Calculator tab
    found_reference = False
    for row in ws.iter_rows():
        for cell in row:
            if (
                cell.value
                and isinstance(cell.value, str)
                and "Calculator" in cell.value
            ):
                found_reference = True
                break

    assert (
        found_reference
    ), "Scenario Comparison tab should contain cell references to Calculator tab"


def test_cell_protection_applied(generated_workbook):
    """Input cells are unlocked; formula cells are locked."""
    ws = generated_workbook["Calculator"]

    # Sheet should have protection enabled
    assert ws.protection.sheet, "Calculator sheet should have protection enabled"

    # Input cell B4 (Building Name) should be unlocked
    input_cell = ws["B4"]
    assert (
        input_cell.protection.locked is False
    ), f"Input cell B4 should be unlocked, got locked={input_cell.protection.locked}"

    # Formula cell D6 (Occupancy Rate) should be locked
    formula_cell = ws["D6"]
    assert (
        formula_cell.protection.locked is True or formula_cell.protection.locked is None
    ), "Formula cell D6 should be locked"


def test_currency_formatting_on_expense_cells(generated_workbook):
    """Dollar amount cells in the expense table use currency number format."""
    ws = generated_workbook["Calculator"]

    # Check at least one expense amount cell has currency formatting
    # Expense amounts start around row 15, column D
    found_currency_format = False
    for row in ws.iter_rows(min_row=15, max_row=25, min_col=4, max_col=4):
        for cell in row:
            if cell.number_format and (
                "$" in cell.number_format or "0.00" in cell.number_format
            ):
                found_currency_format = True
                break

    assert (
        found_currency_format
    ), "Expense amount cells (col D, rows 15-25) should have currency number format"


def test_tenant_allocation_has_formulas(generated_workbook):
    """Tenant Allocation tab has pro-rata and CAM obligation formulas."""
    ws = generated_workbook["Tenant Allocation"]

    # Pro-rata share column (C) and CAM obligation column (D) should have formulas
    found_formula = False
    for row in ws.iter_rows(min_row=10, max_row=20):
        for cell in row:
            if (
                cell.value
                and isinstance(cell.value, str)
                and cell.value.startswith("=")
            ):
                found_formula = True
                break

    assert (
        found_formula
    ), "Tenant Allocation tab rows 10-20 should contain formula cells"
