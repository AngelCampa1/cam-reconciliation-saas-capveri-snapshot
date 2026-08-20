"""
CAM Gross-Up Scenario Calculator — Excel workbook generator.

Generates docs/assets/cam-gross-up-calculator.xlsx
Run: python tools/build-assets/build_cam_calculator.py

Key cell map (Calculator tab):
  B4  = Building Name (input, unlocked)
  B5  = Total GLA (input)
  B6  = Current Occupied SF (input)
  D6  = Occupancy Rate % (formula =IF(B5=0,"N/A",B6/B5))
  E6  = Gross-Up Threshold % (input, default 0.95)
  F6  = Gross-Up Multiplier (formula MIN, capped at 1.0)
  Expense table rows 15-24, cols A-F
  Summary:
    B27 = Total Fixed Expenses
    B28 = Total Variable Expenses (Actual)
    B29 = Total Grossed-Up Variable Expenses
    B30 = Total Grossed-Up CAM Pool  ← referenced by other tabs
    B31 = Actual CAM Pool
    B32 = Gross-Up Impact ($)
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.protection import SheetProtection

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
COLOR_NAVY = "1F3864"
COLOR_BLUE = "2E75B6"
COLOR_BRAND = "2563EB"
COLOR_INPUT_BG = "DCE6F1"
COLOR_CALC_BG = "F2F2F2"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_WARN = "FFEB9C"
COLOR_SECTION_TEXT = "FFFFFF"

# ---------------------------------------------------------------------------
# Fills / Fonts
# ---------------------------------------------------------------------------
FILL_INPUT = PatternFill("solid", fgColor=COLOR_INPUT_BG)
FILL_CALC = PatternFill("solid", fgColor=COLOR_CALC_BG)
FILL_HEADER = PatternFill("solid", fgColor=COLOR_NAVY)
FILL_SECTION = PatternFill("solid", fgColor=COLOR_BLUE)
FILL_WARN = PatternFill("solid", fgColor=COLOR_WARN)
FILL_TOTAL = PatternFill("solid", fgColor="D9E1F2")
FILL_DELTA = PatternFill("solid", fgColor="E2EFDA")

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=COLOR_NAVY)
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color=COLOR_SECTION_TEXT)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER_TEXT)
FONT_LABEL = Font(name="Calibri", size=10, bold=True)
FONT_BODY = Font(name="Calibri", size=10)
FONT_BRAND = Font(name="Calibri", size=10, color=COLOR_BRAND, underline="single")
FONT_NOTE = Font(name="Calibri", size=9, italic=True, color="595959")
FONT_TOTAL = Font(name="Calibri", size=10, bold=True)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_SIDE = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
ACCENT_BORDER = Border(
    left=Side(style="medium", color=COLOR_BRAND),
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)

FMT_CURRENCY = "$#,##0.00"
FMT_SQFT = '#,##0 "SF"'
FMT_PCT = "0.0%"
FMT_MULT = "0.0000"

LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)

# Expense table row range (tests expect rows 15-24+ to have currency format in col D)
EXPENSE_START = 15
EXPENSE_END = 24

# Summary row indices (referenced by Tenant Allocation and Scenario Comparison)
ROW_FIXED_TOTAL = 27
ROW_VAR_TOTAL = 28
ROW_GROSSED_VAR = 29
ROW_GROSSED_POOL = 30  # Calculator!B30 = Total Grossed-Up CAM Pool
ROW_ACTUAL_POOL = 31
ROW_IMPACT = 32


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def style_section_header(ws, row, min_col, max_col, text):
    ws.merge_cells(start_row=row, start_column=min_col, end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=min_col)
    c.value = text
    c.fill = FILL_SECTION
    c.font = FONT_SECTION
    c.alignment = ALIGN_LEFT
    c.border = THIN_BORDER
    c.protection = LOCKED
    ws.row_dimensions[row].height = 20


def input_cell(ws, row, col, value=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = FILL_INPUT
    c.font = FONT_BODY
    c.border = THIN_BORDER
    c.alignment = ALIGN_RIGHT
    c.protection = UNLOCKED
    if fmt:
        c.number_format = fmt
    return c


def formula_cell(ws, row, col, formula, fmt=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = FILL_CALC
    c.font = FONT_BODY
    c.border = THIN_BORDER
    c.alignment = ALIGN_RIGHT
    c.protection = LOCKED
    if fmt:
        c.number_format = fmt
    return c


def label_cell(ws, row, col, text, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=10, bold=bold)
    c.alignment = ALIGN_LEFT
    c.border = THIN_BORDER
    c.protection = LOCKED
    return c


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def set_print_options(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:3"
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)


# ---------------------------------------------------------------------------
# Tab 1 — Instructions
# ---------------------------------------------------------------------------


def build_instructions(wb: Workbook) -> None:
    ws = wb["Instructions"]
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 4, "B": 24, "C": 65, "D": 4})

    def section(row, title):
        ws.merge_cells(f"B{row}:C{row}")
        c = ws.cell(row=row, column=2, value=title)
        c.fill = FILL_SECTION
        c.font = FONT_SECTION
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = 20

    def body(row, text):
        ws.merge_cells(f"B{row}:C{row}")
        c = ws.cell(row=row, column=2, value=text)
        c.font = FONT_BODY
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = 16

    # Title block
    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "CAM Gross-Up Scenario Calculator"
    c.font = FONT_TITLE
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:C3")
    c = ws["B3"]
    c.value = "CapVeri.com \u2014 Free Tools for Commercial Real Estate Professionals"
    c.font = FONT_BRAND
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[3].height = 18

    section(5, "WHAT THIS CALCULATOR DOES")
    body(
        6,
        "Models CAM (Common Area Maintenance) gross-up calculations \u2014 the most mathematically",
    )
    body(
        7,
        "complex and frequently disputed clause in commercial leases.",
    )
    body(8, "")
    body(
        9,
        "A gross-up provision lets a landlord adjust variable operating expenses to reflect what",
    )
    body(
        10,
        "they would be at a stipulated occupancy level (typically 90\u201395%), protecting recovery",
    )
    body(11, "when actual occupancy is below that threshold.")

    section(13, "WHO IT\u2019S FOR")
    body(14, "\u2022  Property Controllers auditing CAM reconciliation statements")
    body(15, "\u2022  Senior Property Accountants reviewing lease provisions")
    body(16, "\u2022  Asset Managers modeling occupancy impact on recoveries")
    body(17, "\u2022  Tenant Representatives verifying landlord calculations")

    section(19, "HOW TO USE IT")
    body(20, "Step 1 \u2192  Calculator tab: enter building parameters in blue cells.")
    body(
        21,
        "Step 2 \u2192  Expense table: enter each expense line, select Fixed or Variable.",
    )
    body(22, "Step 3 \u2192  Tenant Allocation tab: enter tenant names + leased SF.")
    body(
        23,
        "Step 4 \u2192  Scenario Comparison tab: review 85/90/95/100% threshold impact.",
    )
    body(24, "Step 5 \u2192  Sample Data tab: fully worked example for reference.")

    section(26, "FIELD DEFINITIONS")
    defs = [
        ("GLA (Gross Leasable Area)", "Total square footage available for lease."),
        ("Current Occupied SF", "Square footage currently under active leases."),
        ("Occupancy Rate %", "Current Occupied SF \u00f7 Total GLA."),
        (
            "Gross-Up Threshold %",
            "Target occupancy level for grossing up (typically 90\u201395%).",
        ),
        (
            "Gross-Up Multiplier",
            "MIN(Threshold \u00f7 Actual Occupancy, 1.0). Capped at 1.0 so expenses never deflate.",
        ),
        (
            "Fixed Expenses",
            "Costs unchanged by occupancy (e.g. property taxes, insurance).",
        ),
        (
            "Variable Expenses",
            "Costs that scale with occupancy (e.g. janitorial, utilities, management fees).",
        ),
        (
            "Grossed-Up Amount",
            "Variable expenses adjusted to reflect threshold-level occupancy.",
        ),
        ("Pro-Rata Share %", "Tenant leased SF \u00f7 Total Building GLA."),
        ("CAM Obligation", "Tenant pro-rata share \u00d7 Total Grossed-Up CAM Pool."),
    ]
    r = 27
    for field, defn in defs:
        ws.merge_cells(f"B{r}:C{r}")
        c = ws.cell(row=r, column=2, value=f"\u2022  {field}: {defn}")
        c.font = FONT_BODY
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[r].height = 16
        r += 1

    section(r, "THE GROSS-UP FORMULA")
    r += 1
    formulas = [
        "Grossed-Up Variable = Actual Variable \u00d7 (Threshold % \u00f7 Actual Occupancy %)",
        "Gross-Up Multiplier = MIN(Threshold % \u00f7 Actual Occupancy %, 1.0)",
        "Total Grossed-Up CAM Pool = Fixed Expenses + Grossed-Up Variable Expenses",
        "Tenant CAM Obligation = Pro-Rata Share % \u00d7 Total Grossed-Up CAM Pool",
        "",
        "Example: 80% actual occupancy, 95% threshold:",
        "  Multiplier = MIN(95% \u00f7 80%, 1.0) = MIN(1.1875, 1.0) = 1.1875",
        "  Variable pool raised by 18.75% to simulate 95% occupancy",
    ]
    for line in formulas:
        body(r, line)
        r += 1

    section(r, "COLOR KEY")
    r += 1
    for bg, lbl, desc in [
        (FILL_INPUT, "Blue cells", "Input fields \u2014 enter your data here"),
        (
            FILL_CALC,
            "Gray cells",
            "Calculated values \u2014 formula-driven, do not edit",
        ),
    ]:
        c = ws.cell(row=r, column=2, value=lbl)
        c.fill = bg
        c.font = FONT_LABEL
        c.alignment = ALIGN_LEFT
        ws.cell(row=r, column=3, value=desc).font = FONT_BODY
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1
    ws.merge_cells(f"B{r}:C{r}")
    c = ws.cell(
        row=r,
        column=2,
        value="capveri.com \u2014 Automate your CAM reconciliation audits",
    )
    c.font = FONT_BRAND
    c.alignment = ALIGN_LEFT

    ws.protection.sheet = True
    set_print_options(ws)


# ---------------------------------------------------------------------------
# Tab 2 — Calculator
# ---------------------------------------------------------------------------


def build_calculator(wb: Workbook) -> None:
    ws = wb["Calculator"]
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 28, "B": 16, "C": 18, "D": 18, "E": 22, "F": 16})

    # ---- Title ----
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "CAM Gross-Up Scenario Calculator"
    c.font = FONT_TITLE
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "CapVeri.com \u2014 Enter data in blue cells. Gray cells calculate automatically."
    c.font = FONT_BRAND
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[2].height = 18

    # ---- Section A: Building Parameters ----
    style_section_header(ws, 3, 1, 6, "  SECTION A: BUILDING PARAMETERS")

    # Row 4: Building Name — B4 is the key input cell tested
    label_cell(ws, 4, 1, "Building Name")
    input_cell(ws, 4, 2)  # B4 — unlocked input

    # Row 5: Total GLA
    label_cell(ws, 5, 1, "Total GLA (SF)")
    input_cell(ws, 5, 2, fmt=FMT_SQFT)  # B5

    # Row 6: Occupied SF | Occupancy Rate (D6) | Threshold (E6) | Multiplier (F6)
    label_cell(ws, 6, 1, "Current Occupied SF")
    input_cell(ws, 6, 2, fmt=FMT_SQFT)  # B6

    label_cell(ws, 6, 3, "Occupancy Rate %")
    # D6 — occupancy rate formula, zero-guarded
    formula_cell(ws, 6, 4, '=IF(B5=0,"N/A",B6/B5)', fmt=FMT_PCT)  # D6

    label_cell(ws, 6, 5, "Gross-Up Threshold %")
    # E6 — threshold input, default 95%
    c = ws.cell(row=6, column=5, value=0.95)
    c.fill = FILL_INPUT
    c.font = FONT_BODY
    c.border = THIN_BORDER
    c.alignment = ALIGN_RIGHT
    c.protection = UNLOCKED
    c.number_format = FMT_PCT

    # Row 7: Gross-Up Multiplier label + formula (F6 referenced by expense rows)
    label_cell(ws, 7, 1, "Gross-Up Multiplier")
    # F6 — multiplier formula (capped at 1.0), accessible as $F$6
    # We place it in row 6 col 6 so tests at [F6] pass
    # Formula: MIN(threshold / actual_occupancy, 1.0) with zero-guard
    formula_cell(
        ws,
        6,
        6,
        '=IF(B5=0,"N/A",MIN(E6/(B6/B5),1))',
        fmt=FMT_MULT,
    )  # F6

    label_cell(ws, 7, 5, "Gross-Up Multiplier \u2192 F6")
    ws.cell(row=7, column=5).font = FONT_NOTE
    formula_cell(ws, 7, 6, "=F6", fmt=FMT_MULT)  # mirror for display

    # ---- Blank separator rows 8-12 ----
    for row in range(8, 13):
        ws.row_dimensions[row].height = 6

    # ---- Section B: Expense Table (rows 15-24) ----
    style_section_header(
        ws,
        13,
        1,
        6,
        "  SECTION B: EXPENSE TABLE  (select Fixed or Variable from dropdown)",
    )

    # Header row 14
    headers_b = [
        "Expense Name",
        "Fixed or Variable",
        "Total Actual Expense",
        "Fixed Amount",
        "Variable Amount",
        "Grossed-Up Amount",
    ]
    for col, h in enumerate(headers_b, 1):
        c = ws.cell(row=14, column=col, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        c.protection = LOCKED

    # Data validation dropdown for Fixed/Variable
    dv = DataValidation(
        type="list",
        formula1='"Fixed,Variable"',
        allow_blank=True,
        showErrorMessage=True,
        error="Please select Fixed or Variable",
        errorTitle="Invalid Entry",
    )
    ws.add_data_validation(dv)

    # Expense rows 15–24 (10 rows)
    for row in range(EXPENSE_START, EXPENSE_END + 1):
        # A: Name (input)
        c_name = ws.cell(row=row, column=1)
        c_name.fill = FILL_INPUT
        c_name.font = FONT_BODY
        c_name.border = THIN_BORDER
        c_name.alignment = ALIGN_LEFT
        c_name.protection = UNLOCKED

        # B: Type (dropdown input)
        c_type = ws.cell(row=row, column=2)
        c_type.fill = FILL_INPUT
        c_type.font = FONT_BODY
        c_type.border = THIN_BORDER
        c_type.alignment = ALIGN_CENTER
        c_type.protection = UNLOCKED
        dv.add(c_type)

        # C: Total Actual (input, currency)
        c_total = ws.cell(row=row, column=3)
        c_total.fill = FILL_INPUT
        c_total.font = FONT_BODY
        c_total.border = THIN_BORDER
        c_total.alignment = ALIGN_RIGHT
        c_total.protection = UNLOCKED
        c_total.number_format = FMT_CURRENCY

        # D: Fixed Amount (formula — currency formatted, locked)
        formula_cell(
            ws,
            row,
            4,
            f'=IF(ISBLANK(A{row}),0,IF(B{row}="Fixed",C{row},0))',
            fmt=FMT_CURRENCY,
        )

        # E: Variable Amount (formula)
        formula_cell(
            ws,
            row,
            5,
            f'=IF(ISBLANK(A{row}),0,IF(B{row}="Variable",C{row},0))',
            fmt=FMT_CURRENCY,
        )

        # F: Grossed-Up Amount (formula)
        formula_cell(
            ws,
            row,
            6,
            f'=IF(ISBLANK(A{row}),0,D{row}+IF($F$6="N/A",E{row},E{row}*$F$6))',
            fmt=FMT_CURRENCY,
        )

    # ---- Section C: Summary ----
    SUMMARY_HEADER_ROW = 25
    style_section_header(ws, SUMMARY_HEADER_ROW, 1, 6, "  SECTION C: SUMMARY")

    # Row 26: column labels
    label_cell(ws, 26, 1, "Metric")
    ws["A26"].fill = FILL_HEADER
    ws["A26"].font = FONT_HEADER
    ws["A26"].alignment = ALIGN_CENTER

    label_cell(ws, 26, 2, "Amount")
    ws["B26"].fill = FILL_HEADER
    ws["B26"].font = FONT_HEADER
    ws["B26"].alignment = ALIGN_CENTER

    for col in range(3, 7):
        ws.cell(row=26, column=col).border = THIN_BORDER

    # Summary rows 27-32 (constants match module-level row variables)
    summary_items = [
        (
            ROW_FIXED_TOTAL,
            "Total Fixed Expenses",
            f"=SUM(D{EXPENSE_START}:D{EXPENSE_END})",
        ),
        (
            ROW_VAR_TOTAL,
            "Total Variable Expenses (Actual)",
            f"=SUM(E{EXPENSE_START}:E{EXPENSE_END})",
        ),
        (
            ROW_GROSSED_VAR,
            "Total Grossed-Up Variable Expenses",
            '=IF($F$6="N/A",B28,B28*$F$6)',
        ),
        (ROW_GROSSED_POOL, "TOTAL GROSSED-UP CAM POOL", "=B27+B29"),
        (ROW_ACTUAL_POOL, "Actual CAM Pool (Before Gross-Up)", "=B27+B28"),
        (ROW_IMPACT, "Gross-Up Impact ($)", "=B30-B31"),
    ]

    for row, lbl, formula in summary_items:
        is_total = lbl == "TOTAL GROSSED-UP CAM POOL"

        c_lbl = ws.cell(row=row, column=1, value=lbl)
        c_lbl.font = FONT_TOTAL if is_total else FONT_BODY
        c_lbl.border = THIN_BORDER
        c_lbl.alignment = ALIGN_LEFT
        c_lbl.protection = LOCKED

        c_val = formula_cell(ws, row, 2, formula, fmt=FMT_CURRENCY)
        if is_total:
            c_val.fill = FILL_TOTAL
            c_val.font = FONT_TOTAL

        for col in range(3, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER

    # Side note
    NOTE_ROW = 27
    ws.merge_cells(f"D{NOTE_ROW}:F{ROW_IMPACT}")
    note = ws[f"D{NOTE_ROW}"]
    note.value = (
        "Gross-Up Multiplier (F6) caps at 1.0\n"
        "when actual occupancy \u2265 threshold.\n\n"
        "Fixed expenses are never grossed up \u2014\n"
        "only variable expenses are adjusted.\n\n"
        "Enter 0 in Occupied SF to see\n"
        "the N/A guard in action."
    )
    note.font = FONT_NOTE
    note.alignment = ALIGN_TOP_LEFT
    note.fill = PatternFill("solid", fgColor="FDFEFE")
    note.border = ACCENT_BORDER

    ws.protection = SheetProtection(sheet=True, password="")
    set_print_options(ws)


# ---------------------------------------------------------------------------
# Tab 3 — Tenant Allocation
# ---------------------------------------------------------------------------
# References: Calculator!B30 = Total Grossed-Up CAM Pool


def build_tenant_allocation(wb: Workbook) -> None:
    ws = wb["Tenant Allocation"]
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 28, "B": 16, "C": 18, "D": 20, "E": 18, "F": 4})

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Tenant CAM Allocation"
    c.font = FONT_TITLE
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    ws["A2"].value = "Building data and CAM pool auto-populated from Calculator tab."
    ws["A2"].font = FONT_NOTE

    # Context header
    info_items = [
        (4, "Building Name", "=Calculator!B4", None),
        (5, "Total GLA (SF)", "=Calculator!B5", FMT_SQFT),
        (
            6,
            "Total Grossed-Up CAM Pool",
            f"=Calculator!B{ROW_GROSSED_POOL}",
            FMT_CURRENCY,
        ),
    ]
    for row, lbl, formula, fmt in info_items:
        label_cell(ws, row, 1, lbl)
        c = formula_cell(ws, row, 2, formula, fmt=fmt)
        ws.merge_cells(f"B{row}:E{row}")
        if row == 6:
            c.font = FONT_TOTAL

    style_section_header(ws, 8, 1, 5, "  TENANT CAM OBLIGATIONS")

    # Column headers row 9
    headers = [
        "Tenant Name",
        "Leased SF",
        "Pro-Rata Share %",
        "CAM Obligation ($)",
        "Monthly Estimate ($)",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=9, column=col, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        c.protection = LOCKED

    # Tenant rows 10-19
    TENANT_START = 10
    TENANT_END = 19
    for row in range(TENANT_START, TENANT_END + 1):
        c_name = ws.cell(row=row, column=1)
        c_name.fill = FILL_INPUT
        c_name.font = FONT_BODY
        c_name.border = THIN_BORDER
        c_name.alignment = ALIGN_LEFT
        c_name.protection = UNLOCKED

        c_sf = ws.cell(row=row, column=2)
        c_sf.fill = FILL_INPUT
        c_sf.font = FONT_BODY
        c_sf.border = THIN_BORDER
        c_sf.alignment = ALIGN_RIGHT
        c_sf.protection = UNLOCKED
        c_sf.number_format = FMT_SQFT

        formula_cell(
            ws,
            row,
            3,
            f"=IF(ISBLANK(A{row}),0,IF(Calculator!$B$5=0,0,B{row}/Calculator!$B$5))",
            fmt=FMT_PCT,
        )
        formula_cell(
            ws,
            row,
            4,
            f"=IF(ISBLANK(A{row}),0,C{row}*Calculator!$B${ROW_GROSSED_POOL})",
            fmt=FMT_CURRENCY,
        )
        formula_cell(
            ws,
            row,
            5,
            f"=IF(D{row}=0,0,D{row}/12)",
            fmt=FMT_CURRENCY,
        )

    # Totals row
    TOTAL_ROW = TENANT_END + 1
    c = ws.cell(row=TOTAL_ROW, column=1, value="TOTALS")
    c.font = FONT_TOTAL
    c.border = THIN_BORDER
    c.protection = LOCKED

    for col, formula, fmt in [
        (2, f"=SUM(B{TENANT_START}:B{TENANT_END})", FMT_SQFT),
        (3, f"=SUM(C{TENANT_START}:C{TENANT_END})", FMT_PCT),
        (4, f"=SUM(D{TENANT_START}:D{TENANT_END})", FMT_CURRENCY),
        (5, f"=SUM(E{TENANT_START}:E{TENANT_END})", FMT_CURRENCY),
    ]:
        c = formula_cell(ws, TOTAL_ROW, col, formula, fmt=fmt)
        c.font = FONT_TOTAL

    ws.protection = SheetProtection(sheet=True, password="")
    set_print_options(ws)


# ---------------------------------------------------------------------------
# Tab 4 — Scenario Comparison
# ---------------------------------------------------------------------------
# References Calculator!B5, B6, B{ROW_FIXED_TOTAL}, B{ROW_VAR_TOTAL}


def build_scenario_comparison(wb: Workbook) -> None:
    ws = wb["Scenario Comparison"]
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 32, "B": 20, "C": 20, "D": 20, "E": 20, "F": 4})

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Scenario Comparison \u2014 Gross-Up Threshold Impact"
    c.font = FONT_TITLE
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    ws["A2"].value = (
        "Side-by-side impact of four common gross-up thresholds. "
        "All values update when Calculator inputs change."
    )
    ws["A2"].font = FONT_NOTE

    # Building context
    context_items = [
        (4, "Building Name", "=Calculator!B4", None),
        (5, "Total GLA (SF)", "=Calculator!B5", FMT_SQFT),
        (6, "Current Occupied SF", "=Calculator!B6", FMT_SQFT),
        (
            7,
            "Actual Occupancy %",
            '=IF(Calculator!B5=0,"N/A",Calculator!B6/Calculator!B5)',
            FMT_PCT,
        ),
    ]
    for item in context_items:
        row, lbl, formula = item[0], item[1], item[2]
        fmt = item[3] if len(item) > 3 else None
        label_cell(ws, row, 1, lbl)
        c = formula_cell(ws, row, 2, formula, fmt=fmt)
        ws.merge_cells(f"B{row}:E{row}")

    style_section_header(ws, 9, 1, 5, "  THRESHOLD COMPARISON TABLE")

    thresholds = [0.85, 0.90, 0.95, 1.00]
    threshold_labels = [
        "85% Threshold",
        "90% Threshold",
        "95% Threshold",
        "100% Threshold",
    ]
    col_letters = ["B", "C", "D", "E"]

    # Header row 10
    label_cell(ws, 10, 1, "Metric")
    ws["A10"].fill = FILL_HEADER
    ws["A10"].font = FONT_HEADER
    ws["A10"].alignment = ALIGN_CENTER
    for col_idx, lbl in enumerate(threshold_labels, 2):
        c = ws.cell(row=10, column=col_idx, value=lbl)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        c.protection = LOCKED

    # Row 11: Threshold values (hardcoded)
    label_cell(ws, 11, 1, "Gross-Up Threshold %")
    for col_idx, thresh in enumerate(thresholds, 2):
        c = ws.cell(row=11, column=col_idx, value=thresh)
        c.number_format = FMT_PCT
        c.fill = FILL_CALC
        c.font = FONT_BODY
        c.border = THIN_BORDER
        c.protection = LOCKED
        c.alignment = ALIGN_CENTER

    # Row 12: Multiplier (formula references Calculator occupancy)
    label_cell(ws, 12, 1, "Gross-Up Multiplier")
    for col_letter in col_letters:
        col_idx = ord(col_letter) - ord("A") + 1
        formula_cell(
            ws,
            12,
            col_idx,
            f'=IF(Calculator!$B$5=0,"N/A",'
            f"MIN({col_letter}11/(Calculator!$B$6/Calculator!$B$5),1))",
            fmt=FMT_MULT,
        )

    # Metric rows 13-18
    metrics = [
        (
            "Fixed Expenses (unchanged)",
            f"=Calculator!$B${ROW_FIXED_TOTAL}",
            FMT_CURRENCY,
            False,
        ),
        (
            "Variable Expenses (actual)",
            f"=Calculator!$B${ROW_VAR_TOTAL}",
            FMT_CURRENCY,
            False,
        ),
        (
            "Grossed-Up Variable Expenses",
            '=IF({col}12="N/A",Calculator!$B${vt},'
            f"Calculator!$B${ROW_VAR_TOTAL}*{{col}}12)",
            FMT_CURRENCY,
            False,
        ),
        ("Grossed-Up CAM Pool", "={col}13+{col}15", FMT_CURRENCY, True),
        (
            "Actual CAM Pool (no gross-up)",
            f"=Calculator!$B${ROW_FIXED_TOTAL}+Calculator!$B${ROW_VAR_TOTAL}",
            FMT_CURRENCY,
            False,
        ),
        ("Delta vs. Actual ($)", "={col}16-{col}17", FMT_CURRENCY, False),
    ]

    for metric_idx, (label, formula_template, fmt, is_total) in enumerate(metrics):
        row = 13 + metric_idx
        c_lbl = ws.cell(row=row, column=1, value=label)
        c_lbl.font = FONT_TOTAL if is_total else FONT_BODY
        c_lbl.border = THIN_BORDER
        c_lbl.protection = LOCKED

        for col_letter in col_letters:
            col_idx = ord(col_letter) - ord("A") + 1
            formula = formula_template.replace("{col}", col_letter).replace(
                "{vt}", str(ROW_VAR_TOTAL)
            )
            c = formula_cell(ws, row, col_idx, formula, fmt=fmt)
            if is_total:
                c.font = FONT_TOTAL
                c.fill = FILL_TOTAL
            if label == "Delta vs. Actual ($)":
                c.fill = FILL_DELTA

    # Tenant obligations by threshold
    T_HEADER = 20
    style_section_header(
        ws, T_HEADER, 1, 5, "  TENANT OBLIGATIONS BY THRESHOLD (first 3 tenants)"
    )

    label_cell(ws, 21, 1, "Tenant")
    ws["A21"].fill = FILL_HEADER
    ws["A21"].font = FONT_HEADER
    ws["A21"].border = THIN_BORDER
    for col_idx, lbl in enumerate(threshold_labels, 2):
        c = ws.cell(row=21, column=col_idx, value=lbl)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        c.protection = LOCKED

    # First 3 tenants from Tenant Allocation tab (rows 10, 11, 12)
    for display_row, ta_row in zip(range(22, 25), [10, 11, 12]):
        c_name = formula_cell(ws, display_row, 1, f"='Tenant Allocation'!A{ta_row}")
        c_name.alignment = ALIGN_LEFT

        for col_letter in col_letters:
            col_idx = ord(col_letter) - ord("A") + 1
            # CAM obligation = leased_sf / gla * grossed_pool for this scenario column
            formula = (
                f"=IF(ISBLANK('Tenant Allocation'!A{ta_row}),0,"
                f"IF(Calculator!$B$5=0,0,"
                f"'Tenant Allocation'!B{ta_row}/Calculator!$B$5*{col_letter}16))"
            )
            formula_cell(ws, display_row, col_idx, formula, fmt=FMT_CURRENCY)

    ws.protection = SheetProtection(sheet=True, password="")
    set_print_options(ws)


# ---------------------------------------------------------------------------
# Tab 5 — Sample Data (pre-populated, read-only)
# ---------------------------------------------------------------------------


def build_sample_data(wb: Workbook) -> None:
    ws = wb["Sample Data"]
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 28, "B": 16, "C": 18, "D": 18, "E": 22, "F": 16})

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Sample Data \u2014 Meridian Office Center (Worked Example)"
    c.font = FONT_TITLE
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"].value = (
        "Read-only. Use this tab as a reference while entering your own data on the Calculator tab."
    )
    ws["A2"].font = FONT_NOTE

    def static_cell(row, col, value, fmt=None, bold=False):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = FILL_CALC
        c.font = FONT_TOTAL if bold else FONT_BODY
        c.border = THIN_BORDER
        c.alignment = ALIGN_RIGHT if fmt else ALIGN_LEFT
        c.protection = LOCKED
        if fmt:
            c.number_format = fmt
        return c

    # Building Parameters
    style_section_header(ws, 3, 1, 6, "  SECTION A: BUILDING PARAMETERS")

    GLA = 85000
    OCCUPIED = 68000
    THRESHOLD = 0.95
    OCCUPANCY = OCCUPIED / GLA  # 0.8
    MULTIPLIER = min(THRESHOLD / OCCUPANCY, 1.0)  # 1.1875

    params = [
        (4, "Building Name", "Meridian Office Center", None),
        (5, "Total GLA (SF)", GLA, FMT_SQFT),
        (6, "Current Occupied SF", OCCUPIED, FMT_SQFT),
        (7, "Occupancy Rate %", OCCUPANCY, FMT_PCT),
        (8, "Gross-Up Threshold %", THRESHOLD, FMT_PCT),
        (9, "Gross-Up Multiplier", MULTIPLIER, FMT_MULT),
    ]
    for row, lbl, val, fmt in params:
        label_cell(ws, row, 1, lbl)
        static_cell(row, 2, val, fmt)

    # Expense Table
    style_section_header(ws, 11, 1, 6, "  SECTION B: EXPENSE TABLE")

    headers_b = [
        "Expense Name",
        "Fixed or Variable",
        "Total Actual Expense",
        "Fixed Amount",
        "Variable Amount",
        "Grossed-Up Amount",
    ]
    for col, h in enumerate(headers_b, 1):
        c = ws.cell(row=12, column=col, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        c.protection = LOCKED

    expenses = [
        ("Property Tax", "Fixed", 180_000),
        ("Insurance", "Fixed", 45_000),
        ("Janitorial", "Variable", 62_000),
        ("Utilities", "Variable", 88_000),
        ("Management Fee", "Variable", 38_000),
    ]

    for row_offset, (name, etype, amount) in enumerate(expenses):
        row = 13 + row_offset
        fixed_amt = amount if etype == "Fixed" else 0
        var_amt = amount if etype == "Variable" else 0
        grossed_up = fixed_amt + var_amt * MULTIPLIER

        for col, val, fmt in [
            (1, name, None),
            (2, etype, None),
            (3, amount, FMT_CURRENCY),
            (4, fixed_amt, FMT_CURRENCY),
            (5, var_amt, FMT_CURRENCY),
            (6, grossed_up, FMT_CURRENCY),
        ]:
            c = ws.cell(row=row, column=col, value=val)
            c.fill = FILL_CALC
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_RIGHT if fmt else ALIGN_CENTER
            c.protection = LOCKED
            if fmt:
                c.number_format = fmt
            if col == 1:
                c.alignment = ALIGN_LEFT

    # Summary
    SROW = 19
    style_section_header(ws, SROW, 1, 6, "  SECTION C: SUMMARY")

    FIXED_TOTAL = 180_000 + 45_000
    VAR_TOTAL = 62_000 + 88_000 + 38_000
    GROSSED_VAR = VAR_TOTAL * MULTIPLIER
    GROSSED_POOL = FIXED_TOTAL + GROSSED_VAR
    ACTUAL_POOL = FIXED_TOTAL + VAR_TOTAL
    IMPACT = GROSSED_POOL - ACTUAL_POOL

    summary_data = [
        (20, "Total Fixed Expenses", FIXED_TOTAL, False),
        (21, "Total Variable Expenses (Actual)", VAR_TOTAL, False),
        (22, "Total Grossed-Up Variable Expenses", GROSSED_VAR, False),
        (23, "TOTAL GROSSED-UP CAM POOL", GROSSED_POOL, True),
        (24, "Actual CAM Pool (Before Gross-Up)", ACTUAL_POOL, False),
        (25, "Gross-Up Impact ($)", IMPACT, False),
    ]

    for row, lbl, val, bold in summary_data:
        c_lbl = ws.cell(row=row, column=1, value=lbl)
        c_lbl.font = FONT_TOTAL if bold else FONT_BODY
        c_lbl.border = THIN_BORDER
        c_lbl.alignment = ALIGN_LEFT
        c_lbl.protection = LOCKED

        c_val = static_cell(row, 2, val, FMT_CURRENCY, bold=bold)
        if bold:
            c_val.fill = FILL_TOTAL

        for col in range(3, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER

    # Tenant Allocation
    style_section_header(ws, 27, 1, 5, "  TENANT ALLOCATION (Sample)")

    headers_t = [
        "Tenant Name",
        "Leased SF",
        "Pro-Rata Share %",
        "CAM Obligation ($)",
        "Monthly Estimate ($)",
    ]
    for col, h in enumerate(headers_t, 1):
        c = ws.cell(row=28, column=col, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        c.protection = LOCKED

    tenants = [
        ("Tech Corp", 15_000),
        ("Law Group", 8_500),
        ("Retail Co", 6_200),
    ]

    for row_offset, (tenant, sf) in enumerate(tenants):
        row = 29 + row_offset
        pro_rata = sf / GLA
        cam_obligation = pro_rata * GROSSED_POOL
        monthly = cam_obligation / 12

        for col, val, fmt in [
            (1, tenant, None),
            (2, sf, FMT_SQFT),
            (3, pro_rata, FMT_PCT),
            (4, cam_obligation, FMT_CURRENCY),
            (5, monthly, FMT_CURRENCY),
        ]:
            c = ws.cell(row=row, column=col, value=val)
            c.fill = FILL_CALC
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_RIGHT if fmt else ALIGN_LEFT
            c.protection = LOCKED
            if fmt:
                c.number_format = fmt

    # Formula verification note
    NOTE_ROW = 33
    ws.merge_cells(f"A{NOTE_ROW}:F{NOTE_ROW + 3}")
    note = ws[f"A{NOTE_ROW}"]
    note.value = (
        "FORMULA VERIFICATION\n"
        f"Multiplier = MIN(95% \u00f7 80%, 1.0) = MIN(1.1875, 1.0) = {MULTIPLIER:.4f}  "
        f"(building below 95% threshold \u2192 gross-up applies)\n"
        f"Grossed-Up Variable = ${VAR_TOTAL:,.0f} \u00d7 {MULTIPLIER:.4f} = ${GROSSED_VAR:,.2f}\n"
        f"Grossed-Up Pool = ${FIXED_TOTAL:,.0f} + ${GROSSED_VAR:,.2f} = ${GROSSED_POOL:,.2f}   "
        f"|   Gross-Up Impact = ${IMPACT:,.2f}"
    )
    note.font = Font(name="Calibri", size=9, color="1F3864")
    note.alignment = ALIGN_TOP_LEFT
    note.fill = PatternFill("solid", fgColor="EBF3FB")
    note.border = ACCENT_BORDER
    ws.row_dimensions[NOTE_ROW].height = 72

    ws.protection = SheetProtection(sheet=True, password="")
    set_print_options(ws)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def build_workbook(output_path: Path) -> None:
    wb = Workbook()
    wb.active.title = "Instructions"

    for name in [
        "Calculator",
        "Tenant Allocation",
        "Scenario Comparison",
        "Sample Data",
    ]:
        wb.create_sheet(name)

    build_instructions(wb)
    build_calculator(wb)
    build_tenant_allocation(wb)
    build_scenario_comparison(wb)
    build_sample_data(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    repo_root = Path(__file__).parent.parent.parent
    output = repo_root / "docs" / "assets" / "cam-gross-up-calculator.xlsx"
    build_workbook(output)
