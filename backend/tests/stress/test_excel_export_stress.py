"""Property-based serialization roundtrip stress for the audit Excel export.

``export_to_excel`` turns a historical-analysis report dict into the .xlsx file
a landlord hands a tenant during a CAM dispute. A silent corruption here — a
number written to the wrong cell, a row dropped, a Decimal coerced and
truncated, a sparse year defaulting wrong — would put an incorrect figure in
front of a tenant with the company's name on it. So the invariant we stress is
narrow and total: every value the report carries must reappear, unchanged, in
the cell the reader will look at.

We generate arbitrary report dicts (sparse year maps, empty categories, net
credits, no anomalies / many anomalies), serialize to bytes, load the workbook
back with openpyxl, and assert cell-by-cell that:

  * the Year-over-Year header is ``["Expense Pool", *years, "Variance %"]``;
  * each category row lands in order: name in col 1, the per-year amount (or 0
    for a year the category omits) in each year column, ``variance_percent/100``
    in the last column;
  * the totals row sits at ``len(categories) + 2`` with the per-year totals, and
    carries the aggregate ``(last-first)/first`` variance iff >=2 years and a
    positive first-year total;
  * the Anomalies sheet shows the "no anomalies" banner when empty, else one row
    per anomaly with severity upper-cased, type un-snaked/title-cased, the two
    money values, ``variance_percent/100``, and the explanation intact.

Only the value channel is checked — fonts/fills/widths are presentation, not
data, and are out of scope.

Run standalone:
    pytest tests/stress/test_excel_export_stress.py -q
"""

from __future__ import annotations

import math
from decimal import Decimal
from io import BytesIO
from typing import Any

import pytest
from hypothesis import HealthCheck, given, settings
from hypothesis import strategies as st
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from app.services.reports.excel_export import export_to_excel


def _sanitized(text: str) -> str:
    """Mirror the export's control-character stripping for roundtrip comparison."""
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    return ILLEGAL_CHARACTERS_RE.sub("", normalized)


STRESS = settings(
    max_examples=200,
    deadline=None,
    suppress_health_check=[HealthCheck.too_slow, HealthCheck.filter_too_much],
)

_SEVERITIES = ["critical", "warning", "info"]
_ANOMALY_TYPES = [
    "spike",
    "drop",
    "new_category",
    "missing_category",
    "trend_shift",
]


def _money() -> st.SearchStrategy[Decimal]:
    # Net credits (negative) included to stress the sign / variance paths.
    return st.decimals(
        min_value=Decimal("-500000"),
        max_value=Decimal("5000000"),
        places=2,
        allow_nan=False,
        allow_infinity=False,
    )


def _pct() -> st.SearchStrategy[Decimal]:
    return st.decimals(
        min_value=Decimal("-100"),
        max_value=Decimal("1000"),
        places=1,
        allow_nan=False,
        allow_infinity=False,
    )


@st.composite
def _report(draw) -> dict[str, Any]:
    years = sorted(
        draw(
            st.lists(
                st.integers(min_value=2015, max_value=2030),
                min_size=1,
                max_size=4,
                unique=True,
            )
        )
    )

    n_categories = draw(st.integers(min_value=0, max_value=5))
    categories: list[dict[str, Any]] = []
    for i in range(n_categories):
        # Sparse year coverage exercises the missing-year -> 0 default branch.
        present_years = draw(
            st.lists(
                st.sampled_from(years), min_size=0, max_size=len(years), unique=True
            )
        )
        amounts = {y: draw(_money()) for y in present_years}
        categories.append(
            {
                "name": f"pool-{i}",
                "amounts": amounts,
                "variance_percent": draw(_pct()),
            }
        )

    totals = [{"year": y, "total": draw(_money())} for y in years]

    n_anomalies = draw(st.integers(min_value=0, max_value=5))
    anomalies = [
        {
            "severity": draw(st.sampled_from(_SEVERITIES)),
            "pool_name": f"anomaly-pool-{j}",
            "anomaly_type": draw(st.sampled_from(_ANOMALY_TYPES)),
            "current_value": draw(_money()),
            "expected_value": draw(_money()),
            "variance_percent": draw(_pct()),
            "explanation": draw(st.text(min_size=0, max_size=40)),
        }
        for j in range(n_anomalies)
    ]

    return {
        "years_compared": years,
        "year_over_year_comparison": {
            "categories": categories,
            "totals": totals,
        },
        "anomalies": anomalies,
    }


def _close(actual: Any, expected: Decimal) -> bool:
    """openpyxl stores numbers as floats; compare with float tolerance."""
    return math.isclose(float(actual), float(expected), rel_tol=1e-9, abs_tol=1e-6)


@STRESS
@given(report=_report())
def test_yoy_sheet_roundtrips_every_value(report):
    years = report["years_compared"]
    categories = report["year_over_year_comparison"]["categories"]
    totals = report["year_over_year_comparison"]["totals"]

    wb = load_workbook(BytesIO(export_to_excel(report)))
    ws = wb["Year-over-Year Comparison"]

    # Header is exactly the documented shape.
    n_cols = len(years) + 2
    header = [ws.cell(row=1, column=c).value for c in range(1, n_cols + 1)]
    assert header == ["Expense Pool"] + [str(y) for y in years] + ["Variance %"]

    # Each category row, in order: name, per-year amount (0 if absent), pct/100.
    for offset, category in enumerate(categories):
        row = offset + 2
        assert ws.cell(row=row, column=1).value == _sanitized(category["name"])
        amounts = category["amounts"]
        for col_offset, year in enumerate(years):
            expected = amounts.get(year, Decimal("0"))
            assert _close(ws.cell(row=row, column=col_offset + 2).value, expected)
        assert _close(
            ws.cell(row=row, column=n_cols).value,
            category["variance_percent"] / 100,
        )

    # Totals row sits directly below the categories.
    totals_row = len(categories) + 2
    assert ws.cell(row=totals_row, column=1).value == "Total"
    totals_by_year = {t["year"]: t["total"] for t in totals}
    for col_offset, year in enumerate(years):
        assert _close(
            ws.cell(row=totals_row, column=col_offset + 2).value,
            totals_by_year[year],
        )

    # Aggregate variance appears iff >=2 years and a positive first-year total.
    variance_cell = ws.cell(row=totals_row, column=n_cols).value
    first_total = totals_by_year[years[0]]
    last_total = totals_by_year[years[-1]]
    if len(totals) >= 2 and first_total > 0:
        expected = ((last_total - first_total) / first_total) * 100 / 100
        assert _close(variance_cell, expected)
    else:
        # No total-variance written: the cell is empty (or, for a zero-category
        # report, this is also where category rows would have been — still empty).
        assert variance_cell is None


@STRESS
@given(report=_report())
def test_anomalies_sheet_roundtrips_every_value(report):
    anomalies = report["anomalies"]

    wb = load_workbook(BytesIO(export_to_excel(report)))
    ws = wb["Detected Anomalies"]

    expected_header = [
        "Severity",
        "Expense Pool",
        "Type",
        "Current",
        "Expected",
        "Variance %",
        "Explanation",
    ]
    header = [ws.cell(row=1, column=c).value for c in range(1, 8)]
    assert header == expected_header

    if not anomalies:
        assert ws.cell(row=2, column=1).value == (
            "No anomalies detected. All expense patterns appear normal."
        )
        return

    for offset, anomaly in enumerate(anomalies):
        row = offset + 2
        assert ws.cell(row=row, column=1).value == anomaly["severity"].upper()
        assert ws.cell(row=row, column=2).value == _sanitized(anomaly["pool_name"])
        assert ws.cell(row=row, column=3).value == _sanitized(
            anomaly["anomaly_type"].replace("_", " ").title()
        )
        assert _close(ws.cell(row=row, column=4).value, anomaly["current_value"])
        assert _close(ws.cell(row=row, column=5).value, anomaly["expected_value"])
        assert _close(
            ws.cell(row=row, column=6).value, anomaly["variance_percent"] / 100
        )
        # openpyxl returns None for an empty-string cell; both read as "no text".
        explanation = ws.cell(row=row, column=7).value
        assert (explanation or "") == _sanitized(anomaly["explanation"])


if __name__ == "__main__":
    pytest.main([__file__, "-q"])
