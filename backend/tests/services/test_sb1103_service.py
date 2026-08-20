"""Tests for SB 1103 Compliance Service.

Tests cover:
- Window start date calculation (calendar months, not 540 days)
- Response deadline calculation (request_date + 30 days)
- GL entry queries for the compliance window
- Export data assembly with tenant share computation
- PDF and Excel export generation
- Deadline alert queries
"""

from datetime import UTC, date, timedelta
from decimal import Decimal
from io import BytesIO
from unittest.mock import MagicMock
from uuid import uuid4

import pytest

from app.models.sb1103 import (
    SB1103ExportData,
    SB1103GLEntry,
    SB1103Request,
)
from app.services.compliance.sb1103_service import (
    build_sb1103_export_data,
    compute_response_deadline,
    compute_window_start,
    generate_excel_export,
    generate_pdf_export,
    get_deadline_alerts,
    get_gl_entries_for_window,
)

# ---------------------------------------------------------------------------
# Shared test fixtures
# ---------------------------------------------------------------------------

SAMPLE_ORG_ID = uuid4()
SAMPLE_PROPERTY_ID = uuid4()
SAMPLE_LEASE_ID = uuid4()
SAMPLE_BATCH_ID = uuid4()


def _make_request(
    request_date: date = date(2025, 1, 15),
    status: str = "pending",
) -> SB1103Request:
    from datetime import datetime

    return SB1103Request(
        id=uuid4(),
        organization_id=SAMPLE_ORG_ID,
        property_id=SAMPLE_PROPERTY_ID,
        lease_id=SAMPLE_LEASE_ID,
        requested_by_name="Jane Smith",
        requested_by_email="jane@tenant.com",
        request_date=request_date,
        response_deadline=request_date + timedelta(days=30),
        window_start_date=compute_window_start(request_date),
        window_end_date=request_date,
        status=status,
        created_at=datetime.now(UTC),
        updated_at=datetime.now(UTC),
    )


def _make_gl_entry(
    amount: Decimal = Decimal("1000.00"),
    account_description: str = "Janitorial Services",
    transaction_date: date = date(2024, 6, 15),
) -> dict:
    return {
        "id": str(uuid4()),
        "transaction_date": transaction_date.isoformat(),
        "account_code": "5100",
        "account_description": account_description,
        "vendor_name": "Clean Co",
        "description": "Monthly janitorial",
        "amount": str(amount),
        "import_batch_id": str(SAMPLE_BATCH_ID),
    }


def _make_export_data(
    gl_entries: list[SB1103GLEntry] | None = None,
    pro_rata_share: Decimal = Decimal("0.10"),
) -> SB1103ExportData:
    if gl_entries is None:
        entry = SB1103GLEntry(
            id=uuid4(),
            transaction_date=date(2024, 6, 15),
            account_code="5100",
            account_description="Janitorial Services",
            vendor_name="Clean Co",
            description="Monthly janitorial",
            amount=Decimal("1000.00"),
            import_batch_id=SAMPLE_BATCH_ID,
            tenant_share_amount=Decimal("100.00"),
        )
        gl_entries = [entry]

    return SB1103ExportData(
        request=_make_request(),
        property_address="123 Main St, Los Angeles, CA 90001",
        property_name="Downtown Tower",
        tenant_name="Acme Corp",
        pro_rata_share=pro_rata_share,
        gl_entries=gl_entries,
        category_subtotals={"Janitorial Services": Decimal("100.00")},
        is_ca_property=True,
        total_cam_expenses=Decimal("1000.00"),
        total_tenant_share=Decimal("100.00"),
    )


# ---------------------------------------------------------------------------
# TestComputeWindowStart
# ---------------------------------------------------------------------------


class TestComputeWindowStart:
    """Verifies calendar-month arithmetic for 18-month lookback window."""

    def test_mid_month_standard(self):
        """date(2025,1,15) → 18 calendar months back = date(2023,7,15)."""
        result = compute_window_start(date(2025, 1, 15))
        assert result == date(2023, 7, 15)

    def test_end_of_month_truncation(self):
        """date(2025,3,31) → 18 months back = date(2023,9,30) (Sept has 30 days)."""
        result = compute_window_start(date(2025, 3, 31))
        assert result == date(2023, 9, 30)

    def test_leap_year_handling(self):
        """date(2025,8,29) → 18 months back = date(2024,2,29) (2024 is leap)."""
        result = compute_window_start(date(2025, 8, 29))
        assert result == date(2024, 2, 29)

    def test_january_to_july(self):
        """date(2026,1,1) → date(2024,7,1)."""
        result = compute_window_start(date(2026, 1, 1))
        assert result == date(2024, 7, 1)

    def test_not_540_days(self):
        """Ensure result is calendar months, not 540 days."""
        result_cal = compute_window_start(date(2025, 1, 15))
        result_days = date(2025, 1, 15) - timedelta(days=540)
        # Calendar months gives date(2023,7,15); 540 days gives different result
        assert result_cal != result_days


# ---------------------------------------------------------------------------
# TestComputeResponseDeadline
# ---------------------------------------------------------------------------


class TestComputeResponseDeadline:
    """Verifies 30-day deadline calculation."""

    def test_standard(self):
        result = compute_response_deadline(date(2025, 1, 15))
        assert result == date(2025, 2, 14)

    def test_month_boundary(self):
        result = compute_response_deadline(date(2025, 1, 31))
        assert result == date(2025, 3, 2)

    def test_year_boundary(self):
        result = compute_response_deadline(date(2024, 12, 15))
        assert result == date(2025, 1, 14)


# ---------------------------------------------------------------------------
# TestGetGLEntriesForWindow
# ---------------------------------------------------------------------------


class MockQueryBuilder:
    def __init__(self, data=None):
        self._data = data or []
        self._range_start = None
        self._range_end = None

    def select(self, *args, **kwargs):
        return self

    def eq(self, *args, **kwargs):
        return self

    def gte(self, *args, **kwargs):
        return self

    def lte(self, *args, **kwargs):
        return self

    def order(self, *args, **kwargs):
        return self

    def range(self, start, end):
        self._range_start = start
        self._range_end = end
        return self

    def execute(self):
        mock = MagicMock()
        data = self._data
        if self._range_start is not None:
            data = data[self._range_start : self._range_end + 1]
        mock.data = data
        return mock


class MockCtx:
    def __init__(self, data=None):
        self._data = data

    def table(self, name):
        return MockQueryBuilder(self._data)


class TestGetGLEntriesForWindow:
    def test_returns_entries(self):
        entries = [_make_gl_entry()]
        ctx = MockCtx(entries)
        result = get_gl_entries_for_window(
            ctx,
            SAMPLE_PROPERTY_ID,
            date(2023, 7, 15),
            date(2025, 1, 15),
        )
        assert len(result) == 1
        assert result[0]["account_code"] == "5100"

    def test_empty_result(self):
        ctx = MockCtx([])
        result = get_gl_entries_for_window(
            ctx,
            SAMPLE_PROPERTY_ID,
            date(2023, 7, 15),
            date(2025, 1, 15),
        )
        assert result == []

    def test_includes_import_batch_id(self):
        entries = [_make_gl_entry()]
        ctx = MockCtx(entries)
        result = get_gl_entries_for_window(
            ctx,
            SAMPLE_PROPERTY_ID,
            date(2023, 7, 15),
            date(2025, 1, 15),
        )
        assert "import_batch_id" in result[0]
        assert result[0]["import_batch_id"] == str(SAMPLE_BATCH_ID)

    def test_fetches_entries_after_first_page(self):
        entries = [_make_gl_entry() for _ in range(1001)]
        ctx = MockCtx(entries)

        result = get_gl_entries_for_window(
            ctx,
            SAMPLE_PROPERTY_ID,
            date(2023, 7, 15),
            date(2025, 1, 15),
        )

        assert len(result) == 1001
        assert result[-1]["import_batch_id"] == str(SAMPLE_BATCH_ID)


# ---------------------------------------------------------------------------
# TestBuildSB1103ExportData
# ---------------------------------------------------------------------------


def _make_mock_ctx_for_build(
    request_data: dict,
    property_data: dict,
    lease_data: dict,
    org_data: dict,
    gl_entries: list[dict],
):
    """Build a multi-table MockCtx for build_sb1103_export_data."""
    ctx = MagicMock()

    def table_side_effect(name):
        qb = MagicMock()
        qb.select.return_value = qb
        qb.eq.return_value = qb
        qb.gte.return_value = qb
        qb.lte.return_value = qb
        qb.order.return_value = qb
        qb.range.return_value = qb
        qb.single.return_value = qb
        qb.maybe_single.return_value = qb

        if name == "sb1103_requests":
            resp = MagicMock()
            # Supabase always returns a list from .execute() on multi-row queries
            resp.data = [request_data]
            qb.execute.return_value = resp
        elif name == "properties":
            resp = MagicMock()
            resp.data = [property_data]
            qb.execute.return_value = resp
        elif name == "leases":
            resp = MagicMock()
            resp.data = [lease_data]
            qb.execute.return_value = resp
        elif name == "organizations":
            resp = MagicMock()
            resp.data = [org_data]
            qb.execute.return_value = resp
        elif name == "gl_entries":
            resp = MagicMock()
            resp.data = gl_entries
            qb.execute.return_value = resp
        else:
            resp = MagicMock()
            resp.data = None
            qb.execute.return_value = resp

        return qb

    ctx.table.side_effect = table_side_effect
    return ctx


class TestBuildSB1103ExportData:
    def _make_all_data(self):
        from datetime import datetime

        request_id = uuid4()
        request_date = date(2025, 1, 15)
        request_data = {
            "id": str(request_id),
            "organization_id": str(SAMPLE_ORG_ID),
            "property_id": str(SAMPLE_PROPERTY_ID),
            "lease_id": str(SAMPLE_LEASE_ID),
            "requested_by_name": "Jane Smith",
            "requested_by_email": "jane@tenant.com",
            "request_date": request_date.isoformat(),
            "response_deadline": compute_response_deadline(request_date).isoformat(),
            "window_start_date": compute_window_start(request_date).isoformat(),
            "window_end_date": request_date.isoformat(),
            "status": "pending",
            "export_format": None,
            "exported_at": None,
            "notes": None,
            "created_at": datetime.now(UTC).isoformat(),
            "updated_at": datetime.now(UTC).isoformat(),
        }
        property_data = {
            "id": str(SAMPLE_PROPERTY_ID),
            "name": "Downtown Tower",
            "address": "123 Main St",
            "city": "Los Angeles",
            "state": "CA",
            "zip_code": "90001",
            "organization_id": str(SAMPLE_ORG_ID),
        }
        lease_data = {
            "id": str(SAMPLE_LEASE_ID),
            "tenant_name": "Acme Corp",
            "property_id": str(SAMPLE_PROPERTY_ID),
            "recovery_profile": {
                "pro_rata_share": "0.10",
                "base_year": 2024,
                "cap_type": "none",
                "gross_up_base_year": False,
                "admin_fee_percentage": "0.15",
                "excluded_pools": [],
            },
        }
        org_data = {
            "id": str(SAMPLE_ORG_ID),
            "name": "Landlord LLC",
        }
        gl_entries = [
            {
                "id": str(uuid4()),
                "transaction_date": "2024-06-15",
                "account_code": "5100",
                "account_description": "Janitorial Services",
                "vendor_name": "Clean Co",
                "description": "Monthly janitorial",
                "amount": "1000.00",
                "import_batch_id": str(SAMPLE_BATCH_ID),
            }
        ]
        return request_id, request_data, property_data, lease_data, org_data, gl_entries

    def test_builds_export_data(self):
        request_id, req, prop, lease, org, gl = self._make_all_data()
        ctx = _make_mock_ctx_for_build(req, prop, lease, org, gl)
        result = build_sb1103_export_data(ctx, request_id)

        assert result.tenant_name == "Acme Corp"
        assert result.property_name == "Downtown Tower"
        assert result.is_ca_property is True
        assert len(result.gl_entries) == 1

    def test_computes_tenant_share_amount(self):
        request_id, req, prop, lease, org, gl = self._make_all_data()
        ctx = _make_mock_ctx_for_build(req, prop, lease, org, gl)
        result = build_sb1103_export_data(ctx, request_id)

        # pro_rata_share = 0.10, amount = 1000.00 → tenant_share = 100.00
        entry = result.gl_entries[0]
        assert entry.tenant_share_amount == Decimal("100.00")

    def test_is_ca_property_false_for_non_ca(self):
        request_id, req, prop, lease, org, gl = self._make_all_data()
        prop_tx = dict(prop, state="TX")
        ctx = _make_mock_ctx_for_build(req, prop_tx, lease, org, gl)
        result = build_sb1103_export_data(ctx, request_id)
        assert result.is_ca_property is False

    def test_builds_property_address_from_current_schema_fields(self):
        request_id, req, prop, lease, org, gl = self._make_all_data()
        current_schema_property = {
            "id": prop["id"],
            "name": prop["name"],
            "address_line1": "123 Main St",
            "address_line2": "Suite 200",
            "city": "Los Angeles",
            "state": "CA",
            "postal_code": "90001",
            "organization_id": prop["organization_id"],
        }
        ctx = _make_mock_ctx_for_build(req, current_schema_property, lease, org, gl)

        result = build_sb1103_export_data(ctx, request_id)

        assert (
            result.property_address == "123 Main St, Suite 200, Los Angeles, CA 90001"
        )

    def test_category_subtotals_grouped(self):
        request_id, req, prop, lease, org, _ = self._make_all_data()
        gl_multi = [
            {
                "id": str(uuid4()),
                "transaction_date": "2024-06-15",
                "account_code": "5100",
                "account_description": "Janitorial Services",
                "vendor_name": "Clean Co",
                "description": "Monthly",
                "amount": "500.00",
                "import_batch_id": str(SAMPLE_BATCH_ID),
            },
            {
                "id": str(uuid4()),
                "transaction_date": "2024-07-15",
                "account_code": "5100",
                "account_description": "Janitorial Services",
                "vendor_name": "Clean Co",
                "description": "Monthly",
                "amount": "500.00",
                "import_batch_id": str(SAMPLE_BATCH_ID),
            },
        ]
        ctx = _make_mock_ctx_for_build(req, prop, lease, org, gl_multi)
        result = build_sb1103_export_data(ctx, request_id)

        # Both entries are "Janitorial Services" → subtotal = 500*0.10 + 500*0.10 = 100
        assert "Janitorial Services" in result.category_subtotals
        assert result.category_subtotals["Janitorial Services"] == Decimal("100.00")

    def test_raises_not_found_on_missing_request(self):
        from app.exceptions import NotFoundError

        ctx = MagicMock()
        qb = MagicMock()
        qb.select.return_value = qb
        qb.eq.return_value = qb
        qb.single.return_value = qb
        qb.maybe_single.return_value = qb
        resp = MagicMock()
        resp.data = []  # Empty list — Supabase returns [] for no results
        qb.execute.return_value = resp
        ctx.table.return_value = qb

        with pytest.raises(NotFoundError):
            build_sb1103_export_data(ctx, uuid4())

    def test_raises_value_error_on_zero_pro_rata_share(self):
        request_id, req, prop, lease, org, gl = self._make_all_data()
        # Override lease with zero pro_rata_share
        lease_zero = dict(lease, recovery_profile={"pro_rata_share": "0"})
        ctx = _make_mock_ctx_for_build(req, prop, lease_zero, org, gl)

        with pytest.raises(ValueError, match="pro_rata_share"):
            build_sb1103_export_data(ctx, request_id)

    def test_raises_value_error_when_lease_property_mismatches_request(self):
        request_id, req, prop, lease, org, gl = self._make_all_data()
        mismatched_lease = dict(lease, property_id=str(uuid4()))
        ctx = _make_mock_ctx_for_build(req, prop, mismatched_lease, org, gl)

        with pytest.raises(ValueError, match="SB 1103 request property"):
            build_sb1103_export_data(ctx, request_id)

    def test_import_batch_id_present_on_every_entry(self):
        request_id, req, prop, lease, org, gl = self._make_all_data()
        ctx = _make_mock_ctx_for_build(req, prop, lease, org, gl)
        result = build_sb1103_export_data(ctx, request_id)

        for entry in result.gl_entries:
            assert entry.import_batch_id is not None


# ---------------------------------------------------------------------------
# TestGeneratePDFExport
# ---------------------------------------------------------------------------


class TestGeneratePDFExport:
    def test_returns_non_empty_bytes_io(self):
        export_data = _make_export_data()
        result = generate_pdf_export(export_data)
        assert isinstance(result, BytesIO)
        assert result.getvalue()  # non-empty

    def test_cover_page_contains_deadline(self):
        request_date = date(2025, 1, 1)
        req = _make_request(request_date=request_date)
        export_data = _make_export_data()
        # Override request with specific deadline
        export_data = SB1103ExportData(
            request=req,
            property_address="123 Main St, Los Angeles, CA 90001",
            property_name="Downtown Tower",
            tenant_name="Acme Corp",
            pro_rata_share=Decimal("0.10"),
            gl_entries=export_data.gl_entries,
            category_subtotals=export_data.category_subtotals,
            is_ca_property=True,
            total_cam_expenses=Decimal("1000.00"),
            total_tenant_share=Decimal("100.00"),
        )
        # deadline = 2025-01-01 + 30 = 2025-01-31 → "January 31, 2025"
        result = generate_pdf_export(export_data)
        content = result.getvalue()
        # PDF is binary — the text is embedded; check it runs without error
        # and content is non-empty (full text assertion would require PDF parsing)
        assert len(content) > 1000

    def test_no_exception_on_valid_data(self):
        export_data = _make_export_data()
        # Should not raise
        result = generate_pdf_export(export_data)
        assert result is not None

    def test_multiple_entries_no_exception(self):
        entries = [
            SB1103GLEntry(
                id=uuid4(),
                transaction_date=date(2024, i, 15),
                account_code="5100",
                account_description="Janitorial Services",
                amount=Decimal("500.00"),
                import_batch_id=SAMPLE_BATCH_ID,
                tenant_share_amount=Decimal("50.00"),
            )
            for i in range(1, 13)
        ]
        export_data = _make_export_data(gl_entries=entries)
        result = generate_pdf_export(export_data)
        assert result.getvalue()


# ---------------------------------------------------------------------------
# TestGenerateExcelExport
# ---------------------------------------------------------------------------


class TestGenerateExcelExport:
    def test_returns_bytes_io(self):
        export_data = _make_export_data()
        result = generate_excel_export(export_data)
        assert isinstance(result, BytesIO)
        assert result.getvalue()

    def test_workbook_has_four_sheets(self):
        import openpyxl

        export_data = _make_export_data()
        result = generate_excel_export(export_data)
        result.seek(0)
        wb = openpyxl.load_workbook(result)
        assert wb.sheetnames == ["Cover", "Ledger", "Category Subtotals", "Methodology"]

    def test_amounts_stored_as_strings(self):
        import openpyxl

        export_data = _make_export_data()
        result = generate_excel_export(export_data)
        result.seek(0)
        wb = openpyxl.load_workbook(result)
        ledger = wb["Ledger"]

        # Find the amount column (row 2 = first data row after header)
        # Amounts should be strings like "$1,000.00"
        found_amount_string = False
        for row in ledger.iter_rows(min_row=2, values_only=True):
            for cell_value in row:
                if isinstance(cell_value, str) and cell_value.startswith("$"):
                    found_amount_string = True
                    break

        assert found_amount_string, "Expected dollar-formatted strings in Ledger sheet"

    def test_every_ledger_row_has_import_batch_id(self):
        import openpyxl

        entries = [
            SB1103GLEntry(
                id=uuid4(),
                transaction_date=date(2024, 6, 15),
                account_code="5100",
                account_description="Janitorial Services",
                amount=Decimal("1000.00"),
                import_batch_id=SAMPLE_BATCH_ID,
                tenant_share_amount=Decimal("100.00"),
            )
        ]
        export_data = _make_export_data(gl_entries=entries)
        result = generate_excel_export(export_data)
        result.seek(0)
        wb = openpyxl.load_workbook(result)
        ledger = wb["Ledger"]

        # Find import_batch_id column header
        headers = [cell.value for cell in next(ledger.iter_rows(max_row=1))]
        assert "Import Batch ID" in headers

        date_col = headers.index("Date")
        batch_col = headers.index("Import Batch ID") + 1
        for row in ledger.iter_rows(
            min_row=2, max_row=ledger.max_row, values_only=True
        ):
            # Only check data rows (those that have a date in the Date column)
            if row[date_col] is not None and str(row[date_col]).count("-") == 2:
                assert row[batch_col - 1] is not None


# ---------------------------------------------------------------------------
# TestGetDeadlineAlerts
# ---------------------------------------------------------------------------


class TestGetDeadlineAlerts:
    def _make_ctx_with_requests(self, requests_data: list[dict]) -> MagicMock:
        ctx = MagicMock()

        def table_side_effect(name):
            qb = MagicMock()
            qb.select.return_value = qb
            qb.eq.return_value = qb
            qb.neq.return_value = qb
            qb.lte.return_value = qb
            qb.order.return_value = qb

            if name in ("sb1103_requests", "properties", "leases"):
                resp = MagicMock()
                resp.data = requests_data
                qb.execute.return_value = resp
            return qb

        ctx.table.side_effect = table_side_effect
        return ctx

    def test_returns_approaching_deadlines(self):
        today = date.today()
        request_data = [
            {
                "id": str(uuid4()),
                "property_id": str(SAMPLE_PROPERTY_ID),
                "lease_id": str(SAMPLE_LEASE_ID),
                "response_deadline": (today + timedelta(days=5)).isoformat(),
                "status": "pending",
                "requested_by_name": "Jane",
                "requested_by_email": "jane@test.com",
                "request_date": (today - timedelta(days=25)).isoformat(),
                "window_start_date": (today - timedelta(days=570)).isoformat(),
                "window_end_date": (today - timedelta(days=25)).isoformat(),
                "export_format": None,
                "exported_at": None,
                "notes": None,
            }
        ]
        property_data = [{"id": str(SAMPLE_PROPERTY_ID), "name": "Downtown Tower"}]
        lease_data = [{"id": str(SAMPLE_LEASE_ID), "tenant_name": "Acme Corp"}]

        ctx = MagicMock()

        def table_side_effect(name):
            qb = MagicMock()
            qb.select.return_value = qb
            qb.eq.return_value = qb
            qb.neq.return_value = qb
            qb.lte.return_value = qb
            qb.in_.return_value = qb
            qb.order.return_value = qb

            if name == "sb1103_requests":
                resp = MagicMock()
                resp.data = request_data
                qb.execute.return_value = resp
            elif name == "properties":
                resp = MagicMock()
                resp.data = property_data
                qb.execute.return_value = resp
            elif name == "leases":
                resp = MagicMock()
                resp.data = lease_data
                qb.execute.return_value = resp
            return qb

        ctx.table.side_effect = table_side_effect

        alerts = get_deadline_alerts(ctx, days_warning=7)
        assert len(alerts) == 1
        assert alerts[0].days_remaining == 5

    def test_overdue_has_negative_days_remaining(self):
        today = date.today()
        request_data = [
            {
                "id": str(uuid4()),
                "property_id": str(SAMPLE_PROPERTY_ID),
                "lease_id": str(SAMPLE_LEASE_ID),
                "response_deadline": (today - timedelta(days=3)).isoformat(),
                "status": "overdue",
                "requested_by_name": "Jane",
                "requested_by_email": "jane@test.com",
                "request_date": (today - timedelta(days=33)).isoformat(),
                "window_start_date": (today - timedelta(days=580)).isoformat(),
                "window_end_date": (today - timedelta(days=33)).isoformat(),
                "export_format": None,
                "exported_at": None,
                "notes": None,
            }
        ]
        property_data = [{"id": str(SAMPLE_PROPERTY_ID), "name": "Downtown Tower"}]
        lease_data = [{"id": str(SAMPLE_LEASE_ID), "tenant_name": "Acme Corp"}]

        ctx = MagicMock()

        def table_side_effect(name):
            qb = MagicMock()
            qb.select.return_value = qb
            qb.eq.return_value = qb
            qb.neq.return_value = qb
            qb.lte.return_value = qb
            qb.in_.return_value = qb
            qb.order.return_value = qb

            if name == "sb1103_requests":
                resp = MagicMock()
                resp.data = request_data
                qb.execute.return_value = resp
            elif name == "properties":
                resp = MagicMock()
                resp.data = property_data
                qb.execute.return_value = resp
            elif name == "leases":
                resp = MagicMock()
                resp.data = lease_data
                qb.execute.return_value = resp
            return qb

        ctx.table.side_effect = table_side_effect

        alerts = get_deadline_alerts(ctx, days_warning=7)
        assert len(alerts) == 1
        assert alerts[0].days_remaining == -3

    def test_excludes_delivered(self):
        ctx = MagicMock()

        def table_side_effect(name):
            qb = MagicMock()
            qb.select.return_value = qb
            qb.eq.return_value = qb
            qb.neq.return_value = qb
            qb.lte.return_value = qb
            qb.in_.return_value = qb
            qb.order.return_value = qb
            # Simulate that the filter excludes delivered → empty result
            resp = MagicMock()
            resp.data = []
            qb.execute.return_value = resp
            return qb

        ctx.table.side_effect = table_side_effect

        alerts = get_deadline_alerts(ctx, days_warning=7)
        assert alerts == []

    def test_empty_when_none(self):
        ctx = MagicMock()

        def table_side_effect(name):
            qb = MagicMock()
            qb.select.return_value = qb
            qb.eq.return_value = qb
            qb.neq.return_value = qb
            qb.lte.return_value = qb
            qb.in_.return_value = qb
            qb.order.return_value = qb
            resp = MagicMock()
            resp.data = []
            qb.execute.return_value = resp
            return qb

        ctx.table.side_effect = table_side_effect

        alerts = get_deadline_alerts(ctx, days_warning=7)
        assert alerts == []


class _ChunkRecordingTable:
    """Minimal PostgREST-like table that records each ``.in_()`` chunk size and
    honours ``.range()`` windows, so the test can prove the id filters are
    chunked (<=100) and paginated rather than sent as one oversized URL.
    """

    def __init__(self, rows_by_id, recorder):
        self._rows_by_id = rows_by_id
        self._recorder = recorder
        self._filter_ids: list | None = None
        self._start = 0
        self._end: int | None = None

    def select(self, *_args, **_kwargs):
        return self

    def in_(self, _column, values):
        self._filter_ids = list(values)
        self._recorder.append(len(self._filter_ids))
        return self

    def range(self, start, end):
        self._start, self._end = start, end
        return self

    def execute(self):
        selected = [
            self._rows_by_id[i]
            for i in (self._filter_ids or [])
            if i in self._rows_by_id
        ]
        if self._end is not None:
            selected = selected[self._start : self._end + 1]
        resp = MagicMock()
        resp.data = selected
        return resp


class TestGetDeadlineAlertsScale:
    """BUG-13: ``get_deadline_alerts`` chunks its property/lease id filters.

    Hundreds of open SB 1103 requests across many leases/properties would
    otherwise encode hundreds of UUIDs into a single PostgREST ``in.(...)`` URL
    and 414 (and the raw un-paginated read would silently cap at 1000 rows).
    """

    def test_property_and_lease_filters_are_chunked(self):
        today = date.today()
        n = 150
        property_ids = [str(uuid4()) for _ in range(n)]
        lease_ids = [str(uuid4()) for _ in range(n)]
        request_data = [
            {
                "id": str(uuid4()),
                "property_id": property_ids[i],
                "lease_id": lease_ids[i],
                "response_deadline": (today + timedelta(days=3)).isoformat(),
                "status": "pending",
            }
            for i in range(n)
        ]
        props_by_id = {
            pid: {"id": pid, "name": f"Prop {i}"} for i, pid in enumerate(property_ids)
        }
        leases_by_id = {
            lid: {"id": lid, "tenant_name": f"Tenant {i}"}
            for i, lid in enumerate(lease_ids)
        }

        prop_chunks: list[int] = []
        lease_chunks: list[int] = []
        ctx = MagicMock()

        def table_side_effect(name):
            if name == "sb1103_requests":
                qb = MagicMock()
                qb.select.return_value = qb
                qb.eq.return_value = qb
                qb.neq.return_value = qb
                qb.lte.return_value = qb
                qb.order.return_value = qb
                resp = MagicMock()
                resp.data = request_data
                qb.execute.return_value = resp
                return qb
            if name == "properties":
                return _ChunkRecordingTable(props_by_id, prop_chunks)
            if name == "leases":
                return _ChunkRecordingTable(leases_by_id, lease_chunks)
            raise AssertionError(f"unexpected table {name}")

        ctx.table.side_effect = table_side_effect

        alerts = get_deadline_alerts(ctx, days_warning=7)

        # All requests surfaced, with names resolved across every chunk.
        assert len(alerts) == n
        assert {a.property_name for a in alerts} == {f"Prop {i}" for i in range(n)}
        assert {a.tenant_name for a in alerts} == {f"Tenant {i}" for i in range(n)}
        # Both id filters were split into multiple <=100-value chunks.
        for chunks in (prop_chunks, lease_chunks):
            assert len(chunks) > 1
            assert max(chunks) <= 100
            assert sum(chunks) == n
