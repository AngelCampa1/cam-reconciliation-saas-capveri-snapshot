"""Tests for Excel export with actual formatting validation."""

from io import BytesIO

from openpyxl import load_workbook

from app.services.reports.excel_export import export_to_excel


class TestExcelStructure:
    """Test basic Excel workbook structure."""

    def test_export_creates_both_sheets(self):
        """Export creates both required sheets."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [],
                "totals": [],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))

        assert "Year-over-Year Comparison" in wb.sheetnames
        assert "Detected Anomalies" in wb.sheetnames


class TestCurrencyFormatting:
    """Test currency number formatting."""

    def test_currency_format_in_yoy_sheet(self):
        """Amount cells use $#,##0 currency format."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [
                    {
                        "name": "Utilities",
                        "years": [2023, 2024],
                        "amounts": [12345, 23456],
                        "variance_percent": 10.0,
                    }
                ],
                "totals": [
                    {"year": 2023, "total": 12345},
                    {"year": 2024, "total": 23456},
                ],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Year-over-Year Comparison"]

        # Check amount cells have currency format
        amount_cell_2023 = ws.cell(row=2, column=2)
        amount_cell_2024 = ws.cell(row=2, column=3)

        assert amount_cell_2023.number_format == "$#,##0"
        assert amount_cell_2024.number_format == "$#,##0"
        assert amount_cell_2023.value == 12345
        assert amount_cell_2024.value == 23456

    def test_currency_format_in_totals_row(self):
        """Totals row uses currency format."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [
                    {
                        "name": "Utilities",
                        "years": [2023, 2024],
                        "amounts": [10000, 11000],
                        "variance_percent": 10.0,
                    }
                ],
                "totals": [
                    {"year": 2023, "total": 10000},
                    {"year": 2024, "total": 11000},
                ],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Year-over-Year Comparison"]

        # Totals row is at row 3 (1 header + 1 category + 1 totals)
        total_cell_2023 = ws.cell(row=3, column=2)
        total_cell_2024 = ws.cell(row=3, column=3)

        assert total_cell_2023.number_format == "$#,##0"
        assert total_cell_2024.number_format == "$#,##0"

    def test_currency_format_in_anomalies_sheet(self):
        """Anomaly amounts use currency format."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [],
                "totals": [],
            },
            "anomalies": [
                {
                    "severity": "critical",
                    "pool_name": "Repairs",
                    "anomaly_type": "spike",
                    "current_value": 15000,
                    "expected_value": 5000,
                    "variance_percent": 200.0,
                    "explanation": "Unusual spike",
                }
            ],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Detected Anomalies"]

        # Current and expected value cells
        current_cell = ws.cell(row=2, column=4)
        expected_cell = ws.cell(row=2, column=5)

        assert current_cell.number_format == "$#,##0"
        assert expected_cell.number_format == "$#,##0"


class TestPercentageFormatting:
    """Test percentage number formatting."""

    def test_variance_percentage_format(self):
        """Variance cells use 0.0% format."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [
                    {
                        "name": "Utilities",
                        "years": [2023, 2024],
                        "amounts": [10000, 11000],
                        "variance_percent": 10.0,
                    }
                ],
                "totals": [
                    {"year": 2023, "total": 10000},
                    {"year": 2024, "total": 11000},
                ],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Year-over-Year Comparison"]

        variance_cell = ws.cell(row=2, column=4)  # Variance % column

        assert variance_cell.number_format == "0.0%"
        assert variance_cell.value == 0.10  # 10% stored as 0.10


class TestVarianceColorCoding:
    """Test variance color thresholds."""

    def test_variance_red_above_15_percent(self):
        """Variance >15% gets red fill."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [
                    {
                        "name": "Repairs",
                        "years": [2023, 2024],
                        "amounts": [10000, 12000],
                        "variance_percent": 20.0,  # >15%
                    }
                ],
                "totals": [],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Year-over-Year Comparison"]

        variance_cell = ws.cell(row=2, column=4)

        # openpyxl stores colors with alpha channel (00RRGGBB or FFRRGGBB)
        assert variance_cell.fill.start_color.rgb[-6:] == "FFCCCC"  # Red fill

    def test_variance_yellow_between_5_and_15_percent(self):
        """Variance 5-15% gets yellow fill."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [
                    {
                        "name": "Janitorial",
                        "years": [2023, 2024],
                        "amounts": [10000, 11000],
                        "variance_percent": 10.0,  # 5-15%
                    }
                ],
                "totals": [],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Year-over-Year Comparison"]

        variance_cell = ws.cell(row=2, column=4)

        assert variance_cell.fill.start_color.rgb[-6:] == "FFFFCC"  # Yellow fill

    def test_variance_no_fill_below_5_percent(self):
        """Variance <5% has no special fill."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [
                    {
                        "name": "Insurance",
                        "years": [2023, 2024],
                        "amounts": [10000, 10300],
                        "variance_percent": 3.0,  # <5%
                    }
                ],
                "totals": [],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Year-over-Year Comparison"]

        variance_cell = ws.cell(row=2, column=4)

        # No fill means fill_type is None or fill color is default
        assert (
            variance_cell.fill.fill_type is None
            or variance_cell.fill.start_color.rgb
            in [
                "00000000",  # Default/no fill
                None,
            ]
        )


class TestHeaderFormatting:
    """Test header row formatting."""

    def test_headers_bold_with_colored_background(self):
        """Headers are bold with blue text and light blue background."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [],
                "totals": [],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Year-over-Year Comparison"]

        # Check first header cell
        header_cell = ws.cell(row=1, column=1)

        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb[-6:] == "1E3A8A"  # Dark blue text
        assert header_cell.fill.start_color.rgb[-6:] == "E0E7FF"  # Light blue fill
        assert header_cell.alignment.horizontal == "center"
        assert header_cell.alignment.vertical == "center"


class TestSeverityColorCoding:
    """Test anomaly severity color coding."""

    def test_critical_severity_red_styling(self):
        """CRITICAL severity has red background and dark red text."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [],
                "totals": [],
            },
            "anomalies": [
                {
                    "severity": "critical",
                    "pool_name": "Emergency Repairs",
                    "anomaly_type": "spike",
                    "current_value": 50000,
                    "expected_value": 5000,
                    "variance_percent": 900.0,
                    "explanation": "Major system failure",
                }
            ],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Detected Anomalies"]

        severity_cell = ws.cell(row=2, column=1)

        assert severity_cell.value == "CRITICAL"
        assert severity_cell.fill.start_color.rgb[-6:] == "FEE2E2"  # Red background
        assert severity_cell.font.color.rgb[-6:] == "991B1B"  # Dark red text
        assert severity_cell.font.bold is True

    def test_warning_severity_yellow_styling(self):
        """WARNING severity has yellow background and dark yellow text."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [],
                "totals": [],
            },
            "anomalies": [
                {
                    "severity": "warning",
                    "pool_name": "Utilities",
                    "anomaly_type": "trend_break",
                    "current_value": 15000,
                    "expected_value": 10000,
                    "variance_percent": 50.0,
                    "explanation": "Higher than expected",
                }
            ],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Detected Anomalies"]

        severity_cell = ws.cell(row=2, column=1)

        assert severity_cell.value == "WARNING"
        assert severity_cell.fill.start_color.rgb[-6:] == "FEF3C7"  # Yellow background
        assert severity_cell.font.color.rgb[-6:] == "92400E"  # Dark yellow text
        assert severity_cell.font.bold is True

    def test_info_severity_blue_styling(self):
        """INFO severity has blue background and dark blue text."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [],
                "totals": [],
            },
            "anomalies": [
                {
                    "severity": "info",
                    "pool_name": "Landscaping",
                    "anomaly_type": "minor_variance",
                    "current_value": 5500,
                    "expected_value": 5000,
                    "variance_percent": 10.0,
                    "explanation": "Slight increase",
                }
            ],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Detected Anomalies"]

        severity_cell = ws.cell(row=2, column=1)

        assert severity_cell.value == "INFO"
        assert severity_cell.fill.start_color.rgb[-6:] == "DBEAFE"  # Blue background
        assert severity_cell.font.color.rgb[-6:] == "1E3A8A"  # Dark blue text
        assert severity_cell.font.bold is True


class TestColumnWidths:
    """Test column width adjustments."""

    def test_yoy_sheet_column_widths(self):
        """YoY sheet columns have 15-width setting."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [
                    {
                        "name": "Utilities",
                        "years": [2023, 2024],
                        "amounts": [10000, 11000],
                        "variance_percent": 10.0,
                    }
                ],
                "totals": [],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Year-over-Year Comparison"]

        # Check first few columns have width 15
        assert ws.column_dimensions["A"].width == 15
        assert ws.column_dimensions["B"].width == 15
        assert ws.column_dimensions["C"].width == 15

    def test_anomalies_sheet_column_widths(self):
        """Anomalies sheet has specific column widths."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [],
                "totals": [],
            },
            "anomalies": [
                {
                    "severity": "critical",
                    "pool_name": "Repairs",
                    "anomaly_type": "spike",
                    "current_value": 15000,
                    "expected_value": 5000,
                    "variance_percent": 200.0,
                    "explanation": "Major increase",
                }
            ],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Detected Anomalies"]

        # Check specific widths per column
        assert ws.column_dimensions["A"].width == 12  # Severity
        assert ws.column_dimensions["B"].width == 20  # Expense Pool
        assert ws.column_dimensions["C"].width == 18  # Type
        assert ws.column_dimensions["D"].width == 12  # Current
        assert ws.column_dimensions["E"].width == 12  # Expected
        assert ws.column_dimensions["F"].width == 12  # Variance %
        assert ws.column_dimensions["G"].width == 50  # Explanation


class TestEdgeCases:
    """Test edge cases and error handling."""

    def test_empty_anomalies_displays_message(self):
        """No anomalies shows friendly message."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [],
                "totals": [],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Detected Anomalies"]

        message_cell = ws.cell(row=2, column=1)

        assert "No anomalies detected" in message_cell.value
        assert message_cell.alignment.horizontal == "center"

    def test_multiple_categories_all_formatted(self):
        """Multiple categories all get proper formatting."""
        categories = [
            {
                "name": f"Pool {i}",
                "years": [2023, 2024],
                "amounts": [10000 * i, 11000 * i],
                "variance_percent": 10.0,
            }
            for i in range(1, 6)
        ]

        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": categories,
                "totals": [
                    {"year": 2023, "total": 150000},
                    {"year": 2024, "total": 165000},
                ],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Year-over-Year Comparison"]

        # Check all 5 categories have currency format
        for row in range(2, 7):  # Rows 2-6
            amount_cell = ws.cell(row=row, column=2)
            assert amount_cell.number_format == "$#,##0"

    def test_dict_format_for_amounts(self):
        """Dict format for amounts (covers line 77)."""
        report_data = {
            "property": {"id": "prop-123", "name": "Test Property"},
            "years_compared": [2023, 2024],
            "year_over_year_comparison": {
                "categories": [
                    {
                        "name": "Utilities",
                        "years": [2023, 2024],
                        "amounts": {2023: 12000, 2024: 13500},  # Dict format
                        "variance_percent": 12.5,
                    }
                ],
                "totals": [
                    {"year": 2023, "total": 12000},
                    {"year": 2024, "total": 13500},
                ],
            },
            "anomalies": [],
        }

        excel_bytes = export_to_excel(report_data)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Year-over-Year Comparison"]

        # Verify amounts from dict format are written correctly
        amount_cell_2023 = ws.cell(row=2, column=2)
        amount_cell_2024 = ws.cell(row=2, column=3)

        assert amount_cell_2023.value == 12000
        assert amount_cell_2024.value == 13500
        assert amount_cell_2023.number_format == "$#,##0"
        assert amount_cell_2024.number_format == "$#,##0"
