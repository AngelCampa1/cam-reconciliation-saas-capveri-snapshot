import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl
from PyPDF2 import PdfReader

from app.services.leads.asset_registry import (
    ASSETS,
    DOWNLOAD_SLUGS,
    R2_STORAGE_PREFIX,
    get_asset,
)

ROOT = Path(__file__).resolve().parents[2]
MARKETING_REGISTRY = ROOT / "marketing" / "src" / "lib" / "lead-magnets" / "registry.ts"
CLOUDFLARE_REGISTRY = (
    ROOT / "cloudflare-backend" / "src" / "domain" / "leads" / "assets.ts"
)
GENERATED_DIR = ROOT / "generated" / "lead-magnets"
MANIFEST_PATH = GENERATED_DIR / "manifest.json"
sys.path.insert(0, str(ROOT))

from scripts import build_lead_magnets  # noqa: E402


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def test_registry_contains_all_current_lead_magnets() -> None:
    assert len(ASSETS) == 39
    assert get_asset("cam-reconciliation-checklist") is not None


def test_every_asset_is_enabled_and_has_storage_path() -> None:
    for slug, asset in ASSETS.items():
        assert asset.enabled is True, f"{slug} is still disabled"
        assert asset.storage_path, f"{slug} is missing an R2 object key"
        assert asset.storage_path.startswith("lead-magnets/")
        assert asset.storage_path.endswith((".pdf", ".xlsx"))


def test_enabled_assets_match_generated_remote_verified_manifest() -> None:
    manifest = json.loads(MANIFEST_PATH.read_text())
    assert manifest["bucket"] == "capveri-lead-magnets"
    assert manifest["remote_verified"] is True

    manifest_assets = {asset["slug"]: asset for asset in manifest["assets"]}
    assert set(manifest_assets) == set(ASSETS)

    for slug, asset in ASSETS.items():
        manifest_asset = manifest_assets[slug]
        path = GENERATED_DIR / manifest_asset["filename"]

        assert path.exists(), f"{slug} has no generated local asset"
        assert manifest_asset["storage_path"] == asset.storage_path
        assert manifest_asset["sha256"] == sha256_file(path)
        assert manifest_asset["size_bytes"] == path.stat().st_size


def test_generated_assets_pass_builder_verification() -> None:
    specs = {spec.slug: spec for spec in build_lead_magnets.ASSETS}
    assert set(specs) == set(ASSETS)

    for slug, asset in ASSETS.items():
        build_lead_magnets.verify_asset(
            GENERATED_DIR / specs[slug].filename, specs[slug]
        )


def test_generated_assets_are_mobile_readable_and_public_safe() -> None:
    manifest = json.loads(MANIFEST_PATH.read_text())
    retired_phrases = ("lead magnet", "free audit")

    for asset in manifest["assets"]:
        path = GENERATED_DIR / asset["filename"]
        if asset["kind"] == "pdf":
            reader = PdfReader(str(path))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            assert "Phone Summary" in text, f"{asset['slug']} has no phone summary"
            assert "Start a 30-day trial" in text, f"{asset['slug']} lacks trial CTA"
        else:
            workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
            try:
                assert (
                    "Phone Summary" in workbook.sheetnames
                ), f"{asset['slug']} has no phone summary sheet"
                text = "\n".join(
                    str(cell.value)
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows()
                    for cell in row
                    if cell.value is not None
                )
                assert (
                    "Start a 30-day trial" in text
                ), f"{asset['slug']} lacks trial CTA"
            finally:
                workbook.close()

        lowered = text.lower()
        for phrase in retired_phrases:
            assert (
                phrase not in lowered
            ), f"{asset['slug']} still contains retired phrase: {phrase}"


def test_builder_verifier_rejects_invalid_asset(tmp_path: Path) -> None:
    spec = next(spec for spec in build_lead_magnets.ASSETS if spec.kind == "pdf")
    bad_pdf = tmp_path / spec.filename
    bad_pdf.write_bytes(b"%PDF-1.4\n% incomplete\n")

    try:
        build_lead_magnets.verify_asset(bad_pdf, spec)
    except Exception as exc:
        assert "too-small" in str(exc) or "Could not read malformed PDF" in str(exc)
    else:
        raise AssertionError("Malformed lead magnet unexpectedly passed verification")


def test_all_downloadable_non_unlock_assets_are_download_slugs() -> None:
    expected = {
        slug for slug, asset in ASSETS.items() if asset.format in ("pdf", "xlsx")
    }
    assert DOWNLOAD_SLUGS == expected


def test_marketing_registry_matches_backend_registry() -> None:
    marketing_source = MARKETING_REGISTRY.read_text()
    marketing_slugs = set(re.findall(r'slug: "([^"]+)"', marketing_source))
    marketing_paths = set(re.findall(r'storagePath: "([^"]+)"', marketing_source))
    marketing_paths.update(
        f"{R2_STORAGE_PREFIX}{path}"
        for path in re.findall(
            r'leadMagnetStoragePath\(\s*"([^"]+)"\s*,?\s*\)', marketing_source
        )
    )

    assert marketing_slugs == set(ASSETS)
    assert marketing_paths == {asset.storage_path for asset in ASSETS.values()}


def test_cloudflare_registry_matches_backend_registry() -> None:
    cloudflare_source = CLOUDFLARE_REGISTRY.read_text()
    asset_calls = re.findall(
        r'"([^"]+)":\s*asset\(\s*"([^"]+)",\s*"([^"]+)",\s*"([^"]+)",\s*"([^"]+)",',
        cloudflare_source,
        flags=re.MULTILINE,
    )

    cloudflare_assets = {
        key: {
            "slug": slug,
            "display_name": display_name,
            "format": asset_format,
            "storage_path": f"{R2_STORAGE_PREFIX}{filename}",
        }
        for key, slug, display_name, asset_format, filename in asset_calls
    }

    assert set(cloudflare_assets) == set(ASSETS)
    for slug, asset in ASSETS.items():
        cloudflare_asset = cloudflare_assets[slug]
        assert cloudflare_asset["slug"] == asset.slug
        assert cloudflare_asset["display_name"] == asset.display_name
        assert cloudflare_asset["format"] == asset.format
        assert cloudflare_asset["storage_path"] == asset.storage_path
