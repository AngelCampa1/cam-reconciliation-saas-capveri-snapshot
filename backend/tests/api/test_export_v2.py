"""Tests for Export v2 API endpoints (/api/v1/export/*).

These are property-level export endpoints (distinct from snapshot-level
/api/v1/exports/ endpoints). Follows TDD — tests written before implementation.

Test Coverage:
- POST /api/v1/export/pdf/preview   → application/pdf
- POST /api/v1/export/pdf/download  → application/pdf (attachment)
- POST /api/v1/export/pdf/batch     → application/zip
- POST /api/v1/export/erp           → text/csv
- GET  /api/v1/export/history       → JSON list
- POST /api/v1/export/variance/pdf  → application/pdf
- POST /api/v1/export/variance/excel → .xlsx (attachment)
"""

from datetime import UTC, datetime
from io import BytesIO
from uuid import uuid4

import pytest

from tests.conftest import ORG_A_ID, ORG_A_PROPERTY_ID

# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

PROPERTY_ID = str(ORG_A_PROPERTY_ID)
YEAR = 2024


def _seed_finalized_snapshots(client, n: int = 1) -> list[dict]:
    """Seed n finalized snapshots for ORG_A_PROPERTY_ID / 2024."""
    snapshots = []
    for _ in range(n):
        lease_id = str(uuid4())
        snapshot = {
            "id": str(uuid4()),
            "organization_id": str(ORG_A_ID),
            "property_id": PROPERTY_ID,
            "lease_id": lease_id,
            "period_start_date": "2024-01-01",
            "period_end_date": "2024-12-31",
            "status": "finalized",
            "finalized_at": datetime.now(UTC).isoformat(),
            "finalized_by_user_id": str(uuid4()),
            "total_operating_expenses": "150000.00",
            "grossed_up_expenses": "157895.00",
            "base_year_amount": "140000.00",
            "tenant_share_before_cap": "39473.75",
            "tenant_share_after_cap": "38289.54",
            "admin_fee": "5743.43",
            "total_recovery": "44032.97",
            "calculation_trace": [],
            "created_at": datetime.now(UTC).isoformat(),
            "updated_at": datetime.now(UTC).isoformat(),
        }
        snapshots.append(snapshot)
    client.mock_supabase._test_data["reconciliation_snapshots"] = snapshots
    client.mock_supabase._test_data["organizations"] = [
        {"id": str(ORG_A_ID), "name": "Test Org"}
    ]
    client.mock_supabase._test_data["properties"] = [
        {
            "id": PROPERTY_ID,
            "organization_id": str(ORG_A_ID),
            "name": "Test Property",
            "address_line1": "100 Main St",
            "city": "San Francisco",
            "state": "CA",
            "postal_code": "94102",
        }
    ]
    client.mock_supabase._test_data["leases"] = [
        {"id": snapshot["lease_id"], "property_id": PROPERTY_ID}
        for snapshot in snapshots
    ]
    return snapshots


def _pdf_preview_payload(
    *,
    property_id: str = PROPERTY_ID,
    year: int = YEAR,
    include_charts: bool = False,
    include_notes: bool = False,
    tenant_ids: list | None = None,
) -> dict:
    payload: dict = {
        "property_id": property_id,
        "year": year,
        "include_charts": include_charts,
        "include_notes": include_notes,
    }
    if tenant_ids is not None:
        payload["tenant_ids"] = tenant_ids
    return payload


def _erp_payload(erp_system: str = "yardi") -> dict:
    return {
        "property_id": PROPERTY_ID,
        "year": YEAR,
        "erp_system": erp_system,
        "field_mappings": {},
    }


def _variance_payload(threshold: float = 10.0) -> dict:
    return {
        "property_id": PROPERTY_ID,
        "current_year": YEAR,
        "prior_year": YEAR - 1,
        "threshold_percent": threshold,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PDF Preview
# ─────────────────────────────────────────────────────────────────────────────


class TestPdfPreview:
    def test_pdf_preview_returns_pdf_content_type(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/pdf/preview",
            json=_pdf_preview_payload(),
        )
        assert response.status_code == 200
        assert "application/pdf" in response.headers["content-type"]

    def test_pdf_preview_requires_property_id(self, org_a_member_client):
        payload = {"year": YEAR}  # Missing property_id
        response = org_a_member_client.post(
            "/api/v1/export/pdf/preview",
            json=payload,
        )
        assert response.status_code == 422

    def test_pdf_preview_rejects_unsupported_options(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/pdf/preview",
            json=_pdf_preview_payload(include_charts=True, include_notes=True),
        )
        assert response.status_code == 400

    def test_pdf_preview_filters_by_tenant_id(self, org_a_member_client, monkeypatch):
        snapshots = _seed_finalized_snapshots(org_a_member_client, n=2)
        selected = snapshots[1]
        captured = {}

        def fake_generate_property_pdf(ctx, snapshots_for_pdf):
            captured["lease_id"] = snapshots_for_pdf[0]["lease_id"]
            return BytesIO(b"%PDF fake")

        monkeypatch.setattr(
            "app.api.v1.export._generate_property_pdf", fake_generate_property_pdf
        )

        response = org_a_member_client.post(
            "/api/v1/export/pdf/preview",
            json=_pdf_preview_payload(tenant_ids=[selected["lease_id"]]),
        )

        assert response.status_code == 200
        assert captured["lease_id"] == selected["lease_id"]

    def test_pdf_preview_rejects_unknown_tenant_id(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client, n=2)
        response = org_a_member_client.post(
            "/api/v1/export/pdf/preview",
            json=_pdf_preview_payload(tenant_ids=[str(uuid4())]),
        )
        assert response.status_code == 400

    def test_pdf_preview_rejects_multiple_tenant_ids(self, org_a_member_client):
        snapshots = _seed_finalized_snapshots(org_a_member_client, n=2)
        response = org_a_member_client.post(
            "/api/v1/export/pdf/preview",
            json=_pdf_preview_payload(
                tenant_ids=[snapshot["lease_id"] for snapshot in snapshots]
            ),
        )
        assert response.status_code == 400

    def test_pdf_preview_no_snapshots_returns_404(self, org_a_member_client):
        """Returns 404 when no finalized snapshots exist for the property/year."""
        org_a_member_client.mock_supabase._test_data["reconciliation_snapshots"] = []
        response = org_a_member_client.post(
            "/api/v1/export/pdf/preview",
            json=_pdf_preview_payload(),
        )
        assert response.status_code == 404


# ─────────────────────────────────────────────────────────────────────────────
# PDF Download
# ─────────────────────────────────────────────────────────────────────────────


class TestPdfDownload:
    def test_pdf_download_streams_pdf(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/pdf/download",
            json=_pdf_preview_payload(),
        )
        assert response.status_code == 200
        assert "application/pdf" in response.headers["content-type"]

    def test_pdf_download_has_content_disposition(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/pdf/download",
            json=_pdf_preview_payload(),
        )
        assert response.status_code == 200
        assert "attachment" in response.headers.get("content-disposition", "")

    def test_pdf_download_records_export_history(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        org_a_member_client.mock_supabase._test_data["export_history"] = []

        response = org_a_member_client.post(
            "/api/v1/export/pdf/download",
            json=_pdf_preview_payload(),
        )

        assert response.status_code == 200
        history = org_a_member_client.mock_supabase._test_data["export_history"]
        assert len(history) == 1
        row = history[0]
        assert row["organization_id"] == str(ORG_A_ID)
        assert row["property_id"] == PROPERTY_ID
        assert row["format"] == "pdf"
        assert row["file_name"] == "reconciliation-2024-property.pdf"
        assert row["file_size"] > 0
        assert row["status"] == "completed"
        assert row["created_by_name"]
        # F-024: the file is persisted to storage and its path recorded so the
        # export can be re-downloaded later.
        assert row["storage_path"]
        org_a_member_client.mock_supabase_admin.storage.from_.assert_called()
        org_a_member_client.mock_supabase_admin.storage.from_.return_value.upload.assert_called()  # noqa: E501

    def test_pdf_download_requires_auth(self, base_client):
        response = base_client.post(
            "/api/v1/export/pdf/download",
            json=_pdf_preview_payload(),
        )
        assert response.status_code in (401, 403, 422)


# ─────────────────────────────────────────────────────────────────────────────
# Batch PDF
# ─────────────────────────────────────────────────────────────────────────────


class TestBatchPdf:
    def _batch_payload(self, tenant_ids: list, mode: str = "zip") -> dict:
        return {
            "property_id": PROPERTY_ID,
            "year": YEAR,
            "tenant_ids": tenant_ids,
            "mode": mode,
        }

    def test_pdf_batch_returns_zip(self, org_a_member_client):
        snapshots = _seed_finalized_snapshots(org_a_member_client, n=2)
        tenant_ids = [s["lease_id"] for s in snapshots]
        response = org_a_member_client.post(
            "/api/v1/export/pdf/batch",
            json=self._batch_payload(tenant_ids),
        )
        assert response.status_code == 200
        assert "application/zip" in response.headers["content-type"]

    def test_pdf_batch_validates_tenant_ids_required(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        payload = {
            "property_id": PROPERTY_ID,
            "year": YEAR,
            # tenant_ids missing
        }
        response = org_a_member_client.post(
            "/api/v1/export/pdf/batch",
            json=payload,
        )
        assert response.status_code == 422

    def test_pdf_batch_empty_tenant_ids_returns_400(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/pdf/batch",
            json=self._batch_payload([]),
        )
        assert response.status_code == 400

    def test_pdf_batch_unknown_tenant_id_does_not_export_all_tenants(
        self, org_a_member_client
    ):
        _seed_finalized_snapshots(org_a_member_client, n=2)

        response = org_a_member_client.post(
            "/api/v1/export/pdf/batch",
            json=self._batch_payload([str(uuid4())]),
        )

        assert response.status_code == 400

    def test_pdf_batch_individual_mode_rejected(self, org_a_member_client):
        snapshots = _seed_finalized_snapshots(org_a_member_client, n=1)
        response = org_a_member_client.post(
            "/api/v1/export/pdf/batch",
            json=self._batch_payload([snapshots[0]["lease_id"]], mode="individual"),
        )
        assert response.status_code == 400

    def test_pdf_batch_requires_auth(self, base_client):
        response = base_client.post(
            "/api/v1/export/pdf/batch",
            json=self._batch_payload([str(uuid4())]),
        )
        assert response.status_code in (401, 403, 422)


# ─────────────────────────────────────────────────────────────────────────────
# ERP Export
# ─────────────────────────────────────────────────────────────────────────────


class TestERPExport:
    def test_erp_export_yardi_format(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/erp",
            json=_erp_payload("yardi"),
        )
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]

    def test_erp_export_mri_format(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/erp",
            json=_erp_payload("mri"),
        )
        assert response.status_code == 200
        # MRI uses text/plain for fixed-width format
        content_type = response.headers["content-type"]
        assert "text/" in content_type

    def test_erp_export_invalid_system_rejected(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/erp",
            json={
                "property_id": PROPERTY_ID,
                "year": YEAR,
                "erp_system": "sap",  # unsupported
                "field_mappings": {},
            },
        )
        assert response.status_code == 422

    def test_erp_export_no_snapshots_returns_404(self, org_a_member_client):
        org_a_member_client.mock_supabase._test_data["reconciliation_snapshots"] = []
        response = org_a_member_client.post(
            "/api/v1/export/erp",
            json=_erp_payload("yardi"),
        )
        assert response.status_code == 404

    def test_erp_export_rejects_field_mappings(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        payload = _erp_payload("yardi")
        payload["field_mappings"] = {"account": "custom_account"}

        response = org_a_member_client.post("/api/v1/export/erp", json=payload)

        assert response.status_code == 400

    def test_erp_export_has_content_disposition(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/erp",
            json=_erp_payload("yardi"),
        )
        assert response.status_code == 200
        assert "attachment" in response.headers.get("content-disposition", "")

    def test_erp_export_records_export_history(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        org_a_member_client.mock_supabase._test_data["export_history"] = []

        response = org_a_member_client.post(
            "/api/v1/export/erp",
            json=_erp_payload("yardi"),
        )

        assert response.status_code == 200
        history = org_a_member_client.mock_supabase._test_data["export_history"]
        assert len(history) == 1
        row = history[0]
        assert row["organization_id"] == str(ORG_A_ID)
        assert row["property_id"] == PROPERTY_ID
        assert row["format"] == "yardi"
        assert row["file_name"] == "Yardi_CAM_Import_2024.csv"
        assert row["file_size"] > 0
        assert row["status"] == "completed"
        assert row["created_by_name"]

    def test_erp_export_requires_auth(self, base_client):
        response = base_client.post(
            "/api/v1/export/erp",
            json=_erp_payload(),
        )
        assert response.status_code in (401, 403, 422)


# ─────────────────────────────────────────────────────────────────────────────
# Export History
# ─────────────────────────────────────────────────────────────────────────────


class TestExportHistory:
    def _seed_history(self, client, n: int = 2) -> list[dict]:
        items = []
        formats = ["pdf", "excel", "yardi", "mri"]
        for i in range(n):
            items.append(
                {
                    "id": str(uuid4()),
                    "property_id": PROPERTY_ID,
                    "organization_id": str(ORG_A_ID),
                    "format": formats[i % len(formats)],
                    "file_name": f"export-{i}.pdf",
                    "file_size": 512000 + i * 1000,
                    "status": "completed",
                    "created_at": datetime.now(UTC).isoformat(),
                    "created_by_name": "Test User",
                }
            )
        client.mock_supabase._test_data["export_history"] = items
        return items

    def test_export_history_returns_list(self, org_a_member_client):
        self._seed_history(org_a_member_client, n=3)
        response = org_a_member_client.get(
            f"/api/v1/export/history?property_id={PROPERTY_ID}"
        )
        assert response.status_code == 200
        body = response.json()
        assert "items" in body
        assert isinstance(body["items"], list)
        assert body["total"] == 3
        assert body["page"] == 1
        assert body["page_size"] == 25

    def test_export_history_applies_query_pagination(self, org_a_member_client):
        self._seed_history(org_a_member_client, n=30)
        response = org_a_member_client.get(
            f"/api/v1/export/history?property_id={PROPERTY_ID}&page=2&page_size=10"
        )
        assert response.status_code == 200
        body = response.json()
        assert body["total"] == 30
        assert body["page"] == 2
        assert body["page_size"] == 10
        assert len(body["items"]) == 10
        assert body["items"][0]["file_name"] == "export-10.pdf"

    def test_export_history_filtered_by_format(self, org_a_member_client):
        self._seed_history(org_a_member_client, n=4)
        response = org_a_member_client.get(
            f"/api/v1/export/history?property_id={PROPERTY_ID}&format=pdf"
        )
        assert response.status_code == 200
        body = response.json()
        for item in body["items"]:
            assert item["format"] == "pdf"
        assert body["total"] == 1

    def test_export_history_rejects_oversized_page(self, org_a_member_client):
        response = org_a_member_client.get(
            f"/api/v1/export/history?property_id={PROPERTY_ID}&page_size=101"
        )
        assert response.status_code == 422

    def test_export_history_requires_property_id(self, org_a_member_client):
        response = org_a_member_client.get("/api/v1/export/history")
        assert response.status_code == 422

    def test_export_history_requires_auth(self, base_client):
        response = base_client.get(f"/api/v1/export/history?property_id={PROPERTY_ID}")
        assert response.status_code in (401, 403, 422)


# ─────────────────────────────────────────────────────────────────────────────
# Re-download (F-024)
# ─────────────────────────────────────────────────────────────────────────────


class TestDownloadExport:
    def _seed_export(
        self,
        client,
        *,
        storage_path: str | None = "reports/org/prop/abc-export.pdf",
        organization_id: str | None = None,
    ) -> dict:
        export_id = str(uuid4())
        row = {
            "id": export_id,
            "property_id": PROPERTY_ID,
            "organization_id": organization_id or str(ORG_A_ID),
            "format": "pdf",
            "file_name": "reconciliation-2024-property.pdf",
            "file_size": 512000,
            "status": "completed",
            "created_at": datetime.now(UTC).isoformat(),
            "created_by_name": "Test User",
            "storage_path": storage_path,
        }
        client.mock_supabase._test_data["export_history"] = [row]
        return row

    def _stub_signed_url(
        self, client, url: str = "https://signed.example/file"
    ) -> None:
        client.mock_supabase_admin.storage.from_.return_value.create_signed_url.return_value = {  # noqa: E501
            "signedURL": url
        }

    def test_download_returns_signed_url(self, org_a_member_client):
        row = self._seed_export(org_a_member_client)
        self._stub_signed_url(org_a_member_client)

        response = org_a_member_client.get(f"/api/v1/export/download/{row['id']}")

        assert response.status_code == 200
        body = response.json()
        assert body["download_url"] == "https://signed.example/file"
        assert body["file_name"] == "reconciliation-2024-property.pdf"
        assert body["expires_at"]

    def test_download_unknown_id_returns_404(self, org_a_member_client):
        org_a_member_client.mock_supabase._test_data["export_history"] = []
        response = org_a_member_client.get(f"/api/v1/export/download/{uuid4()}")
        assert response.status_code == 404

    def test_download_legacy_row_without_storage_returns_410(self, org_a_member_client):
        row = self._seed_export(org_a_member_client, storage_path=None)
        response = org_a_member_client.get(f"/api/v1/export/download/{row['id']}")
        assert response.status_code == 410

    def test_download_other_org_export_returns_404(self, org_a_member_client):
        # Row exists but belongs to another org; org scoping must hide it.
        row = self._seed_export(org_a_member_client, organization_id=str(uuid4()))
        response = org_a_member_client.get(f"/api/v1/export/download/{row['id']}")
        assert response.status_code == 404

    def test_download_requires_auth(self, base_client):
        response = base_client.get(f"/api/v1/export/download/{uuid4()}")
        assert response.status_code in (401, 403, 422)

    def test_export_then_redownload_round_trip(self, org_a_member_client):
        # End-to-end: generating an export persists a file and records its
        # storage_path, and re-downloading mints a signed URL for that exact
        # path. This guards against any divergence between the key used to
        # upload the object and the key used to sign it.
        _seed_finalized_snapshots(org_a_member_client)
        org_a_member_client.mock_supabase._test_data["export_history"] = []

        generate = org_a_member_client.post(
            "/api/v1/export/pdf/download",
            json=_pdf_preview_payload(),
        )
        assert generate.status_code == 200

        history = org_a_member_client.mock_supabase._test_data["export_history"]
        assert len(history) == 1
        recorded_path = history[0]["storage_path"]
        assert recorded_path

        self._stub_signed_url(org_a_member_client)
        redownload = org_a_member_client.get(
            f"/api/v1/export/download/{history[0]['id']}"
        )
        assert redownload.status_code == 200
        assert redownload.json()["download_url"] == "https://signed.example/file"

        # The signed URL must be minted for the precise object key that was
        # uploaded, not a re-prefixed or mismatched path.
        create_signed_url = (
            org_a_member_client.mock_supabase_admin.storage.from_.return_value.create_signed_url  # noqa: E501
        )
        create_signed_url.assert_called_once()
        called_path = create_signed_url.call_args.args[0]
        assert called_path == recorded_path


# ─────────────────────────────────────────────────────────────────────────────
# Variance PDF
# ─────────────────────────────────────────────────────────────────────────────


class TestVariancePdf:
    def _seed_two_year_snapshots(self, client):
        snapshots = []
        for year_start in ["2023-01-01", "2024-01-01"]:
            year_end = year_start[:4] + "-12-31"
            snapshots.append(
                {
                    "id": str(uuid4()),
                    "organization_id": str(ORG_A_ID),
                    "property_id": PROPERTY_ID,
                    "lease_id": str(uuid4()),
                    "period_start_date": year_start,
                    "period_end_date": year_end,
                    "status": "finalized",
                    "finalized_at": datetime.now(UTC).isoformat(),
                    "total_operating_expenses": "150000.00",
                    "grossed_up_expenses": "157895.00",
                    "base_year_amount": "140000.00",
                    "tenant_share_before_cap": "39473.75",
                    "tenant_share_after_cap": "38289.54",
                    "admin_fee": "5743.43",
                    "total_recovery": "44032.97",
                    "calculation_trace": [],
                    "created_at": datetime.now(UTC).isoformat(),
                    "updated_at": datetime.now(UTC).isoformat(),
                }
            )
        client.mock_supabase._test_data["reconciliation_snapshots"] = snapshots
        client.mock_supabase._test_data["organizations"] = [
            {"id": str(ORG_A_ID), "name": "Test Org"}
        ]
        client.mock_supabase._test_data["properties"] = [
            {
                "id": PROPERTY_ID,
                "organization_id": str(ORG_A_ID),
                "name": "Test Property",
                "address_line1": "100 Main St",
                "city": "San Francisco",
                "state": "CA",
                "postal_code": "94102",
            }
        ]

    def test_variance_pdf_returns_pdf(self, org_a_member_client):
        self._seed_two_year_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/variance/pdf",
            json=_variance_payload(),
        )
        assert response.status_code == 200
        assert "application/pdf" in response.headers["content-type"]

    def test_variance_pdf_escapes_dynamic_property_name(self):
        from app.api.v1.export import _generate_variance_pdf

        pdf_buffer = _generate_variance_pdf(
            snapshots_current=[{"total_recovery": "10000.00"}],
            snapshots_prior=[{"total_recovery": "9000.00"}],
            current_year=2024,
            prior_year=2023,
            threshold_percent=10.0,
            property_data={"name": "AT&T <Main>"},
        )

        assert pdf_buffer.getvalue().startswith(b"%PDF")

    def test_variance_pdf_has_attachment_disposition(self, org_a_member_client):
        self._seed_two_year_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/variance/pdf",
            json=_variance_payload(),
        )
        assert response.status_code == 200
        assert "attachment" in response.headers.get("content-disposition", "")

    def test_variance_pdf_requires_auth(self, base_client):
        response = base_client.post(
            "/api/v1/export/variance/pdf",
            json=_variance_payload(),
        )
        assert response.status_code in (401, 403, 422)

    def test_variance_pdf_no_snapshots_returns_404(self, org_a_member_client):
        org_a_member_client.mock_supabase._test_data["reconciliation_snapshots"] = []
        response = org_a_member_client.post(
            "/api/v1/export/variance/pdf",
            json=_variance_payload(),
        )
        assert response.status_code == 404

    def test_variance_pdf_prior_year_empty_uses_zero(self, org_a_member_client):
        """Prior year with no snapshots => prior_total=0, variance_pct=0 branch."""
        # Only seed current year snapshot, no prior year
        snapshot = {
            "id": str(uuid4()),
            "organization_id": str(ORG_A_ID),
            "property_id": PROPERTY_ID,
            "lease_id": str(uuid4()),
            "period_start_date": "2024-01-01",
            "period_end_date": "2024-12-31",
            "status": "finalized",
            "finalized_at": datetime.now(UTC).isoformat(),
            "total_operating_expenses": "150000.00",
            "grossed_up_expenses": "157895.00",
            "base_year_amount": "140000.00",
            "tenant_share_before_cap": "39473.75",
            "tenant_share_after_cap": "38289.54",
            "admin_fee": "5743.43",
            "total_recovery": "44032.97",
            "calculation_trace": [],
            "created_at": datetime.now(UTC).isoformat(),
            "updated_at": datetime.now(UTC).isoformat(),
        }
        org_a_member_client.mock_supabase._test_data["reconciliation_snapshots"] = [
            snapshot
        ]
        org_a_member_client.mock_supabase._test_data["organizations"] = [
            {"id": str(ORG_A_ID), "name": "Test Org"}
        ]
        org_a_member_client.mock_supabase._test_data["properties"] = [
            {
                "id": PROPERTY_ID,
                "organization_id": str(ORG_A_ID),
                "name": "Test Property",
                "address_line1": "100 Main St",
                "city": "San Francisco",
                "state": "CA",
                "postal_code": "94102",
            }
        ]
        # Request variance for 2024 vs 2023 — only 2024 exists
        response = org_a_member_client.post(
            "/api/v1/export/variance/pdf",
            json=_variance_payload(),
        )
        assert response.status_code == 200
        assert "application/pdf" in response.headers["content-type"]


# ─────────────────────────────────────────────────────────────────────────────
# Variance Excel (F-025)
# ─────────────────────────────────────────────────────────────────────────────


_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class TestVarianceExcel(TestVariancePdf):
    """Excel variance export mirrors the PDF route.

    Inherits ``_seed_two_year_snapshots`` from ``TestVariancePdf``. F-025: the
    frontend's ``useExportVarianceExcel`` hook POSTed to /export/variance/excel
    which had no backend route (only /variance/pdf existed) → 404.
    """

    def test_variance_excel_returns_xlsx(self, org_a_member_client):
        self._seed_two_year_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/variance/excel",
            json=_variance_payload(),
        )
        assert response.status_code == 200
        assert _XLSX_CONTENT_TYPE in response.headers["content-type"]
        # OOXML files are zip archives → start with the PK signature.
        assert response.content.startswith(b"PK")

    def test_variance_excel_has_attachment_disposition(self, org_a_member_client):
        self._seed_two_year_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/variance/excel",
            json=_variance_payload(),
        )
        assert response.status_code == 200
        disposition = response.headers.get("content-disposition", "")
        assert "attachment" in disposition
        assert disposition.endswith('.xlsx"')

    def test_variance_excel_records_export_history(self, org_a_member_client):
        self._seed_two_year_snapshots(org_a_member_client)
        org_a_member_client.mock_supabase._test_data["export_history"] = []
        response = org_a_member_client.post(
            "/api/v1/export/variance/excel",
            json=_variance_payload(),
        )
        assert response.status_code == 200
        history = org_a_member_client.mock_supabase._test_data["export_history"]
        assert len(history) == 1
        assert history[0]["format"] == "variance_excel"
        assert history[0]["file_name"].endswith(".xlsx")

    def test_variance_excel_requires_auth(self, base_client):
        response = base_client.post(
            "/api/v1/export/variance/excel",
            json=_variance_payload(),
        )
        assert response.status_code in (401, 403, 422)

    def test_variance_excel_no_snapshots_returns_404(self, org_a_member_client):
        org_a_member_client.mock_supabase._test_data["reconciliation_snapshots"] = []
        response = org_a_member_client.post(
            "/api/v1/export/variance/excel",
            json=_variance_payload(),
        )
        assert response.status_code == 404

    def test_variance_excel_prior_year_empty_uses_zero(self):
        """prior_total == 0 → variance_pct == 0 branch in the generator."""
        from openpyxl import load_workbook

        from app.api.v1.export import _generate_variance_excel

        buffer = _generate_variance_excel(
            snapshots_current=[{"total_recovery": "10000.00"}],
            snapshots_prior=[],
            current_year=2024,
            prior_year=2023,
            threshold_percent=10.0,
            property_data={"name": "Test Property"},
        )
        workbook = load_workbook(buffer)
        sheet = workbook.active
        # Header row 4; current-year total in B5, prior-year total in B6.
        assert sheet["B5"].value == 10000.0
        assert sheet["B6"].value == 0.0
        # Variance cell stores a fraction formatted as a percentage.
        assert sheet["C6"].value == 0.0

    def test_variance_excel_nonzero_variance_encoded_as_fraction(self):
        """+11.11% increase is stored as the 0.1111 fraction that Excel's
        "0.00%" number format renders back as 11.11%."""
        from openpyxl import load_workbook

        from app.api.v1.export import _generate_variance_excel

        buffer = _generate_variance_excel(
            snapshots_current=[{"total_recovery": "10000.00"}],
            snapshots_prior=[{"total_recovery": "9000.00"}],
            current_year=2024,
            prior_year=2023,
            threshold_percent=10.0,
            property_data={"name": "Test Property"},
        )
        workbook = load_workbook(buffer)
        sheet = workbook.active
        assert sheet["B5"].value == 10000.0
        assert sheet["B6"].value == 9000.0
        # (10000 - 9000) / 9000 * 100 = 11.111…% → stored as 0.1111… fraction.
        assert sheet["C6"].value == pytest.approx(0.1111111, rel=1e-4)
        assert sheet["C6"].number_format == "0.00%"


# ─────────────────────────────────────────────────────────────────────────────
# Additional branch coverage
# ─────────────────────────────────────────────────────────────────────────────


class TestAdditionalBranches:
    def test_pdf_download_no_snapshots_returns_404(self, org_a_member_client):
        org_a_member_client.mock_supabase._test_data["reconciliation_snapshots"] = []
        response = org_a_member_client.post(
            "/api/v1/export/pdf/download",
            json={"property_id": PROPERTY_ID, "year": YEAR},
        )
        assert response.status_code == 404

    def test_pdf_batch_no_snapshots_returns_404(self, org_a_member_client):
        org_a_member_client.mock_supabase._test_data["reconciliation_snapshots"] = []
        response = org_a_member_client.post(
            "/api/v1/export/pdf/batch",
            json={
                "property_id": PROPERTY_ID,
                "year": YEAR,
                "tenant_ids": [str(uuid4())],
            },
        )
        assert response.status_code == 404

    def test_pdf_batch_unmatched_tenants_returns_400(self, org_a_member_client):
        """Requested tenant_ids must match finalized snapshot lease_ids."""
        _seed_finalized_snapshots(org_a_member_client)
        response = org_a_member_client.post(
            "/api/v1/export/pdf/batch",
            json={
                "property_id": PROPERTY_ID,
                "year": YEAR,
                "tenant_ids": [str(uuid4())],
                "mode": "zip",
            },
        )
        assert response.status_code == 400


# ─────────────────────────────────────────────────────────────────────────────
# Board Presentation Export
# ─────────────────────────────────────────────────────────────────────────────


def _board_payload(
    *,
    property_id: str = PROPERTY_ID,
    year: int = YEAR,
    cap_rate: str = "0.07",
) -> dict:
    return {
        "property_id": property_id,
        "year": year,
        "cap_rate": cap_rate,
    }


def _seed_subscription(
    client,
    *,
    plan: str = "professional",
    status: str = "active",
) -> None:
    client.mock_supabase._test_data["subscriptions"] = [
        {
            "id": str(uuid4()),
            "organization_id": str(ORG_A_ID),
            "plan": plan,
            "status": status,
            "stripe_subscription_id": "sub_test123",
        }
    ]


class TestBoardExport:
    def test_board_preview_returns_pdf(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        _seed_subscription(org_a_member_client, plan="professional")
        response = org_a_member_client.post(
            "/api/v1/export/board/preview",
            json=_board_payload(),
        )
        assert response.status_code == 200
        assert "application/pdf" in response.headers["content-type"]

    def test_board_download_returns_pdf_attachment(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        _seed_subscription(org_a_member_client, plan="professional")
        response = org_a_member_client.post(
            "/api/v1/export/board/download",
            json=_board_payload(),
        )
        assert response.status_code == 200
        assert "application/pdf" in response.headers["content-type"]
        assert "attachment" in response.headers["content-disposition"]
        assert "board-presentation" in response.headers["content-disposition"]

    def test_board_preview_no_finalized_snapshots_returns_404(
        self, org_a_member_client
    ):
        _seed_subscription(org_a_member_client, plan="professional")
        # Don't seed snapshots
        response = org_a_member_client.post(
            "/api/v1/export/board/preview",
            json=_board_payload(),
        )
        assert response.status_code == 404

    def test_board_preview_invalid_cap_rate_returns_422(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        _seed_subscription(org_a_member_client, plan="professional")
        response = org_a_member_client.post(
            "/api/v1/export/board/preview",
            json=_board_payload(cap_rate="0.30"),  # above 25% max
        )
        assert response.status_code == 422

    def test_board_preview_cap_rate_below_minimum_returns_422(
        self, org_a_member_client
    ):
        _seed_finalized_snapshots(org_a_member_client)
        _seed_subscription(org_a_member_client, plan="professional")
        response = org_a_member_client.post(
            "/api/v1/export/board/preview",
            json=_board_payload(cap_rate="0.005"),  # below 1% min
        )
        assert response.status_code == 422

    def test_board_preview_missing_property_id_returns_422(self, org_a_member_client):
        response = org_a_member_client.post(
            "/api/v1/export/board/preview",
            json={"year": YEAR, "cap_rate": "0.07"},
        )
        assert response.status_code == 422

    def test_board_preview_multiple_tenants_sums_recovery(self, org_a_member_client):
        """PDF should be generated when multiple tenants are finalized."""
        _seed_finalized_snapshots(org_a_member_client, n=3)
        _seed_subscription(org_a_member_client, plan="professional")
        response = org_a_member_client.post(
            "/api/v1/export/board/preview",
            json=_board_payload(),
        )
        assert response.status_code == 200
        assert len(response.content) > 1000  # Real PDF, not empty

    def test_board_download_filename_includes_year_and_property(
        self, org_a_member_client
    ):
        _seed_finalized_snapshots(org_a_member_client)
        _seed_subscription(org_a_member_client, plan="professional")
        response = org_a_member_client.post(
            "/api/v1/export/board/download",
            json=_board_payload(),
        )
        assert str(YEAR) in response.headers["content-disposition"]

    def test_board_preview_allows_growth_alias_plan(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        _seed_subscription(org_a_member_client, plan="growth")
        response = org_a_member_client.post(
            "/api/v1/export/board/preview",
            json=_board_payload(),
        )
        assert response.status_code == 200

    def test_board_preview_allows_essentials_plan(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        _seed_subscription(org_a_member_client, plan="essentials")
        response = org_a_member_client.post(
            "/api/v1/export/board/preview",
            json=_board_payload(),
        )
        assert response.status_code == 200

    @pytest.mark.real_entitlements
    def test_board_preview_blocks_when_unsubscribed(self, org_a_member_client):
        _seed_finalized_snapshots(org_a_member_client)
        org_a_member_client.mock_supabase._test_data["subscriptions"] = []
        response = org_a_member_client.post(
            "/api/v1/export/board/preview",
            json=_board_payload(),
        )
        assert response.status_code == 402
        assert "payment_required" in response.json()["detail"]

    def test_board_preview_requires_auth(self, base_client):
        response = base_client.post(
            "/api/v1/export/board/preview",
            json=_board_payload(),
        )
        assert response.status_code in (401, 403, 422)

    def test_board_download_requires_auth(self, base_client):
        response = base_client.post(
            "/api/v1/export/board/download",
            json=_board_payload(),
        )
        assert response.status_code in (401, 403, 422)
