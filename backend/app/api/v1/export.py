"""
Export v2 endpoints — property-level exports.

Distinct from /api/v1/exports/ (snapshot-level) — these endpoints accept
property_id + year and operate over all finalized snapshots for that scope.

Endpoints:
  POST /api/v1/export/pdf/preview     → PDF bytes (inline)
  POST /api/v1/export/pdf/download    → PDF attachment
  POST /api/v1/export/pdf/batch       → ZIP of PDFs
  POST /api/v1/export/erp             → CSV/TXT for ERP import
  GET  /api/v1/export/history         → JSON list of past exports
  POST /api/v1/export/variance/pdf    → Variance comparison PDF
  POST /api/v1/export/variance/excel  → Variance comparison .xlsx
  POST /api/v1/export/board/preview   → Board presentation PDF (inline)
  POST /api/v1/export/board/download  → Board presentation PDF attachment
"""

import fnmatch
import logging
import zipfile
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Annotated, Any, Literal
from uuid import UUID, uuid4
from xml.sax.saxutils import escape

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Flowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.api.v1.exports import (
    MRIFormatter,
    TenantPacketGenerator,
    YardiFormatter,
    _load_export_context,
)
from app.auth.dependencies import OrgContext, require_full_access
from app.database.client import get_supabase_admin
from app.database.pagination import fetch_all_pages
from app.exceptions import BadRequestError, NotFoundError
from app.models import ReconciliationStatus
from app.services.analysis.statement_detail_advisor import (
    LineItemEntry,
    PoolLineItemDetail,
    StatementDetailAdvisor,
)
from app.services.billing.entitlements import has_full_access
from app.services.billing.feature_usage import record_feature_use
from app.services.formatting import format_usd

logger = logging.getLogger(__name__)

router = APIRouter()

# Private bucket that holds generated export files for re-download (F-024).
_EXPORT_BUCKET = "reports"
# Lifetime of the signed download URLs minted by GET /export/download/{id}.
_SIGNED_URL_TTL_SECONDS = 3600

# Service-role client used for storage uploads and signed-URL minting (F-024).
# Injected via Depends so tests exercise the mocked admin client rather than a
# real Supabase connection.
AdminClient = Annotated[Any, Depends(get_supabase_admin)]


def _require_professional_feature(
    ctx: OrgContext, feature_key: str, feature_name: str
) -> None:
    if not has_full_access(ctx):
        raise HTTPException(
            status_code=status.HTTP_402_PAYMENT_REQUIRED,
            detail=(
                f"payment_required: {feature_name} " "requires an active subscription."
            ),
        )


def _require_noi_board_access(ctx: OrgContext) -> None:
    _require_professional_feature(
        ctx, "portfolio_board_reports", "NOI Impact and Board Report"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Request / Response schemas
# ─────────────────────────────────────────────────────────────────────────────


class PDFExportRequest(BaseModel):
    property_id: UUID
    year: int
    include_charts: bool = False
    include_notes: bool = False
    tenant_ids: list[UUID] | None = None


class BatchPDFRequest(BaseModel):
    property_id: UUID
    year: int
    tenant_ids: list[UUID]
    mode: Literal["zip", "individual"] = "zip"


class ERPExportRequest(BaseModel):
    property_id: UUID
    year: int
    erp_system: Literal["yardi", "mri"]
    field_mappings: dict[str, str] = Field(default_factory=dict)


class VarianceReportRequest(BaseModel):
    property_id: UUID
    current_year: int
    prior_year: int
    threshold_percent: float = 10.0


class BoardExportRequest(BaseModel):
    property_id: UUID
    year: int
    cap_rate: Decimal = Field(
        default=Decimal("0.07"),
        ge=Decimal("0.01"),
        le=Decimal("0.25"),
        description="Capitalization rate as decimal (e.g. 0.07 for 7%). Must be between 1% and 25%.",
    )


class DetailLevelRequest(BaseModel):
    property_id: UUID
    year: int


class GroupingSuggestionResponse(BaseModel):
    category_name: str
    current_line_count: int
    suggested_label: str
    severity: str
    explanation: str


class ImmaterialItemResponse(BaseModel):
    account_code: str
    account_description: str
    amount: Decimal
    percent_of_total: Decimal
    pool_name: str


class DetailLevelAdvisoryResponse(BaseModel):
    total_line_items: int
    total_categories: int
    overall_severity: str
    summary: str
    grouping_suggestions: list[GroupingSuggestionResponse]
    immaterial_items: list[ImmaterialItemResponse]
    suggested_total_lines: int


class ExportDownloadResponse(BaseModel):
    """Signed URL for re-downloading a previously generated export (F-024)."""

    download_url: str
    file_name: str
    expires_at: str


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────


def _fetch_finalized_snapshots(
    ctx: OrgContext,
    property_id: UUID,
    year: int,
) -> list[dict]:
    """Return all finalized snapshots for a property/year from the DB."""
    property_result = (
        ctx.table("properties")
        .select("id")
        .eq("id", str(property_id))
        .eq("organization_id", str(ctx.organization_id))
        .execute()
    )
    if not property_result.data:
        return []

    year_start = f"{year}-01-01"
    year_end = f"{year}-12-31"
    result = (
        ctx.table("reconciliation_snapshots")
        # Embed property name and lease tenant_name so ERP write-back files
        # carry human-readable Property and Tenant columns rather than internal
        # UUID fragments when re-imported into the landlord's accounting system.
        .select("*, properties!inner(id, name), leases!inner(tenant_name)")
        .eq("organization_id", str(ctx.organization_id))
        .eq("property_id", str(property_id))
        .eq("status", ReconciliationStatus.FINALIZED.value)
        .lte("period_start_date", year_end)
        .gte("period_end_date", year_start)
        .execute()
    )
    return result.data or []


def _generate_property_pdf(
    ctx: OrgContext,
    snapshots: list[dict],
) -> BytesIO:
    """Generate a combined summary PDF for a set of snapshots."""
    if not snapshots:
        raise NotFoundError("reconciliation_snapshots", "property/year")

    # Use the first snapshot for property/org context
    snapshot = snapshots[0]
    lease_data, property_data, org_data = _load_export_context(ctx, snapshot)

    generator = TenantPacketGenerator(
        snapshot_data=snapshot,
        lease_data=lease_data,
        property_data=property_data,
        org_data=org_data,
    )
    return generator.generate()


def _snapshots_for_pdf_request(
    snapshots: list[dict], request: PDFExportRequest
) -> list[dict]:
    """Apply supported PDF request options to finalized snapshots."""
    if request.include_charts:
        raise BadRequestError("include_charts is not supported for PDF exports")
    if request.include_notes:
        raise BadRequestError("include_notes is not supported for PDF exports")
    if not request.tenant_ids:
        return snapshots
    if len(request.tenant_ids) > 1:
        raise BadRequestError("PDF preview/download supports one tenant_id")

    requested_lease_id = str(request.tenant_ids[0])
    filtered = [
        snapshot
        for snapshot in snapshots
        if str(snapshot.get("lease_id", "")) == requested_lease_id
    ]
    if not filtered:
        raise BadRequestError("No finalized snapshot matches the requested tenant_id")
    return filtered


def _generate_variance_pdf(
    snapshots_current: list[dict],
    snapshots_prior: list[dict],
    current_year: int,
    prior_year: int,
    threshold_percent: float,
    property_data: dict,
) -> BytesIO:
    """Generate a PDF comparing current vs prior year totals."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )
    styles = getSampleStyleSheet()
    story: list[Flowable] = []

    prop_name = escape(str(property_data.get("name", "Property")))
    story.append(Paragraph(f"Variance Report — {prop_name}", styles["Title"]))
    story.append(
        Paragraph(
            f"{current_year} vs {prior_year} | Threshold: {threshold_percent}%",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.25 * inch))

    # Aggregate totals across snapshots
    def total_recovery(snapshots: list[dict]) -> Decimal:
        return sum(
            (Decimal(str(s.get("total_recovery", "0"))) for s in snapshots),
            Decimal("0"),
        )

    current_total = total_recovery(snapshots_current)
    prior_total = total_recovery(snapshots_prior)
    if prior_total != 0:
        variance_pct = (current_total - prior_total) / prior_total * 100
    else:
        variance_pct = Decimal("0")

    data = [
        ["Period", "Total Recovery", "Variance"],
        [
            str(current_year),
            format_usd(current_total),
            "",
        ],
        [
            str(prior_year),
            format_usd(prior_total),
            f"{variance_pct:.2f}%",
        ],
    ]

    table = Table(data, colWidths=[2 * inch, 2.5 * inch, 2 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c5282")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 0.2 * inch))

    timestamp = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC")
    story.append(Paragraph(f"Generated: {timestamp}", styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer


def _generate_variance_excel(
    snapshots_current: list[dict],
    snapshots_prior: list[dict],
    current_year: int,
    prior_year: int,
    threshold_percent: float,
    property_data: dict,
) -> BytesIO:
    """Generate an .xlsx workbook comparing current vs prior year totals.

    Mirrors the data model of :func:`_generate_variance_pdf` so the variance
    Excel export matches the variance PDF. Money is summed with ``Decimal`` and
    written as numeric cells so spreadsheet users can re-compute against them.
    """

    def total_recovery(snapshots: list[dict]) -> Decimal:
        return sum(
            (Decimal(str(s.get("total_recovery", "0"))) for s in snapshots),
            Decimal("0"),
        )

    current_total = total_recovery(snapshots_current)
    prior_total = total_recovery(snapshots_prior)
    if prior_total != 0:
        variance_pct = (current_total - prior_total) / prior_total * 100
    else:
        variance_pct = Decimal("0")

    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:  # pragma: no cover - openpyxl always provides an active sheet
        sheet = workbook.create_sheet()
    sheet.title = "Variance"

    # Strip XML-illegal control characters: a property name carrying a stray
    # control byte from a CSV import would otherwise make openpyxl raise
    # IllegalCharacterError and crash the whole export (see excel_export.py).
    prop_name = ILLEGAL_CHARACTERS_RE.sub(
        "", str(property_data.get("name", "Property"))
    )
    header_fill = PatternFill(
        start_color="2C5282", end_color="2C5282", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    right = Alignment(horizontal="right")

    sheet["A1"] = f"Variance Report — {prop_name}"
    sheet["A1"].font = title_font
    sheet["A2"] = f"{current_year} vs {prior_year} | Threshold: {threshold_percent}%"

    header_row = 4
    headers = ["Period", "Total Recovery", "Variance"]
    for col, label in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font

    # Current year row (no variance figure, matching the PDF layout)
    sheet.cell(row=header_row + 1, column=1, value=str(current_year))
    current_cell = sheet.cell(row=header_row + 1, column=2, value=float(current_total))
    current_cell.number_format = "$#,##0.00"
    current_cell.alignment = right

    # Prior year row carries the variance percentage
    sheet.cell(row=header_row + 2, column=1, value=str(prior_year))
    prior_cell = sheet.cell(row=header_row + 2, column=2, value=float(prior_total))
    prior_cell.number_format = "$#,##0.00"
    prior_cell.alignment = right
    variance_cell = sheet.cell(
        row=header_row + 2, column=3, value=float(variance_pct) / 100
    )
    variance_cell.number_format = "0.00%"
    variance_cell.alignment = right

    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 14

    timestamp = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC")
    sheet.cell(row=header_row + 4, column=1, value=f"Generated: {timestamp}")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _upload_export_file(
    admin: Any,
    organization_id: UUID,
    property_id: UUID,
    file_name: str,
    data: bytes,
    content_type: str,
) -> str | None:
    """Upload an export file to the private reports bucket.

    Returns the storage object key on success, or None if the upload failed
    (the export itself is still streamed to the caller, so a storage outage
    must not break the live download — it only means the row is not later
    re-downloadable).
    """
    storage_path = f"reports/{organization_id}/{property_id}/{uuid4()}-{file_name}"
    try:
        admin.storage.from_(_EXPORT_BUCKET).upload(
            storage_path, data, {"content-type": content_type}
        )
        return storage_path
    except Exception:  # pragma: no cover - defensive: storage outage path
        logger.exception(
            "Failed to upload export %s to bucket %s; recording metadata only",
            file_name,
            _EXPORT_BUCKET,
        )
        return None


def _record_export_history(
    ctx: OrgContext,
    *,
    admin: Any,
    property_id: UUID,
    format: str,
    file_name: str,
    file_size: int | None,
    data: bytes,
    content_type: str,
) -> None:
    """Persist a completed export to storage and record its history row.

    The generated file is uploaded to the private reports bucket so it can be
    re-downloaded later via GET /export/download/{id} (F-024). The object key
    is stored on the row; if the upload fails the row is still written with a
    NULL storage_path (metadata only, not re-downloadable).
    """
    created_by_name = ctx.user.full_name or ctx.user.email
    storage_path = _upload_export_file(
        admin, ctx.organization_id, property_id, file_name, data, content_type
    )
    ctx.table("export_history").insert(
        {
            "organization_id": str(ctx.organization_id),
            "property_id": str(property_id),
            "format": format,
            "file_name": file_name,
            "file_size": file_size,
            "status": "completed",
            "created_by_name": created_by_name,
            "storage_path": storage_path,
        }
    ).execute()


def _buffer_size(buffer: BytesIO) -> int:
    return buffer.getbuffer().nbytes


def _text_buffer_size(buffer: StringIO) -> int:
    return len(buffer.getvalue().encode("utf-8"))


# ─────────────────────────────────────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────────────────────────────────────


@router.post("/pdf/preview", dependencies=[Depends(require_full_access)])
async def preview_pdf(
    request: PDFExportRequest,
    ctx: OrgContext,
) -> StreamingResponse:
    """Generate a PDF preview (inline) for a property/year."""
    snapshots = _fetch_finalized_snapshots(ctx, request.property_id, request.year)
    if not snapshots:
        raise NotFoundError("reconciliation_snapshots", str(request.property_id))

    snapshots = _snapshots_for_pdf_request(snapshots, request)
    pdf_buffer = _generate_property_pdf(ctx, snapshots)

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "inline"},
    )


@router.post("/pdf/download", dependencies=[Depends(require_full_access)])
async def download_pdf(
    request: PDFExportRequest,
    ctx: OrgContext,
    admin: AdminClient,
) -> StreamingResponse:
    """Stream a PDF download (attachment) for a property/year."""
    snapshots = _fetch_finalized_snapshots(ctx, request.property_id, request.year)
    if not snapshots:
        raise NotFoundError("reconciliation_snapshots", str(request.property_id))

    snapshots = _snapshots_for_pdf_request(snapshots, request)
    pdf_buffer = _generate_property_pdf(ctx, snapshots)
    filename = f"reconciliation-{request.year}-property.pdf"
    _record_export_history(
        ctx,
        admin=admin,
        property_id=request.property_id,
        format="pdf",
        file_name=filename,
        file_size=_buffer_size(pdf_buffer),
        data=pdf_buffer.getvalue(),
        content_type="application/pdf",
    )
    record_feature_use(admin, str(ctx.organization_id), "pdf_exports")
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.post("/pdf/batch", dependencies=[Depends(require_full_access)])
async def batch_pdf(
    request: BatchPDFRequest,
    ctx: OrgContext,
    admin: AdminClient,
) -> StreamingResponse:
    """Return a ZIP archive of PDFs, one per provided tenant/lease ID."""
    if not request.tenant_ids:
        raise BadRequestError("tenant_ids must not be empty")
    if request.mode != "zip":
        raise BadRequestError("mode='individual' is not supported; use mode='zip'")

    snapshots = _fetch_finalized_snapshots(ctx, request.property_id, request.year)
    if not snapshots:
        raise NotFoundError("reconciliation_snapshots", str(request.property_id))

    # Filter to requested tenants (match by lease_id)
    requested = {str(tid) for tid in request.tenant_ids}
    filtered = [s for s in snapshots if str(s.get("lease_id", "")) in requested]

    if not filtered:
        raise BadRequestError("No finalized snapshots match the requested tenant_ids")

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for snapshot in filtered:
            lease_data, property_data, org_data = _load_export_context(ctx, snapshot)
            pdf_buf = TenantPacketGenerator(
                snapshot_data=snapshot,
                lease_data=lease_data,
                property_data=property_data,
                org_data=org_data,
            ).generate()
            lease_id_short = str(snapshot.get("lease_id", "tenant"))[:8]
            fname = f"reconciliation-{request.year}-{lease_id_short}.pdf"
            zf.writestr(fname, pdf_buf.getvalue())

    zip_buffer.seek(0)
    timestamp = datetime.now(UTC).strftime("%Y%m%d")
    zip_filename = f"reconciliation-{request.year}-batch-{timestamp}.zip"
    _record_export_history(
        ctx,
        admin=admin,
        property_id=request.property_id,
        format="pdf_batch",
        file_name=zip_filename,
        file_size=_buffer_size(zip_buffer),
        data=zip_buffer.getvalue(),
        content_type="application/zip",
    )

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{zip_filename}"',
        },
    )


@router.post("/erp", dependencies=[Depends(require_full_access)])
async def export_erp(
    request: ERPExportRequest,
    ctx: OrgContext,
    admin: AdminClient,
) -> StreamingResponse:
    """Export finalized snapshots as ERP-compatible CSV/TXT."""
    if request.field_mappings:
        raise BadRequestError("field_mappings are not supported for ERP exports")

    snapshots = _fetch_finalized_snapshots(ctx, request.property_id, request.year)
    if not snapshots:
        raise NotFoundError("reconciliation_snapshots", str(request.property_id))

    if request.erp_system == "yardi":
        formatter: YardiFormatter | MRIFormatter = YardiFormatter(snapshots)
    else:
        formatter = MRIFormatter(snapshots)

    export_buffer = formatter.generate()
    filename = formatter.get_filename()
    media_type = formatter.get_media_type()
    _record_export_history(
        ctx,
        admin=admin,
        property_id=request.property_id,
        format=request.erp_system,
        file_name=filename,
        file_size=_text_buffer_size(export_buffer),
        data=export_buffer.getvalue().encode("utf-8"),
        content_type=media_type,
    )

    record_feature_use(admin, str(ctx.organization_id), "excel_exports")
    return StreamingResponse(
        export_buffer,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get("/history")
async def export_history(
    property_id: Annotated[UUID, Query(...)],
    ctx: OrgContext,
    format: Annotated[str | None, Query()] = None,
    page: Annotated[int, Query(ge=1)] = 1,
    page_size: Annotated[int, Query(ge=1, le=100)] = 25,
) -> dict:
    """Return paginated export history for a property."""
    start = (page - 1) * page_size
    end = start + page_size - 1
    query = (
        ctx.table("export_history")
        .select("*", count="exact")
        .eq("property_id", str(property_id))
        .eq("organization_id", str(ctx.organization_id))
        .order("created_at", desc=True)
    )

    if format is not None:
        query = query.eq("format", format)

    result = query.range(start, end).execute()
    items = result.data or []

    return {
        "items": items,
        "total": result.count if result.count is not None else len(items),
        "page": page,
        "page_size": page_size,
    }


@router.get("/download/{export_id}")
async def download_export(
    export_id: UUID,
    ctx: OrgContext,
    admin: AdminClient,
) -> ExportDownloadResponse:
    """Mint a fresh signed URL for re-downloading a past export (F-024).

    The export file is persisted to the private ``reports`` bucket when first
    generated. This endpoint looks up the history row, confirms it belongs to
    the caller's organization, and returns a short-lived signed URL. Legacy
    rows recorded before persisted-file support have a NULL ``storage_path``
    and are reported as no longer downloadable (410 Gone).
    """
    result = (
        ctx.table("export_history")
        .select("*")
        .eq("id", str(export_id))
        .eq("organization_id", str(ctx.organization_id))
        .execute()
    )
    rows = result.data or []
    if not rows:
        raise NotFoundError("export_history", str(export_id))

    row = rows[0]
    storage_path = row.get("storage_path")
    if not storage_path:
        raise HTTPException(
            status_code=status.HTTP_410_GONE,
            detail=(
                "This export is no longer available for download. "
                "Please re-generate it."
            ),
        )

    signed = admin.storage.from_(_EXPORT_BUCKET).create_signed_url(
        storage_path, _SIGNED_URL_TTL_SECONDS
    )
    signed_url = signed.get("signedURL") or signed.get("signedUrl")
    if not signed_url:  # pragma: no cover - defensive: storage outage path
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="Failed to create a download URL. Please try again.",
        )

    expires_at = (
        datetime.now(UTC) + timedelta(seconds=_SIGNED_URL_TTL_SECONDS)
    ).isoformat()
    return ExportDownloadResponse(
        download_url=signed_url,
        file_name=row.get("file_name") or "export",
        expires_at=expires_at,
    )


@router.post("/variance/pdf", dependencies=[Depends(require_full_access)])
async def variance_pdf(
    request: VarianceReportRequest,
    ctx: OrgContext,
    admin: AdminClient,
) -> StreamingResponse:
    """Generate a variance report PDF comparing current vs prior year."""
    current_snapshots = _fetch_finalized_snapshots(
        ctx, request.property_id, request.current_year
    )
    prior_snapshots = _fetch_finalized_snapshots(
        ctx, request.property_id, request.prior_year
    )

    if not current_snapshots and not prior_snapshots:
        raise NotFoundError("reconciliation_snapshots", str(request.property_id))

    # Load property context
    property_result = (
        ctx.table("properties")
        .select("id, name, address_line1, city, state, postal_code")
        .eq("id", str(request.property_id))
        .eq("organization_id", str(ctx.organization_id))
        .execute()
    )
    property_data: dict = {}
    if property_result.data:
        property_data = property_result.data[0]

    pdf_buffer = _generate_variance_pdf(
        snapshots_current=current_snapshots,
        snapshots_prior=prior_snapshots,
        current_year=request.current_year,
        prior_year=request.prior_year,
        threshold_percent=request.threshold_percent,
        property_data=property_data,
    )

    filename = f"variance-report-{request.current_year}-vs-{request.prior_year}.pdf"
    _record_export_history(
        ctx,
        admin=admin,
        property_id=request.property_id,
        format="variance_pdf",
        file_name=filename,
        file_size=_buffer_size(pdf_buffer),
        data=pdf_buffer.getvalue(),
        content_type="application/pdf",
    )
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.post("/variance/excel", dependencies=[Depends(require_full_access)])
async def variance_excel(
    request: VarianceReportRequest,
    ctx: OrgContext,
    admin: AdminClient,
) -> StreamingResponse:
    """Generate a variance report .xlsx comparing current vs prior year."""
    current_snapshots = _fetch_finalized_snapshots(
        ctx, request.property_id, request.current_year
    )
    prior_snapshots = _fetch_finalized_snapshots(
        ctx, request.property_id, request.prior_year
    )

    if not current_snapshots and not prior_snapshots:
        raise NotFoundError("reconciliation_snapshots", str(request.property_id))

    property_result = (
        ctx.table("properties")
        .select("id, name, address_line1, city, state, postal_code")
        .eq("id", str(request.property_id))
        .eq("organization_id", str(ctx.organization_id))
        .execute()
    )
    property_data: dict = {}
    if property_result.data:
        property_data = property_result.data[0]

    excel_buffer = _generate_variance_excel(
        snapshots_current=current_snapshots,
        snapshots_prior=prior_snapshots,
        current_year=request.current_year,
        prior_year=request.prior_year,
        threshold_percent=request.threshold_percent,
        property_data=property_data,
    )

    filename = f"variance-report-{request.current_year}-vs-{request.prior_year}.xlsx"
    _record_export_history(
        ctx,
        admin=admin,
        property_id=request.property_id,
        format="variance_excel",
        file_name=filename,
        file_size=_buffer_size(excel_buffer),
        data=excel_buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    return StreamingResponse(
        excel_buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# Board Presentation PDF generator
# ─────────────────────────────────────────────────────────────────────────────


def _generate_board_presentation_pdf(
    snapshots: list[dict],
    property_data: dict,
    org_data: dict,
    cap_rate: Decimal,
) -> BytesIO:
    """Generate a board presentation PDF showing NOI and asset value impact."""
    from app.services.calculation.noi_impact import NOIImpactInput, calculate_noi_impact

    total_recovery = sum(
        (Decimal(str(s.get("total_recovery", "0"))) for s in snapshots),
        Decimal("0"),
    )

    impact = calculate_noi_impact(
        NOIImpactInput(recovery_amount=total_recovery, cap_rate=cap_rate)
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )
    styles = getSampleStyleSheet()
    story: list[Flowable] = []

    # ── Header ────────────────────────────────────────────────────────────────
    org_name = org_data.get("name", "")
    prop_name = property_data.get("name", "Property")
    year = snapshots[0].get("period_start_date", "")[:4] if snapshots else ""

    story.append(
        Paragraph(
            f"CAM Recovery Impact Report — {prop_name}",
            styles["Title"],
        )
    )
    story.append(
        Paragraph(
            f"{org_name} | {year} Reconciliation Year",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.3 * inch))

    # ── Executive Summary Table ────────────────────────────────────────────────
    story.append(Paragraph("Executive Summary", styles["Heading2"]))
    story.append(Spacer(1, 0.1 * inch))

    summary_data = [
        ["Metric", "Amount"],
        ["CAM Recovery Amount", format_usd(impact.recovery_amount)],
        ["Additional Annual NOI", format_usd(impact.noi_lift)],
        ["Asset Value Increase", format_usd(impact.asset_value_lift)],
    ]

    summary_table = Table(summary_data, colWidths=[3.5 * inch, 3 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c5282")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f7fafc")),
                ("BACKGROUND", (0, 2), (-1, 2), colors.white),
                ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#f7fafc")),
                ("FONTNAME", (0, 3), (-1, 3), "Helvetica-Bold"),
                ("FONTSIZE", (1, 1), (1, 3), 12),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 0.3 * inch))

    # ── Methodology ───────────────────────────────────────────────────────────
    story.append(Paragraph("Calculation Methodology", styles["Heading2"]))
    story.append(Spacer(1, 0.1 * inch))
    story.append(
        Paragraph(
            "Commercial real estate is valued by dividing Net Operating Income (NOI) by the "
            "capitalization rate. CAM reconciliation recoveries represent permanent additional "
            "income — once recovered, this amount recurs annually, making it a direct NOI "
            "increase. Applying the cap rate formula converts this income stream into an "
            "equivalent increase in building market value.",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.2 * inch))
    story.append(
        Paragraph(
            f"Asset Value Lift = CAM Recovery ÷ Cap Rate = "
            f"{format_usd(impact.recovery_amount)} ÷ {cap_rate * 100:.1f}% = "
            f"{format_usd(impact.asset_value_lift)}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.2 * inch))

    # ── Cap Rate Assumption ────────────────────────────────────────────────────
    story.append(Paragraph("Cap Rate Assumption", styles["Heading2"]))
    story.append(Spacer(1, 0.1 * inch))
    story.append(
        Paragraph(
            f"This analysis uses a capitalization rate of <b>{cap_rate * 100:.2f}%</b>. "
            "Cap rates vary by market, asset class, and property quality. Adjust this "
            "assumption to reflect the prevailing cap rate for this asset.",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.3 * inch))

    # ── Footer ────────────────────────────────────────────────────────────────
    timestamp = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC")
    story.append(
        Paragraph(
            f"Prepared by CapVeri | Confidential | Generated: {timestamp}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.15 * inch))
    disclaimer_style = ParagraphStyle(
        "Disclaimer",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
        textColor=colors.HexColor("#6b7280"),
    )
    story.append(
        Paragraph(
            "These figures are generated automatically from data you provided and "
            "the cap rate you selected. They may contain errors. Review and verify "
            "all amounts before presenting them to a board, lender, or investor. "
            "CapVeri is not responsible for errors in outputs you did not "
            "independently verify.",
            disclaimer_style,
        )
    )

    doc.build(story)
    buffer.seek(0)
    return buffer


@router.post("/board/preview")
async def board_preview(
    request: BoardExportRequest,
    ctx: OrgContext,
) -> StreamingResponse:
    """Generate a board presentation PDF (inline) for a property/year."""
    _require_noi_board_access(ctx)
    snapshots = _fetch_finalized_snapshots(ctx, request.property_id, request.year)
    if not snapshots:
        raise NotFoundError("reconciliation_snapshots", str(request.property_id))

    snapshot = snapshots[0]
    _, property_data, org_data = _load_export_context(ctx, snapshot)

    pdf_buffer = _generate_board_presentation_pdf(
        snapshots, property_data, org_data, request.cap_rate
    )

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "inline"},
    )


@router.post("/board/download")
async def board_download(
    request: BoardExportRequest,
    ctx: OrgContext,
    admin: AdminClient,
) -> StreamingResponse:
    """Stream a board presentation PDF download for a property/year."""
    _require_noi_board_access(ctx)
    snapshots = _fetch_finalized_snapshots(ctx, request.property_id, request.year)
    if not snapshots:
        raise NotFoundError("reconciliation_snapshots", str(request.property_id))

    snapshot = snapshots[0]
    _, property_data, org_data = _load_export_context(ctx, snapshot)

    pdf_buffer = _generate_board_presentation_pdf(
        snapshots, property_data, org_data, request.cap_rate
    )
    year = request.year
    filename = f"board-presentation-{year}.pdf"
    _record_export_history(
        ctx,
        admin=admin,
        property_id=request.property_id,
        format="board_pdf",
        file_name=filename,
        file_size=_buffer_size(pdf_buffer),
        data=pdf_buffer.getvalue(),
        content_type="application/pdf",
    )
    record_feature_use(admin, str(ctx.organization_id), "noi_impact_calculator")
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# Detail Level Advisor
# ─────────────────────────────────────────────────────────────────────────────


def _build_pool_details(
    ctx: OrgContext,
    property_id: UUID,
    year: int,
) -> tuple[list[PoolLineItemDetail], Decimal]:
    """Build PoolLineItemDetail list from GL entries and expense pools."""
    # Fetch expense pools for property
    pools_data = fetch_all_pages(
        lambda: ctx.table("expense_pools")
        .select("id, name, pool_type")
        .eq("property_id", str(property_id))
    )
    if not pools_data:
        return [], Decimal("0")

    pool_ids = [p["id"] for p in pools_data]
    pool_info = {p["id"]: p for p in pools_data}

    # Fetch pool mappings
    mappings_data = fetch_all_pages(
        lambda: ctx.table("pool_mappings")
        .select("expense_pool_id, gl_account_pattern, allocation_percentage")
        .in_("expense_pool_id", pool_ids)
    )

    pool_mappings: dict[str, list[dict]] = {pid: [] for pid in pool_ids}
    for m in mappings_data:
        pool_mappings[m["expense_pool_id"]].append(m)

    # Fetch GL entries for property/year
    gl_entries = fetch_all_pages(
        lambda: ctx.table("gl_entries")
        .select("id, account_code, account_description, amount")
        .eq("property_id", str(property_id))
        .eq("period_year", year)
    )

    # Group GL entries into pools
    pool_items: dict[str, list[LineItemEntry]] = {pid: [] for pid in pool_ids}
    grand_total = Decimal("0")

    for entry in gl_entries:
        code = entry.get("account_code", "")
        amount = Decimal(str(entry.get("amount", 0)))
        desc = entry.get("account_description", code)
        grand_total += amount

        for pool_id, mapping_list in pool_mappings.items():
            for mapping in mapping_list:
                pattern = mapping["gl_account_pattern"].replace("%", "*")
                if fnmatch.fnmatch(code, pattern):
                    alloc = Decimal(str(mapping.get("allocation_percentage", 1)))
                    pool_items[pool_id].append(
                        LineItemEntry(
                            account_code=code,
                            account_description=desc,
                            amount=amount * alloc,
                        )
                    )
                    break

    result: list[PoolLineItemDetail] = []
    for pool_id, items in pool_items.items():
        if not items:
            continue
        info = pool_info[pool_id]
        pool_total = sum((i.amount for i in items), Decimal("0"))
        result.append(
            PoolLineItemDetail(
                pool_name=info["name"],
                pool_type=info.get("pool_type", "operating"),
                items=items,
                pool_total=pool_total,
            )
        )

    return result, grand_total


@router.post("/detail-advisor")
async def analyze_detail_level(
    request: DetailLevelRequest,
    ctx: OrgContext,
) -> DetailLevelAdvisoryResponse:
    """Analyze statement detail level and suggest grouping strategies."""
    _require_professional_feature(
        ctx, "statement_detail_advisor", "Statement Detail Advisor"
    )
    pool_details, _ = _build_pool_details(ctx, request.property_id, request.year)

    advisor = StatementDetailAdvisor()
    advisory = advisor.analyze(pool_details)

    return DetailLevelAdvisoryResponse(
        total_line_items=advisory.total_line_items,
        total_categories=advisory.total_categories,
        overall_severity=advisory.overall_severity.value,
        summary=advisory.summary,
        grouping_suggestions=[
            GroupingSuggestionResponse(
                category_name=s.category_name,
                current_line_count=s.current_line_count,
                suggested_label=s.suggested_label,
                severity=s.severity.value,
                explanation=s.explanation,
            )
            for s in advisory.grouping_suggestions
        ],
        immaterial_items=[
            ImmaterialItemResponse(
                account_code=i.account_code,
                account_description=i.account_description,
                amount=i.amount,
                percent_of_total=i.percent_of_total,
                pool_name=i.pool_name,
            )
            for i in advisory.immaterial_items
        ],
        suggested_total_lines=advisory.suggested_total_lines,
    )
