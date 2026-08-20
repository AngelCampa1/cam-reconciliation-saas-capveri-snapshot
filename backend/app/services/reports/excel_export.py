"""Excel export for historical analysis reports."""

import logging
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _safe_text(value: object) -> str:
    """Strip XML-illegal control characters from a string before it is written
    to a worksheet cell.

    Text fields in a report (pool names, anomaly explanations) can originate
    from messy Yardi/MRI CSV exports that carry stray control characters.
    openpyxl rejects those with ``IllegalCharacterError``, which would crash the
    entire tenant-facing report rather than emit one. The export is the last
    line before a document a landlord hands a tenant, so it sanitizes rather
    than fails: the control bytes are non-printable and carry no figure.
    """
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return ILLEGAL_CHARACTERS_RE.sub("", text)


def export_to_excel(report_data: dict[str, Any]) -> bytes:
    """Export historical analysis report to Excel format.

    Args:
        report_data: Dictionary containing report data with structure:
            - property: PropertySummary
            - years_compared: list[int]
            - year_over_year_comparison: dict with categories and totals
            - anomalies: list[DetectedAnomaly]

    Returns:
        Excel file as bytes
    """
    wb = Workbook()

    # Remove default sheet and create named sheets
    if wb.active is not None:
        wb.remove(wb.active)

    # Create Year-over-Year sheet
    _create_yoy_sheet(wb, report_data)

    # Create Anomalies sheet
    _create_anomalies_sheet(wb, report_data)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _create_yoy_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Create year-over-year comparison sheet.

    Args:
        wb: Workbook to add sheet to
        report_data: Report data dictionary
    """
    ws = wb.create_sheet("Year-over-Year Comparison")

    years = report_data["years_compared"]
    categories = report_data["year_over_year_comparison"]["categories"]
    totals = report_data["year_over_year_comparison"]["totals"]

    # Header row
    headers = ["Expense Pool"] + [str(y) for y in years] + ["Variance %"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="1E3A8A")
        cell.fill = PatternFill(start_color="E0E7FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, category in enumerate(categories, 2):
        # Pool name
        ws.cell(row=row_idx, column=1, value=_safe_text(category["name"]))

        # Amounts for each year
        # FIX AS-6: Support both dict and list formats for amounts
        # Dict format: {year: amount} - sparse years work correctly
        # List format: [amount1, amount2, ...] - legacy format, assumes sequential
        amounts_data = category.get("amounts", {})
        for col_idx, year in enumerate(years, 2):
            if isinstance(amounts_data, dict):
                # Preferred format: lookup by year key
                amount = amounts_data.get(year, 0)
            else:
                # Legacy list format: bounds-checked index lookup
                list_idx = col_idx - 2
                amount = amounts_data[list_idx] if list_idx < len(amounts_data) else 0
            cell = ws.cell(row=row_idx, column=col_idx, value=amount)
            cell.number_format = "$#,##0"
            cell.alignment = Alignment(horizontal="right")

        # Variance percentage
        variance = category["variance_percent"]
        variance_cell = ws.cell(row=row_idx, column=len(headers), value=variance / 100)
        variance_cell.number_format = "0.0%"
        variance_cell.alignment = Alignment(horizontal="right")

        # Color based on variance magnitude
        if abs(variance) > 15:
            variance_cell.fill = PatternFill(start_color="FFCCCC", fill_type="solid")
        elif abs(variance) > 5:
            variance_cell.fill = PatternFill(start_color="FFFFCC", fill_type="solid")

    # Totals row
    totals_row = len(categories) + 2
    ws.cell(row=totals_row, column=1, value="Total").font = Font(bold=True)

    totals_dict = {t["year"]: t["total"] for t in totals}
    for col_idx, year in enumerate(years, 2):
        total = totals_dict.get(year, 0)
        cell = ws.cell(row=totals_row, column=col_idx, value=total)
        cell.number_format = "$#,##0"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")

    # Calculate and add total variance
    if len(totals) >= 2:
        first_total = totals_dict.get(years[0], 0)
        last_total = totals_dict.get(years[-1], 0)
        if first_total > 0:
            total_variance = ((last_total - first_total) / first_total) * 100
            variance_cell = ws.cell(
                row=totals_row, column=len(headers), value=total_variance / 100
            )
            variance_cell.number_format = "0.0%"
            variance_cell.font = Font(bold=True)
            variance_cell.alignment = Alignment(horizontal="right")

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Fine-print verification disclaimer below the table
    disclaimer_row = totals_row + 2
    disclaimer_cell = ws.cell(
        row=disclaimer_row,
        column=1,
        value=(
            "This report is generated automatically from data you provided and "
            "may contain errors. Review and verify all figures before relying on "
            "them or billing any tenant. CapVeri is not responsible for errors "
            "in outputs you did not independently verify."
        ),
    )
    disclaimer_cell.font = Font(size=8, italic=True, color="6B7280")
    disclaimer_cell.alignment = Alignment(horizontal="left", vertical="top")
    ws.merge_cells(
        start_row=disclaimer_row,
        start_column=1,
        end_row=disclaimer_row,
        end_column=len(headers),
    )


def _create_anomalies_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Create anomalies sheet.

    Args:
        wb: Workbook to add sheet to
        report_data: Report data dictionary
    """
    ws = wb.create_sheet("Detected Anomalies")

    anomalies = report_data.get("anomalies", [])

    # Header row
    headers = [
        "Severity",
        "Expense Pool",
        "Type",
        "Current",
        "Expected",
        "Variance %",
        "Explanation",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="1E3A8A")
        cell.fill = PatternFill(start_color="E0E7FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if not anomalies:
        # No anomalies message
        ws.cell(
            row=2,
            column=1,
            value="No anomalies detected. All expense patterns appear normal.",
        )
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
        return

    # Data rows
    for row_idx, anomaly in enumerate(anomalies, 2):
        # Severity with color coding
        severity = anomaly["severity"].upper()
        severity_cell = ws.cell(row=row_idx, column=1, value=severity)
        severity_cell.font = Font(bold=True)

        if severity == "CRITICAL":
            severity_cell.fill = PatternFill(start_color="FEE2E2", fill_type="solid")
            severity_cell.font = Font(bold=True, color="991B1B")
        elif severity == "WARNING":
            severity_cell.fill = PatternFill(start_color="FEF3C7", fill_type="solid")
            severity_cell.font = Font(bold=True, color="92400E")
        else:  # INFO
            severity_cell.fill = PatternFill(start_color="DBEAFE", fill_type="solid")
            severity_cell.font = Font(bold=True, color="1E3A8A")

        # Pool name
        ws.cell(row=row_idx, column=2, value=_safe_text(anomaly["pool_name"]))

        # Type
        anomaly_type = anomaly["anomaly_type"].replace("_", " ").title()
        ws.cell(row=row_idx, column=3, value=_safe_text(anomaly_type))

        # Current value
        current_cell = ws.cell(row=row_idx, column=4, value=anomaly["current_value"])
        current_cell.number_format = "$#,##0"

        # Expected value
        expected_cell = ws.cell(row=row_idx, column=5, value=anomaly["expected_value"])
        expected_cell.number_format = "$#,##0"

        # Variance percentage
        variance_cell = ws.cell(
            row=row_idx, column=6, value=anomaly["variance_percent"] / 100
        )
        variance_cell.number_format = "0.0%"

        # Explanation
        ws.cell(row=row_idx, column=7, value=_safe_text(anomaly["explanation"]))

    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 50
