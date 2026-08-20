"""SB 1103 Compliance Service.

California SB 1103 (effective January 1, 2025) requires landlords to provide
Qualified Commercial Tenants (QCTs) with an itemized 18-month historical CAM
expense ledger within 30 days of a written request. Failure to comply gives
tenants the right to rescind their lease.

This module provides:
- Window start / response deadline calculators
- GL entry query for the compliance window
- Export data assembler
- PDF and Excel export generators
- Deadline alert query
"""

import logging
from datetime import date, timedelta
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from typing import Any, cast
from uuid import UUID

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Flowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.database.pagination import fetch_all_pages, fetch_all_pages_chunked_in
from app.exceptions import NotFoundError
from app.models.sb1103 import (
    SB1103DeadlineAlert,
    SB1103ExportData,
    SB1103GLEntry,
    SB1103Request,
)
from app.services.formatting import format_usd

logger = logging.getLogger(__name__)


def _safe_text(value: object) -> str:
    """Strip XML-illegal control characters before writing to a worksheet cell.

    Text fields (pool/property/tenant/vendor names, descriptions) originate from
    messy Yardi/MRI CSV exports that can carry stray control characters.  openpyxl
    rejects those with ``IllegalCharacterError``, crashing the entire export.
    Sanitize rather than fail — the bytes are non-printable and carry no value.
    """
    return ILLEGAL_CHARACTERS_RE.sub("", str(value))


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------


def compute_window_start(request_date: date) -> date:
    """Compute the 18-month lookback window start using calendar months.

    Uses dateutil.relativedelta(months=18) — calendar months, not 540 days.
    This is legally accurate for SB 1103 purposes.
    """
    return cast(date, request_date - relativedelta(months=18))


def compute_response_deadline(request_date: date) -> date:
    """Compute the 30-day response deadline.

    California Civil Code Section 1938.1 (SB 1103) requires the landlord
    to respond within 30 calendar days of the written request.
    """
    return request_date + timedelta(days=30)


# ---------------------------------------------------------------------------
# GL entry query
# ---------------------------------------------------------------------------


def get_gl_entries_for_window(
    ctx: Any,
    property_id: UUID,
    window_start: date,
    window_end: date,
) -> list[dict]:
    """Query GL entries for the given property within the compliance window.

    Selects: id, account_code, account_description, amount, transaction_date,
             vendor_name, description, import_batch_id.
    Ordered by transaction_date ascending.
    """
    return fetch_all_pages(
        lambda: ctx.table("gl_entries")
        .select(
            "id, account_code, account_description, amount, transaction_date,"
            " vendor_name, description, import_batch_id"
        )
        .eq("property_id", str(property_id))
        .gte("transaction_date", window_start.isoformat())
        .lte("transaction_date", window_end.isoformat())
        .order("transaction_date")
    )


# ---------------------------------------------------------------------------
# Export data assembly
# ---------------------------------------------------------------------------


def build_sb1103_export_data(ctx: Any, request_id: UUID) -> SB1103ExportData:
    """Assemble all data needed to generate a complete SB 1103 export.

    Fetches the request, property, lease (with recovery_profile), and GL entries.
    Computes per-entry tenant_share_amount and category_subtotals.

    Raises:
        NotFoundError: If the request, property, or lease cannot be found.
    """
    # Fetch request
    req_result = (
        ctx.table("sb1103_requests").select("*").eq("id", str(request_id)).execute()
    )
    if not req_result.data:
        raise NotFoundError("SB1103Request", str(request_id))
    request = SB1103Request.model_validate(req_result.data[0])

    # Fetch property
    prop_result = (
        ctx.table("properties")
        .select(
            "id, name, address_line1, address_line2, city, state, postal_code,"
            " organization_id"
        )
        .eq("id", str(request.property_id))
        .execute()
    )
    if not prop_result.data:
        raise NotFoundError("Property", str(request.property_id))
    prop = prop_result.data[0]

    # Fetch lease (with recovery_profile for pro_rata_share)
    lease_result = (
        ctx.table("leases")
        .select("id, tenant_name, property_id, recovery_profile")
        .eq("id", str(request.lease_id))
        .execute()
    )
    if not lease_result.data:
        raise NotFoundError("Lease", str(request.lease_id))
    lease = lease_result.data[0]
    if str(lease.get("property_id")) != str(request.property_id):
        raise ValueError("Lease does not belong to the SB 1103 request property")

    # Extract pro_rata_share from recovery_profile
    recovery = lease.get("recovery_profile") or {}
    pro_rata_share = Decimal(str(recovery.get("pro_rata_share", "0")))
    if pro_rata_share <= Decimal("0"):
        raise ValueError(
            f"Lease {request.lease_id} has no valid pro_rata_share in its recovery_profile. "
            "Cannot generate SB 1103 export with a zero or missing pro-rata share."
        )

    # Build property address string
    city = prop.get("city", "")
    state = prop.get("state", "")
    postal_code = prop.get("postal_code", "")
    street_parts = [prop.get("address_line1", ""), prop.get("address_line2", "")]
    address_parts = [", ".join(p for p in street_parts if p)]
    if city or state or postal_code:
        address_parts.append(f"{city}, {state} {postal_code}".strip(", "))
    property_address = ", ".join(p for p in address_parts if p)

    is_ca_property = state.upper() == "CA"

    # Fetch GL entries for window
    raw_entries = get_gl_entries_for_window(
        ctx,
        request.property_id,
        request.window_start_date,
        request.window_end_date,
    )

    # Build SB1103GLEntry list with tenant_share_amount
    gl_entries: list[SB1103GLEntry] = []
    category_subtotals: dict[str, Decimal] = {}
    total_cam_expenses = Decimal("0")
    total_tenant_share = Decimal("0")

    for row in raw_entries:
        amount = Decimal(str(row["amount"])).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        tenant_share = (amount * pro_rata_share).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        entry = SB1103GLEntry(
            id=UUID(row["id"]),
            transaction_date=date.fromisoformat(row["transaction_date"]),
            account_code=row["account_code"],
            account_description=row["account_description"],
            vendor_name=row.get("vendor_name"),
            description=row.get("description"),
            amount=amount,
            import_batch_id=UUID(row["import_batch_id"]),
            tenant_share_amount=tenant_share,
        )
        gl_entries.append(entry)

        desc = row["account_description"]
        category_subtotals[desc] = (
            category_subtotals.get(desc, Decimal("0")) + tenant_share
        )
        total_cam_expenses += amount
        total_tenant_share += tenant_share

    return SB1103ExportData(
        request=request,
        property_address=property_address,
        property_name=prop.get("name", "") if isinstance(prop, dict) else "",
        tenant_name=lease.get("tenant_name", "") if isinstance(lease, dict) else "",
        pro_rata_share=pro_rata_share,
        gl_entries=gl_entries,
        category_subtotals=category_subtotals,
        is_ca_property=is_ca_property,
        total_cam_expenses=total_cam_expenses,
        total_tenant_share=total_tenant_share,
    )


# ---------------------------------------------------------------------------
# PDF Generator
# ---------------------------------------------------------------------------


class SB1103PacketGenerator:
    """ReportLab PDF generator for SB 1103 compliance export packets.

    Standalone (does NOT extend TenantPacketGenerator — incompatible data shapes).
    Follows the same _build_*() section pattern.

    Sections:
    - Cover page with certification language referencing California Civil Code
      Section 1938.1 (SB 1103)
    - Itemized ledger table
    - Category subtotals
    - Pro-rata methodology
    - Gross-up disclosure
    - Footer
    """

    def __init__(self, export_data: SB1103ExportData) -> None:
        self.data = export_data
        self.styles = getSampleStyleSheet()
        self._setup_styles()

    def _setup_styles(self) -> None:
        self.title_style = ParagraphStyle(
            "SB1103Title",
            parent=self.styles["Title"],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor("#1E3A8A"),
        )
        self.heading_style = ParagraphStyle(
            "SB1103Heading",
            parent=self.styles["Heading2"],
            fontSize=12,
            spaceBefore=18,
            spaceAfter=6,
            textColor=colors.HexColor("#1E3A8A"),
        )
        self.body_style = ParagraphStyle(
            "SB1103Body",
            parent=self.styles["Normal"],
            fontSize=9,
            spaceAfter=4,
        )
        self.certification_style = ParagraphStyle(
            "SB1103Cert",
            parent=self.styles["Normal"],
            fontSize=9,
            spaceAfter=6,
            borderPad=8,
            borderColor=colors.HexColor("#1E3A8A"),
        )

    def _fmt_date(self, d: date) -> str:
        # Use day directly to avoid platform-specific %-d vs %#d differences
        return f"{d.strftime('%B')} {d.day}, {d.year}"

    def _fmt_money(self, amount: Decimal) -> str:
        return format_usd(amount)

    def _build_cover(self) -> list[Flowable]:
        req = self.data.request
        elements: list[Flowable] = []

        elements.append(
            Paragraph(
                "California SB 1103 — CAM Expense Disclosure",
                self.title_style,
            )
        )
        elements.append(
            Paragraph(
                "Itemized Common Area Maintenance Ledger",
                self.heading_style,
            )
        )
        elements.append(Spacer(1, 0.2 * inch))

        # Property + tenant info table
        info_data = [
            ["Property:", self.data.property_name],
            ["Address:", self.data.property_address],
            ["Tenant:", self.data.tenant_name],
            ["Requestor Name:", req.requested_by_name],
            ["Requestor Email:", req.requested_by_email],
            ["Request Date:", self._fmt_date(req.request_date)],
            ["Response Deadline:", self._fmt_date(req.response_deadline)],
            [
                "Ledger Period:",
                f"{self._fmt_date(req.window_start_date)} — {self._fmt_date(req.window_end_date)}",
            ],
        ]
        t = Table(info_data, colWidths=[1.8 * inch, 4.5 * inch])
        t.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 0),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F0F4FF")],
                    ),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        elements.append(t)
        elements.append(Spacer(1, 0.3 * inch))

        # Landlord certification block
        elements.append(Paragraph("Landlord Certification", self.heading_style))
        cert_text = (
            "This document is provided pursuant to <b>California Civil Code Section 1938.1 "
            "(SB 1103)</b> in response to a written request from a Qualified Commercial Tenant "
            "(QCT). The landlord certifies that the Common Area Maintenance (CAM) expense "
            "records contained herein are true and accurate to the best of the landlord's "
            "knowledge, drawn from the property's general ledger for the period "
            f"<b>{self._fmt_date(req.window_start_date)}</b> through "
            f"<b>{self._fmt_date(req.window_end_date)}</b>. "
            "This disclosure has been prepared in accordance with the landlord's obligation "
            "under California law to provide itemized CAM expense documentation within "
            "30 days of a qualifying written request."
        )
        elements.append(Paragraph(cert_text, self.certification_style))
        elements.append(Spacer(1, 0.2 * inch))

        if not self.data.is_ca_property:
            elements.append(
                Paragraph(
                    "<i>Note: This property is not recorded as being located in California. "
                    "SB 1103 obligations may not apply. Please verify applicability with "
                    "legal counsel.</i>",
                    self.body_style,
                )
            )
        return elements

    def _build_ledger(self) -> list[Flowable]:
        elements: list[Flowable] = []
        elements.append(Paragraph("Itemized CAM Expense Ledger", self.heading_style))
        elements.append(Spacer(1, 0.1 * inch))

        headers = [
            "Date",
            "Account",
            "Description / Vendor",
            "Full Amount",
            "Your Share",
        ]
        rows = [headers]
        for entry in self.data.gl_entries:
            desc = entry.description or entry.account_description
            vendor = entry.vendor_name or ""
            desc_vendor = f"{desc}\n{vendor}" if vendor else desc
            rows.append(
                [
                    entry.transaction_date.strftime("%m/%d/%Y"),
                    f"{entry.account_code}\n{entry.account_description}",
                    desc_vendor,
                    self._fmt_money(entry.amount),
                    self._fmt_money(entry.tenant_share_amount),
                ]
            )

        # Totals row
        rows.append(
            [
                "",
                "",
                "TOTAL",
                self._fmt_money(self.data.total_cam_expenses),
                self._fmt_money(self.data.total_tenant_share),
            ]
        )

        col_widths = [0.9 * inch, 1.3 * inch, 2.2 * inch, 1.0 * inch, 1.0 * inch]
        t = Table(rows, colWidths=col_widths)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -2),
                        [colors.white, colors.HexColor("#F0F4FF")],
                    ),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E0E7FF")),
                    ("FONTNAME", (2, -1), (-1, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(t)
        return elements

    def _build_category_subtotals(self) -> list[Flowable]:
        elements: list[Flowable] = []
        elements.append(Paragraph("Category Subtotals", self.heading_style))
        elements.append(Spacer(1, 0.1 * inch))

        rows = [["Expense Category", "Tenant Share"]]
        for category, subtotal in sorted(self.data.category_subtotals.items()):
            rows.append([category, self._fmt_money(subtotal)])
        rows.append(["TOTAL", self._fmt_money(self.data.total_tenant_share)])

        t = Table(rows, colWidths=[4.0 * inch, 1.5 * inch])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -2),
                        [colors.white, colors.HexColor("#F0F4FF")],
                    ),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E0E7FF")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        elements.append(t)
        return elements

    def _build_methodology(self) -> list[Flowable]:
        elements: list[Flowable] = []
        elements.append(Paragraph("Pro-Rata Methodology", self.heading_style))

        method_text = (
            f"The tenant's pro-rata share of Common Area Maintenance expenses is "
            f"<b>{self.data.pro_rata_share * Decimal('100'):.4f}%</b> of total CAM charges, "
            f"as specified in the lease agreement. Each line item amount shown under "
            f"'Your Share' is calculated as: Full Amount × Pro-Rata Share, "
            f"rounded to the nearest cent (ROUND_HALF_UP)."
        )
        elements.append(Paragraph(method_text, self.body_style))
        elements.append(Spacer(1, 0.2 * inch))

        elements.append(Paragraph("Gross-Up Disclosure", self.heading_style))
        gross_up_text = (
            "Where applicable, certain variable expenses may have been grossed up to "
            "reflect 100% occupancy, as permitted by the lease. The ledger above reflects "
            "actual expenses as recorded in the general ledger. Any gross-up adjustments "
            "are disclosed separately in the annual CAM reconciliation statement."
        )
        elements.append(Paragraph(gross_up_text, self.body_style))
        return elements

    def generate(self) -> BytesIO:
        """Generate the complete SB 1103 compliance PDF packet."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        elements: list[Flowable] = []
        elements.extend(self._build_cover())
        elements.extend(self._build_ledger())
        elements.extend(self._build_category_subtotals())
        elements.extend(self._build_methodology())

        doc.build(elements)
        buffer.seek(0)
        return buffer


def generate_pdf_export(export_data: SB1103ExportData) -> BytesIO:
    """Generate a SB 1103 compliance PDF packet."""
    return SB1103PacketGenerator(export_data).generate()


# ---------------------------------------------------------------------------
# Excel Generator
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1E3A8A", fill_type="solid")
_ALT_FILL = PatternFill(start_color="F0F4FF", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="E0E7FF", fill_type="solid")


def _auto_width(ws: Any, min_width: int = 10, max_width: int = 50) -> None:
    for col in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(
            max(length + 2, min_width), max_width
        )


def generate_excel_export(export_data: SB1103ExportData) -> BytesIO:
    """Generate a SB 1103 compliance Excel workbook with 4 sheets.

    Sheets: Cover, Ledger, Category Subtotals, Methodology.
    Amounts are stored as strings ("$1,234.56") to avoid float precision issues.
    """
    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    _build_cover_sheet(wb, export_data)
    _build_ledger_sheet(wb, export_data)
    _build_subtotals_sheet(wb, export_data)
    _build_methodology_sheet(wb, export_data)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _fmt_money(amount: Decimal) -> str:
    return format_usd(amount)


def _build_cover_sheet(wb: Workbook, data: SB1103ExportData) -> None:
    ws = wb.create_sheet("Cover")
    req = data.request

    ws["A1"] = "California SB 1103 — CAM Expense Disclosure"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A8A")

    info = [
        ("Property", _safe_text(data.property_name)),
        ("Address", _safe_text(data.property_address)),
        ("Tenant", _safe_text(data.tenant_name)),
        ("Requestor Name", _safe_text(req.requested_by_name)),
        ("Requestor Email", _safe_text(req.requested_by_email)),
        (
            "Request Date",
            f"{req.request_date.strftime('%B')} {req.request_date.day}, {req.request_date.year}",
        ),
        (
            "Response Deadline",
            f"{req.response_deadline.strftime('%B')} {req.response_deadline.day}, {req.response_deadline.year}",
        ),
        (
            "Ledger Period",
            f"{req.window_start_date.strftime('%B')} {req.window_start_date.day}, {req.window_start_date.year}"
            f" — "
            f"{req.window_end_date.strftime('%B')} {req.window_end_date.day}, {req.window_end_date.year}",
        ),
        ("Pro-Rata Share", f"{data.pro_rata_share * Decimal('100'):.4f}%"),
        ("Total CAM Expenses", _fmt_money(data.total_cam_expenses)),
        ("Tenant Total Share", _fmt_money(data.total_tenant_share)),
    ]

    for row_idx, (label, value) in enumerate(info, start=3):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45

    cert_row = len(info) + 5
    ws.cell(row=cert_row, column=1, value="Landlord Certification").font = Font(
        bold=True, size=11, color="1E3A8A"
    )
    cert_text = (
        "This document is provided pursuant to California Civil Code Section 1938.1 "
        "(SB 1103) in response to a written request from a Qualified Commercial Tenant. "
        "The landlord certifies the CAM expense records are true and accurate."
    )
    ws.cell(row=cert_row + 1, column=1, value=cert_text)
    ws.cell(row=cert_row + 1, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(
        start_row=cert_row + 1, start_column=1, end_row=cert_row + 3, end_column=2
    )
    ws.row_dimensions[cert_row + 1].height = 60


def _build_ledger_sheet(wb: Workbook, data: SB1103ExportData) -> None:
    ws = wb.create_sheet("Ledger")

    headers = [
        "Date",
        "Account Code",
        "Account Description",
        "Vendor",
        "Description",
        "Full Amount",
        "Your Share",
        "Import Batch ID",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, entry in enumerate(data.gl_entries, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        values = [
            entry.transaction_date.strftime("%Y-%m-%d"),
            _safe_text(entry.account_code),
            _safe_text(entry.account_description),
            _safe_text(entry.vendor_name or ""),
            _safe_text(entry.description or ""),
            _fmt_money(entry.amount),
            _fmt_money(entry.tenant_share_amount),
            str(entry.import_batch_id),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fill:
                cell.fill = fill

    # Totals row
    total_row = len(data.gl_entries) + 2
    ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=_fmt_money(data.total_cam_expenses)).font = (
        Font(bold=True)
    )
    ws.cell(row=total_row, column=7, value=_fmt_money(data.total_tenant_share)).font = (
        Font(bold=True)
    )
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).fill = _TOTAL_FILL

    _auto_width(ws)


def _build_subtotals_sheet(wb: Workbook, data: SB1103ExportData) -> None:
    ws = wb.create_sheet("Category Subtotals")

    headers = ["Expense Category", "Tenant Share"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (category, subtotal) in enumerate(
        sorted(data.category_subtotals.items()), start=2
    ):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        c1 = ws.cell(row=row_idx, column=1, value=_safe_text(category))
        c2 = ws.cell(row=row_idx, column=2, value=_fmt_money(subtotal))
        if fill:
            c1.fill = fill
            c2.fill = fill

    total_row = len(data.category_subtotals) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=_fmt_money(data.total_tenant_share)).font = (
        Font(bold=True)
    )
    for col in range(1, 3):
        ws.cell(row=total_row, column=col).fill = _TOTAL_FILL

    _auto_width(ws)


def _build_methodology_sheet(wb: Workbook, data: SB1103ExportData) -> None:
    ws = wb.create_sheet("Methodology")

    ws["A1"] = "Pro-Rata Methodology"
    ws["A1"].font = Font(bold=True, size=12, color="1E3A8A")

    ws["A3"] = "Pro-Rata Share"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = f"{data.pro_rata_share * Decimal('100'):.4f}%"

    ws["A4"] = "Calculation Method"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = "Full Amount × Pro-Rata Share, rounded to nearest cent (ROUND_HALF_UP)"

    ws["A6"] = "Gross-Up Disclosure"
    ws["A6"].font = Font(bold=True, size=11, color="1E3A8A")
    ws["A7"] = (
        "Where applicable, certain variable expenses may have been grossed up to reflect "
        "100% occupancy, as permitted by the lease. The ledger reflects actual expenses "
        "as recorded in the general ledger."
    )
    ws["A7"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A7:B9")
    ws.row_dimensions[7].height = 60

    ws["A11"] = "Legal Reference"
    ws["A11"].font = Font(bold=True)
    ws["B11"] = "California Civil Code Section 1938.1 (SB 1103)"

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 65


# ---------------------------------------------------------------------------
# Deadline alerts
# ---------------------------------------------------------------------------


def get_deadline_alerts(ctx: Any, days_warning: int = 7) -> list[SB1103DeadlineAlert]:
    """Return SB 1103 requests with response deadlines within days_warning days.

    Includes overdue requests (negative days_remaining).
    Excludes 'delivered' requests.
    Batch-fetches property and tenant names to minimise round-trips.
    """
    today = date.today()
    cutoff = today + timedelta(days=days_warning)

    result = (
        ctx.table("sb1103_requests")
        .select("id, property_id, lease_id, response_deadline, status")
        .lte("response_deadline", cutoff.isoformat())
        .neq("status", "delivered")
        .order("response_deadline")
        .execute()
    )
    requests = result.data or []
    if not requests:
        return []

    property_ids = list({r["property_id"] for r in requests})
    lease_ids = list({r["lease_id"] for r in requests})

    # Chunk + paginate both id filters. PostgREST encodes ``.in_()`` values into
    # the GET URL, so a property/org with hundreds of open SB 1103 requests
    # overflows the request-line buffer and returns HTTP 414 ("URI too long").
    # The raw un-paginated ``.execute()`` would also silently cap at PostgREST's
    # 1000-row default. ``fetch_all_pages_chunked_in`` splits into <=100-value
    # chunks, each fully paginated, then concatenates.
    props_data = fetch_all_pages_chunked_in(
        lambda chunk: ctx.table("properties").select("id, name").in_("id", chunk),
        property_ids,
    )
    prop_map = {p["id"]: p["name"] for p in props_data}

    leases_data = fetch_all_pages_chunked_in(
        lambda chunk: ctx.table("leases").select("id, tenant_name").in_("id", chunk),
        lease_ids,
    )
    lease_map = {row["id"]: row["tenant_name"] for row in leases_data}

    alerts: list[SB1103DeadlineAlert] = []
    for r in requests:
        deadline = date.fromisoformat(r["response_deadline"])
        days_remaining = (deadline - today).days
        alerts.append(
            SB1103DeadlineAlert(
                request_id=UUID(r["id"]),
                property_id=UUID(r["property_id"]),
                property_name=prop_map.get(r["property_id"], "Unknown Property"),
                tenant_name=lease_map.get(r["lease_id"], "Unknown Tenant"),
                response_deadline=deadline,
                days_remaining=days_remaining,
                status=r["status"],
            )
        )
    return alerts
