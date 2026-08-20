"""Generate the Property Tax Appeal Recovery Calculator workbook.

Output: ``docs/assets/property-tax-appeal-recovery-calculator.xlsx``

Three sheets:
- Inputs: property info, tax values, tenant roster (5 rows), appeal costs
- Calculations: total savings, net savings, per-tenant allocations, summary table
- Instructions: how appeals affect CAM, lease provisions, communicating adjustments
"""

from __future__ import annotations

import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))

from _common import (  # noqa: E402
    app_url,
    apply_header_row,
    bold_font,
    docs_assets_dir,
    fill,
    formula_cell,
    input_cell,
    new_workbook,
    section_label,
    site_url,
    write_footer,
)
from openpyxl.styles import Alignment, Font  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

OUTPUT_PATH = docs_assets_dir() / "property-tax-appeal-recovery-calculator.xlsx"
FOOTER_URL = site_url("/tools/property-tax-appeal-recovery-calculator")

NUM_TENANTS = 5


def build_inputs(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Inputs")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "Property Tax Appeal Recovery Calculator | CapVeri"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 28

    # ── Property / Tax Info ───────────────────────────────────────────────────
    section_label(ws, "A3", "PROPERTY & TAX INFORMATION")
    ws.merge_cells("A3:E3")

    ws.cell(row=4, column=1, value="Property Name").font = bold_font()
    input_cell(ws, 4, 2, "Westside Business Center")
    ws.merge_cells("B4:D4")

    ws.cell(row=5, column=1, value="Tax Year").font = bold_font()
    input_cell(ws, 5, 2, 2024, "0")

    ws.cell(row=6, column=1, value="Assessed Value — Original ($)").font = bold_font()
    input_cell(ws, 6, 2, 8_500_000, "$#,##0")

    ws.cell(row=7, column=1, value="Assessed Value — Appealed/Reduced ($)").font = (
        bold_font()
    )
    input_cell(ws, 7, 2, 7_200_000, "$#,##0")

    ws.cell(row=8, column=1, value="Tax Rate (%)").font = bold_font()
    input_cell(ws, 8, 2, 0.0125, "0.000%")
    ws.cell(row=8, column=3, value="e.g., 1.250% = 0.01250").font = Font(
        italic=True, color="475569"
    )

    ws.cell(row=9, column=1, value="Original Tax Bill ($)").font = bold_font()
    ws.cell(
        row=9,
        column=3,
        value="Auto-calculated below — or override manually",
    ).font = Font(italic=True, color="475569")
    formula_cell(ws, 9, 2, "=B6*B8", "$#,##0")

    ws.cell(row=10, column=1, value="Reduced Tax Bill ($)").font = bold_font()
    formula_cell(ws, 10, 2, "=B7*B8", "$#,##0")

    ws.cell(row=11, column=1, value="Manual Override — Original Bill ($)").font = (
        bold_font()
    )
    input_cell(ws, 11, 2, None, "$#,##0")
    ws.cell(
        row=11,
        column=3,
        value="If entered, overrides formula in B9 (for Calculations)",
    ).font = Font(italic=True, color="475569")

    ws.cell(row=12, column=1, value="Manual Override — Reduced Bill ($)").font = (
        bold_font()
    )
    input_cell(ws, 12, 2, None, "$#,##0")

    # ── Appeal Costs ──────────────────────────────────────────────────────────
    section_label(ws, "A14", "APPEAL COSTS")
    ws.merge_cells("A14:E14")

    ws.cell(row=15, column=1, value="Attorney Fees ($)").font = bold_font()
    input_cell(ws, 15, 2, 12_000, "$#,##0")

    ws.cell(row=16, column=1, value="Filing Fees ($)").font = bold_font()
    input_cell(ws, 16, 2, 500, "$#,##0")

    ws.cell(
        row=17,
        column=1,
        value="Lease Provision — Appeal Cost Recovery",
    ).font = bold_font()
    cost_recovery_cell = input_cell(ws, 17, 2, "Yes")
    dv_cost = DataValidation(
        type="list",
        formula1='"Yes,No,Partial"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_cost)
    dv_cost.add(cost_recovery_cell)
    ws.cell(
        row=17,
        column=3,
        value="Yes = costs offset savings; No = costs absorbed by landlord; Partial = 50%",
    ).font = Font(italic=True, color="475569")
    ws.merge_cells("C17:E17")

    # ── Tenant Roster ─────────────────────────────────────────────────────────
    section_label(ws, "A19", "TENANT ROSTER (up to 5 tenants)")
    ws.merge_cells("A19:E19")

    apply_header_row(
        ws,
        20,
        ["Tenant Name", "Suite", "GLA (sq ft)", "Pro-Rata % (auto)", "Notes"],
        widths=[28, 12, 16, 20, 28],
    )

    sample_tenants = [
        ("Acme Corp", "100", 18_500),
        ("Beta LLC", "200", 12_000),
        ("Gamma Inc", "300", 22_750),
        ("Delta Co", "400", 9_000),
        ("Epsilon Ltd", "500", 15_250),
    ]
    # Total GLA for pro-rata = sum of tenant GLA (using SUM range)
    tenant_start_row = 21
    tenant_end_row = tenant_start_row + NUM_TENANTS - 1

    for i, (name, suite, gla) in enumerate(sample_tenants):
        r = tenant_start_row + i
        input_cell(ws, r, 1, name)
        input_cell(ws, r, 2, suite)
        input_cell(ws, r, 3, gla, "#,##0")
        # Pro-rata = tenant GLA / sum of all tenant GLA
        formula_cell(
            ws,
            r,
            4,
            f"=IF(SUM(C{tenant_start_row}:C{tenant_end_row})>0,"
            f"C{r}/SUM(C{tenant_start_row}:C{tenant_end_row}),0)",
            "0.00%",
        )

    ws.cell(
        row=27,
        column=1,
        value="Pro-rata % is calculated from the tenant GLA rows above (total = 100%).",
    ).font = Font(italic=True, color="475569", size=9)
    ws.merge_cells("A27:E27")

    write_footer(ws, FOOTER_URL, 29)
    ws.sheet_view.showGridLines = False


def build_calculations(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Calculations")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Tax Appeal Savings & Tenant Allocation"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 26

    # ── Savings Summary ───────────────────────────────────────────────────────
    section_label(ws, "A3", "SAVINGS SUMMARY")

    # Effective original bill: manual override if populated, else formula
    ws.cell(row=4, column=1, value="Original Tax Bill ($)").font = bold_font()
    formula_cell(
        ws,
        4,
        2,
        "=IF(ISNUMBER(Inputs!B11),Inputs!B11,Inputs!B9)",
        "$#,##0",
    )

    ws.cell(row=5, column=1, value="Reduced Tax Bill ($)").font = bold_font()
    formula_cell(
        ws,
        5,
        2,
        "=IF(ISNUMBER(Inputs!B12),Inputs!B12,Inputs!B10)",
        "$#,##0",
    )

    ws.cell(row=6, column=1, value="Total Tax Savings ($)").font = bold_font()
    formula_cell(ws, 6, 2, "=MAX(0,B4-B5)", "$#,##0")

    ws.cell(row=7, column=1, value="Total Appeal Costs ($)").font = bold_font()
    formula_cell(ws, 7, 2, "=Inputs!B15+Inputs!B16", "$#,##0")

    ws.cell(row=8, column=1, value="Recoverable Appeal Costs ($)").font = bold_font()
    formula_cell(
        ws,
        8,
        2,
        '=IF(Inputs!B17="Yes",B7,IF(Inputs!B17="Partial",B7*0.5,0))',
        "$#,##0",
    )
    ws.cell(
        row=8, column=3, value="Based on lease provision (Yes/Partial/No in Inputs!B17)"
    ).font = Font(italic=True, color="475569")
    ws.merge_cells("C8:F8")

    ws.cell(row=9, column=1, value="Net Savings after Appeal Costs ($)").font = (
        bold_font()
    )
    formula_cell(ws, 9, 2, "=MAX(0,B6-B8)", "$#,##0")

    # ── Per-Tenant Allocation ─────────────────────────────────────────────────
    section_label(ws, "A11", "PER-TENANT ALLOCATION")

    apply_header_row(
        ws,
        12,
        [
            "Tenant Name",
            "Pro-Rata %",
            "Original Allocated Tax ($)",
            "Reduced Allocated Tax ($)",
            "Savings / Credit ($)",
            "Treatment",
        ],
        widths=[28, 16, 26, 26, 22, 28],
    )

    tenant_start_row = 21  # in Inputs sheet

    for i in range(NUM_TENANTS):
        r = 13 + i
        inp_r = tenant_start_row + i
        formula_cell(ws, r, 1, f"=Inputs!A{inp_r}")
        formula_cell(ws, r, 2, f"=Inputs!D{inp_r}", "0.00%")
        formula_cell(ws, r, 3, f"=B4*C{r}", "$#,##0")
        formula_cell(ws, r, 4, f"=B5*C{r}", "$#,##0")
        formula_cell(ws, r, 5, f"=MAX(0,C{r}-D{r})", "$#,##0")
        # Treatment text
        formula_cell(
            ws,
            r,
            6,
            f'=IF(E{r}>0,"Credit to tenant — apply to next estimate","No adjustment needed")',
        )

    # Totals
    last_row = 13 + NUM_TENANTS - 1
    section_label(ws, f"A{last_row + 2}", "TOTALS")
    formula_cell(ws, last_row + 2, 3, f"=SUM(C13:C{last_row})", "$#,##0")
    formula_cell(ws, last_row + 2, 4, f"=SUM(D13:D{last_row})", "$#,##0")
    formula_cell(ws, last_row + 2, 5, f"=SUM(E13:E{last_row})", "$#,##0")

    ws.cell(
        row=last_row + 4,
        column=1,
        value="Yellow cells (Inputs sheet) are editable. Gray cells are calculated.",
    ).font = Font(italic=True, color="475569", size=9)
    ws.merge_cells(f"A{last_row + 4}:F{last_row + 4}")

    write_footer(ws, FOOTER_URL, last_row + 6)
    ws.freeze_panes = "A13"
    ws.sheet_view.showGridLines = False


def build_instructions(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 110

    rows: list[tuple[str, bool, int]] = [
        ("Property Tax Appeal Recovery Calculator | How to Use", True, 14),
        ("", False, 11),
        ("HOW PROPERTY TAX APPEALS AFFECT CAM", True, 12),
        (
            "When a landlord successfully appeals the assessed value of a property, the resulting "
            "tax bill reduction creates a savings. Under most commercial leases where tenants pay "
            "their pro-rata share of property taxes (as part of CAM or a separate line item), that "
            "savings must be passed through to tenants — either as a credit against future estimates "
            "or as a refund on the reconciliation.",
            False,
            11,
        ),
        ("", False, 11),
        ("LEASE PROVISIONS GOVERNING PASS-THROUGH", True, 12),
        (
            "Most NNN and modified gross leases state that 'Operating Expenses shall not include "
            "the cost of any item to the extent landlord receives reimbursement therefor from "
            "insurance or otherwise.' Courts have generally held that tax refunds fall under this "
            "principle. However, landlords often negotiate the right to offset appeal costs "
            "against the savings before passing through the net.",
            False,
            11,
        ),
        (
            "Key variables: (1) Does the lease give the landlord the right to bring an appeal? "
            "(2) Can the landlord recover attorney fees and filing costs from savings? "
            "(3) Is the savings treated as a current-year credit or a prior-year reconciliation item?",
            False,
            11,
        ),
        ("", False, 11),
        ("COMMON SCENARIOS", True, 12),
        (
            "Scenario A — Appeal decided before annual reconciliation: Reduce the tax line in the "
            "CAM pool to the reduced bill. Tenants see lower actual; no separate credit needed.",
            False,
            11,
        ),
        (
            "Scenario B — Appeal decided after reconciliation sent: Issue a supplemental "
            "reconciliation or credit memo. Use the 'Credit to tenant' treatment in this calculator.",
            False,
            11,
        ),
        (
            "Scenario C — Appeal for a prior year (retroactive refund): Allocate based on each "
            "tenant's pro-rata share during that tax year. Tenants who have vacated may still be "
            "owed a credit per their lease terms.",
            False,
            11,
        ),
        ("", False, 11),
        ("HOW TO COMMUNICATE ADJUSTMENTS", True, 12),
        (
            "1. Send a written notice citing the original assessment, appeal outcome, and new bill.",
            False,
            11,
        ),
        (
            "2. Show the calculation: total savings, less recovered appeal costs (if applicable), "
            "equals net savings passed through.",
            False,
            11,
        ),
        (
            "3. State the tenant's pro-rata percentage and resulting credit amount.",
            False,
            11,
        ),
        (
            "4. Specify how the credit will be applied: against next monthly estimate or "
            "as a check within 30–60 days.",
            False,
            11,
        ),
        ("", False, 11),
        ("HOW TO USE THIS CALCULATOR", True, 12),
        (
            "1. Enter property and tax assessment data in the yellow cells on the Inputs sheet.",
            False,
            11,
        ),
        (
            "2. Enter appeal costs and select the lease cost-recovery provision.",
            False,
            11,
        ),
        (
            "3. Enter each tenant's name and GLA (pro-rata % is auto-calculated).",
            False,
            11,
        ),
        ("4. Review the Calculations sheet for per-tenant credit amounts.", False, 11),
        (
            "5. Use the 'Treatment' column to determine whether to issue a credit or no adjustment.",
            False,
            11,
        ),
        ("", False, 11),
        (
            f"Open the live tool at {site_url('/tools/property-tax-appeal-recovery-calculator')}",
            False,
            11,
        ),
    ]

    for r_idx, (text, is_bold, size) in enumerate(rows, start=1):
        cell = ws.cell(row=r_idx, column=1, value=text)
        cell.font = Font(bold=is_bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_idx].height = 32 if is_bold else 24

    write_footer(ws, FOOTER_URL, len(rows) + 2)
    ws.sheet_view.showGridLines = False


def main() -> None:
    wb = new_workbook()
    build_inputs(wb)
    build_calculations(wb)
    build_instructions(wb)
    wb.active = wb["Calculations"]
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
