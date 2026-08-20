"""Generate the CAM Recovery Ratio Benchmark Worksheet.

Output: ``docs/assets/cam-recovery-ratio-worksheet.xlsx``

Three sheets:
- Inputs: property data, CAM expenses, recoveries, prior years, benchmarks
- Calculations: recovery ratio, variance to benchmark, dollar leakage, YoY trend
- Instructions: what recovery ratio means, causes of low ratios, how to use
"""

from __future__ import annotations

import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))

from _common import (  # noqa: E402
    app_url,
    apply_header_row,
    bold_font,
    docs_assets_dir,
    fill,
    formula_cell,
    input_cell,
    new_workbook,
    section_label,
    site_url,
    write_footer,
)
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

OUTPUT_PATH = docs_assets_dir() / "cam-recovery-ratio-worksheet.xlsx"
FOOTER_URL = site_url("/tools/cam-recovery-ratio-worksheet")


def build_inputs(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Inputs")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 20

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "CAM Recovery Ratio Benchmark Worksheet | CapVeri"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 28

    # ── Property Info ──────────────────────────────────────────────────────────
    section_label(ws, "A3", "PROPERTY INFORMATION")
    ws.merge_cells("A3:F3")

    ws.cell(row=4, column=1, value="Property Name").font = bold_font()
    input_cell(ws, 4, 2, "My Office Park")
    ws.merge_cells("B4:D4")

    ws.cell(row=5, column=1, value="Property Type").font = bold_font()
    prop_type_cell = input_cell(ws, 5, 2, "Office")
    dv_type = DataValidation(
        type="list",
        formula1='"Office,Retail-Strip,Retail-Regional,Industrial,Mixed-Use"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)
    dv_type.add(prop_type_cell)
    ws.cell(row=5, column=3, value="Select from dropdown").font = Font(
        italic=True, color="475569"
    )

    ws.cell(row=6, column=1, value="Total GLA (sq ft)").font = bold_font()
    input_cell(ws, 6, 2, 150000, "#,##0")
    ws.cell(row=7, column=1, value="Occupied GLA (sq ft)").font = bold_font()
    input_cell(ws, 7, 2, 127500, "#,##0")
    ws.cell(row=8, column=1, value="Occupancy %").font = bold_font()
    formula_cell(ws, 8, 2, "=IF(B6>0,B7/B6,0)", "0.0%")
    ws.cell(row=8, column=3, value="Auto-calculated from GLA fields above").font = Font(
        italic=True, color="475569"
    )

    # ── Current Year ───────────────────────────────────────────────────────────
    ws.row_dimensions[10].height = 4
    section_label(ws, "A11", "CURRENT YEAR — CAM FINANCIALS")
    ws.merge_cells("A11:F11")

    ws.cell(row=12, column=1, value="Reconciliation Year").font = bold_font()
    input_cell(ws, 12, 2, 2025, "0")

    ws.cell(row=13, column=1, value="Total CAM Expenses (actual, $)").font = bold_font()
    input_cell(ws, 13, 2, 1_250_000, "$#,##0")

    ws.cell(row=14, column=1, value="Total CAM Recovered from Tenants ($)").font = (
        bold_font()
    )
    input_cell(ws, 14, 2, 975_000, "$#,##0")

    # ── Prior Year 1 ──────────────────────────────────────────────────────────
    ws.row_dimensions[16].height = 4
    section_label(ws, "A17", "PRIOR YEAR 1 — CAM FINANCIALS")
    ws.merge_cells("A17:F17")

    ws.cell(row=18, column=1, value="Year").font = bold_font()
    input_cell(ws, 18, 2, 2024, "0")
    ws.cell(row=19, column=1, value="Total CAM Expenses ($)").font = bold_font()
    input_cell(ws, 19, 2, 1_180_000, "$#,##0")
    ws.cell(row=20, column=1, value="Total CAM Recovered ($)").font = bold_font()
    input_cell(ws, 20, 2, 908_000, "$#,##0")

    # ── Prior Year 2 ──────────────────────────────────────────────────────────
    ws.row_dimensions[22].height = 4
    section_label(ws, "A23", "PRIOR YEAR 2 — CAM FINANCIALS")
    ws.merge_cells("A23:F23")

    ws.cell(row=24, column=1, value="Year").font = bold_font()
    input_cell(ws, 24, 2, 2023, "0")
    ws.cell(row=25, column=1, value="Total CAM Expenses ($)").font = bold_font()
    input_cell(ws, 25, 2, 1_120_000, "$#,##0")
    ws.cell(row=26, column=1, value="Total CAM Recovered ($)").font = bold_font()
    input_cell(ws, 26, 2, 851_000, "$#,##0")

    # ── Market Benchmarks ─────────────────────────────────────────────────────
    ws.row_dimensions[28].height = 4
    section_label(
        ws, "A29", "MARKET BENCHMARKS (pre-populated — edit if you have better data)"
    )
    ws.merge_cells("A29:F29")

    apply_header_row(
        ws,
        30,
        ["Property Type", "Industry Avg Recovery %", "Source / Notes"],
        widths=[30, 28, 42],
    )
    benchmarks = [
        ("Office", 0.78, "BOMA / IREM survey composite"),
        ("Retail-Strip", 0.91, "ICSC operating expense study"),
        ("Retail-Regional", 0.87, "ICSC operating expense study"),
        ("Industrial", 0.83, "CoStar NNN survey data"),
        ("Mixed-Use", 0.82, "Interpolated from component types"),
    ]
    for i, (ptype, pct, note) in enumerate(benchmarks):
        r = 31 + i
        ws.cell(row=r, column=1, value=ptype).font = bold_font()
        input_cell(ws, r, 2, pct, "0.0%")
        ws.cell(row=r, column=3, value=note).font = Font(italic=True, color="475569")

    ws.cell(
        row=37,
        column=1,
        value="TIP: The benchmark for your selected Property Type (cell B5) is "
        "used automatically in the Calculations sheet.",
    ).font = Font(italic=True, color="475569", size=9)
    ws.merge_cells("A37:F37")

    write_footer(ws, FOOTER_URL, 39)
    ws.sheet_view.showGridLines = False


def build_calculations(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Calculations")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 34

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"] = "CAM Recovery Ratio Analysis"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 26

    # ── Benchmark lookup ──────────────────────────────────────────────────────
    # Uses MATCH on property type to pick the right benchmark row (31–35 in Inputs)
    section_label(ws, "A3", "BENCHMARK LOOKUP")

    ws.cell(row=4, column=1, value="Selected Property Type").font = bold_font()
    formula_cell(ws, 4, 2, "=Inputs!B5")

    ws.cell(row=5, column=1, value="Industry Avg Recovery % (benchmark)").font = (
        bold_font()
    )
    # MATCH: Office→31, Retail-Strip→32, Retail-Regional→33, Industrial→34, Mixed-Use→35
    formula_cell(
        ws,
        5,
        2,
        "=INDEX(Inputs!B31:B35,MATCH(Inputs!B5,Inputs!A31:A35,0))",
        "0.0%",
    )

    # ── Current Year Analysis ─────────────────────────────────────────────────
    section_label(ws, "A7", "CURRENT YEAR ANALYSIS")

    ws.cell(row=8, column=1, value="Reconciliation Year").font = bold_font()
    formula_cell(ws, 8, 2, "=Inputs!B12", "0")

    ws.cell(row=9, column=1, value="Total CAM Expenses ($)").font = bold_font()
    formula_cell(ws, 9, 2, "=Inputs!B13", "$#,##0")

    ws.cell(row=10, column=1, value="Total CAM Recovered ($)").font = bold_font()
    formula_cell(ws, 10, 2, "=Inputs!B14", "$#,##0")

    ws.cell(row=11, column=1, value="Recovery Ratio (Recovered / Expenses)").font = (
        bold_font()
    )
    formula_cell(ws, 11, 2, "=IF(B9>0,B10/B9,0)", "0.0%")

    ws.cell(
        row=12, column=1, value="Variance to Benchmark (Recovery - Benchmark)"
    ).font = bold_font()
    formula_cell(ws, 12, 2, "=B11-B5", "0.0%")

    ws.cell(
        row=13, column=1, value="Dollar Leakage = (Benchmark - Actual Ratio) × Expenses"
    ).font = bold_font()
    formula_cell(ws, 13, 2, "=MAX(0,(B5-B11)*B9)", "$#,##0")
    ws.cell(
        row=13,
        column=3,
        value="Amount of expenses landlord failed to recover vs. benchmark",
    ).font = Font(italic=True, color="475569")

    ws.cell(row=14, column=1, value="Rating").font = bold_font()
    formula_cell(
        ws,
        14,
        2,
        '=IF(B11>=B5,"Above benchmark",IF(B11>=(B5-0.05),'
        '"Within 5% of benchmark","Below benchmark — review recommended"))',
    )
    # Conditional formatting via cell color for the rating cell
    from openpyxl.formatting.rule import CellIsRule

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    from openpyxl.formatting.rule import FormulaRule

    ws.conditional_formatting.add(
        "B14",
        FormulaRule(
            formula=['B14="Above benchmark"'],
            fill=green_fill,
            font=Font(color="375623"),
        ),
    )
    ws.conditional_formatting.add(
        "B14",
        FormulaRule(
            formula=['B14="Within 5% of benchmark"'],
            fill=yellow_fill,
            font=Font(color="9C5700"),
        ),
    )
    ws.conditional_formatting.add(
        "B14",
        FormulaRule(
            formula=['LEFT(B14,5)="Below"'],
            fill=red_fill,
            font=Font(color="9C0006"),
        ),
    )

    # ── Year-over-Year Trend ──────────────────────────────────────────────────
    section_label(ws, "A16", "YEAR-OVER-YEAR TREND")

    apply_header_row(
        ws,
        17,
        ["Year", "CAM Expenses ($)", "CAM Recovered ($)", "Recovery Ratio"],
        widths=[16, 22, 22, 20],
    )
    # Prior Year 2
    formula_cell(ws, 18, 1, "=Inputs!B24", "0")
    formula_cell(ws, 18, 2, "=Inputs!B25", "$#,##0")
    formula_cell(ws, 18, 3, "=Inputs!B26", "$#,##0")
    formula_cell(ws, 18, 4, "=IF(B18>0,C18/B18,0)", "0.0%")

    # Prior Year 1
    formula_cell(ws, 19, 1, "=Inputs!B18", "0")
    formula_cell(ws, 19, 2, "=Inputs!B19", "$#,##0")
    formula_cell(ws, 19, 3, "=Inputs!B20", "$#,##0")
    formula_cell(ws, 19, 4, "=IF(B19>0,C19/B19,0)", "0.0%")

    # Current Year
    formula_cell(ws, 20, 1, "=Inputs!B12", "0")
    formula_cell(ws, 20, 2, "=B9", "$#,##0")
    formula_cell(ws, 20, 3, "=B10", "$#,##0")
    formula_cell(ws, 20, 4, "=B11", "0.0%")

    # Trend direction
    ws.cell(row=22, column=1, value="Trend (current vs. 2 years ago)").font = (
        bold_font()
    )
    formula_cell(
        ws,
        22,
        2,
        '=IF(D20>D18,"Improving ↑",IF(D20<D18,"Declining ↓","Flat →"))',
    )
    ws.cell(
        row=22, column=3, value="Compares current year ratio to earliest year in trend"
    ).font = Font(italic=True, color="475569")

    ws.cell(
        row=24,
        column=1,
        value="Yellow cells (Inputs sheet) are editable. Gray cells are calculated.",
    ).font = Font(italic=True, color="475569", size=9)
    ws.merge_cells("A24:C24")

    write_footer(ws, FOOTER_URL, 26)
    ws.freeze_panes = "A8"
    ws.sheet_view.showGridLines = False


def build_instructions(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 110

    rows: list[tuple[str, bool, int]] = [
        ("CAM Recovery Ratio Benchmark Worksheet | How to Use", True, 14),
        ("", False, 11),
        ("WHAT IS CAM RECOVERY RATIO?", True, 12),
        (
            "The CAM recovery ratio is the percentage of total CAM expenses that a landlord "
            "successfully bills back to tenants. Formula: Total CAM Recovered ÷ Total CAM Expenses. "
            "A ratio of 100% means every dollar of expense was passed through; a ratio of 78% means "
            "22 cents of every dollar was absorbed by the landlord.",
            False,
            11,
        ),
        ("", False, 11),
        ("WHAT CAUSES A LOW RECOVERY RATIO?", True, 12),
        (
            "1. Vacant space not grossed up — leases with gross-up clauses allow the landlord to "
            "inflate expenses to a hypothetical full-occupancy level; if gross-up is not applied, "
            "vacant space effectively subsidizes tenants.",
            False,
            11,
        ),
        (
            "2. Exclusions too broad — if the lease excludes capital items, management fees, "
            "insurance, or taxes from the CAM pool, the recoverable base shrinks.",
            False,
            11,
        ),
        (
            "3. Caps limiting recovery — year-over-year caps (5–7% typical) can prevent recovery "
            "of actual expense increases, creating a structural gap.",
            False,
            11,
        ),
        (
            "4. Admin fee below market — a 0% or 3% admin fee when the market supports 10–15% "
            "leaves money on the table.",
            False,
            11,
        ),
        (
            "5. Denominator mismatch — using total project GLA instead of leasable GLA inflates the "
            "denominator, diluting each tenant's pro-rata share.",
            False,
            11,
        ),
        (
            "6. Reconciliation not performed — if annual true-ups are not sent, estimates "
            "accumulate without capturing actuals.",
            False,
            11,
        ),
        ("", False, 11),
        ("INDUSTRY BENCHMARKS (PRE-POPULATED)", True, 12),
        (
            "Office: ~78% — Large exclusion lists and management fee caps common in CBD leases. "
            "Retail-Strip: ~91% — Strong NNN leases with minimal exclusions. "
            "Retail-Regional: ~87% — Anchor tenant exclusions and cap limitations weigh on recovery. "
            "Industrial: ~83% — Triple-net but often exclude landlord admin overhead.",
            False,
            11,
        ),
        ("", False, 11),
        ("HOW TO USE THIS WORKSHEET", True, 12),
        (
            "1. Fill in yellow cells on the Inputs sheet — property info, CAM figures, prior years.",
            False,
            11,
        ),
        ("2. Select your Property Type from the dropdown (cell B5).", False, 11),
        (
            "3. Review the Calculations sheet: recovery ratio, leakage, YoY trend, and rating.",
            False,
            11,
        ),
        (
            "4. If rating is 'Below benchmark — review recommended', audit your lease "
            "exclusion lists, gross-up clause application, and admin fee percentage.",
            False,
            11,
        ),
        (
            "5. Use Dollar Leakage as the business case for improving reconciliation practices.",
            False,
            11,
        ),
        ("", False, 11),
        (
            f"Open the live tool at {site_url('/tools/cam-recovery-ratio-worksheet')}",
            False,
            11,
        ),
    ]

    for r_idx, (text, is_bold, size) in enumerate(rows, start=1):
        cell = ws.cell(row=r_idx, column=1, value=text)
        cell.font = Font(bold=is_bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_idx].height = 32 if is_bold else 24

    write_footer(ws, FOOTER_URL, len(rows) + 2)
    ws.sheet_view.showGridLines = False


def main() -> None:
    wb = new_workbook()
    build_inputs(wb)
    build_calculations(wb)
    build_instructions(wb)
    wb.active = wb["Calculations"]
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
