"""Generate the Admin Fee Calculator workbook.

Output: ``docs/assets/admin-fee-calculator.xlsx``

Three sheets:
- Inputs: gross CAM pool, excluded categories, lease admin fee %, cap settings,
  prior year comparison data
- Calculations: admin fee base, calculated fee, cap check, allowable fee,
  YoY comparison, fee as % of total pool
- Instructions: what admin fee is, how leases cap/exclude it, common disputes
"""

from __future__ import annotations

import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))

from _common import (  # noqa: E402
    app_url,
    apply_header_row,
    bold_font,
    docs_assets_dir,
    fill,
    formula_cell,
    input_cell,
    new_workbook,
    section_label,
    site_url,
    write_footer,
)
from openpyxl.styles import Alignment, Font  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

OUTPUT_PATH = docs_assets_dir() / "admin-fee-calculator.xlsx"
FOOTER_URL = site_url("/tools/admin-fee-calculator")

EXCLUDED_CATEGORIES = [
    ("Insurance", 57_000),
    ("Property Tax", 192_000),
    ("Management Fee", 47_000),
    ("Utilities - Common Area", 0),
    ("Capital Items", 0),
    ("Other Lease-Excluded Item", 0),
]
EXCL_START_ROW = 14
EXCL_END_ROW = EXCL_START_ROW + len(EXCLUDED_CATEGORIES) - 1


def build_inputs(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Inputs")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 42

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"] = "Admin Fee Calculator | CapVeri"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 28

    # ── Property / Context ────────────────────────────────────────────────────
    section_label(ws, "A3", "PROPERTY & RECONCILIATION CONTEXT")
    ws.merge_cells("A3:C3")

    ws.cell(row=4, column=1, value="Property Name").font = bold_font()
    input_cell(ws, 4, 2, "Westside Business Center")

    ws.cell(row=5, column=1, value="Tenant Name").font = bold_font()
    input_cell(ws, 5, 2, "Acme Corp")

    ws.cell(row=6, column=1, value="Reconciliation Year").font = bold_font()
    input_cell(ws, 6, 2, 2025, "0")

    # ── CAM Pool ──────────────────────────────────────────────────────────────
    section_label(ws, "A8", "CAM POOL")
    ws.merge_cells("A8:C8")

    ws.cell(row=9, column=1, value="Gross CAM Pool (total actual expenses, $)").font = (
        bold_font()
    )
    input_cell(ws, 9, 2, 517_300, "$#,##0")
    ws.cell(
        row=9,
        column=3,
        value="Total CAM before admin fee; matches Inputs!C30 in the reconciliation template",
    ).font = Font(italic=True, color="475569")

    # ── Excluded Expenses (from Admin Fee Base) ───────────────────────────────
    section_label(ws, "A11", "EXCLUDED FROM ADMIN FEE BASE (per lease)")
    ws.merge_cells("A11:C11")

    ws.cell(
        row=12,
        column=1,
        value="Enter amounts for each category the lease excludes from the admin fee base. "
        "Leave at $0 if the category is included.",
    ).font = Font(italic=True, color="475569", size=9)
    ws.merge_cells("A12:C12")

    apply_header_row(
        ws,
        13,
        ["Excluded Category", "Amount ($)", "Notes"],
        widths=[36, 20, 42],
    )

    for i, (cat, amt) in enumerate(EXCLUDED_CATEGORIES):
        r = EXCL_START_ROW + i
        ws.cell(row=r, column=1, value=cat).font = bold_font()
        input_cell(ws, r, 2, amt, "$#,##0")

    excl_tot_row = EXCL_END_ROW + 1
    ws.cell(row=excl_tot_row, column=1, value="Total Excluded ($)").font = bold_font()
    formula_cell(
        ws,
        excl_tot_row,
        2,
        f"=SUM(B{EXCL_START_ROW}:B{EXCL_END_ROW})",
        "$#,##0",
    )

    ws.cell(row=excl_tot_row + 1, column=1, value="Admin Fee Base ($)").font = (
        bold_font(12)
    )
    formula_cell(ws, excl_tot_row + 1, 2, f"=B9-B{excl_tot_row}", "$#,##0")

    # ── Lease Admin Fee % ─────────────────────────────────────────────────────
    fee_start = excl_tot_row + 4
    section_label(ws, f"A{fee_start}", "LEASE ADMIN FEE TERMS")
    ws.merge_cells(f"A{fee_start}:C{fee_start}")

    ws.cell(row=fee_start + 1, column=1, value="Lease Admin Fee % (of base)").font = (
        bold_font()
    )
    input_cell(ws, fee_start + 1, 2, 0.10, "0.00%")
    ws.cell(
        row=fee_start + 1, column=3, value="Typical commercial range: 5–15%"
    ).font = Font(italic=True, color="475569")

    ws.cell(row=fee_start + 2, column=1, value="Is Admin Fee Capped?").font = (
        bold_font()
    )
    cap_cell = input_cell(ws, fee_start + 2, 2, "N")
    dv_cap = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    ws.add_data_validation(dv_cap)
    dv_cap.add(cap_cell)

    ws.cell(row=fee_start + 3, column=1, value="Cap Amount ($) — if capped").font = (
        bold_font()
    )
    input_cell(ws, fee_start + 3, 2, None, "$#,##0")
    ws.cell(
        row=fee_start + 3,
        column=3,
        value="Enter a dollar cap. Leave blank if lease uses a % cap instead.",
    ).font = Font(italic=True, color="475569")

    ws.cell(
        row=fee_start + 4, column=1, value="Cap % of Total CAM — if capped"
    ).font = bold_font()
    input_cell(ws, fee_start + 4, 2, None, "0.00%")
    ws.cell(
        row=fee_start + 4,
        column=3,
        value="Enter a % cap of Gross CAM Pool. Leave blank if using dollar cap.",
    ).font = Font(italic=True, color="475569")

    # ── Prior Year Comparison ─────────────────────────────────────────────────
    prior_start = fee_start + 7
    section_label(ws, f"A{prior_start}", "PRIOR YEAR DATA (for trend)")
    ws.merge_cells(f"A{prior_start}:C{prior_start}")

    ws.cell(
        row=prior_start + 1, column=1, value="Prior Year Reconciliation Year"
    ).font = bold_font()
    input_cell(ws, prior_start + 1, 2, 2024, "0")

    ws.cell(
        row=prior_start + 2, column=1, value="Prior Year Admin Fee Billed ($)"
    ).font = bold_font()
    input_cell(ws, prior_start + 2, 2, 22_100, "$#,##0")

    ws.cell(
        row=prior_start + 3, column=1, value="Prior Year Gross CAM Pool ($)"
    ).font = bold_font()
    input_cell(ws, prior_start + 3, 2, 490_000, "$#,##0")

    write_footer(ws, FOOTER_URL, prior_start + 6)
    ws.sheet_view.showGridLines = False

    # Store dynamic row references as named ranges so Calculations can use them
    # (We'll reference by hardcoded computed values instead — see build_calculations)
    ws._fee_start = fee_start  # type: ignore[attr-defined]
    ws._excl_tot_row = excl_tot_row  # type: ignore[attr-defined]
    ws._prior_start = prior_start  # type: ignore[attr-defined]


def build_calculations(wb, fee_start: int, excl_tot_row: int, prior_start: int) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Calculations")
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 42

    ws.merge_cells("A1:C1")
    ws["A1"] = "Admin Fee Calculation"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 26

    # ── Base & Calculated Fee ─────────────────────────────────────────────────
    section_label(ws, "A3", "ADMIN FEE CALCULATION")

    ws.cell(row=4, column=1, value="Gross CAM Pool ($)").font = bold_font()
    formula_cell(ws, 4, 2, "=Inputs!B9", "$#,##0")

    ws.cell(row=5, column=1, value="Total Excluded from Admin Fee Base ($)").font = (
        bold_font()
    )
    formula_cell(ws, 5, 2, f"=Inputs!B{excl_tot_row}", "$#,##0")

    ws.cell(row=6, column=1, value="Admin Fee Base ($)").font = bold_font()
    formula_cell(ws, 6, 2, f"=Inputs!B{excl_tot_row + 1}", "$#,##0")

    ws.cell(row=7, column=1, value="Lease Admin Fee %").font = bold_font()
    formula_cell(ws, 7, 2, f"=Inputs!B{fee_start + 1}", "0.00%")

    ws.cell(row=8, column=1, value="Calculated Admin Fee ($) = Base × %").font = (
        bold_font()
    )
    formula_cell(ws, 8, 2, "=B6*B7", "$#,##0")

    # ── Cap Check ─────────────────────────────────────────────────────────────
    section_label(ws, "A10", "CAP CHECK")

    ws.cell(row=11, column=1, value="Is Admin Fee Capped?").font = bold_font()
    formula_cell(ws, 11, 2, f"=Inputs!B{fee_start + 2}")

    ws.cell(row=12, column=1, value="Dollar Cap Amount ($)").font = bold_font()
    formula_cell(
        ws,
        12,
        2,
        f'=IF(ISNUMBER(Inputs!B{fee_start + 3}),Inputs!B{fee_start + 3},"N/A")',
    )

    ws.cell(row=13, column=1, value="Percentage Cap Amount ($)").font = bold_font()
    formula_cell(
        ws,
        13,
        2,
        f'=IF(ISNUMBER(Inputs!B{fee_start + 4}),Inputs!B4*Inputs!B{fee_start + 4},"N/A")',
        "$#,##0",
    )

    ws.cell(row=14, column=1, value="Effective Cap ($)").font = bold_font()
    # Use smallest of dollar cap and percentage cap (whichever is provided)
    formula_cell(
        ws,
        14,
        2,
        f'=IF(B11="Y",'
        f"MIN(IF(ISNUMBER(Inputs!B{fee_start + 3}),Inputs!B{fee_start + 3},B8),"
        f"IF(ISNUMBER(Inputs!B{fee_start + 4}),B4*Inputs!B{fee_start + 4},B8)),"
        f'"No cap")',
    )

    ws.cell(row=15, column=1, value="Allowable Admin Fee ($)").font = bold_font(12)
    formula_cell(
        ws,
        15,
        2,
        '=IF(B11="Y",MIN(B8,IF(ISNUMBER(B14),B14,B8)),B8)',
        "$#,##0",
    )
    ws.cell(
        row=15,
        column=3,
        value="This is the amount to bill — calculated fee or cap, whichever is lower.",
    ).font = Font(italic=True, color="475569")

    # ── As % of Total Pool ────────────────────────────────────────────────────
    section_label(ws, "A17", "BENCHMARKING")

    ws.cell(row=18, column=1, value="Admin Fee as % of Gross CAM Pool").font = (
        bold_font()
    )
    formula_cell(ws, 18, 2, "=IF(B4>0,B15/B4,0)", "0.00%")
    ws.cell(
        row=18,
        column=3,
        value="Benchmark: 5–15% typical for commercial properties",
    ).font = Font(italic=True, color="475569")

    ws.cell(row=19, column=1, value="Benchmark Lower Bound (5%)").font = bold_font()
    formula_cell(ws, 19, 2, "=B4*0.05", "$#,##0")

    ws.cell(row=20, column=1, value="Benchmark Upper Bound (15%)").font = bold_font()
    formula_cell(ws, 20, 2, "=B4*0.15", "$#,##0")

    ws.cell(row=21, column=1, value="Position vs. Benchmark").font = bold_font()
    formula_cell(
        ws,
        21,
        2,
        '=IF(B18<0.05,"Below market — consider reviewing lease",'
        'IF(B18>0.15,"Above market — may face tenant challenge","Within market range"))',
    )

    # ── Prior Year Comparison ─────────────────────────────────────────────────
    section_label(ws, "A23", "YEAR-OVER-YEAR COMPARISON")

    ws.cell(row=24, column=1, value="Prior Year Admin Fee Billed ($)").font = (
        bold_font()
    )
    formula_cell(ws, 24, 2, f"=Inputs!B{prior_start + 2}", "$#,##0")

    ws.cell(row=25, column=1, value="Current Year Allowable Admin Fee ($)").font = (
        bold_font()
    )
    formula_cell(ws, 25, 2, "=B15", "$#,##0")

    ws.cell(row=26, column=1, value="Change ($)").font = bold_font()
    formula_cell(ws, 26, 2, "=B25-B24", "$#,##0")

    ws.cell(row=27, column=1, value="Change (%)").font = bold_font()
    formula_cell(ws, 27, 2, "=IF(B24>0,(B25-B24)/B24,0)", "0.0%")

    ws.cell(
        row=28, column=1, value="Prior Year Admin Fee as % of Prior CAM Pool"
    ).font = bold_font()
    formula_cell(
        ws,
        28,
        2,
        f"=IF(Inputs!B{prior_start + 3}>0,Inputs!B{prior_start + 2}/Inputs!B{prior_start + 3},0)",
        "0.00%",
    )

    ws.cell(
        row=30,
        column=1,
        value="Yellow cells (Inputs sheet) are editable. Gray cells are calculated.",
    ).font = Font(italic=True, color="475569", size=9)
    ws.merge_cells("A30:C30")

    write_footer(ws, FOOTER_URL, 32)
    ws.sheet_view.showGridLines = False


def build_instructions(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 110

    rows: list[tuple[str, bool, int]] = [
        ("Admin Fee Calculator | How to Use", True, 14),
        ("", False, 11),
        ("WHAT IS THE ADMIN FEE?", True, 12),
        (
            "The admin fee (also called the administrative overhead charge) is a landlord's "
            "charge for administering the CAM process — maintaining records, preparing statements, "
            "managing vendor contracts, and overseeing common area operations. It is typically "
            "expressed as a percentage of the total CAM pool or of a subset (the 'admin fee base').",
            False,
            11,
        ),
        (
            "The commercial market range is 5–15% of the CAM pool. Most office leases land at "
            "10–15%; many industrial and retail leases are at 5–10%. Fees above 15% are uncommon "
            "and may trigger tenant challenges.",
            False,
            11,
        ),
        ("", False, 11),
        ("HOW LEASES TYPICALLY CAP OR EXCLUDE IT", True, 12),
        (
            "1. Excluded categories: Many leases explicitly state that the admin fee shall not "
            "apply to insurance, taxes, management fees, or capital expenditures. The fee is "
            "then applied only to the remaining 'admin fee base.'",
            False,
            11,
        ),
        (
            "2. Dollar cap: Some leases set a maximum dollar amount for the admin fee, "
            "regardless of the pool size (e.g., 'not to exceed $50,000 per year').",
            False,
            11,
        ),
        (
            "3. Percentage cap of total CAM: A few leases cap the admin fee at a percentage "
            "of the gross CAM pool (e.g., 'not to exceed 10% of total Operating Expenses').",
            False,
            11,
        ),
        (
            "4. Separate management fee vs. admin fee: Many leases allow BOTH a management fee "
            "(paid to a property manager, e.g., 4–6% of gross revenues) AND a separate admin fee. "
            "Confirm whether the lease includes both, and whether the management fee counts as "
            "an excluded item from the admin fee base.",
            False,
            11,
        ),
        ("", False, 11),
        ("COMMON ADMIN FEE DISPUTES", True, 12),
        (
            "1. Double-charging: Landlord charges both a management fee (to the PM) and an admin "
            "fee — but the lease only allows one of them in the CAM pool.",
            False,
            11,
        ),
        (
            "2. Wrong base: Admin fee applied to gross CAM pool including insurance and taxes, "
            "when the lease says the base excludes those items.",
            False,
            11,
        ),
        (
            "3. Above-market fee: Admin fee of 15–20% on a simple industrial lease where 5% "
            "is the market norm. Tenant audits frequently catch this.",
            False,
            11,
        ),
        (
            "4. Fee increasing faster than pool: Admin fee growing at a rate inconsistent with "
            "the underlying work required (red flag in year-over-year analysis).",
            False,
            11,
        ),
        ("", False, 11),
        ("HOW TO USE THIS CALCULATOR", True, 12),
        (
            "1. Enter Gross CAM Pool (total actual CAM expenses before admin fee).",
            False,
            11,
        ),
        (
            "2. Enter dollar amounts for each category excluded from the admin fee base per your lease. "
            "Leave at $0 if the category is included in the base.",
            False,
            11,
        ),
        ("3. Enter the Lease Admin Fee % from the lease agreement.", False, 11),
        (
            "4. If the fee is capped, set the toggle to Y and enter either a dollar cap, "
            "a percentage cap, or both. The lower of the two will be used.",
            False,
            11,
        ),
        (
            "5. Review the Calculations sheet: Allowable Admin Fee, benchmark position, "
            "and year-over-year trend.",
            False,
            11,
        ),
        ("", False, 11),
        (
            f"Open the live tool at {site_url('/tools/admin-fee-calculator')}",
            False,
            11,
        ),
    ]

    for r_idx, (text, is_bold, size) in enumerate(rows, start=1):
        cell = ws.cell(row=r_idx, column=1, value=text)
        cell.font = Font(bold=is_bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_idx].height = 32 if is_bold else 24

    write_footer(ws, FOOTER_URL, len(rows) + 2)
    ws.sheet_view.showGridLines = False


def main() -> None:
    wb = new_workbook()
    build_inputs(wb)
    # Read back the dynamic row offsets from the Inputs sheet object
    inputs_ws = wb["Inputs"]
    fee_start: int = inputs_ws._fee_start  # type: ignore[attr-defined]
    excl_tot_row: int = inputs_ws._excl_tot_row  # type: ignore[attr-defined]
    prior_start: int = inputs_ws._prior_start  # type: ignore[attr-defined]
    build_calculations(wb, fee_start, excl_tot_row, prior_start)
    build_instructions(wb)
    wb.active = wb["Calculations"]
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
