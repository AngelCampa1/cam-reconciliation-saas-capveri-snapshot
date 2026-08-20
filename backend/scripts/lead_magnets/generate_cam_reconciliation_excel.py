"""Generate the full CAM Reconciliation Excel Template workbook.

Output: ``docs/assets/cam-reconciliation-excel.xlsx``

Four sheets:
- Inputs: property/tenant info, CAM pool expense table, gross-up settings, cap settings
- Calculations: pool total, exclusions, gross-up, cap, pro-rata, true-up
- Statement: print-ready tenant statement pulling from Calculations
- Instructions: how to fill inputs, when gross-up applies, how caps work
"""

from __future__ import annotations

import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))

from _common import (  # noqa: E402
    app_url,
    apply_header_row,
    bold_font,
    docs_assets_dir,
    fill,
    formula_cell,
    input_cell,
    new_workbook,
    section_label,
    site_url,
    write_footer,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

OUTPUT_PATH = docs_assets_dir() / "cam-reconciliation-excel.xlsx"
FOOTER_URL = site_url("/tools/cam-reconciliation-excel")

EXPENSE_CATEGORIES = [
    "Janitorial",
    "Landscaping",
    "HVAC Maintenance",
    "Utilities - Common Area",
    "Security",
    "Insurance",
    "Management Fee",
    "Admin Fee",
    "Property Tax",
    "Parking Lot Maintenance",
    "Miscellaneous",
]

EXPENSE_BUDGETS = [
    42_000,
    28_500,
    35_000,
    67_000,
    18_000,
    55_000,
    45_000,
    0,
    185_000,
    12_000,
    8_500,
]
EXPENSE_ACTUALS = [
    44_200,
    27_800,
    38_500,
    71_000,
    17_200,
    57_000,
    47_000,
    0,
    192_000,
    13_500,
    9_100,
]

EXP_START_ROW = 18
EXP_END_ROW = EXP_START_ROW + len(EXPENSE_CATEGORIES) - 1


def build_inputs(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Inputs")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 36

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "CAM Reconciliation Template | CapVeri"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 28

    # ── Property / Tenant Info ────────────────────────────────────────────────
    section_label(ws, "A3", "PROPERTY & TENANT INFORMATION")
    ws.merge_cells("A3:E3")

    info_rows = [
        ("Property Name", "Westside Business Center", None),
        ("Property Address", "123 Commerce Blvd, Suite 100", None),
        ("Reconciliation Year", 2025, "0"),
        ("Tenant Name", "Acme Corp", None),
        ("Suite / Unit", "200", None),
        ("Tenant GLA (sq ft)", 18_500, "#,##0"),
        ("Total Project GLA (sq ft)", 150_000, "#,##0"),
        ("Occupied GLA (sq ft)", 127_500, "#,##0"),
        ("Occupancy % (calculated)", None, "0.0%"),
        ("Monthly CAM Estimate Paid ($)", 4_200, "$#,##0"),
    ]

    for i, (label, value, fmt) in enumerate(info_rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = bold_font()
        if label == "Occupancy % (calculated)":
            formula_cell(ws, r, 2, "=IF(B9>0,B11/B9,0)", "0.0%")
        elif value is not None:
            if fmt:
                input_cell(ws, r, 2, value, fmt)
            else:
                input_cell(ws, r, 2, value)

    # ── CAM Pool Expense Table ────────────────────────────────────────────────
    section_label(ws, "A16", "CAM POOL EXPENSES")
    ws.merge_cells("A16:E16")

    apply_header_row(
        ws,
        17,
        ["Expense Category", "Budget ($)", "Actual ($)", "Variance ($)", "Notes"],
        widths=[34, 20, 20, 20, 36],
    )

    for i, (cat, budget, actual) in enumerate(
        zip(EXPENSE_CATEGORIES, EXPENSE_BUDGETS, EXPENSE_ACTUALS)
    ):
        r = EXP_START_ROW + i
        ws.cell(row=r, column=1, value=cat).font = bold_font()
        input_cell(ws, r, 2, budget, "$#,##0")
        input_cell(ws, r, 3, actual, "$#,##0")
        formula_cell(ws, r, 4, f"=C{r}-B{r}", "$#,##0")

    # Totals row
    tot_row = EXP_END_ROW + 1
    ws.cell(row=tot_row, column=1, value="TOTAL CAM POOL").font = bold_font(12)
    formula_cell(ws, tot_row, 2, f"=SUM(B{EXP_START_ROW}:B{EXP_END_ROW})", "$#,##0")
    formula_cell(ws, tot_row, 3, f"=SUM(C{EXP_START_ROW}:C{EXP_END_ROW})", "$#,##0")
    formula_cell(ws, tot_row, 4, f"=C{tot_row}-B{tot_row}", "$#,##0")

    # ── Gross-Up Settings ─────────────────────────────────────────────────────
    gu_start = tot_row + 3
    section_label(ws, f"A{gu_start}", "GROSS-UP SETTINGS")
    ws.merge_cells(f"A{gu_start}:E{gu_start}")

    ws.cell(row=gu_start + 1, column=1, value="Apply Gross-Up?").font = bold_font()
    gu_cell = input_cell(ws, gu_start + 1, 2, "Y")
    dv_gu = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    ws.add_data_validation(dv_gu)
    dv_gu.add(gu_cell)

    ws.cell(row=gu_start + 2, column=1, value="Occupancy Threshold (%)").font = (
        bold_font()
    )
    input_cell(ws, gu_start + 2, 2, 0.90, "0%")
    ws.cell(
        row=gu_start + 2,
        column=3,
        value="Gross-up kicks in when occupancy falls below this threshold",
    ).font = Font(italic=True, color="475569")

    ws.cell(row=gu_start + 3, column=1, value="Gross-Up Target (%)").font = bold_font()
    input_cell(ws, gu_start + 3, 2, 0.95, "0%")
    ws.cell(
        row=gu_start + 3,
        column=3,
        value="Variable expenses grossed up to this occupancy level",
    ).font = Font(italic=True, color="475569")

    # ── Cap Settings ──────────────────────────────────────────────────────────
    cap_start = gu_start + 6
    section_label(ws, f"A{cap_start}", "CAP SETTINGS")
    ws.merge_cells(f"A{cap_start}:E{cap_start}")

    ws.cell(row=cap_start + 1, column=1, value="Cap Type").font = bold_font()
    cap_cell = input_cell(ws, cap_start + 1, 2, "Cumulative")
    dv_cap = DataValidation(
        type="list",
        formula1='"Cumulative,Non-Cumulative,None"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_cap)
    dv_cap.add(cap_cell)

    ws.cell(row=cap_start + 2, column=1, value="Cap Base Year").font = bold_font()
    input_cell(ws, cap_start + 2, 2, 2023, "0")

    ws.cell(row=cap_start + 3, column=1, value="Cap % (per year)").font = bold_font()
    input_cell(ws, cap_start + 3, 2, 0.05, "0%")

    ws.cell(row=cap_start + 4, column=1, value="Prior Year Ceiling ($)").font = (
        bold_font()
    )
    input_cell(ws, cap_start + 4, 2, 490_000, "$#,##0")
    ws.cell(
        row=cap_start + 4,
        column=3,
        value="Enter actual prior year tenant ceiling (from prior reconciliation)",
    ).font = Font(italic=True, color="475569")

    write_footer(ws, FOOTER_URL, cap_start + 7)
    ws.sheet_view.showGridLines = False


def build_calculations(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Calculations")
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 40

    ws.merge_cells("A1:C1")
    ws["A1"] = "CAM Reconciliation — Step-by-Step Calculations"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 26

    # Resolve row references from Inputs
    tot_row = EXP_END_ROW + 1
    gu_start = tot_row + 3
    cap_start = gu_start + 6

    section_label(ws, "A3", "STEP 1 — CAM POOL")

    ws.cell(row=4, column=1, value="Gross CAM Pool (sum of actuals)").font = bold_font()
    formula_cell(ws, 4, 2, f"=Inputs!C{tot_row}", "$#,##0")

    ws.cell(
        row=5, column=1, value="Exclusion Adjustment (manual, negative number)"
    ).font = bold_font()
    input_cell(ws, 5, 2, 0, "$#,##0")
    ws.cell(
        row=5,
        column=3,
        value="Enter expenses excluded per lease (e.g., -15000). Default = 0.",
    ).font = Font(italic=True, color="475569")

    ws.cell(row=6, column=1, value="Net Recoverable Pool ($)").font = bold_font()
    formula_cell(ws, 6, 2, "=B4+B5", "$#,##0")

    section_label(ws, "A8", "STEP 2 — GROSS-UP (if applicable)")

    ws.cell(row=9, column=1, value="Apply Gross-Up?").font = bold_font()
    formula_cell(ws, 9, 2, f"=Inputs!B{gu_start + 1}")

    ws.cell(row=10, column=1, value="Current Occupancy %").font = bold_font()
    formula_cell(ws, 10, 2, "=IF(Inputs!B9>0,Inputs!B11/Inputs!B9,0)", "0.0%")

    ws.cell(row=11, column=1, value="Occupancy Threshold").font = bold_font()
    formula_cell(ws, 11, 2, f"=Inputs!B{gu_start + 2}", "0%")

    ws.cell(row=12, column=1, value="Gross-Up Target").font = bold_font()
    formula_cell(ws, 12, 2, f"=Inputs!B{gu_start + 3}", "0%")

    ws.cell(
        row=13, column=1, value="Grossed-Up Pool (variable expenses scaled)"
    ).font = bold_font()
    # If apply gross-up = Y AND occupancy < threshold: pool / occupancy * target
    # Otherwise: use net recoverable pool
    formula_cell(
        ws,
        13,
        2,
        '=IF(AND(B9="Y",B10<B11),B6/B10*B12,B6)',
        "$#,##0",
    )
    ws.cell(
        row=13,
        column=3,
        value="=Net Pool ÷ Actual Occ% × Target Occ% (only when below threshold)",
    ).font = Font(italic=True, color="475569")

    section_label(ws, "A15", "STEP 3 — CAP CALCULATION")

    ws.cell(row=16, column=1, value="Cap Type").font = bold_font()
    formula_cell(ws, 16, 2, f"=Inputs!B{cap_start + 1}")

    ws.cell(row=17, column=1, value="Prior Year Ceiling ($)").font = bold_font()
    formula_cell(ws, 17, 2, f"=Inputs!B{cap_start + 4}", "$#,##0")

    ws.cell(row=18, column=1, value="Cap %").font = bold_font()
    formula_cell(ws, 18, 2, f"=Inputs!B{cap_start + 3}", "0%")

    ws.cell(row=19, column=1, value="This Year Cap Ceiling ($)").font = bold_font()
    formula_cell(ws, 19, 2, "=B17*(1+B18)", "$#,##0")

    ws.cell(row=20, column=1, value="Recoverable Amount ($) — after cap").font = (
        bold_font()
    )
    formula_cell(
        ws,
        20,
        2,
        '=IF(B16="None",B13,MIN(B13,B19))',
        "$#,##0",
    )
    ws.cell(
        row=20,
        column=3,
        value='If cap type = "None", full grossed-up pool recoverable. '
        "Otherwise capped at ceiling.",
    ).font = Font(italic=True, color="475569")

    section_label(ws, "A22", "STEP 4 — TENANT PRO-RATA SHARE")

    ws.cell(row=23, column=1, value="Tenant GLA (sq ft)").font = bold_font()
    formula_cell(ws, 23, 2, "=Inputs!B6", "#,##0")

    ws.cell(row=24, column=1, value="Denominator GLA (sq ft)").font = bold_font()
    formula_cell(ws, 24, 2, "=Inputs!B9", "#,##0")
    ws.cell(
        row=24,
        column=3,
        value="Uses Total Project GLA from Inputs. Adjust if lease uses leased GLA.",
    ).font = Font(italic=True, color="475569")

    ws.cell(row=25, column=1, value="Tenant Pro-Rata % (GLA method)").font = bold_font()
    formula_cell(ws, 25, 2, "=IF(B24>0,B23/B24,0)", "0.0000%")

    ws.cell(row=26, column=1, value="Tenant Annual CAM Share ($)").font = bold_font()
    formula_cell(ws, 26, 2, "=B20*B25", "$#,##0")

    section_label(ws, "A28", "STEP 5 — TRUE-UP")

    ws.cell(row=29, column=1, value="Monthly Estimate Paid ($)").font = bold_font()
    formula_cell(ws, 29, 2, f"=Inputs!B{4 + 9}", "$#,##0")

    ws.cell(row=30, column=1, value="Total Estimated Payments (×12)").font = bold_font()
    formula_cell(ws, 30, 2, "=B29*12", "$#,##0")

    ws.cell(row=31, column=1, value="True-Up Balance (Due / Credit)").font = bold_font()
    formula_cell(ws, 31, 2, "=B26-B30", "$#,##0")
    ws.cell(
        row=31,
        column=3,
        value="Positive = amount due from tenant. Negative = credit to tenant.",
    ).font = Font(italic=True, color="475569")

    ws.cell(
        row=33,
        column=1,
        value="Yellow cells (Inputs sheet) are editable. Gray cells are calculated.",
    ).font = Font(italic=True, color="475569", size=9)
    ws.merge_cells("A33:C33")

    write_footer(ws, FOOTER_URL, 35)
    ws.sheet_view.showGridLines = False


def build_statement(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Statement")
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 26

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(bottom=thin)

    # Title block
    ws.merge_cells("A1:B1")
    ws["A1"] = "CAM RECONCILIATION STATEMENT"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:B2")
    ws["A2"] = f"CapVeri · {site_url('')}"
    ws["A2"].font = Font(color="0066FF", italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Property / Tenant info
    stmt_rows: list[tuple[str, str]] = [
        ("Property:", "=Inputs!B4"),
        ("Tenant:", "=Inputs!B7"),
        ("Suite:", "=Inputs!B8"),
        ("Reconciliation Year:", '=TEXT(Inputs!B6,"0")'),
    ]

    for i, (label, formula) in enumerate(stmt_rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = bold_font()
        cell = ws.cell(row=r, column=2, value=formula)
        cell.border = border

    # Divider
    ws.row_dimensions[9].height = 4

    # Expense summary
    ws.cell(row=10, column=1, value="EXPENSE SUMMARY").font = bold_font(12)
    ws["A10"].fill = fill("D9E8FF")
    ws.merge_cells("A10:B10")

    for i, cat in enumerate(EXPENSE_CATEGORIES):
        r = 11 + i
        ws.cell(row=r, column=1, value=cat).font = Font(size=10)
        formula_cell(
            ws,
            r,
            2,
            f"=Inputs!C{EXP_START_ROW + i}",
            "$#,##0",
        )

    tot_stmt = 11 + len(EXPENSE_CATEGORIES)
    ws.cell(row=tot_stmt, column=1, value="Gross CAM Pool").font = bold_font()
    formula_cell(ws, tot_stmt, 2, "=Calculations!B4", "$#,##0")

    ws.cell(row=tot_stmt + 1, column=1, value="Less: Exclusions").font = Font(size=10)
    formula_cell(ws, tot_stmt + 1, 2, "=Calculations!B5", "$#,##0")

    ws.cell(row=tot_stmt + 2, column=1, value="Net Recoverable Pool").font = bold_font()
    formula_cell(ws, tot_stmt + 2, 2, "=Calculations!B6", "$#,##0")

    ws.cell(row=tot_stmt + 3, column=1, value="Grossed-Up Pool (if applied)").font = (
        Font(size=10)
    )
    formula_cell(ws, tot_stmt + 3, 2, "=Calculations!B13", "$#,##0")

    ws.cell(row=tot_stmt + 4, column=1, value="Recoverable After Cap").font = (
        bold_font()
    )
    formula_cell(ws, tot_stmt + 4, 2, "=Calculations!B20", "$#,##0")

    ws.row_dimensions[tot_stmt + 5].height = 4

    # Tenant allocation
    ws.cell(row=tot_stmt + 6, column=1, value="TENANT ALLOCATION").font = bold_font(12)
    ws[f"A{tot_stmt + 6}"].fill = fill("D9E8FF")
    ws.merge_cells(f"A{tot_stmt + 6}:B{tot_stmt + 6}")

    ws.cell(row=tot_stmt + 7, column=1, value="Tenant Pro-Rata %").font = Font(size=10)
    formula_cell(ws, tot_stmt + 7, 2, "=Calculations!B25", "0.0000%")

    ws.cell(row=tot_stmt + 8, column=1, value="Tenant Annual CAM Share ($)").font = (
        bold_font()
    )
    formula_cell(ws, tot_stmt + 8, 2, "=Calculations!B26", "$#,##0")

    ws.cell(
        row=tot_stmt + 9, column=1, value="Less: Monthly Estimates Paid (×12)"
    ).font = Font(size=10)
    formula_cell(ws, tot_stmt + 9, 2, "=Calculations!B30", "$#,##0")

    ws.cell(row=tot_stmt + 10, column=1, value="BALANCE DUE / (CREDIT)").font = (
        bold_font(12)
    )
    ws[f"A{tot_stmt + 10}"].fill = fill("0066FF")
    ws[f"A{tot_stmt + 10}"].font = Font(color="FFFFFF", bold=True, size=12)
    formula_cell(ws, tot_stmt + 10, 2, "=Calculations!B31", "$#,##0")
    ws[f"B{tot_stmt + 10}"].fill = fill("0066FF")
    ws[f"B{tot_stmt + 10}"].font = Font(color="FFFFFF", bold=True, size=12)

    ws.cell(
        row=tot_stmt + 12,
        column=1,
        value="This statement was prepared using CapVeri CAM Reconciliation Software.",
    ).font = Font(italic=True, color="475569", size=9)
    ws.merge_cells(f"A{tot_stmt + 12}:B{tot_stmt + 12}")

    ws.print_area = f"A1:B{tot_stmt + 12}"

    write_footer(ws, FOOTER_URL, tot_stmt + 14)
    ws.sheet_view.showGridLines = False


def build_instructions(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 110

    rows: list[tuple[str, bool, int]] = [
        ("CAM Reconciliation Excel Template | How to Use", True, 14),
        ("", False, 11),
        ("OVERVIEW", True, 12),
        (
            "This workbook walks through a complete CAM reconciliation for a single tenant. "
            "Fill in the yellow cells on the Inputs sheet. The Calculations sheet performs "
            "each step in order, and the Statement sheet generates a print-ready summary.",
            False,
            11,
        ),
        ("", False, 11),
        ("STEP 1 — FILL IN INPUTS SHEET", True, 12),
        (
            "Property & Tenant: Enter property name, address, year, tenant details, GLA, "
            "and the monthly CAM estimate the tenant has been paying.",
            False,
            11,
        ),
        (
            "CAM Pool Expenses: Enter Budget and Actual for each category. Variance is "
            "auto-calculated. Add categories or rename rows as needed. The Admin Fee row "
            "should be left at $0 if admin fee is calculated as a % in the Calculations sheet.",
            False,
            11,
        ),
        (
            "Gross-Up Settings: Set to Y if your lease has a gross-up clause. "
            "Occupancy Threshold (typically 90%) is the trigger. Target % (typically 95%) "
            "is the hypothetical occupancy to which variable expenses are scaled.",
            False,
            11,
        ),
        (
            "Cap Settings: Select cap type (Cumulative / Non-Cumulative / None). "
            "Enter the prior year ceiling from last year's reconciliation. "
            "The cap % is typically 5–7% for controllable expenses.",
            False,
            11,
        ),
        ("", False, 11),
        ("WHEN GROSS-UP APPLIES", True, 12),
        (
            "Gross-up addresses the economics of vacancy. When a building is below the threshold "
            "occupancy, fixed common-area costs (lighting, security, HVAC in hallways) are spread "
            "over fewer tenants — making each tenant's share artificially high. The gross-up clause "
            "says: we'll pretend the building is X% occupied and scale variable costs accordingly, "
            "so tenants aren't penalized for the landlord's vacancy.",
            False,
            11,
        ),
        (
            "Important: only variable expenses should be grossed up (janitorial, landscaping, "
            "utilities). Fixed costs (insurance, taxes, management fee) should NOT be grossed up. "
            "This template grosses up the total pool for simplicity; adjust the Calculations sheet "
            "formulas to separate fixed vs. variable if your lease distinguishes them.",
            False,
            11,
        ),
        ("", False, 11),
        ("HOW CAPS WORK", True, 12),
        (
            "The cap limits how much a tenant's CAM share can increase year-over-year on "
            "controllable expenses. The calculation: This Year Ceiling = Prior Year Ceiling × "
            "(1 + Cap%). Recoverable = MIN(Grossed-Up Pool, Ceiling). Under a Non-Cumulative cap, "
            "unused room is lost. Under a Cumulative cap, unused room may carry forward (this "
            "template uses a simplified single-year cap — see the Cumulative Cap Bank Calculator "
            "for multi-year modeling).",
            False,
            11,
        ),
        ("", False, 11),
        ("WHAT TO REVIEW BEFORE SENDING", True, 12),
        (
            "1. Confirm all expense categories are correctly included or excluded per the lease.",
            False,
            11,
        ),
        (
            "2. Verify gross-up was applied (or not) correctly based on actual occupancy.",
            False,
            11,
        ),
        (
            "3. Confirm the pro-rata denominator matches the lease (total GLA vs. leased GLA).",
            False,
            11,
        ),
        ("4. Check that the prior year ceiling in cap settings is correct.", False, 11),
        (
            "5. Verify the monthly estimate amount matches actual payments received.",
            False,
            11,
        ),
        (
            "6. Review the Statement sheet for formatting before sending to tenant.",
            False,
            11,
        ),
        ("", False, 11),
        (
            f"Open the live tool at {site_url('/tools/cam-reconciliation-excel')}",
            False,
            11,
        ),
    ]

    for r_idx, (text, is_bold, size) in enumerate(rows, start=1):
        cell = ws.cell(row=r_idx, column=1, value=text)
        cell.font = Font(bold=is_bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_idx].height = 32 if is_bold else 24

    write_footer(ws, FOOTER_URL, len(rows) + 2)
    ws.sheet_view.showGridLines = False


def main() -> None:
    wb = new_workbook()
    build_inputs(wb)
    build_calculations(wb)
    build_statement(wb)
    build_instructions(wb)
    wb.active = wb["Inputs"]
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
