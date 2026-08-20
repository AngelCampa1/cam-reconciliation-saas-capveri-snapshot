"""Generate the Lease Clause Extraction Matrix workbook.

Output: ``docs/assets/lease-clause-extraction-matrix.xlsx``

Three sheets:
- Matrix: structured grid with all key CAM lease clause fields; 5 sample rows;
  conditional formatting for missing caps and short audit windows
- Summary: count by cap type, audit rights, avg audit window, cumulative cap count
- Instructions: how to use the matrix, what to look for, red flags
"""

from __future__ import annotations

import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))

from _common import (  # noqa: E402
    app_url,
    apply_header_row,
    bold_font,
    docs_assets_dir,
    fill,
    formula_cell,
    new_workbook,
    site_url,
    write_footer,
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402

OUTPUT_PATH = docs_assets_dir() / "lease-clause-extraction-matrix.xlsx"
FOOTER_URL = site_url("/tools/lease-clause-extraction-matrix")

HEADERS = [
    "Property Name",  # A  1
    "Tenant Name",  # B  2
    "Suite",  # C  3
    "Lease Start",  # D  4
    "Lease End",  # E  5
    "CAM Definition (excerpt)",  # F  6
    "Exclusions (list)",  # G  7
    "Cap Type",  # H  8
    "Cap % YoY",  # I  9
    "Base Year",  # J  10
    "Gross-Up Threshold %",  # K  11
    "Gross-Up Method",  # L  12
    "Admin Fee %",  # M  13
    "Mgmt Fee % Cap",  # N  14
    "Audit Rights (Y/N)",  # O  15
    "Audit Window (days)",  # P  16
    "Recon Deadline (mo)",  # Q  17
    "Pro-Rata Method",  # R  18
    "Denominator Definition",  # S  19
    "Notes / Flags",  # T  20
]

SAMPLE_DATA = [
    [
        "Westside Office Park",
        "Acme Corp",
        "100",
        "2021-01-01",
        "2026-12-31",
        "All costs of operating, maintaining, and repairing the Common Areas",
        "Capital items >$10k, executive salaries, leasing commissions",
        "Cumulative",
        "5%",
        "2021",
        "90%",
        "Grossed to 95%",
        "12%",
        "3%",
        "Y",
        "90",
        "4",
        "GLA",
        "Leased only",
        "Strong lease — good benchmark",
    ],
    [
        "Westside Office Park",
        "Beta LLC",
        "200",
        "2020-06-01",
        "2025-05-31",
        "CAM as defined in Exhibit C, excluding items listed in Section 7.4",
        "Insurance, property taxes, management fee, capital expenditures",
        "Non-Cumulative",
        "4%",
        "2020",
        "85%",
        "Actual",
        "10%",
        "None",
        "Y",
        "60",
        "3",
        "GLA",
        "Total project GLA",
        "Denominator uses total project — review for anchor exclusions",
    ],
    [
        "Eastside Retail Strip",
        "Gamma Inc",
        "300",
        "2022-03-15",
        "2027-03-14",
        "All costs of common area maintenance per Schedule A",
        "None listed",
        "None",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "15%",
        "5%",
        "Y",
        "45",
        "3",
        "BOMA",
        "Leased only",
        "FLAG: No cap — costs passed through uncapped. Audit window < 60 days.",
    ],
    [
        "Eastside Retail Strip",
        "Delta Co",
        "400",
        "2019-07-01",
        "2024-06-30",
        "Operating costs as described in Landlord's annual statement",
        "Taxes, insurance, debt service",
        "Cumulative",
        "6%",
        "2019",
        "80%",
        "Grossed to 90%",
        "8%",
        "None",
        "N",
        "N/A",
        "4",
        "GLA",
        "Total project GLA",
        "No audit rights — tenant exposed. Verify admin fee is within lease range.",
    ],
    [
        "Northgate Industrial",
        "Epsilon Ltd",
        "500",
        "2023-01-01",
        "2028-12-31",
        "Net operating costs of the Building and Property",
        "Capital improvements, ground lease payments, financing costs",
        "Non-Cumulative",
        "5%",
        "2023",
        "95%",
        "Grossed to 95%",
        "5%",
        "2%",
        "Y",
        "120",
        "4",
        "Custom",
        "Leasable area per Exhibit B",
        "Custom denominator — confirm Exhibit B is current",
    ],
]

DATA_ROW_START = 2
DATA_ROW_END = DATA_ROW_START + len(SAMPLE_DATA) - 1

COLUMN_WIDTHS = [
    22,
    18,
    10,
    14,
    14,
    36,
    36,
    18,
    12,
    12,
    20,
    20,
    14,
    16,
    18,
    20,
    20,
    18,
    28,
    36,
]


def build_matrix(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Matrix")

    apply_header_row(ws, 1, HEADERS, widths=COLUMN_WIDTHS)
    ws.row_dimensions[1].height = 42

    input_yellow = PatternFill("solid", fgColor="FFF2CC")
    text_font = Font(size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    for row_idx, row_data in enumerate(SAMPLE_DATA, start=DATA_ROW_START):
        ws.row_dimensions[row_idx].height = 60
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = input_yellow
            cell.font = text_font
            cell.alignment = wrap_align

    # ── Conditional Formatting ────────────────────────────────────────────────
    # Red fill on Cap Type (col H = 8) = "None"
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    ws.conditional_formatting.add(
        f"H{DATA_ROW_START}:H{DATA_ROW_END}",
        CellIsRule(operator="equal", formula=['"None"'], fill=red_fill),
    )

    # Yellow fill on Audit Window (col P = 16) < 60 (numeric check)
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    ws.conditional_formatting.add(
        f"P{DATA_ROW_START}:P{DATA_ROW_END}",
        FormulaRule(
            formula=[f"AND(ISNUMBER(P{DATA_ROW_START}),P{DATA_ROW_START}<60)"],
            fill=yellow_fill,
        ),
    )

    # Note row at bottom
    note_row = DATA_ROW_END + 2
    ws.cell(
        row=note_row,
        column=1,
        value="Red cells in Cap Type column = no cap in place (flag for review). "
        "Yellow cells in Audit Window column = window < 60 days (tenant risk). "
        "Add rows as needed — yellow cells are editable.",
    ).font = Font(italic=True, color="475569", size=9)
    ws.merge_cells(f"A{note_row}:T{note_row}")

    write_footer(ws, FOOTER_URL, note_row + 2)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_summary(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 36

    ws.merge_cells("A1:C1")
    ws["A1"] = "Portfolio Summary"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 26

    # ── Cap Type Counts ───────────────────────────────────────────────────────
    from _common import section_label  # noqa: E402

    section_label(ws, "A3", "CAP TYPE DISTRIBUTION")

    ws.cell(row=4, column=1, value="Cumulative").font = bold_font()
    formula_cell(
        ws,
        4,
        2,
        f'=COUNTIF(Matrix!H{DATA_ROW_START}:H{DATA_ROW_END},"Cumulative")',
        "0",
    )

    ws.cell(row=5, column=1, value="Non-Cumulative").font = bold_font()
    formula_cell(
        ws,
        5,
        2,
        f'=COUNTIF(Matrix!H{DATA_ROW_START}:H{DATA_ROW_END},"Non-Cumulative")',
        "0",
    )

    ws.cell(row=6, column=1, value="None (no cap — flag for review)").font = bold_font()
    formula_cell(
        ws,
        6,
        2,
        f'=COUNTIF(Matrix!H{DATA_ROW_START}:H{DATA_ROW_END},"None")',
        "0",
    )

    ws.cell(row=7, column=1, value="Total leases in matrix").font = bold_font()
    formula_cell(
        ws,
        7,
        2,
        f"=COUNTA(Matrix!B{DATA_ROW_START}:B{DATA_ROW_END})",
        "0",
    )

    # ── Audit Rights ──────────────────────────────────────────────────────────
    from _common import section_label  # noqa: E402

    section_label(ws, "A9", "AUDIT RIGHTS")

    ws.cell(row=10, column=1, value="Leases with audit rights (Y)").font = bold_font()
    formula_cell(
        ws,
        10,
        2,
        f'=COUNTIF(Matrix!O{DATA_ROW_START}:O{DATA_ROW_END},"Y")',
        "0",
    )

    ws.cell(row=11, column=1, value="Leases without audit rights (N)").font = (
        bold_font()
    )
    formula_cell(
        ws,
        11,
        2,
        f'=COUNTIF(Matrix!O{DATA_ROW_START}:O{DATA_ROW_END},"N")',
        "0",
    )

    ws.cell(
        row=12, column=1, value="Average audit window (days, numeric only)"
    ).font = bold_font()
    formula_cell(
        ws,
        12,
        2,
        f'=AVERAGEIF(Matrix!P{DATA_ROW_START}:P{DATA_ROW_END},">0")',
        "0.0",
    )
    ws.cell(row=12, column=3, value="Excludes N/A rows (no audit rights)").font = Font(
        italic=True, color="475569"
    )

    ws.cell(row=13, column=1, value="Audit windows < 60 days (tenant risk)").font = (
        bold_font()
    )
    formula_cell(
        ws,
        13,
        2,
        f'=COUNTIFS(Matrix!P{DATA_ROW_START}:P{DATA_ROW_END},"<60",'
        f'Matrix!P{DATA_ROW_START}:P{DATA_ROW_END},">0")',
        "0",
    )

    section_label(ws, "A15", "GROSS-UP & ADMIN FEE")

    ws.cell(row=16, column=1, value="Leases with gross-up clause applied").font = (
        bold_font()
    )
    formula_cell(
        ws,
        16,
        2,
        f'=COUNTIF(Matrix!L{DATA_ROW_START}:L{DATA_ROW_END},"Grossed*")',
        "0",
    )

    ws.cell(row=17, column=1, value="Average admin fee %").font = bold_font()
    formula_cell(
        ws,
        17,
        2,
        f'=AVERAGEIF(Matrix!M{DATA_ROW_START}:M{DATA_ROW_END},">0%")',
        "0.0%",
    )

    ws.cell(
        row=19,
        column=1,
        value="Counts reference the Matrix sheet. Add rows to the Matrix to "
        "automatically update this summary.",
    ).font = Font(italic=True, color="475569", size=9)
    ws.merge_cells("A19:C19")

    write_footer(ws, FOOTER_URL, 21)
    ws.sheet_view.showGridLines = False


def build_instructions(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 110

    rows: list[tuple[str, bool, int]] = [
        ("Lease Clause Extraction Matrix | How to Use", True, 14),
        ("", False, 11),
        ("PURPOSE", True, 12),
        (
            "This matrix is a structured template for extracting and comparing the key CAM-related "
            "lease clauses across a portfolio. By standardizing how you capture lease provisions, "
            "you can quickly identify inconsistencies, flag risk, and prioritize lease audits.",
            False,
            11,
        ),
        ("", False, 11),
        ("HOW TO USE THE MATRIX SHEET", True, 12),
        (
            "1. Copy each tenant's lease clauses into the corresponding columns. "
            "Use direct excerpt language for CAM Definition and Exclusions columns.",
            False,
            11,
        ),
        (
            "2. For Cap Type, use exactly: Cumulative / Non-Cumulative / None "
            "(triggering the red conditional formatting for no-cap leases).",
            False,
            11,
        ),
        (
            "3. For Audit Window, enter the numeric number of days only. "
            "Enter 'N/A' if the tenant has no audit right.",
            False,
            11,
        ),
        ("4. Review Summary sheet for portfolio-level statistics.", False, 11),
        ("", False, 11),
        ("WHAT TO LOOK FOR — COLUMN BY COLUMN", True, 12),
        (
            "CAM Definition: Does it include or exclude management fees? Is it 'controllable "
            "expenses' or all operating expenses? Broad definitions favor landlords.",
            False,
            11,
        ),
        (
            "Exclusions: The longer the list, the smaller the recoverable pool. "
            "Watch for: capital improvements, executive salaries, prior owner costs, "
            "lease-up costs, financing costs.",
            False,
            11,
        ),
        (
            "Cap Type / Cap %: Cumulative caps let landlords bank unused room — "
            "more aggressive than non-cumulative. No cap is unusual and should be flagged.",
            False,
            11,
        ),
        (
            "Gross-Up: Without a gross-up clause, vacant space effectively subsidizes occupying tenants "
            "by inflating the denominator. Threshold typically 90–95%; method should specify "
            "variable vs. fixed expense treatment.",
            False,
            11,
        ),
        (
            "Admin Fee %: Commercial market range is 5–15%. Below 5% may undercharge; "
            "above 15% may be challenged by tenants. Confirm whether taxes, insurance, "
            "and management fee are in or out of the admin fee base.",
            False,
            11,
        ),
        (
            "Audit Rights / Audit Window: Fewer than 60 days after receipt of reconciliation "
            "is a risk for tenants with complex portfolios. No audit right means tenants "
            "have no recourse.",
            False,
            11,
        ),
        (
            "Pro-Rata Method / Denominator: 'Total project GLA' vs. 'leased area only' "
            "can swing recovery by 10–20% at high vacancy. Confirm what the lease says "
            "vs. what is being billed.",
            False,
            11,
        ),
        ("", False, 11),
        ("RED FLAGS TO ESCALATE", True, 12),
        ("- Cap Type = None (no protection against expense spikes)", False, 11),
        ("- Audit window < 60 days (operational risk for tenants)", False, 11),
        ("- No audit rights at all", False, 11),
        ("- Admin fee > 15% or < 3%", False, 11),
        ("- Denominator = Total project GLA with high vacancy", False, 11),
        ("- No gross-up clause on a building with < 80% occupancy", False, 11),
        ("", False, 11),
        (
            f"Open the live tool at {site_url('/tools/lease-clause-extraction-matrix')}",
            False,
            11,
        ),
    ]

    for r_idx, (text, is_bold, size) in enumerate(rows, start=1):
        cell = ws.cell(row=r_idx, column=1, value=text)
        cell.font = Font(bold=is_bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_idx].height = 32 if is_bold else 24

    write_footer(ws, FOOTER_URL, len(rows) + 2)
    ws.sheet_view.showGridLines = False


def main() -> None:
    wb = new_workbook()
    build_matrix(wb)
    build_summary(wb)
    build_instructions(wb)
    wb.active = wb["Matrix"]
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
