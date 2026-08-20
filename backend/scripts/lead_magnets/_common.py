"""Shared brand styling helpers for lead magnet generators.

Reuses the visual conventions established in
``backend/scripts/generate_cam_calculator.py`` and ``generate_lease_matrix.py``:

- Header fill ``#0066FF`` (CapVeri blue) with bold white header font
- Input cells filled ``#FFF2CC`` (yellow) so users know what to edit
- Calculated cells filled ``#F2F2F2`` (gray) so users do not overwrite formulas
- Footer URL link on every PDF page plus a page counter

The helpers are stateless wrappers around ``openpyxl`` and ``reportlab``
primitives. They are intentionally narrow so each generator can mix in custom
layout while still sharing a consistent look.
"""

from __future__ import annotations

import json
import pathlib
from collections.abc import Iterable
from urllib.parse import urljoin

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen.canvas import Canvas

# Brand palette
HEADER_BLUE = "0066FF"
INPUT_YELLOW = "FFF2CC"
CALCULATED_GRAY = "F2F2F2"
SECTION_BLUE = "D9E8FF"
WHITE = "FFFFFF"

DOCS_ASSETS_DIR = (
    pathlib.Path(__file__).resolve().parent.parent.parent.parent / "docs" / "assets"
)
PUBLIC_KNOWLEDGE_PATH = (
    pathlib.Path(__file__).resolve().parents[3]
    / "knowledge"
    / "generated"
    / "public-knowledge.json"
)


def _company_knowledge() -> dict[str, object]:
    return json.loads(PUBLIC_KNOWLEDGE_PATH.read_text(encoding="utf-8"))["company"]


def site_url(path: str = "/") -> str:
    """Build a canonical marketing-site URL from public knowledge."""
    base_url = str(_company_knowledge()["siteUrl"])
    if path in {"", "/"}:
        return base_url
    return urljoin(f"{base_url}/", path.lstrip("/"))


def app_url(path: str = "/") -> str:
    """Build a canonical app URL from public knowledge."""
    base_url = str(_company_knowledge()["appUrl"])
    if path in {"", "/"}:
        return base_url
    return urljoin(f"{base_url}/", path.lstrip("/"))


def docs_assets_dir() -> pathlib.Path:
    """Return the absolute path to ``docs/assets/``, creating it if needed."""
    DOCS_ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    return DOCS_ASSETS_DIR


# ─────────────────────────────────────────────
# openpyxl helpers
# ─────────────────────────────────────────────
def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def header_font() -> Font:
    return Font(color=WHITE, bold=True, size=11)


def bold_font(size: int = 11) -> Font:
    return Font(bold=True, size=size)


def apply_header_row(
    ws: Worksheet,
    row: int,
    values: list[str],
    widths: list[float] | None = None,
) -> None:
    """Write a CapVeri-blue header row with white bold text and optional widths."""
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = fill(HEADER_BLUE)
        cell.font = header_font()
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    if widths:
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def input_cell(ws: Worksheet, row: int, col: int, value=None, number_format: str = ""):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(INPUT_YELLOW)
    if number_format:
        cell.number_format = number_format
    return cell


def formula_cell(
    ws: Worksheet, row: int, col: int, formula: str, number_format: str = ""
):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.fill = fill(CALCULATED_GRAY)
    if number_format:
        cell.number_format = number_format
    return cell


def section_label(ws: Worksheet, cell_ref: str, text: str) -> None:
    cell = ws[cell_ref]
    cell.value = text
    cell.font = bold_font(12)
    cell.fill = fill(SECTION_BLUE)


def write_footer(ws: Worksheet, url: str, row: int) -> None:
    """Write a hyperlinked footer URL on a worksheet."""
    cell = ws.cell(row=row, column=1, value=url)
    cell.hyperlink = url
    cell.font = Font(color="0066FF", underline="single", italic=True, size=9)


def new_workbook() -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    return wb


# ─────────────────────────────────────────────
# reportlab helpers
# ─────────────────────────────────────────────
NAVY = colors.HexColor("#0A2540")
BRAND_BLUE = colors.HexColor("#0066FF")
SLATE = colors.HexColor("#475569")
LIGHT_GRAY = colors.HexColor("#F1F5F9")


def make_page_decorator(footer_url: str, doc_title: str, last_updated: str):
    """Return an ``onPage`` callable that draws a header, footer URL, and page number."""

    def _decorator(canvas: Canvas, doc) -> None:  # type: ignore[no-untyped-def]
        canvas.saveState()
        width, height = letter

        # Top brand bar
        canvas.setFillColor(BRAND_BLUE)
        canvas.rect(0, height - 0.18 * inch, width, 0.18 * inch, fill=1, stroke=0)

        # Footer separator
        canvas.setStrokeColor(LIGHT_GRAY)
        canvas.setLineWidth(0.5)
        canvas.line(
            doc.leftMargin,
            0.6 * inch,
            width - doc.rightMargin,
            0.6 * inch,
        )

        # Footer URL (left) and page number (right)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(BRAND_BLUE)
        canvas.drawString(doc.leftMargin, 0.42 * inch, footer_url)
        canvas.linkURL(
            footer_url,
            (
                doc.leftMargin,
                0.36 * inch,
                doc.leftMargin + 4.5 * inch,
                0.5 * inch,
            ),
            relative=0,
        )

        canvas.setFillColor(SLATE)
        canvas.drawRightString(
            width - doc.rightMargin,
            0.42 * inch,
            f"Page {doc.page}",
        )
        canvas.drawString(
            doc.leftMargin,
            0.28 * inch,
            f"{doc_title}  |  Last updated {last_updated}  |  CapVeri",
        )

        canvas.restoreState()

    return _decorator


def deterministic_canvas(*args, **kwargs) -> Canvas:  # type: ignore[no-untyped-def]
    """Reportlab canvas factory that produces byte-stable PDFs across runs."""
    kwargs["invariant"] = 1
    return Canvas(*args, **kwargs)


def join_paragraphs(parts: Iterable[str]) -> str:
    """Join paragraphs with double line breaks, dropping empty entries."""
    return "<br/><br/>".join(p for p in parts if p)
