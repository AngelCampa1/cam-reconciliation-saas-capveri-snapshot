"""Generate the Cumulative CAM Cap Bank Calculator workbook.

Output: ``docs/assets/cumulative-cap-bank-calculator.xlsx``

The workbook helps an asset manager model an 8-year horizon of CAM caps where
unused cap room from one year may (or may not) be carried forward to absorb
overages in later years. Three sheets:

- Inputs: editable lease parameters (base year CAM, cap %, cap type, actuals)
- Calculations: per-year ceiling, recoverable amount, unused cap room, bank
  balance carried forward
- Instructions: domain primer on cumulative vs non-cumulative caps and how
  bank balances are typically drafted in real leases
"""

from __future__ import annotations

import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))

from _common import (  # noqa: E402
    app_url,
    apply_header_row,
    bold_font,
    docs_assets_dir,
    fill,
    formula_cell,
    input_cell,
    new_workbook,
    section_label,
    site_url,
    write_footer,
)
from openpyxl.styles import Alignment, Font  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

OUTPUT_PATH = docs_assets_dir() / "cumulative-cap-bank-calculator.xlsx"
FOOTER_URL = site_url("/tools/cumulative-cap-bank-calculator")


def build_inputs(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Inputs")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 38

    ws.merge_cells("A1:C1")
    ws["A1"] = "Cumulative CAM Cap Bank Calculator | CapVeri"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 28

    section_label(ws, "A3", "LEASE INPUTS")
    ws.merge_cells("A3:C3")

    rows = [
        ("Base year (e.g., 2024)", 2024, "Year the cap baseline is set."),
        ("Base year CAM ($)", 250000, "Audited recoverable CAM in base year."),
        (
            "Annual cap rate (%)",
            0.05,
            "Lease cap on year-over-year increase (typical 4–7%).",
        ),
    ]
    for offset, (label, value, note) in enumerate(rows):
        r = 4 + offset
        ws.cell(row=r, column=1, value=label).font = bold_font()
        if isinstance(value, float):
            input_cell(ws, r, 2, value, "0.00%")
        else:
            input_cell(ws, r, 2, value, "0")
        ws.cell(row=r, column=3, value=note).font = Font(italic=True, color="475569")

    # Cap type dropdown
    ws.cell(row=7, column=1, value="Cap type").font = bold_font()
    cap_type_cell = input_cell(ws, 7, 2, "Cumulative")
    ws.cell(
        row=7,
        column=3,
        value="Cumulative banks unused room; Non-cumulative resets each year.",
    ).font = Font(italic=True, color="475569")
    dv = DataValidation(
        type="list",
        formula1='"Cumulative,Non-Cumulative"',
        allow_blank=False,
    )
    ws.add_data_validation(dv)
    dv.add(cap_type_cell)

    section_label(ws, "A9", "ACTUAL CAM BY YEAR (8-YEAR HORIZON)")
    ws.merge_cells("A9:C9")
    apply_header_row(
        ws, 10, ["Year offset", "Year", "Actual CAM ($)"], widths=[18, 14, 22]
    )
    actuals = [255000, 268000, 295000, 280000, 315000, 330000, 348000, 360000]
    for i, amount in enumerate(actuals):
        r = 11 + i
        ws.cell(row=r, column=1, value=f"Year {i + 1}").font = bold_font()
        formula_cell(ws, r, 2, f"=$B$4+{i + 1}", "0")
        input_cell(ws, r, 3, amount, "$#,##0")

    write_footer(ws, FOOTER_URL, 22)
    ws.sheet_view.showGridLines = False


def build_calculations(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Calculations")
    ws.merge_cells("A1:H1")
    ws["A1"] = "Year-by-Year Cap Bank Reconciliation"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill("0066FF")
    ws.row_dimensions[1].height = 26

    apply_header_row(
        ws,
        3,
        [
            "Year",
            "Actual CAM ($)",
            "Prior Ceiling ($)",
            "Current Ceiling ($)",
            "Recoverable ($)",
            "Unused Room ($)",
            "Bank Balance In ($)",
            "Bank Balance Out ($)",
        ],
        widths=[10, 16, 18, 20, 18, 18, 20, 20],
    )

    # Year 1: ceiling = base * (1 + cap)
    # Year N: prior ceiling = previous current ceiling
    #         current ceiling = prior ceiling * (1 + cap)
    #         recoverable_uncapped = MIN(actual, current ceiling + bank_in)  if cumulative
    #                              = MIN(actual, current ceiling)            if non-cumulative
    #         unused_room = MAX(0, current ceiling - actual)
    #         bank_out = bank_in + unused_room - over_cap_drawn   (cumulative only)
    for i in range(8):
        r = 4 + i
        year_offset = i + 1
        actual_ref = f"Inputs!$C${10 + year_offset}"

        # Year label
        formula_cell(ws, r, 1, f"=Inputs!$B$4+{year_offset}", "0")
        # Actual
        formula_cell(ws, r, 2, f"={actual_ref}", "$#,##0")

        if i == 0:
            # Prior ceiling = base year CAM
            formula_cell(ws, r, 3, "=Inputs!$B$5", "$#,##0")
            # Current ceiling = base * (1 + cap)
            formula_cell(ws, r, 4, "=Inputs!$B$5*(1+Inputs!$B$6)", "$#,##0")
            # Bank balance in = 0
            formula_cell(ws, r, 7, "=0", "$#,##0")
        else:
            # Prior ceiling = previous current ceiling
            formula_cell(ws, r, 3, f"=D{r - 1}", "$#,##0")
            # Current ceiling = prior * (1 + cap)
            formula_cell(ws, r, 4, f"=C{r}*(1+Inputs!$B$6)", "$#,##0")
            # Bank in = previous bank out (only if cumulative)
            formula_cell(
                ws,
                r,
                7,
                f'=IF(Inputs!$B$7="Cumulative",H{r - 1},0)',
                "$#,##0",
            )

        # Recoverable: cumulative uses ceiling + bank_in; non-cumulative uses ceiling only
        formula_cell(
            ws,
            r,
            5,
            f'=IF(Inputs!$B$7="Cumulative",MIN(B{r},D{r}+G{r}),MIN(B{r},D{r}))',
            "$#,##0",
        )
        # Unused room = MAX(0, current ceiling - actual)
        formula_cell(ws, r, 6, f"=MAX(0,D{r}-B{r})", "$#,##0")
        # Bank out:
        #   cumulative: bank_in + unused_room - amount_drawn
        #   amount_drawn = MAX(0, recoverable - current_ceiling)
        #   non-cumulative: 0 (resets each year)
        formula_cell(
            ws,
            r,
            8,
            f'=IF(Inputs!$B$7="Cumulative",MAX(0,G{r}+F{r}-MAX(0,E{r}-D{r})),0)',
            "$#,##0",
        )

    # Totals row
    section_label(ws, "A13", "8-YEAR TOTALS")
    formula_cell(ws, 13, 2, "=SUM(B4:B11)", "$#,##0")
    formula_cell(ws, 13, 5, "=SUM(E4:E11)", "$#,##0")
    formula_cell(ws, 13, 6, "=SUM(F4:F11)", "$#,##0")

    # Lost recovery row
    ws.cell(row=14, column=1, value="Unrecoverable overage ($)").font = bold_font()
    formula_cell(ws, 14, 5, "=SUM(B4:B11)-SUM(E4:E11)", "$#,##0")

    ws.cell(
        row=16,
        column=1,
        value="Yellow cells (Inputs sheet) are editable. Gray cells are calculated.",
    ).font = Font(italic=True, color="475569", size=9)
    ws.merge_cells("A16:H16")

    write_footer(ws, FOOTER_URL, 18)
    ws.freeze_panes = "A4"


def build_instructions(wb) -> None:  # type: ignore[no-untyped-def]
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 100

    rows = [
        ("Cumulative CAM Cap Bank Calculator | How it works", True, 14),
        ("", False, 11),
        ("WHAT IS A CAM CAP?", True, 12),
        (
            "Most office and retail leases that pass through controllable CAM include a cap on the "
            "year-over-year increase. Typical cap language: 'Tenant's share of Controllable Operating "
            "Expenses shall not increase by more than five percent (5%) per Lease Year over the prior "
            "Lease Year's actual Controllable Operating Expenses.'",
            False,
            11,
        ),
        ("", False, 11),
        ("CUMULATIVE VS NON-CUMULATIVE", True, 12),
        (
            "A NON-CUMULATIVE cap resets each year. If actual CAM is below the cap one year, that "
            "unused 'room' is lost; the next year's ceiling is still calculated only off the prior "
            "year's ceiling. Tenant-friendly drafting.",
            False,
            11,
        ),
        (
            "A CUMULATIVE cap allows unused room to BANK. If Year 1 ceiling is $262,500 and actual is "
            "only $255,000, the $7,500 of unused room carries into Year 2. If Year 2 actual exceeds "
            "the Year 2 ceiling, the bank can absorb the overage up to the bank balance. Landlord-"
            "friendly drafting; common in NNN retail.",
            False,
            11,
        ),
        ("", False, 11),
        ("HOW THE CALCULATIONS SHEET WORKS", True, 12),
        (
            "Year 1 ceiling = Base Year CAM x (1 + cap rate). Each subsequent year's ceiling = prior "
            "year's ceiling x (1 + cap rate). Recoverable amount is the lesser of actual CAM or the "
            "ceiling (plus bank balance, if cumulative). Unused room = MAX(0, ceiling - actual). Bank "
            "out = bank in + unused room - amount drawn against bank.",
            False,
            11,
        ),
        ("", False, 11),
        ("COMMON LEASE LANGUAGE PATTERNS TO SCRUB", True, 12),
        (
            "1. 'Compounding cap': always cumulative; ceiling resets to actuals if actual is below "
            "ceiling. Aggressive landlord drafting — model both ways and discuss.",
            False,
            11,
        ),
        (
            "2. 'Cumulative compounding cap': what this calculator models by default. Bank carries; "
            "ceiling compounds off prior ceiling.",
            False,
            11,
        ),
        (
            "3. 'Non-cumulative non-compounding': ceiling each year = base year x (1 + cap)^N. "
            "Tenant-friendly. Modify the Year-N ceiling formula in the Calculations sheet if your "
            "lease uses this structure.",
            False,
            11,
        ),
        (
            "4. 'Cap excludes uncontrollables': insurance, taxes, utilities, and snow removal are "
            "frequently carved out. Run the calculator on the controllable pool only.",
            False,
            11,
        ),
        ("", False, 11),
        ("HOW TO USE", True, 12),
        ("1. Edit yellow cells on the Inputs sheet only.", False, 11),
        ("2. Pick Cumulative or Non-Cumulative from the dropdown.", False, 11),
        ("3. Review the 8-year horizon on the Calculations sheet.", False, 11),
        (
            "4. Read the 'Unrecoverable overage' line at the bottom — that is the dollar amount the "
            "landlord cannot pass through under the cap structure modeled.",
            False,
            11,
        ),
        ("", False, 11),
        (
            f"Open the live tool at {site_url('/tools/cumulative-cap-bank-calculator')}",
            False,
            11,
        ),
    ]

    for r_idx, (text, is_bold, size) in enumerate(rows, start=1):
        cell = ws.cell(row=r_idx, column=1, value=text)
        cell.font = Font(bold=is_bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_idx].height = 30 if is_bold else 22

    write_footer(ws, FOOTER_URL, len(rows) + 2)
    ws.sheet_view.showGridLines = False


def main() -> None:
    wb = new_workbook()
    build_inputs(wb)
    build_calculations(wb)
    build_instructions(wb)
    wb.active = wb["Calculations"]
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
