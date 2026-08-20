"""
Generate the Lease Abstract Discrepancy Matrix Excel workbook.

Produces: docs/assets/lease-abstract-matrix.xlsx

Run from repo root:
    pip install openpyxl
    python backend/scripts/generate_lease_matrix.py
"""

import pathlib
from datetime import date

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = (
    pathlib.Path(__file__).parent.parent.parent
    / "docs"
    / "assets"
    / "lease-abstract-matrix.xlsx"
)

# Colors
DARK_BLUE = "1F3864"
LIGHT_BLUE_HEADER = "D9E1F2"
CALCULATED_GRAY = "F2F2F2"
INPUT_YELLOW = "FFF2CC"
FLAG_RED = "FF0000"
FLAG_YELLOW = "FFFF00"
FLAG_ORANGE = "FFA500"


def header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def white_bold_font() -> Font:
    return Font(color="FFFFFF", bold=True, size=11)


def bold_font() -> Font:
    return Font(bold=True)


def set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def apply_header_row(ws, row: int, values: list, widths: list | None = None) -> None:
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = header_fill(DARK_BLUE)
        cell.font = white_bold_font()
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    if widths:
        for col_idx, w in enumerate(widths, start=1):
            set_col_width(ws, col_idx, w)
    ws.row_dimensions[row].height = 40


def yellow_input(ws, row: int, col: int, value=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = header_fill(INPUT_YELLOW)


def gray_cell(ws, row: int, col: int, value=None, number_format: str = "") -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = header_fill(CALCULATED_GRAY)
    if number_format:
        cell.number_format = number_format


# ─────────────────────────────────────────────
# TAB 1: Instructions
# ─────────────────────────────────────────────
def build_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 95

    rows = [
        ("CapVeri — Free Lease Abstract Discrepancy Matrix", True, 14),
        ("", False, 11),
        ("WHAT IS A LEASE ABSTRACT?", True, 12),
        (
            "A lease abstract is a summarized version of key lease terms that affect CAM "
            "reconciliation: expense stops, CAM caps, pro-rata share, excluded categories, and "
            "admin fees. Property controllers use abstracts to quickly verify tenant billings "
            "without re-reading full lease documents.",
            False,
            11,
        ),
        ("", False, 11),
        ("WHAT THIS MATRIX TRACKS", True, 12),
        ("• Lease type: NNN (triple net), Gross, Modified Gross", False, 11),
        ("• Base year and expense stop amounts (used in gross leases)", False, 11),
        (
            "• CAM cap type and cap percentage (limits annual CAM increases for tenants)",
            False,
            11,
        ),
        (
            "• Admin fee percentage (management overhead charged on top of CAM expenses)",
            False,
            11,
        ),
        (
            "• Excluded expense categories (insurance, taxes, cap-ex items that are carved out)",
            False,
            11,
        ),
        (
            "• Anchor carve-outs (anchor tenants sometimes excluded from pro-rata pools)",
            False,
            11,
        ),
        ("• Reconciliation due dates and last reconciled year", False, 11),
        ("", False, 11),
        ("DISCREPANCY FLAGS", True, 12),
        (
            "The matrix auto-flags common data entry errors and high-risk conditions:",
            False,
            11,
        ),
        (
            "🔴 RED: CAM Cap Type is set but CAM Cap % is blank — cap will be ignored in billing.",
            False,
            11,
        ),
        (
            "🟡 YELLOW: Both Base Year and Expense Stop are populated — verify which applies "
            "(typically it is one or the other, not both).",
            False,
            11,
        ),
        (
            "🟠 ORANGE: Last Reconciled Year is more than 12 months ago — reconciliation is overdue.",
            False,
            11,
        ),
        ("", False, 11),
        ("HOW TO USE", True, 12),
        ("1. Go to 'Portfolio Overview' to enter building-level details.", False, 11),
        (
            "2. Go to 'Tenant Matrix' to enter one row per tenant. Yellow cells = inputs.",
            False,
            11,
        ),
        (
            "3. The Discrepancy Flag column auto-populates. Check 'Discrepancy Summary' "
            "for a list of all flagged rows.",
            False,
            11,
        ),
        (
            "4. Use the Notes column to record follow-up actions.",
            False,
            11,
        ),
        ("", False, 11),
        (
            "For full automation across your portfolio, visit capveri.com",
            False,
            11,
        ),
    ]

    for r_idx, (text, bold, size) in enumerate(rows, start=1):
        cell = ws.cell(row=r_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_idx].height = 30 if bold else 20

    ws.sheet_view.showGridLines = False


# ─────────────────────────────────────────────
# TAB 2: Portfolio Overview
# ─────────────────────────────────────────────
def build_portfolio_overview(wb: Workbook) -> None:
    ws = wb.create_sheet("Portfolio Overview")

    ws.merge_cells("A1:F1")
    ws["A1"] = "capveri.com — Portfolio Overview"
    ws["A1"].fill = header_fill(DARK_BLUE)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    apply_header_row(
        ws,
        3,
        [
            "Property Name",
            "Address",
            "GLA (SF)",
            "# Tenants",
            "Management System",
            "Fiscal Year End",
        ],
        widths=[28, 36, 14, 12, 22, 16],
    )

    mgmt_dv = DataValidation(
        type="list",
        formula1='"Yardi,MRI,RealPage,AppFolio,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(mgmt_dv)

    # Sample property
    yellow_input(ws, 4, 1, "Lakeside Commerce Center")
    yellow_input(ws, 4, 2, "123 Commerce Blvd, Austin, TX 78701")
    yellow_input(ws, 4, 3, 50000)
    ws.cell(row=4, column=3).number_format = "#,##0"
    yellow_input(ws, 4, 4, 3)
    yellow_input(ws, 4, 5, "Yardi")
    mgmt_dv.add(ws.cell(row=4, column=5))
    yellow_input(ws, 4, 6, "December 31")

    # Blank rows for more properties
    for r in range(5, 12):
        for col in [1, 2, 3, 4, 5, 6]:
            yellow_input(ws, r, col)
            if col == 3:
                ws.cell(row=r, column=col).number_format = "#,##0"
        mgmt_dv.add(ws.cell(row=r, column=5))


# ─────────────────────────────────────────────
# TAB 3: Tenant Matrix (main)
# ─────────────────────────────────────────────

TENANT_MATRIX_HEADERS = [
    "Tenant Name",  # A = col 1
    "Suite / Unit",  # B = col 2
    "Leased SF",  # C = col 3
    "Pro-Rata %",  # D = col 4
    "Lease Type",  # E = col 5
    "Base Year",  # F = col 6
    "Expense Stop ($)",  # G = col 7
    "CAM Cap Type",  # H = col 8
    "CAM Cap %",  # I = col 9
    "Admin Fee %",  # J = col 10
    "Excluded Categories",  # K = col 11
    "Anchor Carve-Out",  # L = col 12
    "Recon Due Date",  # M = col 13
    "Last Reconciled Yr",  # N = col 14
    "Discrepancy Flag",  # O = col 15
    "Notes",  # P = col 16
]

TENANT_MATRIX_WIDTHS = [
    24,
    14,
    12,
    12,
    14,
    10,
    16,
    18,
    12,
    12,
    26,
    16,
    16,
    18,
    20,
    28,
]

SAMPLE_TENANTS = [
    {
        "name": "Acme Corp",
        "suite": "100",
        "sf": 8500,
        "lease_type": "NNN",
        "base_year": "",
        "expense_stop": "",
        "cap_type": "Non-Cumulative",
        "cap_pct": 0.05,
        "admin_fee": 0.15,
        "excluded": "Capital Expenditures",
        "anchor": "No",
        "recon_due": "March 31",
        "last_recon": 2024,
    },
    {
        "name": "BlueSky LLC",
        "suite": "200",
        "sf": 12000,
        "lease_type": "Gross",
        "base_year": 2019,
        "expense_stop": 8.50,
        "cap_type": "",
        "cap_pct": "",
        "admin_fee": 0.10,
        "excluded": "Insurance, Property Taxes",
        "anchor": "No",
        "recon_due": "April 30",
        "last_recon": 2022,  # deliberately stale to trigger orange flag
    },
    {
        "name": "Metro Dental",
        "suite": "105",
        "sf": 3200,
        "lease_type": "Modified Gross",
        "base_year": "",
        "expense_stop": "",
        "cap_type": "Cumulative",
        "cap_pct": "",  # deliberately blank to trigger red flag
        "admin_fee": 0.15,
        "excluded": "",
        "anchor": "No",
        "recon_due": "March 31",
        "last_recon": 2024,
    },
]


def build_tenant_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("Tenant Matrix")

    # Branding header
    ws.merge_cells("A1:P1")
    ws["A1"] = "capveri.com — Lease Abstract Discrepancy Matrix"
    ws["A1"].fill = header_fill(DARK_BLUE)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # GLA input for pro-rata calculation
    ws["A2"] = "Building GLA (SF):"
    ws["A2"].font = bold_font()
    yellow_input(ws, 2, 2, 50000)
    ws["B2"].number_format = "#,##0"

    header_row = 4
    apply_header_row(ws, header_row, TENANT_MATRIX_HEADERS, TENANT_MATRIX_WIDTHS)

    # Data validations
    lease_dv = DataValidation(
        type="list", formula1='"NNN,Gross,Modified Gross"', allow_blank=True
    )
    cap_dv = DataValidation(
        type="list",
        formula1='"Non-Cumulative,Cumulative,Lesser Of,None"',
        allow_blank=True,
    )
    anchor_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(lease_dv)
    ws.add_data_validation(cap_dv)
    ws.add_data_validation(anchor_dv)

    data_start = header_row + 1

    for i, t in enumerate(SAMPLE_TENANTS):
        r = data_start + i
        ws.cell(row=r, column=1, value=t["name"])
        yellow_input(ws, r, 2, t["suite"])
        yellow_input(ws, r, 3, t["sf"])
        ws.cell(row=r, column=3).number_format = "#,##0"

        # Pro-Rata % = leased SF / GLA
        gray_cell(
            ws,
            r,
            4,
            f"=IF($B$2=0,0,C{r}/$B$2)",
            "0.00%",
        )

        yellow_input(ws, r, 5, t["lease_type"])
        lease_dv.add(ws.cell(row=r, column=5))

        yellow_input(ws, r, 6, t["base_year"] or None)
        yellow_input(ws, r, 7, t["expense_stop"] or None)
        if t["expense_stop"]:
            ws.cell(row=r, column=7).number_format = "$#,##0.00"

        yellow_input(ws, r, 8, t["cap_type"] or None)
        cap_dv.add(ws.cell(row=r, column=8))

        yellow_input(ws, r, 9, t["cap_pct"] or None)
        if t["cap_pct"]:
            ws.cell(row=r, column=9).number_format = "0%"

        yellow_input(ws, r, 10, t["admin_fee"] or None)
        if t["admin_fee"]:
            ws.cell(row=r, column=10).number_format = "0%"

        yellow_input(ws, r, 11, t["excluded"] or None)
        yellow_input(ws, r, 12, t["anchor"])
        anchor_dv.add(ws.cell(row=r, column=12))
        yellow_input(ws, r, 13, t["recon_due"])
        yellow_input(ws, r, 14, t["last_recon"] or None)

        # Discrepancy Flag formula
        # Red: CAM Cap Type set but Cap % blank
        # Yellow: Both Base Year AND Expense Stop set
        # Orange: Last Reconciled Year <= current year - 2 (more than 12 months)
        current_year = date.today().year
        flag_formula = (
            f'=IF(AND(H{r}<>"",H{r}<>"None",I{r}=""),"RED: CAM Cap % missing",'
            f'IF(AND(F{r}<>"",G{r}<>""),"YELLOW: Base year + stop both set",'
            f'IF(AND(N{r}<>"",N{r}<={current_year - 2}),"ORANGE: Reconciliation overdue","")'
            f"))"
        )
        gray_cell(ws, r, 15, flag_formula)

        yellow_input(ws, r, 16, None)  # Notes

    # Blank rows 4-15
    for i in range(len(SAMPLE_TENANTS), 15):
        r = data_start + i
        for col in [1, 2, 3, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 16]:
            yellow_input(ws, r, col)
        ws.cell(row=r, column=3).number_format = "#,##0"
        ws.cell(row=r, column=7).number_format = "$#,##0.00"
        ws.cell(row=r, column=9).number_format = "0%"
        ws.cell(row=r, column=10).number_format = "0%"
        lease_dv.add(ws.cell(row=r, column=5))
        cap_dv.add(ws.cell(row=r, column=8))
        anchor_dv.add(ws.cell(row=r, column=12))
        gray_cell(ws, r, 4, f"=IF($B$2=0,0,C{r}/$B$2)", "0.00%")
        current_year = date.today().year
        flag_formula = (
            f'=IF(AND(H{r}<>"",H{r}<>"None",I{r}=""),"RED: CAM Cap % missing",'
            f'IF(AND(F{r}<>"",G{r}<>""),"YELLOW: Base year + stop both set",'
            f'IF(AND(N{r}<>"",N{r}<={current_year - 2}),"ORANGE: Reconciliation overdue","")'
            f"))"
        )
        gray_cell(ws, r, 15, flag_formula)

    # Totals row
    totals_row = data_start + 15
    ws.cell(row=totals_row, column=1, value="TOTALS").font = bold_font()
    gray_cell(ws, totals_row, 3, f"=SUM(C{data_start}:C{totals_row-1})", "#,##0")
    gray_cell(ws, totals_row, 4, f"=SUM(D{data_start}:D{totals_row-1})", "0.00%")

    ws.freeze_panes = "A5"

    # Conditional formatting for flag column (O)
    flag_col = f"O{data_start}:O{totals_row-1}"
    ws.conditional_formatting.add(
        flag_col,
        FormulaRule(
            formula=[f'LEFT(O{data_start},3)="RED"'],
            fill=PatternFill(bgColor="FFC7CE", fill_type="solid"),
            font=Font(color="9C0006"),
        ),
    )
    ws.conditional_formatting.add(
        flag_col,
        FormulaRule(
            formula=[f'LEFT(O{data_start},6)="YELLOW"'],
            fill=PatternFill(bgColor="FFEB9C", fill_type="solid"),
            font=Font(color="9C6500"),
        ),
    )
    ws.conditional_formatting.add(
        flag_col,
        FormulaRule(
            formula=[f'LEFT(O{data_start},6)="ORANGE"'],
            fill=PatternFill(bgColor="FFCC99", fill_type="solid"),
            font=Font(color="7F3F00"),
        ),
    )


# ─────────────────────────────────────────────
# TAB 4: Discrepancy Summary
# ─────────────────────────────────────────────
def build_discrepancy_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Discrepancy Summary")

    ws.merge_cells("A1:D1")
    ws["A1"] = "capveri.com — Discrepancy Summary (Auto-Generated)"
    ws["A1"].fill = header_fill(DARK_BLUE)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = (
        "This sheet summarizes all flagged rows from the Tenant Matrix tab. "
        "Review and resolve each flag before finalizing reconciliations."
    )
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 30

    apply_header_row(
        ws,
        4,
        ["Tenant Name", "Suite", "Discrepancy Flag", "Notes"],
        widths=[28, 12, 40, 36],
    )

    # Pull flagged rows from Tenant Matrix using IFERROR + IF formulas
    # Rows 5-19 in Tenant Matrix (data_start=5 to 5+15=19)
    data_start = 5  # Tenant Matrix data start row
    output_row = 5
    for i in range(15):
        src_r = data_start + i
        r = output_row + i
        # Only show row if flag is not empty
        ws.cell(
            row=r,
            column=1,
            value=f"=IF('Tenant Matrix'!O{src_r}<>\"\",'Tenant Matrix'!A{src_r},\"\")",
        )
        ws.cell(
            row=r,
            column=2,
            value=f"=IF('Tenant Matrix'!O{src_r}<>\"\",'Tenant Matrix'!B{src_r},\"\")",
        )
        ws.cell(row=r, column=3, value=f"='Tenant Matrix'!O{src_r}")
        ws.cell(row=r, column=4, value=f"='Tenant Matrix'!P{src_r}")

        # Style flag cell
        ws.cell(row=r, column=3).fill = header_fill(CALCULATED_GRAY)

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = True


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────
def main() -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    build_instructions(wb)
    build_portfolio_overview(wb)
    build_tenant_matrix(wb)
    build_discrepancy_summary(wb)

    wb.active = wb["Tenant Matrix"]

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
