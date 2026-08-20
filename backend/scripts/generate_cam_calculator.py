"""
Generate the CAM Gross-Up Scenario Calculator Excel workbook.

Produces: docs/assets/cam-gross-up-calculator.xlsx

Run from repo root:
    pip install openpyxl
    python backend/scripts/generate_cam_calculator.py
"""

import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = (
    pathlib.Path(__file__).parent.parent.parent
    / "docs"
    / "assets"
    / "cam-gross-up-calculator.xlsx"
)

# Colors
DARK_BLUE = "1F3864"
INPUT_YELLOW = "FFF2CC"
CALCULATED_GRAY = "F2F2F2"
WHITE = "FFFFFF"
LIGHT_BLUE_HEADER = "D9E1F2"


def header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def white_font() -> Font:
    return Font(color="FFFFFF", bold=True, size=11)


def bold_font() -> Font:
    return Font(bold=True)


def set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def apply_header_row(ws, row: int, values: list, widths: list | None = None) -> None:
    """Write a dark-blue header row with white bold text."""
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = header_fill(DARK_BLUE)
        cell.font = white_font()
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    if widths:
        for col_idx, w in enumerate(widths, start=1):
            set_col_width(ws, col_idx, w)


def yellow_input(ws, row: int, col: int, value=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = header_fill(INPUT_YELLOW)


def gray_formula(ws, row: int, col: int, formula: str, number_format: str = "") -> None:
    cell = ws.cell(row=row, column=col, value=formula)
    cell.fill = header_fill(CALCULATED_GRAY)
    if number_format:
        cell.number_format = number_format


# ─────────────────────────────────────────────
# TAB 1: Instructions
# ─────────────────────────────────────────────
def build_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 90

    rows = [
        ("CapVeri — Free CAM Gross-Up Scenario Calculator", True, 14),
        ("", False, 11),
        ("WHAT IS GROSS-UP?", True, 12),
        (
            "When a building is not fully occupied, variable CAM expenses (like janitorial, "
            "utilities, and landscaping) are artificially low. Gross-up provisions allow landlords "
            "to 'gross up' those variable expenses to what they would have been at a specified "
            "occupancy level (usually 90–95%). This protects landlords from under-recovering "
            "expenses when occupancy is below threshold.",
            False,
            11,
        ),
        ("", False, 11),
        ("VOCABULARY", True, 12),
        (
            "• GLA (Gross Leasable Area): The total rentable square footage of the building.",
            False,
            11,
        ),
        (
            "• Occupied SF: Total square footage currently under executed leases.",
            False,
            11,
        ),
        ("• Occupancy Rate: Occupied SF ÷ GLA", False, 11),
        (
            "• Gross-Up Threshold: The minimum occupancy level at which expenses are grossed-up "
            "(typically 90% or 95%).",
            False,
            11,
        ),
        (
            "• Fixed Expenses: Do not vary with occupancy (e.g., insurance, property taxes).",
            False,
            11,
        ),
        (
            "• Variable Expenses: Scale with occupancy (e.g., janitorial, utilities, landscaping).",
            False,
            11,
        ),
        (
            "• Pro-Rata Share: Each tenant's proportionate share = tenant SF ÷ total occupied SF.",
            False,
            11,
        ),
        ("", False, 11),
        ("HOW TO USE THIS CALCULATOR", True, 12),
        ("1. Go to the 'Calculator' tab.", False, 11),
        (
            "2. Fill in the YELLOW cells: Building GLA, Occupied SF, Gross-Up Threshold.",
            False,
            11,
        ),
        (
            "3. In the expense table, enter each CAM expense line, its total amount, and whether "
            "it is Fixed or Variable using the dropdown.",
            False,
            11,
        ),
        (
            "4. The gray cells auto-calculate the grossed-up pool and per-tenant obligations.",
            False,
            11,
        ),
        (
            "5. Go to 'Tenant Allocation' to see each tenant's share of the grossed-up pool.",
            False,
            11,
        ),
        (
            "6. Go to 'Scenario Comparison' to compare gross-up at 85%, 90%, 95%, and 100% "
            "occupancy thresholds.",
            False,
            11,
        ),
        ("", False, 11),
        ("IMPORTANT NOTES", True, 12),
        (
            "• This calculator uses native Excel formulas only — no macros, no VBA.",
            False,
            11,
        ),
        (
            "• Compatible with Excel 2016+ and Google Sheets.",
            False,
            11,
        ),
        (
            "• If occupancy = 0%, gross-up displays 'N/A' to prevent division-by-zero errors.",
            False,
            11,
        ),
        (
            "• If occupancy >= threshold, no gross-up is applied (gross-up = actual).",
            False,
            11,
        ),
        ("", False, 11),
        (
            "For questions or to automate this calculation for your entire portfolio, visit capveri.com",
            False,
            11,
        ),
    ]

    for r_idx, (text, bold, size) in enumerate(rows, start=1):
        cell = ws.cell(row=r_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_idx].height = 30 if bold else 20

    ws.sheet_view.showGridLines = False


# ─────────────────────────────────────────────
# TAB 2: Calculator (main)
# ─────────────────────────────────────────────
def build_calculator(wb: Workbook) -> None:
    ws = wb.create_sheet("Calculator")

    # Branding header
    ws.merge_cells("A1:H1")
    brand_cell = ws["A1"]
    brand_cell.value = "capveri.com — Free CAM Gross-Up Calculator"
    brand_cell.fill = header_fill(DARK_BLUE)
    brand_cell.font = Font(color="FFFFFF", bold=True, size=13)
    brand_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Building Inputs ──
    ws["A3"] = "BUILDING INPUTS"
    ws["A3"].font = bold_font()
    ws["A3"].fill = header_fill(LIGHT_BLUE_HEADER)
    ws.merge_cells("A3:B3")

    ws["A4"] = "Building GLA (SF)"
    yellow_input(ws, 4, 2, 50000)
    ws["B4"].number_format = "#,##0"

    ws["A5"] = "Occupied SF"
    yellow_input(ws, 5, 2, 43000)
    ws["B5"].number_format = "#,##0"

    ws["A6"] = "Gross-Up Threshold"
    yellow_input(ws, 6, 2, 0.95)
    ws["B6"].number_format = "0%"

    ws["A7"] = "Occupancy Rate (auto)"
    gray_formula(ws, 7, 2, '=IF(B4=0,"N/A",B5/B4)', "0.0%")

    # ── Expense Table ──
    ws["A9"] = "EXPENSE TABLE"
    ws["A9"].font = bold_font()
    ws["A9"].fill = header_fill(LIGHT_BLUE_HEADER)
    ws.merge_cells("A9:G9")

    header_row = 10
    apply_header_row(
        ws,
        header_row,
        [
            "Expense Name",
            "Total Actual ($)",
            "Fixed or Variable",
            "Fixed Portion ($)",
            "Variable Portion ($)",
            "Grossed-Up Variable ($)",
            "Grossed-Up Total ($)",
        ],
        widths=[24, 18, 20, 18, 18, 22, 20],
    )

    # DataValidation for Fixed/Variable dropdown
    dv = DataValidation(type="list", formula1='"Fixed,Variable"', allow_blank=False)
    ws.add_data_validation(dv)

    # Sample data: 5 expense rows
    sample_expenses = [
        ("Property Insurance", 45000, "Fixed"),
        ("Property Taxes", 120000, "Fixed"),
        ("Landscaping", 28000, "Variable"),
        ("Janitorial Services", 52000, "Variable"),
        ("Utilities - Common Area", 35000, "Variable"),
    ]

    for i, (name, amount, exp_type) in enumerate(sample_expenses):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=name)
        yellow_input(ws, r, 2, amount)
        ws.cell(row=r, column=2).number_format = "$#,##0.00"
        yellow_input(ws, r, 3, exp_type)
        dv.add(ws.cell(row=r, column=3))

        # Fixed Portion = IF(C=Fixed, B, 0)
        gray_formula(ws, r, 4, f'=IF(C{r}="Fixed",B{r},0)', "$#,##0.00")
        # Variable Portion = IF(C=Variable, B, 0)
        gray_formula(ws, r, 5, f'=IF(C{r}="Variable",B{r},0)', "$#,##0.00")
        # Grossed-Up Variable = IF(C=Variable, IF(B7=0,"N/A", IF(B7>=B6, E, E*(B6/B7))), 0)
        gray_formula(
            ws,
            r,
            6,
            f'=IF(C{r}="Variable",IF($B$7=0,"N/A",IF($B$7>=$B$6,E{r},E{r}*($B$6/$B$7))),0)',
            "$#,##0.00",
        )
        # Grossed-Up Total = Fixed + Grossed-Up Variable
        gray_formula(ws, r, 7, f'=IF(F{r}="N/A","N/A",D{r}+F{r})', "$#,##0.00")

    # Blank rows 6–10 for additional expenses
    for i in range(5, 10):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value="")
        yellow_input(ws, r, 2, None)
        ws.cell(row=r, column=2).number_format = "$#,##0.00"
        yellow_input(ws, r, 3, None)
        dv.add(ws.cell(row=r, column=3))
        gray_formula(ws, r, 4, f'=IF(C{r}="Fixed",B{r},0)', "$#,##0.00")
        gray_formula(ws, r, 5, f'=IF(C{r}="Variable",B{r},0)', "$#,##0.00")
        gray_formula(
            ws,
            r,
            6,
            f'=IF(C{r}="Variable",IF($B$7=0,"N/A",IF($B$7>=$B$6,E{r},E{r}*($B$6/$B$7))),0)',
            "$#,##0.00",
        )
        gray_formula(ws, r, 7, f'=IF(F{r}="N/A","N/A",D{r}+F{r})', "$#,##0.00")

    # Summary row
    last_expense_row = header_row + 10
    summary_row = last_expense_row + 2
    ws.cell(row=summary_row, column=1, value="TOTALS").font = bold_font()
    gray_formula(
        ws, summary_row, 2, f"=SUM(B{header_row+1}:B{last_expense_row})", "$#,##0.00"
    )
    gray_formula(
        ws, summary_row, 4, f"=SUM(D{header_row+1}:D{last_expense_row})", "$#,##0.00"
    )
    gray_formula(
        ws, summary_row, 5, f"=SUM(E{header_row+1}:E{last_expense_row})", "$#,##0.00"
    )
    gray_formula(
        ws,
        summary_row,
        7,
        f'=SUMPRODUCT((G{header_row+1}:G{last_expense_row}<>"N/A")*IFERROR(G{header_row+1}:G{last_expense_row},0))',
        "$#,##0.00",
    )

    ws.cell(row=summary_row, column=1).fill = header_fill(LIGHT_BLUE_HEADER)
    ws.cell(row=summary_row, column=2).fill = header_fill(LIGHT_BLUE_HEADER)
    ws.cell(row=summary_row, column=4).fill = header_fill(LIGHT_BLUE_HEADER)
    ws.cell(row=summary_row, column=5).fill = header_fill(LIGHT_BLUE_HEADER)
    ws.cell(row=summary_row, column=7).fill = header_fill(LIGHT_BLUE_HEADER)

    # Delta row
    delta_row = summary_row + 1
    ws.cell(row=delta_row, column=1, value="Gross-Up Delta ($)").font = bold_font()
    gray_formula(
        ws,
        delta_row,
        7,
        f"=G{summary_row}-B{summary_row}",
        "$#,##0.00",
    )

    # Freeze top rows
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    # Instructions note
    note_row = delta_row + 2
    ws.cell(
        row=note_row,
        column=1,
        value="Yellow cells = inputs  |  Gray cells = auto-calculated  |  See 'Tenant Allocation' for per-tenant share",
    )
    ws.cell(row=note_row, column=1).font = Font(italic=True, color="666666", size=9)
    ws.merge_cells(f"A{note_row}:G{note_row}")


# ─────────────────────────────────────────────
# TAB 3: Tenant Allocation
# ─────────────────────────────────────────────
def build_tenant_allocation(wb: Workbook) -> None:
    ws = wb.create_sheet("Tenant Allocation")

    ws.merge_cells("A1:E1")
    ws["A1"] = "capveri.com — Tenant Pro-Rata Allocation"
    ws["A1"].fill = header_fill(DARK_BLUE)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Grossed-Up CAM Pool (from Calculator tab):"
    ws["A2"].font = bold_font()
    ws.merge_cells("A2:C2")
    gray_formula(ws, 2, 4, "=Calculator!G23", "$#,##0.00")  # points to summary row

    apply_header_row(
        ws,
        4,
        ["Tenant Name", "Leased SF", "Pro-Rata % (auto)", "CAM Obligation ($)"],
        widths=[28, 14, 18, 20],
    )

    sample_tenants = [
        ("Acme Corp", 8500),
        ("BlueSky LLC", 12000),
        ("Metro Dental", 3200),
    ]

    for i, (name, sf) in enumerate(sample_tenants):
        r = 5 + i
        ws.cell(row=r, column=1, value=name)
        yellow_input(ws, r, 2, sf)
        ws.cell(row=r, column=2).number_format = "#,##0"
        # Pro-rata = leased SF / sum of all leased SF
        gray_formula(
            ws, r, 3, f"=IF(SUM($B$5:$B$14)=0,0,B{r}/SUM($B$5:$B$14))", "0.00%"
        )
        # CAM obligation = pro-rata * grossed-up pool
        gray_formula(ws, r, 4, f"=IFERROR(C{r}*$D$2,0)", "$#,##0.00")

    # Blank rows
    for i in range(3, 10):
        r = 5 + i
        yellow_input(ws, r, 2, None)
        ws.cell(row=r, column=2).number_format = "#,##0"
        gray_formula(
            ws, r, 3, f"=IF(SUM($B$5:$B$14)=0,0,B{r}/SUM($B$5:$B$14))", "0.00%"
        )
        gray_formula(ws, r, 4, f"=IFERROR(C{r}*$D$2,0)", "$#,##0.00")

    # Totals
    ws.cell(row=15, column=1, value="TOTALS").font = bold_font()
    gray_formula(ws, 15, 2, "=SUM(B5:B14)", "#,##0")
    gray_formula(ws, 15, 3, "=SUM(C5:C14)", "0.00%")
    gray_formula(ws, 15, 4, "=SUM(D5:D14)", "$#,##0.00")

    ws.freeze_panes = "A5"


# ─────────────────────────────────────────────
# TAB 4: Scenario Comparison
# ─────────────────────────────────────────────
def build_scenario_comparison(wb: Workbook) -> None:
    ws = wb.create_sheet("Scenario Comparison")

    ws.merge_cells("A1:I1")
    ws["A1"] = "capveri.com — Gross-Up Scenario Comparison (85% / 90% / 95% / 100%)"
    ws["A1"].fill = header_fill(DARK_BLUE)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column layout: expense name, actual, then 4 threshold scenarios
    apply_header_row(
        ws,
        3,
        [
            "Expense Name",
            "Actual ($)",
            "Fixed/Var",
            "Grossed-Up @ 85%",
            "Grossed-Up @ 90%",
            "Grossed-Up @ 95%",
            "Grossed-Up @ 100%",
            "Delta @ 85%",
            "Delta @ 95%",
        ],
        widths=[24, 14, 12, 18, 18, 18, 18, 14, 14],
    )

    ws["A2"] = "Building Occupancy (from Calculator tab):"
    ws["A2"].font = Font(italic=True, size=9)
    gray_formula(ws, 2, 4, "=Calculator!B7", "0.0%")
    ws.merge_cells("A2:C2")

    # Expense rows reference Calculator tab
    thresholds = [0.85, 0.90, 0.95, 1.00]
    for i in range(10):
        src_r = 11 + i  # Calculator tab rows 11-20
        r = 4 + i
        ws.cell(row=r, column=1, value=f"=Calculator!A{src_r}").number_format = "@"
        # Actual
        cell = ws.cell(row=r, column=2, value=f"=Calculator!B{src_r}")
        cell.fill = header_fill(CALCULATED_GRAY)
        cell.number_format = "$#,##0.00"
        # Type
        cell2 = ws.cell(row=r, column=3, value=f"=Calculator!C{src_r}")
        cell2.fill = header_fill(CALCULATED_GRAY)

        for t_idx, threshold in enumerate(thresholds):
            col = 4 + t_idx
            # Grossed-up formula: IF(Variable AND occupancy>0, IF(occ>=threshold, actual, actual*(threshold/occ)), actual)
            occ = "Calculator!$B$7"
            formula = (
                f'=IF(Calculator!C{src_r}="Variable",'
                f'IF({occ}=0,"N/A",'
                f"IF({occ}>={threshold},Calculator!B{src_r},Calculator!B{src_r}*({threshold}/{occ}))),"
                f"Calculator!B{src_r})"
            )
            gray_formula(ws, r, col, formula, "$#,##0.00")

        # Delta @ 85% = col4 - col2
        gray_formula(ws, r, 8, f'=IFERROR(D{r}-B{r},"N/A")', "$#,##0.00")
        # Delta @ 95% = col6 - col2
        gray_formula(ws, r, 9, f'=IFERROR(F{r}-B{r},"N/A")', "$#,##0.00")

    # Summary row
    s_row = 14
    ws.cell(row=s_row, column=1, value="POOL TOTALS").font = bold_font()
    for col in [2, 4, 5, 6, 7, 8, 9]:
        gray_formula(
            ws,
            s_row,
            col,
            f'=SUMPRODUCT(({chr(64+col)}4:{chr(64+col)}13<>"N/A")*IFERROR({chr(64+col)}4:{chr(64+col)}13,0))',
            "$#,##0.00",
        )
    for col in [2, 4, 5, 6, 7, 8, 9]:
        ws.cell(row=s_row, column=col).fill = header_fill(LIGHT_BLUE_HEADER)

    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────
def main() -> None:
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    build_instructions(wb)
    build_calculator(wb)
    build_tenant_allocation(wb)
    build_scenario_comparison(wb)

    # Set active sheet to Calculator
    wb.active = wb["Calculator"]

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
